import pandas as pd   #pre final output  >>>>>>> move to new output
import numpy as np
import lxml
import os
from os import getcwd
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
from io import StringIO
import csv as csv
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
from io import BytesIO
import time
import sys
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
import PySimpleGUI as sg
import subprocess
import threading
import urllib
import urllib.parse
import sqlalchemy 
import sqlalchemy_access as sq_a
import sqlalchemy_pyodbc_mssql as sqlalchemy
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
import shutil
import pyodbc
import pypyodbc
import odbc
import psycopg2
import mysql.connector as sql #1
import MySQLdb as sql
from sqlite3 import dbapi2 as sqlite
import sqlite3
from datetime import datetime #2
import re


os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

# Load the Excel file
excel_file_path = 'BOM_List_OP.xlsx'  # Replace with your file path

# Define your column lists
column_list = ["PartNumber", "Group", 'Priority', 'RefList', 'Qty', 'Shape', 'Long Des']

df = pd.read_excel(excel_file_path, sheet_name='BOM' ,usecols=column_list, index_col= False)

print(df)

df['Description'] = df['Long Des']

data = df['Description']

'''data = {
"Description": [
        "SMT,H/IND 1.2uH 8.6A 4.3X4.3X2.0 20% BOURNS",
        "FERRITE BEAD SMD 1.8K OHM R 0402 1LN",
        "Capacitor,SMT,1206,10uF,16V,10%,Tantalum",
        "Cap SMD Aluminum Lytic 4700uF 16V 20% (18 X 21.5mm) 1750mA 2000h 85°C Automotive T/R",
        "Cap Super THT 3.6-3.69V Impedance:≤ 500 Mω 21.5x9mm",
        "CAP,MLCC,SMD,220PF,50V,±5%,0402,C0G,(NP0)",
        "RES SMD 3.3M OHM 1% 1/10W 0603",
        "RES,THICK,FILM,SMD,47K,OHM,1%,0.0625W,0402,100ppm",
        "EMI,filter,bead,SMT,0402,1k,ohms,Tape,GHz,Band,Gen,Use",
        "FERRITE BEAD SMD 1K OHM 0603 1LN",
        "Ferrite,Chip,SMT,0805,Filter,330R",
        "IND CMC 2A 2LN 500 OHM SMD",
        "ALLUM CAP SMD 100uF 6.3volts 20 % AEC-Q200",
        "ALUM CAP SMD 100UF 6.3V 20% (6.60mm x 6.60mm)",
        "Aluminum Electrolytic Capacitors - SMD 1500uF 35V",
        "CAP ALUM 1500UF 20% 35V SMD",
        "CAP TAN SMD 4.7UF 10V ±20% 0805",
        "CAP ALU SMD 330uF ± 20% 35V",
        "A/CAP 2.2uF 0402 6.3V K(+-10%) X5R",
        "A/CAP 4.7uF 0402 ≧6.3V ≦M X5R",
        "SMT,A/CAP 10uF 0603 ≧25V ≦M X5R",
        "SMT,A/CAP 22uF 0805 ≧25V ≦M X5R",
        "H/CAP 47uF 0603 ≧6.3V ≦M X5R",
        "H/RES 15Ω 0402 ≦J 1/16W",
        "H/RES 47Ω 0402 ≦J 1/16W",
        "H/RES 240Ω 0201 F(+-1%) 1/20W",
        "H/RES 240Ω 0402 ≦F 1/16W",
        "SMT,H/RES 249Ω 0603 melf F(+-1%) 1/10W",
        "smt,ZEN,MELF,DIODE"
    ]
}'''
df = pd.DataFrame(data)
# Define desired shapes
desired_shapes = ("0201", "0402", "0603", "0805", "1206")

# Custom function to extract and separate the shape
def extract_shape(description):
    shape_match = re.search(r'\b\d{4}\b', description)
    if shape_match:
        shape = shape_match.group()
        if shape in desired_shapes:
            return shape
    return None

# Apply the custom function to create the Shape column
df['Shape'] = df['Description'].apply(extract_shape)

# Display the result
print(df[['Description', 'Shape']])

# Define desired component types
desired_COMP = ("CAP", "RES", "IND")

# Custom function to extract and separate the component type
def extract_COMPs(description):
    for comp_type in desired_COMP:
            if comp_type.lower() in description.lower():
                return comp_type
    return None

# Apply the custom function to create the LCRTYPE column
df['LCRTYPE'] = df['Description'].apply(extract_COMPs)

# Define desired Special component types
desired_SPERESCOMP = ["MELF"]

# Custom function to check if the description contains special components
def contains_special_res_component(description):
    for serescomp_type in desired_SPERESCOMP:
        if serescomp_type.lower() in description.lower():
            return True
    return False

# Apply the custom function to create the SPERESCOMP column
df['SPERESCOMP'] = df['Description'].apply(contains_special_res_component)

# Transform 'SPERESCOMP' column to 'MELF' when it's True
df['SPERESCOMP'] = np.where(df['SPERESCOMP'], 'MELF', '')

# Define desired Special component types
desired_SPETHTCOMP = ["THT"]

# Custom function to check if the description contains special components
def contains_special_tht_component(description):
    for sethtcomp_type in desired_SPETHTCOMP:
        if sethtcomp_type.lower() in description.lower():
            return True
    return False

# Apply the custom function to create the SPERESCOMP column
df['SPETHTCOMP'] = df['Description'].apply(contains_special_tht_component)

# Transform 'SPERESCOMP' column to 'MELF' when it's True
df['SPETHTCOMP'] = np.where(df['SPETHTCOMP'], 'THT', '')

# Define desired Special component types
desired_SPESODCOMP = ("ZENER", "DIODE", "SOD")

# Custom function to extract and separate the component type
def extract_SPESODCOMP(description):
    for sesodcomp_type in desired_SPESODCOMP:
        if sesodcomp_type.lower() in description.lower():
            return sesodcomp_type
    return None

# Apply the custom function to create the LCRTYPE column
df['SPESODCOMP'] = df['Description'].apply(extract_SPESODCOMP)

# Define desired Special component types
desired_SPECAPCOMP = ("TAN", "Tantalum", "Aluminium","ALLUM", "ALUM", "Electrolytic" ,"ALU")

# Custom function to extract and separate the component type
def extract_SPECAPCOMP(description):
    for setancomp_type in desired_SPECAPCOMP:
        if setancomp_type.lower() in description.lower():
            return setancomp_type
    return None

# Apply the custom function to create the LCRTYPE column
df['SPECAPCOMP'] = df['Description'].apply(extract_SPECAPCOMP)


# Assuming df is your DataFrame
df['PACKAGE'] = df['Shape'].replace({'0201': '0802P', '0402': '0802P', '0603': '0804P', '0805': '0804E', '1206': '0804E'})


#LCR Wrork#


def extract_component_info(component_line):
    # Regular expressions for CAPACITOR
    capacitor_type_match = re.search(r'\b(CAP)\b', component_line, re.IGNORECASE)
    capacitor_value_match = re.search(r'(\d+(\.\d+)?)([pnuμmkKMΩ]?)\s*(\d+%)?', component_line)

    # Regular expressions for RESISTOR
    resistor_type_match = re.search(r'\b(Res|Resistor|RESISTOR)\b', component_line, re.IGNORECASE)
    resistor_value_match = re.search(r'(\d+(\.\d+)?)([pnuμmkKMΩ]?)\s*(\d+%)?', component_line)

    # Assign default values
    LCR_Type = None
    LCR_Value = None
    LCR_Unit = None
    LCR_Tolerance = None

    if capacitor_type_match:
        LCR_Type = capacitor_type_match.group(1)

    if capacitor_value_match:
        groups = capacitor_value_match.groups()
        LCR_Value = groups[0] if groups[0] is not None else None
        LCR_Unit = groups[2] if groups[2] is not None else None

    if resistor_type_match:
        LCR_Type = resistor_type_match.group(1)

    if resistor_value_match:
        groups = resistor_value_match.groups()
        LCR_Value = groups[0] if groups[0] is not None else None
        LCR_Unit = groups[2] if groups[2] is not None else None

    return LCR_Type, LCR_Value, LCR_Unit, LCR_Tolerance

component_lines =df['Description']

'''component_lines = [
    "SMT,H/RES 249Ω 0603 F(+-1%) 1/10W",
    "RES SMD 560 OHM 1% 0.1W 0603",
    "Thick,Film,Resistors,-,SMD,10,OHMS,1%,0.1W,0603,100PPM",
    "Thick,Film,Resistors,-,SMD,15,kOhms,100-200,mW,0603,1%",
    "CAP,MLCC,SMD,220PF,50V,±5%,0402,C0G,(NP0)",
    "SMT,A/CAP 10uF 0603 ≧25V ≦M X5R",
    "SMT,A/CAP 22uF 0805 ≧25V ≦M X5R",
    "H/CAP 47uF 0603 ≧6.3V ≦M X5R",
    "H/RES 15Ω 0402 ≦J 1/16W",
    "H/RES 47Ω 0402 ≦J 1/16W",
]'''

data = {'LCR Type': [], 'LCR Value': [], 'LCR Unit': [], 'LCR Tolerance': []}

for component_line in component_lines:
    LCR_Type, LCR_Value, LCR_Unit, LCR_Tolerance = extract_component_info(component_line)
    data['LCR Type'].append(LCR_Type)
    data['LCR Value'].append(LCR_Value)
    data['LCR Unit'].append(LCR_Unit)
    data['LCR Tolerance'].append(LCR_Tolerance)

df = pd.DataFrame(data)
print(df)


# Display the result
print(df[['Description', 'Shape', 'LCRTYPE', 'SPERESCOMP', 'SPETHTCOMP', 'SPESODCOMP', 'SPECAPCOMP', 'PACKAGE']])

# Display the result
#print(df[['Description', 'Shape', 'LCRTYPE', 'SPERESCOMP', 'SPETHTCOMP', 'SPESODCOMP', 'SPECAPCOMP', 'PACKAGE', 'LCR Type', 'LCR Value', 'LCR Unit', 'LCR Tolerance']])

# Display the result PRINT THE RESULT IN SINGLE TAKE
#print(df[['Description', 'Shape', 'LCRTYPE', 'SPERESCOMP', 'SPETHTCOMP', 'SPESODCOMP', 'SPECAPCOMP', 'PACKAGE', 'LCR Type', 'LCR Value', 'LCR Unit', 'LCR Tolerance']])

print(df.columns)

