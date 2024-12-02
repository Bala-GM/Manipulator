import pandas as pd
import numpy as np
import os
from os import getcwd
import csv as csv
from io import StringIO
import re
import datetime
from io import BytesIO
import time
import sys
from datetime import datetime
import shutil
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import subprocess
import threading
import urllib
import urllib.parse
from sqlite3 import dbapi2 as sqlite
import sqlite3
import lxml
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
import PySimpleGUI as sg
import sqlalchemy
from sqlalchemy import create_engine
import sqlalchemy_access as sq_a
import sqlalchemy_pyodbc_mssql as sqlalchemy
from flask_sqlalchemy import SQLAlchemy
import pyodbc
import pypyodbc
import odbc
import psycopg2
import mysql.connector as sql
import MySQLdb as sql #pip install mysqlclient
from plyer import notification
import xlrd
import xml.etree.ElementTree as ET
from openpyxl.utils.exceptions import InvalidFileException
import psutil

# Program 2: Database inspection interface_GUI/J0124-89P13
def program_2():

    #bil5 = pyfiglet.figlet_format("BOM Manipulation", width = 100)
    print('\n')
    print("\033[92;4m******BOM Manipulation******\033[0m")
    print('\n')
    ##########################################################################################################################################

    ##########################################################################################################################################

    #BOM MANIPULATE

    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Upload"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)
            
    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)
    
    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\PartNuber"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)

    try:
        # BOM MANIPULATION
        os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
        file_path = 'BOM.xlsx'

        if os.path.isfile(file_path):
            ds1 = pd.read_excel(file_path, index_col=False)
        else:
            # Try reading as '.xls' format if '.xlsx' fails
            file_path = 'BOM.xls'
            ds1 = pd.read_excel(file_path, index_col=False)

        dfbom1 = ds1

    except ValueError:
        dfbom1 = pd.read_excel(file_path,index_col=False) 

    except Exception as e:
        # Handle the exception gracefully
        error_message = f"An error occurred: {e}"

        # Show error message in a pop-up box
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    # Define your column lists
    column_list_1 = ['Material', 'AltItemGroup', 'Priority', 'Long. Description', 'Ref.Designator/Circuit Reference', 'Quantity', 'Material Group']
    column_list_2 = ['Internal P/N', 'Group', 'Priority', 'Description', 'Ref.Designator', 'Qty', 'SMT/THT/Mech']

    # Check which column list is present in the DataFrame
    if all(column in ds1.columns for column in column_list_1):
        columns_to_use = column_list_1
    elif all(column in ds1.columns for column in column_list_2):
        columns_to_use = column_list_2
    else:
        # Show error message if none of the column lists is present
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        missing_columns = [column for column_list in [column_list_1, column_list_2] for column in column_list if column not in ds1.columns]
        error_message = f"The following columns are missing: {', '.join(missing_columns)}"
        error_msgbm1 = f"The following columns are missing: in SAP BOM\n'Material'\n'AltItemGroup'\n'Priority'\n'Long. Description'\n'Ref.Designator/Circuit Reference'\n'Quantity'\n'Material Group'"
        error_msgbm2 = f"The following columns are missing: in Internal BOM\n'Internal P/N'\n'Group'\n'Priority'\n'Description'\n'Ref.Designator'\n'Qty'\n'SMT/THT/Mech'"
        messagebox.showerror("Error", error_message)
        messagebox.showerror("Error", error_msgbm1)
        messagebox.showerror("Error", error_msgbm2)
        sys.exit(1)  # Exit the program with an error code

    # Continue with the rest of your code using 'columns_to_use'
    print(f"Using columns: {columns_to_use}")

    # Rest of your code here
    # ...

    ds1.rename(
        columns={'Material':"PartNumber", 'AltItemGroup':"Group", 'Priority':'Priority', 'Long. Description':'Long Des', 'Ref.Designator/Circuit Reference':'RefList', 'Quantity':'Qty','Material Group':'Shape'},
        inplace=True,
    )

    ds1.rename(
        columns={'Internal P/N':"PartNumber", 'Group':"Group", 'Priority':'Priority', 'Description':'Long Des', 'Ref.Designator':'RefList', 'Qty':'Qty','SMT/THT/Mech':'Shape'},
        inplace=True,
    )

    print(ds1)

    ds2 = ds1[ds1['Priority'].isin([0, 1])]

    # Assuming ds2 is your DataFrame and 'PartNumber' and 'RefList' are the columns you want to check
    part_number_column = ds2['PartNumber']
    ref_list_column = ds2['RefList']

    # Flag to check if an empty value is found
    empty_value_found = False

    # Iterate through both columns simultaneously using iterrows
    for index, (part_number_value, ref_list_value) in ds2[['PartNumber', 'RefList']].iterrows():
        # Check if the 'RefList' value is empty (NaN or None)
        if pd.isna(ref_list_value):
            print(f"Error: Empty value found in 'RefList' for 'PartNumber' {part_number_value}")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = f"Empty value found in 'RefList' for 'PartNumber' {part_number_value}. Program will stop."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code
            #raise ValueError("Empty value found in 'RefList'")
            #empty_value_found = True
            #break  # Stop the iteration when the first empty value is found

    # If no empty values are found, print the 'PartNumber' column
    if not empty_value_found:
        print(part_number_column)
        # Continue with the rest of your program

    #file_name ="output.xlsx"
    #ds1.to_excel(file_name)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM/BOM_List_OP.xlsx") as writer:
        ds1.to_excel(writer, sheet_name="Orginal_BOM", index=False)
        ds2.to_excel(writer, sheet_name="BOM", index=False)

        pass
        print('The file does not exist.')

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
    Chd = os.getcwd()
    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        ds1 = pd.read_excel(file_path, sheet_name="BOM", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False) 
        dsn1 = pd.read_excel(file_path, sheet_name="Orginal_BOM", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False)
        
        ds1 = ds1[['PartNumber', 'Group','Priority','Long Des','Qty','Shape','RefList']]
        ds1['RefList'] = ds1['RefList'].str.replace("_x000D_","")
        ds1['RefList'] = ds1['RefList'].str.replace(" ","")
        ds1['RefList'] = ds1['RefList'].str.replace("\n","")
        print(ds1)

    #ds1 = pd.read_excel('Filename_OP.xlsx','BOM', index_col=False)

        ds2 = ds1.explode('RefList')

        ds2['RefList'] = ds2['RefList'].str.replace(" "," ")

    #ds2.drop(ds2.iloc[:, 1:6], inplace=True, axis=1)

        print(ds2)

        ds2.rename(columns = {'PartNumber':'B_Part_No'}, inplace = True)

        ds2.rename(columns = {'RefList':'B_Ref_List'}, inplace = True)

        ds2['B_Ref_List'] = ds2['B_Ref_List'] .str.strip('[]').str.split(',')

        print(ds2)

        ds2.to_dict()

        ds2.explode ('B_Ref_List',ignore_index=True)

        ds3 = ds2.explode('B_Ref_List',ignore_index=True) # split the Ref below example code

        '''import pandas as pd

        # Sample DataFrame
        data = {'ID': [1, 2], 'B_Ref_List': [['R1', 'R2'], ['R3', 'R4', 'R5']]}

        ds2 = pd.DataFrame(data)

        # Explode 'B_Ref_List'
        ds3 = ds2.explode('B_Ref_List', ignore_index=True)

        # Display the result
        print(ds3)
        Output:
            ID B_Ref_List
        0   1         R1
        1   1         R2
        2   2         R3
        3   2         R4
        4   2         R5'''

        ds2 = ds2[['Group','Priority','B_Part_No']]
        dc1 = ds2[['B_Part_No']]
        dc1.rename(columns = {'B_Part_No':'PBARNO'}, inplace = True)
        dc1['PBARPTN'] = dc1['PBARNO']
        dc1['PBARBAR'] = dc1['PBARNO']
        dc1.insert(3,'PBARQTY', 10000)
        dc1.insert(4,'PBARFTYP', 3)

    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>#
        dfs2 = ds2[['Group','Priority','B_Part_No']]
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("15","A")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("14","B")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("13","C")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("12","D")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("11","E")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("10","F")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("9","PTN_9")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("8","PTN_8")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("7","PTN_7")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("6","PTN_6")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("5","PTN_5")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("4","PTN_4")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("3","PTN_3")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("2","PTN_2")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("1","PTN_1")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("0","PTN_0")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("A","PTN_15")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("B","PTN_14")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("C","PTN_13")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("D","PTN_12")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("E","PTN_11")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("F","PTN_10")
        dfs2.dropna(subset=['Group'], inplace=True)
        #df2 = dfs2.pivot(index='Group',columns='Priority',values='B_Part_No')

            # Assuming 'dfs2' is the DataFrame with 'Group', 'Priority', and 'B_Part_No' columns
        # Check for duplicate entries in 'Group' and 'Priority'
        duplicate_entries = dfs2[dfs2.duplicated(subset=['Group', 'Priority'], keep=False)]

        if not duplicate_entries.empty:
            # Show an error message if duplicates are found
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = f"Duplicate entries found in 'Group' and 'Priority':\n{duplicate_entries}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        # If no duplicates, proceed with pivoting
        try:
            df2 = dfs2.pivot(index='Group', columns='Priority', values='B_Part_No')
        except ValueError as e:
            # Handle the exception gracefully
            error_message = f"An error occurred during pivoting: {e}"
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        ds3.head()

        T10_col = ds3.pop('B_Ref_List') # col-1

        ds3.insert(0, 'B_Ref_List', T10_col)

        ds3 = ds3[['B_Ref_List','B_Part_No','Long Des']]

        ds1.dropna(subset=['RefList'], inplace=True)
        ds3.dropna(subset=['B_Ref_List'], inplace=True)

    #ONLY AVL PARTMASTER AND GOUPING
        dsn1 = dsn1[['PartNumber', 'Group','Priority','Long Des','Qty','Shape','RefList']]
        dsn1['RefList'] = dsn1['RefList'].str.replace("_x000D_","")
        dsn1['RefList'] = dsn1['RefList'].str.replace(" ","")
        dsn1['RefList'] = dsn1['RefList'].str.replace("\n","")
        print(dsn1)

    #ds1 = pd.read_excel('Filename_OP.xlsx','BOM', index_col=False)

        dsn2 = dsn1.explode('RefList')

        dsn2['RefList'] = dsn2['RefList'].str.replace(" "," ")

    #ds2.drop(ds2.iloc[:, 1:6], inplace=True, axis=1)

        print(dsn2)

        dsn2.rename(columns = {'PartNumber':'B_Part_No'}, inplace = True)

        dsn2.rename(columns = {'RefList':'B_Ref_List'}, inplace = True)

        dsn2['B_Ref_List'] = dsn2['B_Ref_List'] .str.strip('[]').str.split(',')

        dsn2.to_dict()

        dsn2.explode ('B_Ref_List',ignore_index=True)

        dsn3 = dsn2.explode('B_Ref_List',ignore_index=True)

        dsn2 = dsn2[['Group','Priority','B_Part_No']]

        # Condition: Check if Priority is only 0
        if (dsn2['Priority'] == 0).all():
        # Check if 1 and 2 are not present
            if not ((dsn2['Priority'] == 1) | (dsn2['Priority'] == 2)).any():
        # Add 1, 2, 3 in Priority column
                
                dsn2['Priority'] = dsn2['Priority']
                
        # Append corresponding Dummy_Part rows
                dummy_data = {'Group': ['B89P13', 'B89P13', 'B89P13'],
                            'Priority': [1, 2, 3],
                            'B_Part_No': ['Dummy_Part1', 'Dummy_Part2', 'Dummy_Part3']}
                
                dummy_df = pd.DataFrame(dummy_data)
                dsn2 = pd.concat([dsn2, dummy_df], ignore_index=True)

        # Continue with the rest of your code
        print(dsn2[['Group', 'Priority', 'B_Part_No']])

        dcn1 = dsn2[['B_Part_No']]
        duplicate_rows = dcn1[dcn1.duplicated(subset=['B_Part_No'], keep=False)]
        
        if not duplicate_rows.empty:
            # Show an error message if duplicates are found
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = f"Duplicate entries found in 'B_Part_No':\nCheck the BOM! PartNo Col.\n{duplicate_rows}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code
        
        print(dcn1)

        # Assuming 'B_Part_No' contains values like 'Dummy_Part1', 'Dummy_Part2', 'Dummy_Part3'
        dummy_values = [f'Dummy_Part{i}' for i in range(1, 4)]

        # Remove rows where 'B_Part_No' contains dummy values
        dcn1 = dcn1[~dcn1['B_Part_No'].isin(dummy_values)]

        dcn1.rename(columns = {'B_Part_No':'PBARNO'}, inplace = True)
        dcn1['PBARPTN'] = dcn1['PBARNO']
        dcn1['PBARBAR'] = dcn1['PBARNO']
        dcn1.insert(3,'PBARQTY', 10000)
        dcn1.insert(4,'PBARFTYP', 3)

        #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>#
        dfsn2 = dsn2[['Group','Priority','B_Part_No']]
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("15","A")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("14","B")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("13","C")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("12","D")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("11","E")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("10","F")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("9","PTN_9")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("8","PTN_8")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("7","PTN_7")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("6","PTN_6")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("5","PTN_5")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("4","PTN_4")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("3","PTN_3")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("2","PTN_2")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("1","PTN_1")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("0","PTN_0")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("A","PTN_15")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("B","PTN_14")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("C","PTN_13")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("D","PTN_12")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("E","PTN_11")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("F","PTN_10")
        dfsn2.dropna(subset=['Group'], inplace=True)
        #dfn2 = dfsn2.pivot(index='Group',columns='Priority',values='B_Part_No')

        # Assuming 'dfs2' is the DataFrame with 'Group', 'Priority', and 'B_Part_No' columns
        # Check for duplicate entries in 'Group' and 'Priority'
        duplicate_entries = dfsn2[dfsn2.duplicated(subset=['Group', 'Priority'], keep=False)]

        if not duplicate_entries.empty:
            # Show an error message if duplicates are found
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = f"Duplicate entries found in 'Group' and 'Priority':\n{duplicate_entries}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        # If no duplicates, proceed with pivoting
        try:
            dfn2 = dfsn2.pivot(index='Group', columns='Priority', values='B_Part_No')
        except ValueError as e:
            # Handle the exception gracefully
            error_message = f"An error occurred during pivoting: {e}"
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

                # Desired column order
        desired_order = ['Group Name','PTN_1','PTN_2','PTN_3','PTN_4','PTN_5','PTN_6','PTN_7','PTN_8','PTN_9','PTN_10','PTN_11','PTN_12','PTN_13','PTN_14','PTN_15']
        #desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10','PTN_11','P_11','PTN_12','P_12','PTN_13','P_13','PTN_14','P_14','PTN_15','P_15']

        # Create a list of columns present in both DataFrame and desired_order
        common_columns = [col for col in desired_order if col in dfn2.columns]

        # Reorder the DataFrame based on the desired_order
        df_AL1 = dfn2[common_columns]

        '''    # Assuming df is your DataFrame
            column_to_check = 'PTN_15'

            # Check if the column is present
            if column_to_check not in df_AL1.columns:
                # Show a pop-up message if the column is not present
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showinfo("Notification", f"The column '{column_to_check}' is missing.")

            # Continue with the rest of your code
            print("Continuing with the rest of the code...")
            # Your next line of code here'''

        # Assuming df is your DataFrame
        column_to_check = 'PTN_11'

        # Check if the column is present
        if column_to_check in df_AL1.columns:
            # Show a pop-up message if the column is present
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showinfo("Notification", f"The column '{column_to_check}' is present.")

        # Continue with the rest of your code
        print("Continuing with the rest of the code...")
        # Your next line of code here

        dsn3.head()

        T10_col = dsn3.pop('B_Ref_List') # col-1

        dsn3.insert(0, 'B_Ref_List', T10_col)

        dsn3 = dsn3[['B_Ref_List','B_Part_No','Long Des']]

        dsn1.dropna(subset=['RefList'], inplace=True)
        dsn3.dropna(subset=['B_Ref_List'], inplace=True)

        #NEW PN# PARTNO
        ds1['PartNO'] = "PN#"
        ds1["PartNumber"] = ds1['PartNO'].astype(str) +""+ ds1['PartNumber'].astype(str)
        del ds1['PartNO']
        ds3['PartNO'] = "PN#"
        ds3["B_Part_No"] = ds3['PartNO'].astype(str) +""+ ds3['B_Part_No'].astype(str)
        del ds3['PartNO']

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx") as writer:

        #dt_H.to_excel(writer, sheet_name="Home", index=False)  
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        #df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
        #df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            ds1.to_excel(writer, sheet_name="BOM", index=False)
            dsn2.to_excel(writer, sheet_name="AVL GROUP", index=False)
            dcn1.to_excel(writer, sheet_name="Part Master", index=False)
            #dfn2.to_excel(writer, sheet_name="AVL_SHEET", index=True)
            df_AL1.to_excel(writer, sheet_name="AVL_SHEET", index=True)
            ds3.to_excel(writer, sheet_name="BOM_Data", index=False)
        #df2.to_excel(writer, sheet_name="AVL_SHEET", index=True) this line record upto 1 & 0
        #dc1.to_excel(writer, sheet_name="Part Master", index=False) this line record uoto 1 & 0 
        #ds2.to_excel(writer, sheet_name="AVL GROUP", index=False) this line record upto PTN1
        pass
        print('The file does not exist.')

    #########################################################################################################################################################################
    #########################################################################################################################################################################
        #@@ CRD Inspection @@#
        print('\n')
        print("\033[92;4m******CRD CHECK******\033[0m")
        print('\n')
    #########################################################################################################################################################################
    #########################################################################################################################################################################
        Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
        Chd = os.getcwd()
        file_path = 'BOM_List_OP.xlsx'
        directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

        # Assuming 'BOM DATA' sheet contains a column named 'Bom Ref'
        df_bom_data = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM_Data")
        duplicates_bom_data = df_bom_data[df_bom_data.duplicated(subset='B_Ref_List', keep=False)]
        # Assuming 'XY DATA' sheet contains a column named 'R'
        #df_xy_data = pd.read_excel("BOM_List_OP.xlsx", sheet_name="XY DATA")

        # Function to check for duplicates and display an error message
        def check_and_show_duplicates(df, column_name, sheet_name):
            duplicates = df[df.duplicated(subset=column_name, keep=False)]
            if not duplicates.empty:
                print("Duplicate values in 'B_Ref_List' column of 'BOM_Data':")
                print(duplicates_bom_data['B_Ref_List'].tolist())
                root = tk.Tk()
                root.withdraw()
                message = f"Duplicate values found in '{column_name}' column of '{sheet_name}' sheet:\n{duplicates[column_name].tolist()}"
                messagebox.showerror("Error", message)
                sys.exit()

        # Check for duplicates in 'Bom Ref' column of 'BOM DATA'
        check_and_show_duplicates(df_bom_data, 'B_Ref_List', 'BOM_Data')

        # Check for duplicates in 'R' column of 'XY DATA'
        #check_and_show_duplicates(df_xy_data, 'R', 'XY DATA')

    #########################################################################################################################################################################
    #########################################################################################################################################################################

    ##########################################################################################################################################
    ##########################################################################################################################################
        #@@ AVL Inspection @@#
    ##########################################################################################################################################

    print('\n')
    print("\033[92;4m******AVL LINE INSPECTION******\033[0m")
    print('\n')

    ##########################################################################################################################################

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()
    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            df_Iav1 = pd.read_csv(file_path)

    except ValueError:
        df_Iav1 = pd.read_excel(file_path, sheet_name="AVL_SHEET", index_col=False) 

        print(df_Iav1)

        # Function to check for missing values between two strings
    def check_missing_values(row):
        start_index = None
        end_index = None

        for i in range(1, len(row) + 1):  # Check up to the last column
            col_name = f'PTN_{i}'
            if col_name in row.index:  # Check if the column exists
                current_value = row[col_name]

                if pd.isna(current_value):
                    if start_index is None:
                        start_index = i
                    end_index = i
                else:
                    if start_index is not None and end_index is not None:
                        show_error(row['Group'], start_index, end_index)
                        start_index = None
                        end_index = None

    # Function to show pop-up error message
    def show_error(group, start_index, end_index):
        root = tk.Tk()
        root.withdraw()
        error_message = f"Error: Missing values between PTN_{start_index} and PTN_{end_index} in group '{group}'."
        messagebox.showerror("Error", error_message)

    # Check for missing values row-wise
    for index, row in df_Iav1.iterrows():
        check_missing_values(row)

    # Display the DataFrame with styling
    print(df_Iav1)

    # Function to check for missing values between two strings
    def check_missing_values(row):
        start_index = None
        end_index = None

        for i in range(1, len(row) + 1):  # Check up to the last column
            col_name = f'PTN_{i}'
            if col_name in row.index:  # Check if the column exists
                current_value = row[col_name]

                if pd.isna(current_value):
                    if start_index is None:
                        start_index = i
                    end_index = i
                else:
                    if start_index is not None and end_index is not None:
                        show_error(row.get('Group', 'Unknown Group'), start_index, end_index)
                        start_index = None
                        end_index = None

    # Function to show pop-up error message
    def show_error(group, start_index, end_index):
        root = tk.Tk()
        root.withdraw()
        
        error_message = f"Error: Missing values between PTN_{start_index} and PTN_{end_index} in group '{group}'.\nDo you want to stop the program?"
        response = messagebox.askquestion("Error", error_message)

        if response == 'yes':
            sys.exit(1)

    # ...

    # Check for missing values row-wise
    for index, row in df_Iav1.iterrows():
        check_missing_values(row)

    # Display the DataFrame with styling
    print(df_Iav1)

    ##########################################################################################################################################

    ##########################################################################################################################################

    #bil6 = pyfiglet.figlet_format("Part Master Process", width = 100)
    print('\n')
    print("\033[92;4m******Part Master Process******\033[0m")
    print('\n')
    ##########################################################################################################################################

    ##########################################################################################################################################
        
    #PART MASTER

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()

    # Excel file path
    df_PM1 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="Part Master")
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/PartMaster.xlsx") as writer:
        df_PM1.to_excel(writer, sheet_name="T_PBAR", index=False)
    excel_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/PartMaster.xlsx'

    # Access database connection parameters
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_Part Master/MODEL.mdb', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb')
    access_db_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb'
    driver = 'Microsoft Access Driver (*.mdb, *.accdb)'
    user = ''
    password = ''

    # Set up the connection string
    conn_str = f"DRIVER={{{driver}}};DBQ={access_db_path};UID={user};PWD={password};"

    # Connect to the Access database
    print(pyodbc.drivers())
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    # Read Excel data into a pandas DataFrame
    print('Open Excel....')
    df = pd.read_excel(excel_file_path)
    print(df.head(10))

    # Define the table name in the Access database
    print('open MS Access....')
    table_name = 'T_PBAR'

    # Check if the table exists
    existing_tables = [table[2] for table in cursor.tables(tableType='TABLE')]
    if table_name in existing_tables:
        # Append data to the existing table
        for _, row in df.iterrows():
            insert_query = f'''
            INSERT INTO {table_name} ({', '.join(df.columns)})
            VALUES ({', '.join(map(lambda x: f"'{row[x]}'", df.columns))})
            '''
            cursor.execute(insert_query)
            conn.commit()
            print('writing to access')
    else:
        print(f"The table '{table_name}' does notYT exist in the Access database.")

    # Close the database connection
    conn.close()
    print('write complete')

    ##########################################################################################################################################

    ##########################################################################################################################################

    #bil7 = pyfiglet.figlet_format("AVL Progress", width = 100)
    print('\n')
    print("\033[92;4m******AVL Progress******\033[0m")
    print('\n')
    ##########################################################################################################################################

    ##########################################################################################################################################

    #AVL#@@#

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')

    print(pyodbc.drivers())

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()

    xls = pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
    df1 = pd.read_excel('BOM_List_OP.xlsx', sheet_name='AVL_SHEET')

    print(df1.head(10))

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
        df1.to_excel(writer, sheet_name="AVL_SHEET", index=False)

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()
    df_AL1 = pd.read_excel('AVL.xlsx', sheet_name='AVL_SHEET')

    #new_column_name  = df_AL1.insert(1,'Group Name', '') #new_column_name = ('Group Name') # Get user input for the new column name
    # Desired column name
    desired_column_name = 'Group'

    # Get user input for the new column name
    new_column_name = ("Group Name")

    # Get user input for the new column value
    new_column_value = input(f"\033[93mEnter the value for the new column '{new_column_name}': \033[0m")
    dL22 = new_column_value
    #dL1 = new_column_value
    # Check if the desired column name exists
    if desired_column_name in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name)
        
        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value) #dL1 = new_column_value

    print(df_AL1)

    try:

        df_AL1['AVL Name']=df_AL1['PTN_1']
        # Replace values in 'AVL Name' with values from 'PTN_1' where 'PTN_1' is not empty
        #df_AL1['AVL Name'] = df_AL1['PTN_1'].fillna(df_AL1['AVL Name'])

    except Exception as e:
        # Handle the exception gracefully
        error_message = f"An error occurred:\nSomething went wrong while assigning AVL values {e}"
        error_msg1 = f"Check SF-02 is deleted\nCheck AVL Priority assign Properly {e}"
        # Show error message in a pop-up box
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showerror("Error", error_message)
        messagebox.showerror("AVL Error", error_msg1)
        #sys.exit(1)  # Exit the program with an error code

    first_column = df_AL1.pop('AVL Name')

    df_AL1.insert(2, 'AVL Name', first_column)

    df_AL1.insert(3, 'Comment', '')

    df_AL1['Comment'] = df_AL1['Group']

    #PTN_1>>

    # Desired column name
    desired_column_name1 = 'PTN_1'

    # New column to insert
    new_column_name = 'P_1'
    new_column_value = '1'

    # Check if the desired column name exists
    if desired_column_name1 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name1)
        
        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value)

    #PTN_2>>

    desired_column_name2 = 'PTN_2'

    # New column to insert
    new_column_name = 'P_2'
    new_column_value = '0'

    # Check if the desired column name exists
    if desired_column_name2 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name2)
        
        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value)

    #PTN_3>>

    desired_column_name3 = 'PTN_3'
    column_to_replace3 = 'PTN_3'

    # Check if the column exists in the DataFrame
    if column_to_replace3 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace3] = df_AL1[column_to_replace3].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name3 = 'P_3'
        new_column_value3 = '0'

        # Check if the desired column name exists
        if desired_column_name3 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name3)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name3,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name3].any():
                print(f"Values are present in '{desired_column_name3}' column:")
                print(df_AL1[desired_column_name3])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name3] != "$", new_column_name3] = 0
            else:
                print(f"No values are present in '{desired_column_name3}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name3}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace3}' does not exist in the DataFrame.")

    column_to_replace3 = 'PTN_3'

    # Check if the column exists in the DataFrame
    if column_to_replace3 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace3] = df_AL1[column_to_replace3].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace3}' does not exist in the DataFrame.")

    #PTN_4>>

    desired_column_name4 = 'PTN_4'
    column_to_replace4 = 'PTN_4'

    # Check if the column exists in the DataFrame
    if column_to_replace4 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace4] = df_AL1[column_to_replace4].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name4 = 'P_4'
        new_column_value4 = '0'

        # Check if the desired column name exists
        if desired_column_name4 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name4)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name4,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name4].any():
                print(f"Values are present in '{desired_column_name4}' column:")
                print(df_AL1[desired_column_name4])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name4] != "$", new_column_name4] = 0
            else:
                print(f"No values are present in '{desired_column_name4}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name4}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace4}' does not exist in the DataFrame.")

    column_to_replace4 = 'PTN_4'

    # Check if the column exists in the DataFrame
    if column_to_replace4 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace4] = df_AL1[column_to_replace4].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace4}' does not exist in the DataFrame.")

    #PTN_5>>

    desired_column_name5 = 'PTN_5'
    column_to_replace5 = 'PTN_5'

    # Check if the column exists in the DataFrame
    if column_to_replace5 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace5] = df_AL1[column_to_replace5].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name5 = 'P_5'
        new_column_value5 = '0'

        # Check if the desired column name exists
        if desired_column_name5 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name5)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name5,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name5].any():
                print(f"Values are present in '{desired_column_name5}' column:")
                print(df_AL1[desired_column_name5])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name5] != "$", new_column_name5] = 0
            else:
                print(f"No values are present in '{desired_column_name5}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name5}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace5}' does not exist in the DataFrame.")

    column_to_replace5 = 'PTN_5'

    # Check if the column exists in the DataFrame
    if column_to_replace5 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace5] = df_AL1[column_to_replace5].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace5}' does not exist in the DataFrame.")

    #PTN_6>>

    desired_column_name6 = 'PTN_6'
    column_to_replace6 = 'PTN_6'

    # Check if the column exists in the DataFrame
    if column_to_replace6 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace6] = df_AL1[column_to_replace6].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name6 = 'P_6'
        new_column_value6 = '0'

        # Check if the desired column name exists
        if desired_column_name6 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name6)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name6,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name6].any():
                print(f"Values are present in '{desired_column_name6}' column:")
                print(df_AL1[desired_column_name6])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name6] != "$", new_column_name6] = 0
            else:
                print(f"No values are present in '{desired_column_name6}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name6}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace6}' does not exist in the DataFrame.")

    column_to_replace6 = 'PTN_6'

    # Check if the column exists in the DataFrame
    if column_to_replace6 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace6] = df_AL1[column_to_replace6].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace6}' does not exist in the DataFrame.")

    #PTN_7>>

    desired_column_name7 = 'PTN_7'
    column_to_replace7 = 'PTN_7'

    # Check if the column exists in the DataFrame
    if column_to_replace7 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace7] = df_AL1[column_to_replace7].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name7 = 'P_7'
        new_column_value7 = '0'

        # Check if the desired column name exists
        if desired_column_name7 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name7)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name7,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name7].any():
                print(f"Values are present in '{desired_column_name7}' column:")
                print(df_AL1[desired_column_name7])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name7] != "$", new_column_name7] = 0
            else:
                print(f"No values are present in '{desired_column_name7}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name7}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace7}' does not exist in the DataFrame.")

    column_to_replace7 = 'PTN_7'

    # Check if the column exists in the DataFrame
    if column_to_replace7 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace7] = df_AL1[column_to_replace7].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace7}' does not exist in the DataFrame.")

    #PTN_8>>

    desired_column_name8 = 'PTN_8'
    column_to_replace8 = 'PTN_8'

    # Check if the column exists in the DataFrame
    if column_to_replace8 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace8] = df_AL1[column_to_replace8].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name8 = 'P_8'
        new_column_value8 = '0'

        # Check if the desired column name exists
        if desired_column_name8 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name8)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name8,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name8].any():
                print(f"Values are present in '{desired_column_name8}' column:")
                print(df_AL1[desired_column_name8])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name8] != "$", new_column_name8] = 0
            else:
                print(f"No values are present in '{desired_column_name8}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name8}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace8}' does not exist in the DataFrame.")

    column_to_replace8 = 'PTN_8'

    # Check if the column exists in the DataFrame
    if column_to_replace8 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace8] = df_AL1[column_to_replace8].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace8}' does not exist in the DataFrame.")

    #PTN_9>>

    desired_column_name9 = 'PTN_9'
    column_to_replace9 = 'PTN_9'

    # Check if the column exists in the DataFrame
    if column_to_replace9 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace9] = df_AL1[column_to_replace9].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name9 = 'P_9'
        new_column_value9 = '0'

        # Check if the desired column name exists
        if desired_column_name9 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name9)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name9,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name9].any():
                print(f"Values are present in '{desired_column_name9}' column:")
                print(df_AL1[desired_column_name9])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name9] != "$", new_column_name9] = 0
            else:
                print(f"No values are present in '{desired_column_name9}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name9}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace9}' does not exist in the DataFrame.")

    column_to_replace9 = 'PTN_9'

    # Check if the column exists in the DataFrame
    if column_to_replace9 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace9] = df_AL1[column_to_replace9].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace9}' does not exist in the DataFrame.")

    #PTN_10>>

    desired_column_name10 = 'PTN_10'
    column_to_replace10 = 'PTN_10'

    # Check if the column exists in the DataFrame
    if column_to_replace10 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace10] = df_AL1[column_to_replace10].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name10 = 'P_10'
        new_column_value10 = '0'

        # Check if the desired column name exists
        if desired_column_name10 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name10)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name10,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name10].any():
                print(f"Values are present in '{desired_column_name10}' column:")
                print(df_AL1[desired_column_name10])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name10] != "$", new_column_name10] = 0
            else:
                print(f"No values are present in '{desired_column_name10}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name10}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace10}' does not exist in the DataFrame.")

    column_to_replace10 = 'PTN_10'

    # Check if the column exists in the DataFrame
    if column_to_replace10 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace10] = df_AL1[column_to_replace10].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace10}' does not exist in the DataFrame.")

    #PTN_11>>

    desired_column_name11 = 'PTN_11'
    column_to_replace11 = 'PTN_11'

    # Check if the column exists in the DataFrame
    if column_to_replace11 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace11] = df_AL1[column_to_replace11].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name11 = 'P_11'
        new_column_value11 = '0'

        # Check if the desired column name exists
        if desired_column_name11 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name11)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name11,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name11].any():
                print(f"Values are present in '{desired_column_name11}' column:")
                print(df_AL1[desired_column_name11])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name11] != "$", new_column_name11] = 0
            else:
                print(f"No values are present in '{desired_column_name11}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name11}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace11}' does not exist in the DataFrame.")

    column_to_replace11 = 'PTN_11'

    # Check if the column exists in the DataFrame
    if column_to_replace11 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace11] = df_AL1[column_to_replace11].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace11}' does not exist in the DataFrame.")

    #PTN_12>>

    desired_column_name12 = 'PTN_12'
    column_to_replace12 = 'PTN_12'

    # Check if the column exists in the DataFrame
    if column_to_replace12 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace12] = df_AL1[column_to_replace12].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name12 = 'P_12'
        new_column_value12 = '0'

        # Check if the desired column name exists
        if desired_column_name12 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name12)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name12,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name12].any():
                print(f"Values are present in '{desired_column_name12}' column:")
                print(df_AL1[desired_column_name12])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name12] != "$", new_column_name12] = 0
            else:
                print(f"No values are present in '{desired_column_name12}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name12}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace12}' does not exist in the DataFrame.")

    column_to_replace12 = 'PTN_12'

    # Check if the column exists in the DataFrame
    if column_to_replace12 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace12] = df_AL1[column_to_replace12].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace12}' does not exist in the DataFrame.")

    #PTN_13>>

    desired_column_name13 = 'PTN_13'
    column_to_replace13 = 'PTN_13'

    # Check if the column exists in the DataFrame
    if column_to_replace13 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace13] = df_AL1[column_to_replace13].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name13 = 'P_13'
        new_column_value13 = '0'

        # Check if the desired column name exists
        if desired_column_name13 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name13)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name13,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name13].any():
                print(f"Values are present in '{desired_column_name13}' column:")
                print(df_AL1[desired_column_name13])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name13] != "$", new_column_name13] = 0
            else:
                print(f"No values are present in '{desired_column_name13}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name13}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace13}' does not exist in the DataFrame.")

    column_to_replace13 = 'PTN_13'

    # Check if the column exists in the DataFrame
    if column_to_replace13 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace13] = df_AL1[column_to_replace13].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace13}' does not exist in the DataFrame.")

    #PTN_14>>

    desired_column_name14 = 'PTN_14'
    column_to_replace14 = 'PTN_14'

    # Check if the column exists in the DataFrame
    if column_to_replace14 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace14] = df_AL1[column_to_replace14].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name14 = 'P_14'
        new_column_value14 = '0'

        # Check if the desired column name exists
        if desired_column_name14 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name14)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name14,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name14].any():
                print(f"Values are present in '{desired_column_name14}' column:")
                print(df_AL1[desired_column_name14])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name14] != "$", new_column_name14] = 0
            else:
                print(f"No values are present in '{desired_column_name14}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name14}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace14}' does not exist in the DataFrame.")

    column_to_replace14 = 'PTN_14'

    # Check if the column exists in the DataFrame
    if column_to_replace14 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace14] = df_AL1[column_to_replace14].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace14}' does not exist in the DataFrame.")

    #PTN_15>>

    desired_column_name15 = 'PTN_15'
    column_to_replace15 = 'PTN_15'

    # Check if the column exists in the DataFrame
    if column_to_replace15 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace15] = df_AL1[column_to_replace15].replace(np.nan, "$")
        print("\nDataFrame after replacement:")
        print(df_AL1)

        # New column to insert
        new_column_name15 = 'P_15'
        new_column_value15 = '0'

        # Check if the desired column name exists
        if desired_column_name15 in df_AL1.columns:
            # Find the index of the desired column
            index_of_desired_column = df_AL1.columns.get_loc(desired_column_name15)

            # Insert the new column next to the desired column
            df_AL1.insert(index_of_desired_column + 1, new_column_name15,  '')

            # Check if values are present in the desired column
            if df_AL1[desired_column_name15].any():
                print(f"Values are present in '{desired_column_name15}' column:")
                print(df_AL1[desired_column_name15])

            # Set the values in the new column to zero for cells where values are present in the desired column
                df_AL1.loc[df_AL1[desired_column_name15] != "$", new_column_name15] = 0
            else:
                print(f"No values are present in '{desired_column_name15}' column.")

            print("\nUpdated DataFrame:")
            print(df_AL1)
        else:
            print(f"Column '{desired_column_name15}' does not exist in the DataFrame.")
    else:
        print(f"Column '{column_to_replace15}' does not exist in the DataFrame.")

    column_to_replace15 = 'PTN_15'

    # Check if the column exists in the DataFrame
    if column_to_replace15 in df_AL1.columns:
        # Replace NaN values with "$" in the specified column
        df_AL1[column_to_replace15] = df_AL1[column_to_replace15].replace("$", np.NaN)
        print("\nDataFrame after replacement:")
        print(df_AL1)
    else:
        print(f"Column '{column_to_replace15}' does not exist in the DataFrame.")

    print(df_AL1.head(5))

    del df_AL1['Group']

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
        df_AL1.to_excel(writer, sheet_name="AVL_SHEET", index=False)

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()

    # Example DataFrame
    data = pd.read_excel('AVL.xlsx', sheet_name='AVL_SHEET')
    df_AL1 = pd.DataFrame(data)

    # Desired column order
    desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10']
    #desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10','PTN_11','P_11','PTN_12','P_12','PTN_13','P_13','PTN_14','P_14','PTN_15','P_15']

    # Create a list of columns present in both DataFrame and desired_order
    common_columns = [col for col in desired_order if col in df_AL1.columns]

    # Reorder the DataFrame based on the desired_order
    df_AL1 = df_AL1[common_columns]

    # Display the reordered DataFrame
    print(df_AL1)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
        df_AL1.to_excel(writer, sheet_name="AVL_SHEET", index=False)
        df_AL1.T.reset_index().T.to_excel(writer, sheet_name="AVL_SHEET", header=False ,index=False)
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')

    #read_file = pd.read_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.xlsx',skiprows=0)

    read_file = pd.read_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.xlsx')

    read_file.to_csv (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.txt', index = None, header= None)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()

    # Replace '0.0' with '0' in the content
    modified_content = content.replace('0.0', '0')

    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(modified_content)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC20 = content.replace(',,,,,,,,,,,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC20)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC19 = content.replace(',,,,,,,,,,,,,,,,,,,', '')
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC19)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC18 = content.replace(',,,,,,,,,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC18)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC17 = content.replace(',,,,,,,,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC17)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC16 = content.replace(',,,,,,,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC16)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC15 = content.replace(',,,,,,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC15)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC14 = content.replace(',,,,,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC14)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC13 = content.replace(',,,,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC13)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC12 = content.replace(',,,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC12)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC11 = content.replace(',,,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC11)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC10 = content.replace(',,,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC10)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC9 = content.replace(',,,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC9)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC8 = content.replace(',,,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC8)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC7 = content.replace(',,,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC7)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC6 = content.replace(',,,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC6)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC5 = content.replace(',,,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC5)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC4 = content.replace(',,,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC4)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC3 = content.replace(',,,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC3)

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC2 = content.replace(',,', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC2)
        #output_file.write('MC2','MC3','MC4','MC5','MC6','MC7','MC8','MC9','MC10','MC11','MC12','MC13','MC14','MC15','MC16','MC17','MC18','MC19','MC20')

    '''# Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC1 = content.replace('+', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC1)'''

    # Specify the path to your text file
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

    # Read the content of the text file
    with open(txt_file_path, 'r') as file:
        content = file.read()
    MC0 = content.replace('.0', '')
    # Write the modified content back to the text file
    with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
        output_file.write(MC0)

    print(f"AVL CREATED: D:/NX_BACKWORK/r'AVL.txt")

    # Specify the path to your text file
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()
    txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
    df = pd.read_table(txt_file_path, delimiter='\t', quoting=3)  # 3 corresponds to QUOTE_NONE
    csv_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.csv'
    df.to_csv(csv_file_path, index=False, sep='\t')  # 0 corresponds to QUOTE_NONE

    #read_file = pd.read_table (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.txt', sep='"')
    #read_file.to_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVList.xlsx', index=None)
    #read_file.to_csv (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.csv', index = None, header= None)
    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')

    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Line_X"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)

    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/MODEL.mdb')
    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.csv', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/AVL.CSV')

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X')
    Chd = os.getcwd()

    src_3 = 'MODEL.mdb'
    os.rename(src_3, dL22 +"_PM-Model"+".mdb")

    src_7 = 'AVL.csv'
    os.rename(src_7, dL22 +"_AVL"+".csv")

    # Specify the current name of the folder
    cfn1 = "Line_X"

    # Rename the folder
    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
    Chd = os.getcwd()
    from datetime import datetime

    # Assuming datetime is a datetime object
    datetime_str = datetime.now().strftime("%Y-%m-%d %I-%M-%S %p")

    os.rename(cfn1, f"{dL22}-{datetime_str}")

    print(f"\033[92mFolder '{cfn1}' renamed successfully to '{dL22}'.\033[0m")

    #program_6()
    # Assuming feeder verification is completed
    avpm = True

    if avpm:
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        time.sleep (5)
        messagebox.showinfo("AVL & PART MASTER", "AVL & Part Master File has been Generated!")
        
        sys.exit() #avpm