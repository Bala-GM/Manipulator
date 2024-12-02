import os
import sys
import openpyxl
import csv
from tkinter import filedialog, Tk
from tkinter import filedialog
import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

# Program 3: Database inspection interface_GUI/J0124-89P13
def program_3():
    # Function to write the CSV file in the format "Namechanger","B_Part_No"
    def save_csv_with_format(data, csv_filename):
        with open(csv_filename, mode='w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL)  # This will wrap values in quotes
            writer.writerows(data)

    # Load the workbook and the sheet
    file_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['AVL GROUP']

    # Add a new column 'Namechanger' and assign 'Dump1', 'Dump2', etc.
    namechanger_column_index = sheet.max_column + 1
    sheet.cell(row=1, column=namechanger_column_index).value = 'Namechanger'  # Add header

    for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=1):
        sheet.cell(row=idx+1, column=namechanger_column_index).value = f'Dump{idx}'

    # Create another sheet for Priority > 2, with only Namechanger and B_Part_No columns
    if 'AVL_Namechanger' not in wb.sheetnames:
        avl_namechanger_sheet = wb.create_sheet('AVL_Namechanger')
    else:
        avl_namechanger_sheet = wb['AVL_Namechanger']

    # Write header to the AVL_Namechanger sheet
    avl_namechanger_sheet.append(['Namechanger', 'B_Part_No'])

    # Collect data for Priority > 2
    priority_greater_than_01_data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        group, priority, b_part_no = row[0], row[1], row[2]
        namechanger = row[namechanger_column_index - 1]  # Accessing the Namechanger column
        
        if priority > 1:
            avl_namechanger_sheet.append([namechanger, b_part_no])
            priority_greater_than_01_data.append([namechanger, b_part_no])

    # Save the workbook with the new sheet
    wb.save(file_path)

    # Get CSV filename from user input
    csv_filename = input("Please provide a name for the CSV file (without extension): ") + ".csv"
    csv_full_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\\' + csv_filename

    # Save the data in the format "Namechanger","B_Part_No" to a CSV file
    save_csv_with_format(priority_greater_than_01_data, csv_full_path)

    print(f"Data saved successfully to {csv_filename} and in the sheet 'AVL_Namechanger'.")
    
    print("All processes completed successfully!")

    # Popup message to notify that the operation is complete
    sg.Popup('Operation Completed', 'All processes have been successfully completed.')

    # Exit the program after completion
    sys.exit()

#pyinstaller -F -i "AVL.ico" --noconsole AVL-Checker.py