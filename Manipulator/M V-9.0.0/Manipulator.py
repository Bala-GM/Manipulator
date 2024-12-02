import pandas as pd
import numpy as np
import os
from os import getcwd
import csv as csv
from io import StringIO
import re
import datetime
from io import BytesIO
import time
import sys
from datetime import datetime
import shutil
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import subprocess
import threading
import urllib
import urllib.parse
from sqlite3 import dbapi2 as sqlite
import sqlite3
import lxml
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
import PySimpleGUI as sg
import sqlalchemy
from sqlalchemy import create_engine
import sqlalchemy_access as sq_a
import sqlalchemy_pyodbc_mssql as sqlalchemy
from flask_sqlalchemy import SQLAlchemy
import pyodbc
import pypyodbc
import odbc
import psycopg2
import mysql.connector as sql
import MySQLdb as sql #pip install mysqlclient
from plyer import notification
import xlrd
import xml.etree.ElementTree as ET
from openpyxl.utils.exceptions import InvalidFileException
import psutil

from Program1 import program_1
from Program2 import program_2
from Program3 import program_3
from Program4 import program_4
from Program5 import program_5
from ProgramA import program_A
from ProgramB import program_B
from ProgramC import program_C
#from ProgramD import program_D
from ProgramE import program_E

def main():
# Display menu
    print("\033[92;40mFeederSetup & BOM Manipulator\033[0m \033[1;34;40mSYRMA\033[0m \033[1;36;40mSGS\033[0m \n\n\033[92;40mManipulator PY V-9 J26924 Select a program: \033[0m")
    print("\n")
    print("\033[1;36;40mPrograming Steps\033[0m")
    print("\n")
    print("1. BOM & XY DATA Manipulation: V-1.3") #89P13
    print("2. Part Master & AVL") #89P13
    print("3. Part Number(NameChanger): V-1.0") #89P13
    print("4. NeximPartNumber: V-1.0") #89P13
    print("5. LCR Tol. Help: V-1.0") #89P13
    print("\n")
    print("\033[1;36;40mFeederSetup Verification\033[0m")
    print("\n")
    print("A. FeederSetup: V-2.4") #89P13
    print("B. FeederSetup: V-2.4-X") #89P13
    print("C. LoadingList: V-2.6-X") #89P13
    print("D. Feeder Setup Comparison Support {LLV-2.6-X}") #89P13
    print("E. NPI-FeederList: V-1.0-X") #89P13
    print("\n")
    print("X. \033[1;31;40mExit Program\033[0m") #70599
    print("\n")

# Get user choice
    choice = input("\033[1;36;40mChoose the program number: \033[0m")

    # Run the selected program
    if choice == '1':
        program_1()
    elif choice == '2':
        program_2()
    elif choice == '3':
        program_3()
    elif choice == '4':
        program_4() 
    elif choice == '5':
        program_5()
    elif choice == 'A':
        program_A()
    elif choice == 'B':
        program_B()
    elif choice == 'C':
        program_C()
    #elif choice == 'D':
        #program_D()
    elif choice == 'E':
        program_E()
    elif choice == 'X':

        print("\n")
        print("\033[1;31;40mExiting the program.\033[0m")
        print("\nThank You")
    
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showinfo("Program Terminated", "Exiting the Program")
        sys.exit()
    else:
        print("Invalid choice. Exiting.")
        print("\nThank You")

    root = tk.Tk()
    root.withdraw()  # Hide the main window
    messagebox.showinfo("Program Terminated", "Enter Invalid choice!")
    messagebox.showinfo("Program Dev", "SIG#00110111 00110000 00110101 00111001 00111001\n01000010 00111000 00111001 01010000 00110001 00110011")
    sys.exit()
    
if __name__ == "__main__":
    main()
    
    #pyinstaller -F -i SYRMA.ico Manipulator.py