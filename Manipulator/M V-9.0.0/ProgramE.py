def program_E():
    import os
    import sys
    import shutil
    import time
    import datetime
    import openpyxl
    import pandas as pd
    import tkinter as tk
    from tkinter import messagebox
    from tkinter import filedialog
    import PySimpleGUI as sg
    from openpyxl import load_workbook
    import xml.etree.ElementTree as ET
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    # Print a message indicating the program is running
    print("\033[32;4m*******AVL FeederSetup Running*******\033[0m")

    # Define the root directory
    root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

    # Function to find all FeederSetup.xml files
    def find_feeder_setup_files(root_directory):
        feeder_setup_files = []
        for root, dirs, files in os.walk(root_directory):
            for file in files:
                if file == "FeederSetup.xml":
                    feeder_setup_files.append(os.path.join(root, file))
        return feeder_setup_files

    # Function to find the setup description from the file path
    def find_setup_description(file_path):
        parts = file_path.split(";")
        setup_description = parts[-2].strip()
        return setup_description

    # Function to rename FeederSetup.xml with a description-based name
    def rename_feeder_setup_with_description(file_path):
        setup_description = find_setup_description(file_path)
        if "[Top] Line1" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line1" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        elif "[Top] Line2" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line2" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        elif "[Top] Line3" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line3" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        elif "[Top] Line4" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line4" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        elif "[Top] Line-1" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line-1" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        else:
            print(f"Unrecognized description in {file_path}. Skipping rename.")
            return None
        
        new_path = os.path.join(os.path.dirname(file_path), new_file_name)
        os.rename(file_path, new_path)
        print(f"FeederSetup.xml renamed to: {new_path}")
        return new_path

    # Function to convert XML to DataFrame
    def xml_to_dataframe(xml_file):
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        data = []
        for feeder in root.findall('Feeder'):
            feeder_dict = {}
            for child in feeder.iter():
                if child.tag != 'Feeder':
                    if child.text:
                        feeder_dict[child.tag] = child.text.strip()
            if feeder_dict:
                data.append(feeder_dict)
        
        df = pd.DataFrame(data)
        
        # Remove rows where PartNumber is blank
        df = df[df['PartNumber'].notna() & (df['PartNumber'].str.strip() != '')]

        # Sort DataFrame by PartNumber column
        df.sort_values(by=['PartNumber'], inplace=True)
        
        return df

    # Function to convert DataFrame to Excel file
    def dataframe_to_excel(df, excel_file):
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False)

    # Function to process the Excel file and keep specified columns
    def process_excel_file(file_path):
        df = pd.read_excel(file_path)
        
        # Define the columns to keep
        columns_to_keep = ['PartNumber', 'AVLNAME', 'AVL', 'PartShapeName', 'PartNumberImage', 'QTY']
        
        # Check if 'AVLNAME' and 'AVL' are present in the DataFrame
        missing_columns = [col for col in ['AVLNAME', 'AVL'] if col not in df.columns]
        
        if missing_columns:
            root = tk.Tk()
            root.withdraw()  # Hide the main Tkinter window
            answer = messagebox.askyesno("Missing Columns", f"Columns {missing_columns} are missing. Do you want to create them?")
            
            if answer:
                for col in missing_columns:
                    df[col] = ""
            else:
                print("User chose not to create missing columns. Exiting the program.")
                sys.exit(1)  # Exit the program after user refuses to create the columns
        
        # Keep only the specified columns
        df = df[columns_to_keep]
        
        # Add the 'Side' column based on the file name
        side = 'BOT' if 'AVL_B' in file_path else 'TOP'
        df['Side'] = side
        
        # Save the modified DataFrame back to Excel
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False)

    # Main processing loop
    feeder_setup_files = find_feeder_setup_files(root_directory)
    for feeder_setup_file in feeder_setup_files:
        renamed_xml_file = rename_feeder_setup_with_description(feeder_setup_file)
        
        if renamed_xml_file:
            # Convert XML to DataFrame
            df = xml_to_dataframe(renamed_xml_file)
            
            # Convert DataFrame to Excel
            output_excel = os.path.splitext(renamed_xml_file)[0] + '.xlsx'
            dataframe_to_excel(df, output_excel)
            
            # Move the Excel file to the destination directory
            destination_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
            shutil.move(output_excel, destination_directory)
            
            # Process the Excel file to keep specific columns and add the Side column
            final_excel_path = os.path.join(destination_directory, os.path.basename(output_excel))
            process_excel_file(final_excel_path)
        else:
            print(f"Skipping {feeder_setup_file} due to missing description.")
    ##################################################################         
     # Ensure the output folder exists
    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
    if not os.path.isdir(yourfolder):
        print('Folder does not exist, creating it now...')
        os.makedirs(yourfolder)

    time.sleep(2)  # Pause to show completion message
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    # Define the source and destination paths for each file
    
    def copy_file_if_exists(source, destination):
        if os.path.exists(source):
            try:
                shutil.copy(source, destination)
                print(f"Successfully copied {source} to {destination}")
            except PermissionError:
                print(f"Permission error: Could not copy {source}. File might be in use.")
            except Exception as e:
                print(f"An error occurred while copying the file: {e}")
        else:
            print(f"File {source} not found. Skipping this file.")

    # Copy FeederSetup_AVL_B.xlsx and FeederSetup_AVL_T.xlsx if they exist
    source_file_b = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/FeederSetup_AVL_B.xlsx"
    dest_file_b = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/FeederSetup_AVL_B.xlsx"
    copy_file_if_exists(source_file_b, dest_file_b)

    source_file_t = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/FeederSetup_AVL_T.xlsx"
    dest_file_t = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/FeederSetup_AVL_T.xlsx"
    copy_file_if_exists(source_file_t, dest_file_t)

    # Combine the files
    combined_file = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"

    def combine_and_remove(file_b, file_t, output_file):
        dfs = []
        if os.path.exists(file_b):
            df_b = pd.read_excel(file_b)
            dfs.append(df_b)
            print(f"Loaded {file_b}")
        if os.path.exists(file_t):
            df_t = pd.read_excel(file_t)
            dfs.append(df_t)
            print(f"Loaded {file_t}")

        if dfs:
            combined_df = pd.concat(dfs, ignore_index=True)
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                combined_df.to_excel(writer, index=False, sheet_name="AVL_Combine")
            print(f"Combined data saved to {output_file} with sheet name 'AVL_Combine'")

            if os.path.exists(file_b):
                os.remove(file_b)
                print(f"Removed {file_b}")
            if os.path.exists(file_t):
                os.remove(file_t)
                print(f"Removed {file_t}")
        else:
            print("No files to combine.")

    combine_and_remove(dest_file_b, dest_file_t, combined_file)

    # Verify combined file exists
    if not os.path.exists(combined_file):
        print(f"File {combined_file} not found. Exiting.")
        exit()

    # Reorder columns
    wb = openpyxl.load_workbook(combined_file)
    avl_combine_sheet = wb["AVL_Combine"]

    new_column_order = ["PartNumber", "AVLNAME", "PartShapeName", "PartNumberImage", "QTY", "Side", "AVL"]
    header = next(avl_combine_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    print(f"Header: {header}")

    col_map = {col: header.index(col) for col in new_column_order if col in header}

    avl_combine_reordered = wb.create_sheet("AVL_Combine_Reordered")
    avl_combine_reordered.append(new_column_order)

    for row in avl_combine_sheet.iter_rows(min_row=2, values_only=True):
        reordered_row = [row[col_map[col]] for col in new_column_order if col in col_map]
        avl_combine_reordered.append(reordered_row)

    wb.remove(avl_combine_sheet)
    wb["AVL_Combine_Reordered"].title = "AVL_Combine"
    wb.save(combined_file)
    wb.close()

    print("Columns repositioned and saved successfully.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    # Function to delimit the AVL column by commas and remove double quotes
    def delimit_avl_column(input_file, delimiters=[',']):
        # Load the existing Excel file, specifying the engine explicitly
        xls = pd.ExcelFile(input_file, engine='openpyxl')

        # Read the data from the first sheet (assuming the data is there)
        df = pd.read_excel(xls, sheet_name=0)

        # Ensure that the 'AVL' column exists
        if 'AVL' in df.columns:
            # Remove double quotes and split by commas
            avl_cleaned = df['AVL'].fillna("").replace('"', '')  # Remove all double quotes before splitting
            avl_split = avl_cleaned.str.split(delimiters[0], expand=True)

            # Rename the columns resulting from the split (AVL_1, AVL_2, AVL_3, etc.)
            avl_split.columns = [f'AVL_{i+1}' for i in range(avl_split.shape[1])]

            # Ensure there are no remaining double quotes
            avl_split = avl_split.apply(lambda col: col.str.replace('"', '', regex=False))

            # Add the other columns from the original DataFrame, excluding the 'AVL' column
            result_df = pd.concat([df.drop(columns=['AVL']), avl_split], axis=1)

            # Use ExcelWriter with mode='a' to append to the existing workbook
            with pd.ExcelWriter(input_file, engine='openpyxl', mode='a') as writer:
                # Write the new DataFrame to a new sheet called 'AVL_Delimited'
                result_df.to_excel(writer, sheet_name='AVL_Delimited', index=False)

            print(f"New sheet 'AVL_Delimited' added to {input_file}.")
        else:
            print("The 'AVL' column does not exist in the file.")

    # Call the function to delimit the 'AVL' column and create a new sheet
    delimit_avl_column(combined_file, delimiters=[','])
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #AVL_Delimited cpy to AVL_Compare

    # File path to the Excel file
    file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"
    sheet_name = "AVL_Delimited"
    new_sheet_name = "AVL_Compare"

    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Check if the sheet exists
    if sheet_name in wb.sheetnames:
        # Get the sheet to copy
        sheet_to_copy = wb[sheet_name]
        
        # Create a new sheet with a copy of the original sheet's content
        new_sheet = wb.copy_worksheet(sheet_to_copy)
        
        # Rename the new sheet
        new_sheet.title = new_sheet_name
        
        # Save the workbook with the new sheet
        wb.save(file_path)
        print(f"Sheet '{sheet_name}' successfully copied and renamed to '{new_sheet_name}'.")
    else:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    # Function to Sort and QTY "0" and deletion and save

    # Path to the Excel file
    file_path = r'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'

    # Load the Excel file
    if os.path.exists(file_path):
        # Open the Excel file
        excel_file = pd.ExcelFile(file_path)
        
        # Check for the sheets
        sheets_to_check = ['AVL_Compare'] #'AVL_Delimited_T', 'AVL_Delimited_B'

        # Create a dictionary to store DataFrames
        sheet_data = {}

        for sheet in sheets_to_check:
            if sheet in excel_file.sheet_names:
                # Load the sheet into a DataFrame
                df = pd.read_excel(file_path, sheet_name=sheet)
                
                # Check if 'QTY' column exists and filter out rows where QTY is 0
                if 'QTY' in df.columns:
                    df = df[df['QTY'] != 0]
                
                # Store the modified DataFrame
                sheet_data[sheet] = df

        # Save the modified sheets back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for sheet, df in sheet_data.items():
                df.to_excel(writer, sheet_name=sheet, index=False)
    else:
        print(f"The file at {file_path} does not exist.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    # BOM AVL import to AVL_Polarity_Check Define file paths
    source_file = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx"
    target_file = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"

    def copy_avl_to_combined(source_file, target_file, new_sheet_name='BOM_AVL'):
        # Load the source Excel file
        avl_data = pd.read_excel(source_file)

        # Load the target Excel file and open it in append mode
        with pd.ExcelWriter(target_file, engine='openpyxl', mode='a') as writer:
            # Write the AVL data to a new sheet called 'BOM_AVL'
            avl_data.to_excel(writer, sheet_name=new_sheet_name, index=False)

        print(f"Data from {source_file} has been copied to {target_file} in a new sheet named '{new_sheet_name}'.")

    # Call the function to copy AVL data to the combined file
    copy_avl_to_combined(source_file, target_file)
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #Program AVL and BOM AVL checking Step 1

    # Case 1: Checking Group Name in AVL_SHEET
    def check_group_name_avl_sheet(source_file, sheet_name):
        df = pd.read_excel(source_file, sheet_name=sheet_name)
        
        # Check if "Group Name" column contains 'B89P13'
        if 'B89P13' in df['Group Name'].values:
            return True  # BOM AVL needs to be skipped
        else:
            return False  # BOM AVL exists

    # Case 2: Checking AVLNAME and AVL in AVL_Combine sheet
    def check_avl_columns_avl_combine(source_file, sheet_name):
        df = pd.read_excel(source_file, sheet_name=sheet_name)

        # Check if both 'AVLNAME' and 'AVL' columns are completely empty
        if df['AVLNAME'].isna().all() and df['AVL'].isna().all():
            return True  # Program AVL is empty
        else:
            return False  # Program AVL has data

    # Case 3: Combined logic to decide overall processing
    def combined_check():
        # Define source files and sheet names for Case 1 and Case 2
        bom_source_file = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx"
        bom_sheet_name = "AVL_SHEET"
        
        avl_source_file = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"
        avl_sheet_name = "AVL_Combine"

        # Case 1: Check Group Name in BOM AVL (AVL_SHEET)
        case_1_result = check_group_name_avl_sheet(bom_source_file, bom_sheet_name)

        # Case 2: Check AVLNAME and AVL in Program AVL (AVL_Combine)
        case_2_result = check_avl_columns_avl_combine(avl_source_file, avl_sheet_name)

        # Tkinter setup for message boxes
        root = tk.Tk()
        root.withdraw()  # Hide the main Tkinter window

        # Conditional logic based on the results of Case 1 and Case 2
        if case_1_result and case_2_result:
            message = "NO BOM AVL and NO Program AVL. Skipping the code and proceeding to the next."
            print(message)
            messagebox.showinfo("Skip", message)
            return  # Skip further processing

        elif case_1_result and not case_2_result:
            message = "NO BOM AVL but there is Program AVL. Please Check."
            print(message)
            messagebox.showinfo("Abort", message)
            raise SystemExit  # Abort the program

        elif not case_1_result and case_2_result:
            message = "BOM AVL Found but Program AVL not Found."
            print(message)
            messagebox.showinfo("Abort", message)
            raise SystemExit  # Abort the program

        else:
            message = "BOM AVL Found and Program AVL Found. Verifying the AVL data."
            print(message)
            messagebox.showinfo("Verification", message)
    ################################################################################################################################
    ################################################################################################################################
    #Program AVL and BOM AVL checking SKIP if it True or else it need to run

        # Function to process and clean the data
        def process_and_clean_avl_bom(df_delimited, df_bom):
            # Define the columns order for AVL_Compare
            avl_columns_delimited = ['PartNumber', 'AVLNAME', 'QTY', 'Side'] + [col for col in df_delimited.columns if col.startswith('AVL_')]
            df_delimited_filtered = df_delimited[avl_columns_delimited]

            # Print initial DataFrame before cleaning
            print("Initial AVL Compare DataFrame:")
            print(df_delimited_filtered.head())

            # Clean the AVL columns by adding 'PN#' and removing '.0'
            for col in ['PartNumber', 'AVLNAME'] + [col for col in avl_columns_delimited if col.startswith('AVL_')]:
                df_delimited_filtered[col] = df_delimited_filtered[col].apply(
                    lambda x: f'PN#{int(x)}' if pd.notna(x) and str(x).replace('.0', '').isdigit() else x)  # Remove .0 and add 'PN#'
            
            # Print DataFrame after prefix addition and cleaning
            print("\nCleaned AVL Compare DataFrame:")
            print(df_delimited_filtered.head())

            # Define the columns order for BOM_AVL
            bom_columns = ['Group Name', 'AVL Name', 'Comment'] + [col for col in df_bom.columns if col.startswith('PTN_')]
            df_bom_filtered = df_bom[bom_columns]

            # Clean the PTN columns by adding 'PN#' and removing '.0'
            for col in ['AVL Name'] + [col for col in bom_columns if col.startswith('PTN_')]:
                df_bom_filtered[col] = df_bom_filtered[col].apply(
                    lambda x: f'PN#{int(x)}' if pd.notna(x) and str(x).replace('.0', '').isdigit() else x)  # Remove .0 and add 'PN#'
            
            # Print BOM DataFrame after prefix addition and cleaning
            print("\nCleaned BOM AVL DataFrame:")
            print(df_bom_filtered.head())

            return df_delimited_filtered, df_bom_filtered

        # Function to check for strict bidirectional match (set-based)
        def check_strict_bidirectional_match(row, bom_columns, avl_columns):
            ptn_values = set()
            for col in bom_columns[3:]:  # PTN_ columns start from index 3
                if pd.notna(row[col]) and row[col] != '':
                    ptn_values.add(row[col].strip())

            avl_values = set()
            for col in avl_columns[4:]:  # AVL_ columns start from index 4
                if pd.notna(row[col]) and row[col] != '':
                    avl_values.add(row[col].strip())

            return 'Match' if avl_values == ptn_values else 'Miss'

        # Function to align, compare, and save results
        def align_compare_and_save(input_file, output_sheet_name):
            # Load data from AVL_Compare and BOM_AVL sheets
            df_delimited = pd.read_excel(input_file, sheet_name='AVL_Compare')
            df_bom = pd.read_excel(input_file, sheet_name='BOM_AVL')

            # Process and clean the data
            df_delimited_filtered, df_bom_filtered = process_and_clean_avl_bom(df_delimited, df_bom)

            # Rename 'AVL Name' to 'AVLNAME' for merging
            df_bom_filtered.rename(columns={'AVL Name': 'AVLNAME'}, inplace=True)

            # Merge the two DataFrames on 'AVLNAME'
            df_combined = pd.merge(df_bom_filtered, df_delimited_filtered, how='outer', on='AVLNAME', indicator=True)

            # Drop rows where 'AVLNAME' is missing
            df_combined = df_combined.dropna(subset=['AVLNAME'])

            # Check for strict bidirectional match and create the 'Result' column
            bom_columns = df_bom_filtered.columns
            avl_columns = df_delimited_filtered.columns

            df_combined['Result'] = df_combined.apply(
                lambda row: check_strict_bidirectional_match(row, bom_columns, avl_columns), axis=1
            )

            # Sort the rows with unmatched ones first
            unmatched_rows = df_combined[df_combined['Result'] == 'Miss']
            matched_rows = df_combined[df_combined['Result'] == 'Match']
            df_combined_sorted = pd.concat([unmatched_rows, matched_rows], ignore_index=True)

            # Print the final sorted DataFrame for debugging
            print("\nFinal Sorted DataFrame:")
            print(df_combined_sorted.head())

            # Write the sorted DataFrame to the 'AVL_Compare_Results' sheet
            with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_combined_sorted.to_excel(writer, sheet_name=output_sheet_name, index=False)

            # Reload workbook to apply styling
            wb = load_workbook(input_file)
            ws = wb[output_sheet_name]

            # Define fill colors for match/mismatch
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Highlight match and mismatch rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                result_value = row[-1].value
                fill = green_fill if result_value == 'Match' else yellow_fill
                for cell in row:
                    cell.fill = fill

            # Save the workbook
            wb.save(input_file)
            print(f"Data processed and saved to '{output_sheet_name}'.")

        # Function call to process and save the result
        input_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'
        align_compare_and_save(input_file, 'AVL_Compare_Results')
        ################################################################################################################################
        ################################################################################################################################
        #miss log capture
        # Function to log misses to a file
        def log_misses_to_file(input_file, sheet_name, log_file):
            # Read the Excel file into a DataFrame
            df = pd.read_excel(input_file, sheet_name=sheet_name)

            # Check if 'Result' column contains the word "miss"
            miss_rows = df[df['Result'].str.contains('miss', case=False, na=False)]

            # If there are rows with "miss", proceed to log them
            if not miss_rows.empty:
                with open(log_file, 'a') as f:
                    # Get the current date and time correctly
                    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    f.write(f"\n\nMissed Rows Logged at: {current_time}\n")

                    # Write each missed row to the log file
                    miss_rows.to_string(f, index=False)
                    f.write("\n")  # Add a newline after the logged data

                print(f"{len(miss_rows)} missed rows have been logged.")
            else:
                print("No 'miss' found in the 'Result' column.")

        # Specify input file, sheet name, and log file path
        input_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'
        sheet_name = 'AVL_Compare_Results'
        log_file = 'D:/NX_BACKWORK/Database_File/SMT_AVL/AVL-Check logfile.txt'

        # Call the function to log misses
        log_misses_to_file(input_file, sheet_name, log_file)
        ################################################################################################################################
        ################################################################################################################################
        #100% match checker
        def check_match_percentage(input_file, sheet_name):
            # Load the Excel sheet
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            
            # Count the total number of rows
            total_rows = len(df)
            
            # Count the rows where the 'Result' column is 'Match'
            match_count = len(df[df['Result'] == 'Match'])
            
            # Calculate the match percentage
            match_percentage = (match_count / total_rows) * 100

            # Check if match percentage is 100%
            if match_percentage == 100:
                sg.popup('100% match success, moving to next flow!')
                print('100% match success, moving to next flow!')
            else:
                sg.popup(f'Match less than 100%, current match: {match_percentage:.2f}%. Aborting the program!')
                print(f'Match less than 100%, current match: {match_percentage:.2f}%. Aborting the program!')
                exit()  # Abort the program

        # File path and sheet name
        input_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'
        sheet_name = 'AVL_Compare_Results'

        # Call the function to check the match percentage
        check_match_percentage(input_file, sheet_name)
        ################################################################################################################################
        ################################################################################################################################

    # --- Continue processing here if none of the skip or abort conditions were met ---
    print("Proceeding with the rest of the processing...")

    # Call the combined check function to handle all cases
    combined_check()
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #NO AVL then start from here
    # BOM import to AVL_Polarity_Check Define file paths
    source_file = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx"
    target_file = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"

    def copy_bomlcr_to_combined(source_file, target_file, new_sheet_name='BOM'):
        # Load the source Excel file
        bomlcr_data = pd.read_excel(source_file)

        # Load the target Excel file and open it in append mode
        with pd.ExcelWriter(target_file, engine='openpyxl', mode='a') as writer:
            # Write the bom data to a new sheet called 'BOM'
            bomlcr_data.to_excel(writer, sheet_name=new_sheet_name, index=False)

        print(f"Data from {source_file} has been copied to {target_file} in a new sheet named '{new_sheet_name}'.")

    # Call the function to copy AVL data to the combined file
    copy_bomlcr_to_combined(source_file, target_file)
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    import pandas as pd
    from openpyxl import load_workbook
    import re

    # Function to extract LCR information from each component line
    def extract_component_info(component_line):
        # Regular expressions for CAPACITOR
        capacitor_type_match = re.search(r'\b(CAP|Capacitor|MLCC|CAPACITOR)\b', component_line, re.IGNORECASE)
        capacitor_value_match = re.search(r'(\d+(\.\d+)?)\s*([pnuμmkPUNKM]?\w?)F', component_line)
        capacitor_tolerance_match = re.search(r'[±](\d+)%', component_line)
        
        # Regular expressions for RESISTOR
        resistor_type_match = re.search(r'\b(Res|Resistor|FLIM|RESISTOR)\b', component_line, re.IGNORECASE)
        resistor_value_match = re.search(r'(\d+(\.\d+)?)\s*([pnuμmkPUNKM]?\w?)\s*(\d+%)?', component_line)
        resistor_tolerance_match = re.search(r'[±](\d+)%', component_line)

        # Default values
        LCR_Type = None
        LCR_Value = None
        LCR_Unit = None
        LCR_Tolerance = None

        # Extract capacitor details
        if capacitor_type_match:
            LCR_Type = capacitor_type_match.group(1)
        if capacitor_value_match:
            LCR_Value = capacitor_value_match.group(1)
            LCR_Unit = capacitor_value_match.group(3)
        if capacitor_tolerance_match:
            LCR_Tolerance = capacitor_tolerance_match.group(1)

        # Extract resistor details
        if resistor_type_match:
            LCR_Type = resistor_type_match.group(1)
        if resistor_value_match:
            LCR_Value = resistor_value_match.group(1)
            LCR_Unit = resistor_value_match.group(3)
        if resistor_tolerance_match:
            LCR_Tolerance = resistor_tolerance_match.group(1)

        return LCR_Type, LCR_Value, LCR_Unit, LCR_Tolerance

    # Function to process the 'Long Des' column and save results
    def process_bom_file(input_file, sheet_name):
        # Load the BOM sheet into a DataFrame
        df = pd.read_excel(input_file, sheet_name=sheet_name)

        # Create new columns to store the extracted data
        df['LCR_Type'] = None
        df['LCR_Value'] = None
        df['LCR_Unit'] = None
        df['LCR_Tolerance'] = None

        # Apply the extraction function to each row in the 'Long Des' column
        for index, row in df.iterrows():
            long_des = row['Long Des']
            if pd.notna(long_des):  # Ensure 'Long Des' is not NaN
                lcr_type, lcr_value, lcr_unit, lcr_tolerance = extract_component_info(long_des)
                df.at[index, 'LCR_Type'] = lcr_type
                df.at[index, 'LCR_Value'] = lcr_value
                df.at[index, 'LCR_Unit'] = lcr_unit
                df.at[index, 'LCR_Tolerance'] = lcr_tolerance

        # Load the workbook and save the updated DataFrame back into the 'BOM' sheet
        with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Processed data saved in '{sheet_name}' sheet of {input_file}")

    # File path and sheet name
    input_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'
    sheet_name = 'BOM'

    # Call the function to process and save the BOM sheet
    process_bom_file(input_file, sheet_name)
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #file move to another file

    import openpyxl

    # File paths
    source_file = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx"
    target_file = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"

    # Load the workbooks
    source_wb = openpyxl.load_workbook(source_file)
    target_wb = openpyxl.load_workbook(target_file)

    # Load the sheet from the source file
    source_sheet = source_wb["Upload_data"]

    # Check if 'Upload_data' sheet already exists in target, if so remove it
    if "Upload_data" in target_wb.sheetnames:
        target_wb.remove(target_wb["Upload_data"])

    # Create a new sheet in the target workbook
    target_sheet = target_wb.create_sheet("Upload_data")

    # Copy data from the source sheet to the target sheet
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

    # Save the updated target file
    target_wb.save(target_file)

    # Close the workbooks
    source_wb.close()
    target_wb.close()

    print("Data moved successfully.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #upload Part Number Column to PartNumber
    import openpyxl

    # File path
    file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"

    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Upload_data"]

    # Find the index of the "Part Number" column
    part_number_col = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "Part Number":
            part_number_col = col
            break

    if part_number_col:
        # Determine the new column position at the end
        new_col_index = sheet.max_column + 1

        # Rename the new column header to "PartNumber" and move the column data
        sheet.cell(row=1, column=new_col_index).value = "PartNumber"  # Rename the header
        for row_index, cell in enumerate(part_number_col[1:], start=2):  # Copy data excluding header
            sheet.cell(row=row_index, column=new_col_index).value = cell.value

        # Save the workbook
        wb.save(file_path)
        print("Column moved and renamed successfully.")
    else:
        print("Part Number column not found.")

    # Close the workbook
    wb.close()
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #Combine LCR into Upload
    import pandas as pd

    # Function to move BOM data to Upload_data sheet
    def move_bom_to_upload_data(input_file):
        # Load both sheets into DataFrames
        df_bom = pd.read_excel(input_file, sheet_name='BOM')
        df_upload = pd.read_excel(input_file, sheet_name='Upload_data')

        # Columns to move from BOM to Upload_data
        columns_to_move = ['Shape', 'LCR_Type', 'LCR_Value', 'LCR_Unit', 'LCR_Tolerance']

        # Ensure that BOM has these columns
        if not set(columns_to_move).issubset(df_bom.columns):
            raise ValueError(f"One or more of the required columns {columns_to_move} not found in BOM sheet.")

        # Merge based on 'PartNumber'
        merged_data = pd.merge(df_upload, df_bom[['PartNumber'] + columns_to_move], how='left', on='PartNumber')

        # Check for duplicate PartNumbers in BOM
        duplicate_partnumbers = df_bom[df_bom.duplicated(subset=['PartNumber'], keep=False)]

        if not duplicate_partnumbers.empty:
            # For each duplicate PartNumber, append the data
            for partnumber in duplicate_partnumbers['PartNumber'].unique():
                duplicate_rows = duplicate_partnumbers[duplicate_partnumbers['PartNumber'] == partnumber][columns_to_move]
                # Append these rows to Upload_data
                for _, row in duplicate_rows.iterrows():
                    new_row = df_upload[df_upload['PartNumber'] == partnumber].iloc[0].copy()
                    new_row.update(row)
                    merged_data = pd.concat([merged_data, pd.DataFrame([new_row])], ignore_index=True)

        # Write the updated data to the same sheet 'Upload_data'
        with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            merged_data.to_excel(writer, sheet_name='Upload_data', index=False)

        print(f"Data successfully moved from BOM to Upload_data and saved in {input_file}")

    # File path and sheet names
    input_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'

    # Call the function to move the BOM data to Upload_data
    move_bom_to_upload_data(input_file)
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #Combine LCR shape
    import pandas as pd

    def combine_and_rearrange_columns(input_file):
        # Load the Upload_data sheet into a DataFrame
        df_upload = pd.read_excel(input_file, sheet_name='Upload_data')

        # Ensure the necessary columns are present in the sheet
        if 'Shape' not in df_upload.columns or 'LCR_Type' not in df_upload.columns:
            raise ValueError("Both 'Shape' and 'LCR_Type' columns must be present in the Upload_data sheet.")

        # Combine 'Shape' and 'LCR_Type' into a new column 'LCRShape_Type'
        df_upload['LCRShape_Type'] = df_upload['Shape'].fillna('') + '_' + df_upload['LCR_Type'].fillna('')

        # Rearrange the columns in the specified order
        column_order = [
            'Feeder Location', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 
            'Part Number', 'Part Description', 'Reference', 'QTY', 'LCRShape_Type', 
            'LCR_Value', 'LCR_Unit', 'LCR_Tolerance', 'Side', 'ModelName', 'PartNumber'
        ]

        # Ensure all columns are present before rearranging
        missing_columns = set(column_order) - set(df_upload.columns)
        if missing_columns:
            raise ValueError(f"Missing columns in the Upload_data sheet: {missing_columns}")

        # Reorder the DataFrame according to the specified column sequence
        df_upload = df_upload[column_order]

        # Write the updated data back to the 'Upload_data' sheet
        with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_upload.to_excel(writer, sheet_name='Upload_data', index=False)

        print(f"Columns combined and rearranged successfully in {input_file}")

    # File path to the Excel file
    input_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'

    # Call the function to combine columns and rearrange the sequence
    combine_and_rearrange_columns(input_file)
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #upload to excel top and bot in seprate

    # Load the Excel file and the AVL_Delimited sheet
    file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"
    sheet_name = "Upload_data"
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Filter rows where 'Side' is 'Top' or 'Bot'
    top_df = df[df['Side'] == 'TOP']
    bot_df = df[df['Side'] == 'BOT']

    # Load the existing Excel file to write into it
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Write to 'Top' sheet if there are rows for 'TOP'
        if not top_df.empty:
            top_df.to_excel(writer, sheet_name='Upload_data_T', index=False)
        
        # Write to 'Bot' sheet if there are rows for 'BOT'
        if not bot_df.empty:
            bot_df.to_excel(writer, sheet_name='Upload_data_B', index=False)

    print("Sorting completed and sheets created successfully (if applicable).")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    import openpyxl

    # File path
    file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"

    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)
    avl_combine_sheet = wb["AVL_Combine"]

    # Create a new sheet called AVL_UP
    if "AVL_UP" in wb.sheetnames:
        avl_up_sheet = wb["AVL_UP"]
    else:
        avl_up_sheet = wb.create_sheet("AVL_UP")

    # Copy data from AVL_Combine to AVL_UP and modify column headers
    for row_idx, row in enumerate(avl_combine_sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # Replace 'QTY' with 'Q1Z' and 'Side' with 'S1TB' in the header
            new_row = [cell if cell not in ["QTY", "Side"] else ("Q1Z" if cell == "QTY" else "S1TB") for cell in row]
        else:
            # Copy the data as is for non-header rows
            new_row = row
        
        avl_up_sheet.append(new_row)

    # Save the workbook
    wb.save(file_path)
    wb.close()

    print("Data copied and modified successfully, saved in AVL_UP.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    import openpyxl

    # File path
    file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx"

    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Define function to process each sheet
    def process_sheet(upload_data_sheet_name, avl_upload_sheet_name):
        # Check if the Upload_data sheet exists
        if upload_data_sheet_name not in wb.sheetnames:
            print(f"{upload_data_sheet_name} sheet is missing. Skipping.")
            return

        # Load sheets
        upload_data_sheet = wb[upload_data_sheet_name]
        avl_up_sheet = wb["AVL_UP"]

        # Create or get the AVL_Upload sheet
        if avl_upload_sheet_name in wb.sheetnames:
            avl_upload_sheet = wb[avl_upload_sheet_name]
        else:
            avl_upload_sheet = wb.create_sheet(avl_upload_sheet_name)

        # Create a dictionary to store AVL_UP data based on PartNumber
        avl_up_data = {}
        for row in avl_up_sheet.iter_rows(min_row=2, values_only=True):
            part_number = row[0]  # Assuming 'PartNumber' is in the first column
            avl_up_data[part_number] = row

        # Define the new column headers
        new_headers = ["Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number", 
                    "Part Description", "Reference", "QTY", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance",
                        "Side", "ModelName", "PartNumber", "AVLNAME", "PartShapeName", "PartNumberImage", "Q1Z", "S1TB", "AVL"]

        # Write headers to AVL_Upload sheet
        avl_upload_sheet.append(new_headers)

        # Find PartNumber column index in Upload_data
        upload_partnumber_col_idx = None
        for col_idx, col in enumerate(upload_data_sheet.iter_cols(1, upload_data_sheet.max_column, 1, 1), start=1):
            if col[0].value == "PartNumber":
                upload_partnumber_col_idx = col_idx - 1
                break

        if upload_partnumber_col_idx is None:
            print(f"PartNumber column not found in {upload_data_sheet_name}.")
            return

        # Merge data from Upload_data and AVL_UP
        for row in upload_data_sheet.iter_rows(min_row=2, values_only=True):
            upload_row_data = list(row)
            part_number = upload_row_data[upload_partnumber_col_idx]
            
            if part_number in avl_up_data:
                avl_row = avl_up_data[part_number]
                combined_row = upload_row_data + list(avl_row[1:])  # Combine upload_data row with AVL_UP (excluding duplicate PartNumber)
            else:
                combined_row = upload_row_data + [None] * 6  # Fill empty columns for missing AVL_UP data
            
            avl_upload_sheet.append(combined_row)

    # Process each sheet
    process_sheet("Upload_data_T", "AVL_Upload_T")
    process_sheet("Upload_data_B", "AVL_Upload_B")

    # Save the workbook
    wb.save(file_path)
    wb.close()

    print("Data combined and saved in AVL_Upload_T and AVL_Upload_B successfully.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    import pandas as pd
    from openpyxl import load_workbook

    # File path
    file_path = r'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'

    # Load the workbook and check if the required sheets exist
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames

    # Initialize an empty list to hold dataframes
    df_list = []

    # Check and load data from AVL_Upload_T if it exists
    if 'AVL_Upload_T' in sheet_names:
        df_t = pd.read_excel(file_path, sheet_name='AVL_Upload_T')
        df_list.append(df_t)

    # Check and load data from AVL_Upload_B if it exists
    if 'AVL_Upload_B' in sheet_names:
        df_b = pd.read_excel(file_path, sheet_name='AVL_Upload_B')
        df_list.append(df_b)

    # If either sheet is present, combine the data
    if df_list:
        # Concatenate the dataframes
        combined_df = pd.concat(df_list, ignore_index=True)

        # Save the combined data to a new sheet 'AVL_Upload_Combined'
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            combined_df.to_excel(writer, sheet_name='AVL_Upload', index=False)

    print(f"File saved to {file_path} and closed.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #CREATE & SEPRATE FEEDER LOADING LIST DATA

    os.getcwd()

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check')
    Chd = os.getcwd()

    df1 = pd.read_excel("AVL_Polarity_Check.xlsx", sheet_name="AVL_Upload")

    df1.sort_values(by='Side', inplace=True, ascending=True)
    df2_1 = df1
    #del.TOP
    df1 = df1[df1["Side"].str.contains("TOP")==False]
    df2 = df1[df1["ModelName"].str.contains("AIMEX2|AIMEX3|AIMEX-IIIC_2|AIMEX-IIIC_3")==False]
    df2.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df3 = df1[df1["ModelName"].str.contains("NXT|AIMEX3|AIMEX-IIIC_1|AIMEX-IIIC_3")==False]
    df3.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df4 = df1[df1["ModelName"].str.contains("NXT|AIMEX2|AIMEX-IIIC_1|AIMEX-IIIC_2")==False]
    df4.sort_values(by='Feeder Location', inplace=True, ascending=True)
    #del.BOT
    df2_1 = df2_1[df2_1["Side"].str.contains("BOT")==False]
    df2_2 = df2_1[df2_1["ModelName"].str.contains("AIMEX2|AIMEX3|AIMEX-IIIC_2|AIMEX-IIIC_3")==False]
    df2_2.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df2_3 = df2_1[df2_1["ModelName"].str.contains("NXT|AIMEX3|AIMEX-IIIC_1|AIMEX-IIIC_3")==False]
    df2_3.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df2_4 = df2_1[df2_1["ModelName"].str.contains("NXT|AIMEX2|AIMEX-IIIC_1|AIMEX-IIIC_2")==False]
    df2_4.sort_values(by='Feeder Location', inplace=True, ascending=True)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/Upload-Data.xlsx") as writer:
        df2.to_excel(writer, sheet_name="NXT&AMX1_B", index=False)
        df3.to_excel(writer, sheet_name="AMX2_B", index=False)
        df4.to_excel(writer, sheet_name="AMX3_B", index=False)

        df2_2.to_excel(writer, sheet_name="NXT&AMX1_T", index=False)
        df2_3.to_excel(writer, sheet_name="AMX2_T", index=False)
        df2_4.to_excel(writer, sheet_name="AMX3_T", index=False)
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    import os
    import pandas as pd

    # Set working directory
    os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check')

    # Load data and process each sheet
    def process_sheet(sheet_name):
        df = pd.read_excel("Upload-Data.xlsx", sheet_name=sheet_name)
        df["ModuleSide"] = df['Side'].astype(str) + "--" + df['ModelName']
        del df['Side']
        del df['ModelName']
        del df['Q1Z']
        del df['S1TB']
        del df['PartNumber']
        #del df['PartShapeName']
        # Define the desired column order
        columns = [
            "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
            "Part Description", "Reference", "QTY", "AVLNAME", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "PartShapeName", "PartNumberImage",
            "ModuleSide", "AVL"]
        
        df = df[columns]
        return df

    # Process each sheet
    df1 = process_sheet("NXT&AMX1_B")
    df2 = process_sheet("AMX2_B")
    df3 = process_sheet("AMX3_B")
    df4 = process_sheet("NXT&AMX1_T")
    df5 = process_sheet("AMX2_T")
    df6 = process_sheet("AMX3_T")

    # Write the processed data back to the Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/Upload-Data.xlsx") as writer:
        df1.to_excel(writer, sheet_name="NXT&AMX1_B", index=False)
        df2.to_excel(writer, sheet_name="AMX2_B", index=False)
        df3.to_excel(writer, sheet_name="AMX3_B", index=False)
        df4.to_excel(writer, sheet_name="NXT&AMX1_T", index=False)
        df5.to_excel(writer, sheet_name="AMX2_T", index=False)
        df6.to_excel(writer, sheet_name="AMX3_T", index=False)

    print("Sheets processed and saved successfully.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #NXT&AMX1_B

    # Path to FeederSetup.xlsx file
    input_file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/Upload-Data.xlsx"

    # Load the Excel file to get sheet names
    excel_file = pd.ExcelFile(input_file_path)

    # Check if the 'NXT&AMX1_B' sheet exists
    if 'NXT&AMX1_B' in excel_file.sheet_names:
        # Load data from Excel file into a DataFrame
        df = pd.read_excel(input_file_path, sheet_name='NXT&AMX1_B')

        # Check if the DataFrame has at least two rows (data apart from the header)
        if len(df) > 0:
            # Check if the second row is empty
            if df.iloc[0].isnull().all():
                print("No data present in the second row. Skipping processing.")
            else:
                # Replace NaN values with empty string
                df = df.fillna('')

                # Remove "AVL" text and quotes from values in the 'AVL' column
                df['AVL'] = df['AVL'].str.replace('"', '').str.replace('\n', ',')

                # Print to verify data
                print("H")
                print(df)
                print("End")

                # Split the 'AVL' column into individual rows
                df_split = df.set_index([
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "PartShapeName", "PartNumberImage",
                    "ModuleSide"
                ]).apply(lambda x: x.str.split(',').explode()).reset_index()

                # Print to verify split DataFrame
                print("H1")
                print(df_split)
                print("End")

                # Define the desired column order and add new columns
                desired_columns = [
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "AVL", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance",
                    "OCV", "Remarks1", "PartShapeName", "PartNumberImage", "ModuleSide", "Remarks2"
                ]

                # Add new columns with empty values if they don't exist in the DataFrame "LCR-Type", "LCR-Value", "LCR-Tolc", "OCV", "Remarks1", "Remarks2"
                for col in ["LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "OCV", "Remarks1", "Remarks2"]:
                    if col not in df_split.columns:
                        df_split[col] = ''

                # Rearrange the columns in the desired order
                df_split = df_split[desired_columns]

                # Print to verify rearranged DataFrame
                print("H2")
                print(df_split)
                print("End")

                # Identify start cells for merging the 'Feeder Location' column
                startCells = [1]
                for row in range(2, len(df_split) + 1):
                    if df_split.loc[row - 1, 'Feeder Location'] != df_split.loc[row - 2, 'Feeder Location']:
                        startCells.append(row)

                # Create a Pandas Excel writer using XlsxWriter as the engine
                output_excel = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/NXT&AMX1_B.xlsx'
                with pd.ExcelWriter(output_excel, engine='xlsxwriter') as writer:
                    df_split.to_excel(writer, sheet_name='NXT&AMX1_B', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['NXT&AMX1_B']
                    merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 2})

                    lastRow = len(df_split)

                    # Iterate over start cells to merge all columns except 'AVL'
                    for row in startCells:
                        try:
                            endRow = startCells[startCells.index(row) + 1] - 1
                            if row == endRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, endRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, endRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, endRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, endRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, endRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, endRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, endRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, endRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, endRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, endRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, endRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, endRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, endRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, endRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, endRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, endRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, endRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, endRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, endRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, endRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, endRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, endRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                        except IndexError:
                            # Handle cases where row exceeds the last row
                            if row == lastRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, lastRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, lastRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, lastRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, lastRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, lastRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, lastRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, lastRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, lastRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, lastRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, lastRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, lastRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, lastRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, lastRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, lastRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, lastRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, lastRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, lastRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, lastRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, lastRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, lastRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, lastRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, lastRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                    print("Merged data written to Excel successfully!")
        else:
            print("The DataFrame has fewer than two rows. Skipping processing.")
    else:
        print("Sheet 'NXT&AMX1_B' not found in the Excel file.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #AMX2_B
    import pandas as pd

    # Path to FeederSetup.xlsx file
    input_file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/Upload-Data.xlsx"

    # Load the Excel file to get sheet names
    excel_file = pd.ExcelFile(input_file_path)

    # Check if the 'AMX2_B' sheet exists
    if 'AMX2_B' in excel_file.sheet_names:
        # Load data from Excel file into a DataFrame
        df = pd.read_excel(input_file_path, sheet_name='AMX2_B')

        # Check if the DataFrame has at least two rows (data apart from the header)
        if len(df) > 0:
            # Check if the second row is empty
            if df.iloc[0].isnull().all():
                print("No data present in the second row. Skipping processing.")
            else:
                # Replace NaN values with empty string
                df = df.fillna('')

                # Remove "AVL" text and quotes from values in the 'AVL' column
                df['AVL'] = df['AVL'].str.replace('"', '').str.replace('\n', ',')

                # Print to verify data
                print("H")
                print(df)
                print("End")

                # Split the 'AVL' column into individual rows
                df_split = df.set_index([
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "PartShapeName", "PartNumberImage",
                    "ModuleSide"
                ]).apply(lambda x: x.str.split(',').explode()).reset_index()

                # Print to verify split DataFrame
                print("H1")
                print(df_split)
                print("End")

                # Define the desired column order and add new columns
                desired_columns = [
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "AVL", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance",
                    "OCV", "Remarks1", "PartShapeName", "PartNumberImage", "ModuleSide", "Remarks2"
                ]

                # Add new columns with empty values if they don't exist in the DataFrame "LCR-Type", "LCR-Value", "LCR-Tolc", "OCV", "Remarks1", "Remarks2"
                for col in ["LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "OCV", "Remarks1", "Remarks2"]:
                    if col not in df_split.columns:
                        df_split[col] = ''

                # Rearrange the columns in the desired order
                df_split = df_split[desired_columns]

                # Print to verify rearranged DataFrame
                print("H2")
                print(df_split)
                print("End")

                # Identify start cells for merging the 'Feeder Location' column
                startCells = [1]
                for row in range(2, len(df_split) + 1):
                    if df_split.loc[row - 1, 'Feeder Location'] != df_split.loc[row - 2, 'Feeder Location']:
                        startCells.append(row)

                # Create a Pandas Excel writer using XlsxWriter as the engine
                output_excel = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AMX2_B.xlsx'
                with pd.ExcelWriter(output_excel, engine='xlsxwriter') as writer:
                    df_split.to_excel(writer, sheet_name='AMX2_B', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['AMX2_B']
                    merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 2})

                    lastRow = len(df_split)

                    # Iterate over start cells to merge all columns except 'AVL'
                    for row in startCells:
                        try:
                            endRow = startCells[startCells.index(row) + 1] - 1
                            if row == endRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, endRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, endRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, endRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, endRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, endRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, endRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, endRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, endRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, endRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, endRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, endRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, endRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, endRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, endRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, endRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, endRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, endRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, endRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, endRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, endRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, endRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, endRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                        except IndexError:
                            # Handle cases where row exceeds the last row
                            if row == lastRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, lastRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, lastRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, lastRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, lastRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, lastRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, lastRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, lastRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, lastRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, lastRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, lastRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, lastRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, lastRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, lastRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, lastRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, lastRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, lastRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, lastRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, lastRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, lastRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, lastRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, lastRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, lastRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                    print("Merged data written to Excel successfully!")
        else:
            print("The DataFrame has fewer than two rows. Skipping processing.")
    else:
        print("Sheet 'AMX2_B' not found in the Excel file.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #AMX3_B
    import pandas as pd

    # Path to FeederSetup.xlsx file
    input_file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/Upload-Data.xlsx"

    # Load the Excel file to get sheet names
    excel_file = pd.ExcelFile(input_file_path)

    # Check if the 'AMX3_B' sheet exists
    if 'AMX3_B' in excel_file.sheet_names:
        # Load data from Excel file into a DataFrame
        df = pd.read_excel(input_file_path, sheet_name='AMX3_B')

        # Check if the DataFrame has at least two rows (data apart from the header)
        if len(df) > 0:
            # Check if the second row is empty
            if df.iloc[0].isnull().all():
                print("No data present in the second row. Skipping processing.")
            else:
                # Replace NaN values with empty string
                df = df.fillna('')

                # Remove "AVL" text and quotes from values in the 'AVL' column
                df['AVL'] = df['AVL'].str.replace('"', '').str.replace('\n', ',')

                # Print to verify data
                print("H")
                print(df)
                print("End")

                # Split the 'AVL' column into individual rows
                df_split = df.set_index([
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "PartShapeName", "PartNumberImage",
                    "ModuleSide"
                ]).apply(lambda x: x.str.split(',').explode()).reset_index()

                # Print to verify split DataFrame
                print("H1")
                print(df_split)
                print("End")

                # Define the desired column order and add new columns
                desired_columns = [
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "AVL", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance",
                    "OCV", "Remarks1", "PartShapeName", "PartNumberImage", "ModuleSide", "Remarks2"
                ]

                # Add new columns with empty values if they don't exist in the DataFrame "LCR-Type", "LCR-Value", "LCR-Tolc", "OCV", "Remarks1", "Remarks2"
                for col in ["LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "OCV", "Remarks1", "Remarks2"]:
                    if col not in df_split.columns:
                        df_split[col] = ''

                # Rearrange the columns in the desired order
                df_split = df_split[desired_columns]

                # Print to verify rearranged DataFrame
                print("H2")
                print(df_split)
                print("End")

                # Identify start cells for merging the 'Feeder Location' column
                startCells = [1]
                for row in range(2, len(df_split) + 1):
                    if df_split.loc[row - 1, 'Feeder Location'] != df_split.loc[row - 2, 'Feeder Location']:
                        startCells.append(row)

                # Create a Pandas Excel writer using XlsxWriter as the engine
                output_excel = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AMX3_B.xlsx'
                with pd.ExcelWriter(output_excel, engine='xlsxwriter') as writer:
                    df_split.to_excel(writer, sheet_name='AMX3_B', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['AMX3_B']
                    merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 2})

                    lastRow = len(df_split)

                    # Iterate over start cells to merge all columns except 'AVL'
                    for row in startCells:
                        try:
                            endRow = startCells[startCells.index(row) + 1] - 1
                            if row == endRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, endRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, endRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, endRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, endRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, endRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, endRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, endRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, endRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, endRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, endRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, endRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, endRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, endRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, endRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, endRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, endRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, endRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, endRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, endRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, endRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, endRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, endRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                        except IndexError:
                            # Handle cases where row exceeds the last row
                            if row == lastRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, lastRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, lastRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, lastRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, lastRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, lastRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, lastRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, lastRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, lastRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, lastRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, lastRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, lastRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, lastRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, lastRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, lastRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, lastRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, lastRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, lastRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, lastRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, lastRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, lastRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, lastRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, lastRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                    print("Merged data written to Excel successfully!")
        else:
            print("The DataFrame has fewer than two rows. Skipping processing.")
    else:
        print("Sheet 'AMX3_B' not found in the Excel file.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #NXT&AMX1_T
    import pandas as pd

    # Path to FeederSetup.xlsx file
    input_file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/Upload-Data.xlsx"

    # Load the Excel file to get sheet names
    excel_file = pd.ExcelFile(input_file_path)

    # Check if the 'NXT&AMX1_T' sheet exists
    if 'NXT&AMX1_T' in excel_file.sheet_names:
        # Load data from Excel file into a DataFrame
        df = pd.read_excel(input_file_path, sheet_name='NXT&AMX1_T')

        # Check if the DataFrame has at least two rows (data apart from the header)
        if len(df) > 0:
            # Check if the second row is empty
            if df.iloc[0].isnull().all():
                print("No data present in the second row. Skipping processing.")
            else:
                # Replace NaN values with empty string
                df = df.fillna('')

                # Remove "AVL" text and quotes from values in the 'AVL' column
                df['AVL'] = df['AVL'].str.replace('"', '').str.replace('\n', ',')

                # Print to verify data
                print("H")
                print(df)
                print("End")

                # Split the 'AVL' column into individual rows
                df_split = df.set_index([
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "PartShapeName", "PartNumberImage",
                    "ModuleSide"
                ]).apply(lambda x: x.str.split(',').explode()).reset_index()

                # Print to verify split DataFrame
                print("H1")
                print(df_split)
                print("End")

                # Define the desired column order and add new columns
                desired_columns = [
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "AVL", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance",
                    "OCV", "Remarks1", "PartShapeName", "PartNumberImage", "ModuleSide", "Remarks2"
                ]

                # Add new columns with empty values if they don't exist in the DataFrame "LCR-Type", "LCR-Value", "LCR-Tolc", "OCV", "Remarks1", "Remarks2"
                for col in ["LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "OCV", "Remarks1", "Remarks2"]:
                    if col not in df_split.columns:
                        df_split[col] = ''

                # Rearrange the columns in the desired order
                df_split = df_split[desired_columns]

                # Print to verify rearranged DataFrame
                print("H2")
                print(df_split)
                print("End")

                # Identify start cells for merging the 'Feeder Location' column
                startCells = [1]
                for row in range(2, len(df_split) + 1):
                    if df_split.loc[row - 1, 'Feeder Location'] != df_split.loc[row - 2, 'Feeder Location']:
                        startCells.append(row)

                # Create a Pandas Excel writer using XlsxWriter as the engine
                output_excel = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/NXT&AMX1_T.xlsx'
                with pd.ExcelWriter(output_excel, engine='xlsxwriter') as writer:
                    df_split.to_excel(writer, sheet_name='NXT&AMX1_T', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['NXT&AMX1_T']
                    merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 2})

                    lastRow = len(df_split)

                    # Iterate over start cells to merge all columns except 'AVL'
                    for row in startCells:
                        try:
                            endRow = startCells[startCells.index(row) + 1] - 1
                            if row == endRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, endRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, endRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, endRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, endRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, endRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, endRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, endRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, endRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, endRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, endRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, endRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, endRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, endRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, endRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, endRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, endRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, endRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, endRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, endRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, endRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, endRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, endRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                        except IndexError:
                            # Handle cases where row exceeds the last row
                            if row == lastRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, lastRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, lastRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, lastRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, lastRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, lastRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, lastRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, lastRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, lastRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, lastRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, lastRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, lastRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, lastRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, lastRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, lastRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, lastRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, lastRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, lastRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, lastRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, lastRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, lastRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, lastRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, lastRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                    print("Merged data written to Excel successfully!")
        else:
            print("The DataFrame has fewer than two rows. Skipping processing.")
    else:
        print("Sheet 'NXT&AMX1_T' not found in the Excel file.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #AMX2_T
    import pandas as pd

    # Path to FeederSetup.xlsx file
    input_file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/Upload-Data.xlsx"

    # Load the Excel file to get sheet names
    excel_file = pd.ExcelFile(input_file_path)

    # Check if the 'AMX2_T' sheet exists
    if 'AMX2_T' in excel_file.sheet_names:
        # Load data from Excel file into a DataFrame
        df = pd.read_excel(input_file_path, sheet_name='AMX2_T')

        # Check if the DataFrame has at least two rows (data apart from the header)
        if len(df) > 0:
            # Check if the second row is empty
            if df.iloc[0].isnull().all():
                print("No data present in the second row. Skipping processing.")
            else:
                # Replace NaN values with empty string
                df = df.fillna('')

                # Remove "AVL" text and quotes from values in the 'AVL' column
                df['AVL'] = df['AVL'].str.replace('"', '').str.replace('\n', ',')

                # Print to verify data
                print("H")
                print(df)
                print("End")

                # Split the 'AVL' column into individual rows
                df_split = df.set_index([
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "PartShapeName", "PartNumberImage",
                    "ModuleSide"
                ]).apply(lambda x: x.str.split(',').explode()).reset_index()

                # Print to verify split DataFrame
                print("H1")
                print(df_split)
                print("End")

                # Define the desired column order and add new columns
                desired_columns = [
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "AVL", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance",
                    "OCV", "Remarks1", "PartShapeName", "PartNumberImage", "ModuleSide", "Remarks2"
                ]

                # Add new columns with empty values if they don't exist in the DataFrame "LCR-Type", "LCR-Value", "LCR-Tolc", "OCV", "Remarks1", "Remarks2"
                for col in ["LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "OCV", "Remarks1", "Remarks2"]:
                    if col not in df_split.columns:
                        df_split[col] = ''

                # Rearrange the columns in the desired order
                df_split = df_split[desired_columns]

                # Print to verify rearranged DataFrame
                print("H2")
                print(df_split)
                print("End")

                # Identify start cells for merging the 'Feeder Location' column
                startCells = [1]
                for row in range(2, len(df_split) + 1):
                    if df_split.loc[row - 1, 'Feeder Location'] != df_split.loc[row - 2, 'Feeder Location']:
                        startCells.append(row)

                # Create a Pandas Excel writer using XlsxWriter as the engine
                output_excel = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AMX2_T.xlsx'
                with pd.ExcelWriter(output_excel, engine='xlsxwriter') as writer:
                    df_split.to_excel(writer, sheet_name='AMX2_T', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['AMX2_T']
                    merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 2})

                    lastRow = len(df_split)

                    # Iterate over start cells to merge all columns except 'AVL'
                    for row in startCells:
                        try:
                            endRow = startCells[startCells.index(row) + 1] - 1
                            if row == endRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, endRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, endRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, endRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, endRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, endRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, endRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, endRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, endRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, endRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, endRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, endRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, endRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, endRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, endRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, endRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, endRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, endRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, endRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, endRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, endRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, endRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, endRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                        except IndexError:
                            # Handle cases where row exceeds the last row
                            if row == lastRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, lastRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, lastRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, lastRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, lastRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, lastRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, lastRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, lastRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, lastRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, lastRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, lastRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, lastRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, lastRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, lastRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, lastRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, lastRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, lastRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, lastRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, lastRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, lastRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, lastRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, lastRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, lastRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                    print("Merged data written to Excel successfully!")
        else:
            print("The DataFrame has fewer than two rows. Skipping processing.")
    else:
        print("Sheet 'AMX2_T' not found in the Excel file.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    #AMX3_T
    import pandas as pd

    # Path to FeederSetup.xlsx file
    input_file_path = r"D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/Upload-Data.xlsx"

    # Load the Excel file to get sheet names
    excel_file = pd.ExcelFile(input_file_path)

    # Check if the 'AMX3_T' sheet exists
    if 'AMX3_T' in excel_file.sheet_names:
        # Load data from Excel file into a DataFrame
        df = pd.read_excel(input_file_path, sheet_name='AMX3_T')

        # Check if the DataFrame has at least two rows (data apart from the header)
        if len(df) > 0:
            # Check if the second row is empty
            if df.iloc[0].isnull().all():
                print("No data present in the second row. Skipping processing.")
            else:
                # Replace NaN values with empty string
                df = df.fillna('')

                # Remove "AVL" text and quotes from values in the 'AVL' column
                df['AVL'] = df['AVL'].str.replace('"', '').str.replace('\n', ',')

                # Print to verify data
                print("H")
                print(df)
                print("End")

                # Split the 'AVL' column into individual rows
                df_split = df.set_index([
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "PartShapeName", "PartNumberImage",
                    "ModuleSide"
                ]).apply(lambda x: x.str.split(',').explode()).reset_index()

                # Print to verify split DataFrame
                print("H1")
                print(df_split)
                print("End")

                # Define the desired column order and add new columns
                desired_columns = [
                    "Feeder Location", "FeederName", "Type", "Size", "FeedPitch", "Part Height", "Part Number",
                    "Part Description", "Reference", "QTY", "AVLNAME", "AVL", "LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance",
                    "OCV", "Remarks1", "PartShapeName", "PartNumberImage", "ModuleSide", "Remarks2"
                ]

                # Add new columns with empty values if they don't exist in the DataFrame "LCR-Type", "LCR-Value", "LCR-Tolc", "OCV", "Remarks1", "Remarks2"
                for col in ["LCRShape_Type", "LCR_Value", "LCR_Unit", "LCR_Tolerance", "OCV", "Remarks1", "Remarks2"]:
                    if col not in df_split.columns:
                        df_split[col] = ''

                # Rearrange the columns in the desired order
                df_split = df_split[desired_columns]

                # Print to verify rearranged DataFrame
                print("H2")
                print(df_split)
                print("End")

                # Identify start cells for merging the 'Feeder Location' column
                startCells = [1]
                for row in range(2, len(df_split) + 1):
                    if df_split.loc[row - 1, 'Feeder Location'] != df_split.loc[row - 2, 'Feeder Location']:
                        startCells.append(row)

                # Create a Pandas Excel writer using XlsxWriter as the engine
                output_excel = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AMX3_T.xlsx'
                with pd.ExcelWriter(output_excel, engine='xlsxwriter') as writer:
                    df_split.to_excel(writer, sheet_name='AMX3_T', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['AMX3_T']
                    merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 2})

                    lastRow = len(df_split)

                    # Iterate over start cells to merge all columns except 'AVL'
                    for row in startCells:
                        try:
                            endRow = startCells[startCells.index(row) + 1] - 1
                            if row == endRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, endRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, endRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, endRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, endRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, endRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, endRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, endRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, endRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, endRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, endRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, endRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, endRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, endRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, endRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, endRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, endRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, endRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, endRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, endRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, endRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, endRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, endRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                        except IndexError:
                            # Handle cases where row exceeds the last row
                            if row == lastRow:
                                worksheet.write(row, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.write(row, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.write(row, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.write(row, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.write(row, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.write(row, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.write(row, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.write(row, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.write(row, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.write(row, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.write(row, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.write(row, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.write(row, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.write(row, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.write(row, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.write(row, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.write(row, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.write(row, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.write(row, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.write(row, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.write(row, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.write(row, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                            else:
                                worksheet.merge_range(row, 0, lastRow, 0, df_split.loc[row - 1, 'Feeder Location'], merge_format)
                                worksheet.merge_range(row, 1, lastRow, 1, df_split.loc[row - 1, 'FeederName'], merge_format)
                                worksheet.merge_range(row, 2, lastRow, 2, df_split.loc[row - 1, 'Type'], merge_format)
                                worksheet.merge_range(row, 3, lastRow, 3, df_split.loc[row - 1, 'Size'], merge_format)
                                worksheet.merge_range(row, 4, lastRow, 4, df_split.loc[row - 1, 'FeedPitch'], merge_format)
                                worksheet.merge_range(row, 5, lastRow, 5, df_split.loc[row - 1, 'Part Height'], merge_format)
                                worksheet.merge_range(row, 6, lastRow, 6, df_split.loc[row - 1, 'Part Number'], merge_format)
                                worksheet.merge_range(row, 7, lastRow, 7, df_split.loc[row - 1, 'Part Description'], merge_format)
                                worksheet.merge_range(row, 8, lastRow, 8, df_split.loc[row - 1, 'Reference'], merge_format)
                                worksheet.merge_range(row, 9, lastRow, 9, df_split.loc[row - 1, 'QTY'], merge_format)
                                worksheet.merge_range(row, 10, lastRow, 10, df_split.loc[row - 1, 'AVLNAME'], merge_format)
                                #worksheet.merge_range(row, 11, lastRow, 11, df_split.loc[row - 1, 'AVL'], merge_format)
                                worksheet.merge_range(row, 12, lastRow, 12, df_split.loc[row - 1, 'LCRShape_Type'], merge_format)
                                worksheet.merge_range(row, 13, lastRow, 13, df_split.loc[row - 1, 'LCR_Value'], merge_format)
                                worksheet.merge_range(row, 14, lastRow, 14, df_split.loc[row - 1, 'LCR_Unit'], merge_format)
                                worksheet.merge_range(row, 15, lastRow, 15, df_split.loc[row - 1, 'LCR_Tolerance'], merge_format)
                                #worksheet.merge_range(row, 16, lastRow, 16, df_split.loc[row - 1, 'OCV'], merge_format)
                                #worksheet.merge_range(row, 17, lastRow, 17, df_split.loc[row - 1, 'Remarks1'], merge_format)
                                worksheet.merge_range(row, 18, lastRow, 18, df_split.loc[row - 1, 'PartShapeName'], merge_format)
                                worksheet.merge_range(row, 19, lastRow, 19, df_split.loc[row - 1, 'PartNumberImage'], merge_format)
                                worksheet.merge_range(row, 20, lastRow, 20, df_split.loc[row - 1, 'ModuleSide'], merge_format)
                                #worksheet.merge_range(row, 21, lastRow, 21, df_split.loc[row - 1, 'Remarks2'], merge_format)
                    print("Merged data written to Excel successfully!")
        else:
            print("The DataFrame has fewer than two rows. Skipping processing.")
    else:
        print("Sheet 'AMX3_T' not found in the Excel file.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    from openpyxl import load_workbook

    # Load the workbook
    file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'
    wb = load_workbook(file_path)

    # List of sheets to remove
    sheets_to_remove = ['AVL_Combine', 'AVL_Delimited_T', 'AVL_Delimited_B', 'Upload_data_T', 'Upload_data_B', 'AVL_UP', 'AVL_Upload_T', 'AVL_Upload_B','AVL_Stack_T','AVL_Stack_B']

    # Loop through the sheets and remove them if they exist
    for sheet in sheets_to_remove:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            wb.remove(ws)
            print(f"Removed sheet: {sheet}")
        else:
            print(f"Sheet {sheet} not found, skipping...")

    # Save the workbook after removing the sheets
    wb.save(file_path)

    print("Operation completed.")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    import os
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
    from shutil import copyfile

    # Paths and files
    output_dir = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
    database_file = r"D:\NX_BACKWORK\Database_File\SMT_FeederSetup\Line X Sample - AVL.xlsx"
    top_files = ["NXT&AMX1_T.xlsx", "AMX2_T.xlsx", "AMX3_T.xlsx"]
    bottom_files = ["NXT&AMX1_B.xlsx", "AMX2_B.xlsx", "AMX3_B.xlsx"]

    # Define destination file paths
    avl_t_file = os.path.join(output_dir, "Line X Sample - AVL_T.xlsx")
    avl_b_file = os.path.join(output_dir, "Line X Sample - AVL_B.xlsx")

    # Check for top and bottom files
    has_top_files = any([os.path.exists(os.path.join(output_dir, f)) for f in top_files])
    has_bottom_files = any([os.path.exists(os.path.join(output_dir, f)) for f in bottom_files])

    # Copy the database file and rename accordingly
    if has_top_files:
        copyfile(database_file, avl_t_file)
    if has_bottom_files:
        copyfile(database_file, avl_b_file)

    # Function to copy merged cell ranges
    def copy_merged_cells(src_ws, dest_ws, row_offset):
        for merged_range in src_ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            dest_ws.merge_cells(start_row=min_row + row_offset, start_column=min_col,
                                end_row=max_row + row_offset, end_column=max_col)

    # Function to check if a cell is the top-left of a merged range
    def is_top_left_of_merged_range(cell, merged_ranges):
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if cell.row == min_row and cell.column == min_col:
                return True
        return False

    # Function to copy data with row offset and handle merged cells
    def copy_data_with_offset(src_file, dest_file, src_sheet_name, dest_sheet_name, row_offset=5):
        # Load source and destination workbooks
        src_wb = load_workbook(src_file)
        dest_wb = load_workbook(dest_file)
        
        # Get source and destination sheets
        src_ws = src_wb[src_sheet_name]
        dest_ws = dest_wb[dest_sheet_name]
        
        # Step 1: Copy merged cell ranges first
        copy_merged_cells(src_ws, dest_ws, row_offset)
        
        merged_ranges = list(src_ws.merged_cells.ranges)  # Get list of merged ranges
        
        # Step 2: Copy data with row offset
        for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
            for cell in row:
                # Check if the cell is part of a merged range
                if cell.coordinate in src_ws.merged_cells:
                    # Only copy value if the cell is the top-left of the merged range
                    if is_top_left_of_merged_range(cell, merged_ranges):
                        dest_cell = dest_ws.cell(row=cell.row + row_offset, column=cell.column)
                        dest_cell.value = cell.value
                else:
                    # Non-merged cell, copy value directly
                    dest_cell = dest_ws.cell(row=cell.row + row_offset, column=cell.column)
                    dest_cell.value = cell.value
        
        # Save the destination workbook
        dest_wb.save(dest_file)

    # Process TOP files
    if has_top_files:
        top_file_map = {
            "NXT&AMX1_T.xlsx": ("NXT&AMX1_T", "NXT"),
            "AMX2_T.xlsx": ("AMX2_T", "AIMEX 2"),
            "AMX3_T.xlsx": ("AMX3_T", "AIMEX 3"),
        }
        for file, (src_sheet, dest_sheet) in top_file_map.items():
            src_file_path = os.path.join(output_dir, file)
            if os.path.exists(src_file_path):
                copy_data_with_offset(src_file_path, avl_t_file, src_sheet, dest_sheet, row_offset=5)

    # Process BOT files
    if has_bottom_files:
        bottom_file_map = {
            "NXT&AMX1_B.xlsx": ("NXT&AMX1_B", "NXT"),
            "AMX2_B.xlsx": ("AMX2_B", "AIMEX 2"),
            "AMX3_B.xlsx": ("AMX3_B", "AIMEX 3"),
        }
        for file, (src_sheet, dest_sheet) in bottom_file_map.items():
            src_file_path = os.path.join(output_dir, file)
            if os.path.exists(src_file_path):
                copy_data_with_offset(src_file_path, avl_b_file, src_sheet, dest_sheet, row_offset=5)

    print("Data copied successfully!")
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    ################################################################################################################################
    import openpyxl
    import os
    from tqdm import tqdm

    # Paths for T and B files
    output_dir = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
    file_T = os.path.join(output_dir, "Line X Sample - AVL_T.xlsx")
    file_B = os.path.join(output_dir, "Line X Sample - AVL_B.xlsx")

    # Target sheets to process
    TARGET_SHEETS = ["NXT", "AIMEX 2", "AIMEX 3"]

    def hide_empty_rows(file_path):
        if os.path.exists(file_path):
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)
            worksheet_names = workbook.sheetnames

            # Iterate over each targeted sheet in the workbook
            for sheet_name in worksheet_names:
                if sheet_name in TARGET_SHEETS:
                    worksheet = workbook[sheet_name]

                    # Initialize progress bar for rows
                    with tqdm(total=worksheet.max_row, desc=f"Hiding empty rows in {sheet_name}") as pbar:
                        # Iterate over all rows in the worksheet
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                            # Check if the row is empty (all cells are None)
                            if all(cell.value is None for cell in row):
                                # If the row is empty, hide it
                                worksheet.row_dimensions[row[0].row].hidden = True

                            # Update the progress bar
                            pbar.update(1)

            # Save and close the workbook after hiding rows
            workbook.save(file_path)
            workbook.close()
            print(f"Processed and saved: {file_path}")
        else:
            print(f"File does not exist: {file_path}")

    # Process T and B files
    hide_empty_rows(file_T)
    hide_empty_rows(file_B)
    ################################################################################################################################
    ################################################################################################################################
    #ims for images 8th code
    import os
    import PySimpleGUI as sg  
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    '''# Confirmation prompt before running the code
    if sg.popup_yes_no("Do you want to run the IMAGE Excel Processor?") == 'No':
        sg.popup("Operation cancelled.")
        exit()  # Exit the program if the user chooses 'No'.'''

    # GUI Layout for file and directory selection
    layout = [
        [sg.Checkbox('FeederSetup_AVL_B.xlsx', default=True, key='-USE_FILE_B-'),
        sg.Checkbox('FeederSetup_AVL_T.xlsx', default=True, key='-USE_FILE_T-')],
        [sg.Text('Select Image Directory for AVL_B', size=(30, 1)), 
        sg.Input(key='-DIR_B-', size=(60, 1), readonly=True), 
        sg.FolderBrowse(initial_folder=r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup")],
        [sg.Text('Select Image Directory for AVL_T', size=(30, 1)), 
        sg.Input(key='-DIR_T-', size=(60, 1), readonly=True), 
        sg.FolderBrowse(initial_folder=r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup")],
        [sg.Submit(), sg.Cancel()]
    ]

    # Create the window
    window = sg.Window('Excel Processor', layout)

    # Confirmation prompt before running the Excel processing code
    if sg.popup_yes_no("Do you want to run the Excel Processor?") == 'Yes':
        run_excel_processing = True
    else:
        run_excel_processing = False
        sg.popup("Excel processing skipped.")

    # Define headers and column widths
    headers = ['Feeder Location', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Part Number', 'Part Description', 
            'Reference', 'QTY', 'AVLNAME', 'AVL', 'LCR-Type', 'LCR-Value', 'LCR-Tolc', 'OCV', 'Remarks1', 
            'PartShapeName', 'PartNumberImage', 'ModuleSide', 'Remarks2']

    # Define column widths
    column_widths = {
        'A': 25, 'B': 25, 'C': 20, 'D': 15, 'E': 15, 'F': 15, 'G': 30, 'H': 50, 'I': 50, 'J': 10,
        'K': 20, 'L': 20, 'M': 20, 'N': 20, 'O': 20, 'P': 25, 'Q': 30, 'R': 20, 'S': 25, 'T': 40, 'U': 20, 'V':20
    }

    # Headers that should not have their height or width changed
    secondary_headers = ['S.NO', 'Date', 'Revision', 'Description of Changes', 'Revised By', 'Approved By']

    def adjust_column_widths_and_heights(ws, data_start_row, data_end_row):
        """Adjusts column widths and row heights for specific columns while skipping rows with secondary headers."""
        for row in range(data_start_row, data_end_row + 1):
            if row != data_start_row - 1 and row != data_end_row + 1:  # Skip secondary header rows
                ws.row_dimensions[row].height = 100  # Adjust as needed

        # Set column widths
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Apply text wrapping and center alignment for data rows
        for row in ws.iter_rows(min_row=data_start_row, max_row=data_end_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    def insert_images(ws, data, image_directory, part_number_image_col, data_row_start):
        """Inserts images into the 'PartNumberImage' column."""
        for idx, row in data.iterrows():
            image_name = row['PartNumberImage']
            if pd.notna(image_name):
                image_path = os.path.join(image_directory, os.path.basename(image_name))
                if os.path.exists(image_path):
                    img = Image(image_path)
                    img.width, img.height = 210, 130  # Adjust the size as needed
                    cell_location = ws.cell(row=data_row_start + idx, column=part_number_image_col).coordinate
                    ws.add_image(img, cell_location)

    def process_excel_file(file_path, image_dir, skip_rows=5):
        """Processes the Excel file: adjusts formatting, row heights, column widths, and inserts images."""
        try:
            # Load the workbook
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Load data into DataFrame for processing
                df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows)

                # Find the starting and ending rows for data
                data_row_start = 7
                data_row_end = len(df) + data_row_start - 1  # End at the last data row

                # Adjust formatting for rows and columns
                adjust_column_widths_and_heights(ws, data_row_start, data_row_end)

                # Find the column index for 'PartNumberImage'
                part_number_image_col = find_column_by_header(ws, 'PartNumberImage', header_row=6)
                if part_number_image_col:
                    insert_images(ws, df, image_dir, part_number_image_col, data_row_start)

            # Save the workbook
            wb.save(file_path)
            print(f"Processed {file_path} successfully.")
        
        except Exception as e:
            print(f"Error processing {file_path}: {e}")

    def find_column_by_header(ws, header_name, header_row=6):
        """Finds the column index by the header name in the specified header row."""
        for cell in ws[header_row]:
            if cell.value == header_name:
                return cell.column
        return None

    # Event loop for the GUI
    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            break

        # Retrieve user inputs
        use_file_B = values['-USE_FILE_B-']
        use_file_T = values['-USE_FILE_T-']
        dir_B = values['-DIR_B-']
        dir_T = values['-DIR_T-']

        # Paths to Excel files
        file_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check\Line X Sample - AVL_B.xlsx"
        file_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check\Line X Sample - AVL_T.xlsx"

        # Process selected files
        if use_file_B and dir_B:
            process_excel_file(file_B, dir_B)
        if use_file_T and dir_T:
            process_excel_file(file_T, dir_T)

        sg.popup("Operation completed.")
        time.sleep (10)
        window.close()
        time.sleep (30)
        break
    ################################################################################################################################
    ################################################################################################################################
    import os
    import openpyxl
    import shutil

    # Specify the folder containing your Excel files
    excel_folder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
    line_x_folder = os.path.join(excel_folder, "Line_X")

    # Create the Line_X folder if it doesn't exist
    if not os.path.isdir(line_x_folder):
        print('Folder does not exist. Creating now...')
        os.makedirs(line_x_folder)

    # Change the current working directory to the Excel folder
    os.chdir(excel_folder)

    # Check if the files exist
    file_T = 'Line X Sample - AVL_T.xlsx'
    file_B = 'Line X Sample - AVL_B.xlsx'

    workbooks = {}
    if os.path.isfile(file_T):
        workbooks['T'] = openpyxl.load_workbook(file_T)
    else:
        print(f"File not found: {file_T}")

    if os.path.isfile(file_B):
        workbooks['B'] = openpyxl.load_workbook(file_B)
    else:
        print(f"File not found: {file_B}")

    # Specify the worksheet names
    worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

    # Input values for cell B3 and Revision A1
    value_B3 = input("\033[93mEnter Feeder Name: \033[0m").strip()[:12]  # Take only the first 12 characters
    new_name = value_B3[:12]
    Revision = input("\033[93mEnter Revision: \033[0m")
    dLine123 = input("\033[93mEnter Line no: \033[0m")  # Assuming you need this input as well

    # User input for renaming the files
    #new_name = input("\033[93mEnter new name for the Line X Sample files (without extension): \033[0m").strip()

    # Update file names
    new_file_T = f"{new_name} - AVL_T.xlsx"
    new_file_B = f"{new_name} - AVL_B.xlsx"

    # Iterate over each workbook
    for location, workbook in workbooks.items():
        # Iterate over each sheet in the workbook
        for sheet_name in worksheet_names:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Hide empty rows
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                    if all(cell.value is None for cell in row):
                        worksheet.row_dimensions[row[0].row].hidden = True

                # Combine the input values for B3 & K4 cell
                worksheet['B3'] = f"{value_B3} {location} {Revision}"
                worksheet['K4'] = dLine123

    # Save the workbooks with the new names
    if 'T' in workbooks:
        workbooks['T'].save(new_file_T)
    if 'B' in workbooks:
        workbooks['B'].save(new_file_B)

    # Move files to the Line_X folder with new names
    if 'T' in workbooks:
        shutil.move(new_file_T, os.path.join(line_x_folder, new_file_T))
    if 'B' in workbooks:
        shutil.move(new_file_B, os.path.join(line_x_folder, new_file_B))

    print("Process completed.")
    
    sg.Popup('Operation Completed', 'All processes have been successfully completed!')

    sys.exit()  # Exit the program
