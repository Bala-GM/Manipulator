import os
import sys
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
import os
import csv  # Import the standard Python csv module
import xml.etree.ElementTree as ET
from tkinter import filedialog
import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# Program 4: Database inspection interface_GUI/J0124-89P13
def program_4():

    # Function to write the CSV file in the format "Namechanger","B_Part_No"
    def save_csv_with_format(data, csv_filename):
        with open(csv_filename, mode='w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL)  # This will wrap values in quotes
            writer.writerows(data)

    # Load the workbook and the sheet
    file_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['AVL GROUP']

    # Add a new column 'Namechanger' and assign 'Dump1', 'Dump2', etc.
    namechanger_column_index = sheet.max_column + 1
    sheet.cell(row=1, column=namechanger_column_index).value = 'Namechanger'  # Add header

    for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=1):
        sheet.cell(row=idx+1, column=namechanger_column_index).value = f'Dump{idx}'

    # Create another sheet for Priority > 2, with only Namechanger and B_Part_No columns
    if 'AVL_Namechanger' not in wb.sheetnames:
        avl_namechanger_sheet = wb.create_sheet('AVL_Namechanger')
    else:
        avl_namechanger_sheet = wb['AVL_Namechanger']

    # Write header to the AVL_Namechanger sheet
    avl_namechanger_sheet.append(['Namechanger', 'B_Part_No'])

    # Collect data for Priority > 2
    priority_greater_than_01_data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        group, priority, b_part_no = row[0], row[1], row[2]
        namechanger = row[namechanger_column_index - 1]  # Accessing the Namechanger column
        
        if priority > 1:
            avl_namechanger_sheet.append([namechanger, b_part_no])
            priority_greater_than_01_data.append([namechanger, b_part_no])

    # Save the workbook with the new sheet
    wb.save(file_path)

    # Get CSV filename from user input
    csv_filename = input("Please provide a name for the CSV file (without extension): ") + ".csv"
    csv_full_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\\' + csv_filename

    # Save the data in the format "Namechanger","B_Part_No" to a CSV file
    save_csv_with_format(priority_greater_than_01_data, csv_full_path)

    print(f"Data saved successfully to {csv_filename} and in the sheet 'AVL_Namechanger'.")
    
    print("All processes completed successfully!")
    ################################################################################################################################
    # Define paths
    csv_directory = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\PartNumber'
    xlsx_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'

    # Step 1: Find the latest CSV file starting with "NeximPartNumber_"
    csv_files = [f for f in os.listdir(csv_directory) if f.startswith("NeximPartNumber_") and f.endswith(".csv")]
    if not csv_files:
        print("No CSV files found with prefix 'NeximPartNumber_'.")
        exit()

    # Sort CSV files by creation/modification date and select the latest
    latest_csv_file = max([os.path.join(csv_directory, f) for f in csv_files], key=os.path.getctime)
    csv_file_name = os.path.splitext(os.path.basename(latest_csv_file))[0]  # Get the file name without extension

    # Step 2: Read the data from the latest CSV file
    csv_data = pd.read_csv(latest_csv_file)

    # Step 3: Open the Excel workbook and create a sheet with the CSV file name
    wb = openpyxl.load_workbook(xlsx_path)

    # If the sheet name based on the CSV filename exists, remove it and create a new one
    if csv_file_name in wb.sheetnames:
        del wb[csv_file_name]

    ws = wb.create_sheet(title=csv_file_name)

    # Step 4: Write the CSV data, including the header, into the new sheet
    for idx, col in enumerate(csv_data.columns, start=1):
        ws.cell(row=1, column=idx, value=col)  # Write header

    for row_idx, row_data in enumerate(csv_data.values, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Step 5: Save the Excel workbook
    wb.save(xlsx_path)

    print(f"Data from {latest_csv_file} saved successfully into sheet '{csv_file_name}' in {xlsx_path}.")
    ################################################################################################################################
    # Load the workbook and the specific sheet
    xlsx_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb['AVL GROUP']

    # Step 1: Extract the columns in the correct order (Group, Priority, Namechanger, B_Part_No)
    header_row = [cell.value for cell in sheet[1]]  # Get the header row
    columns = ['Group', 'Priority', 'Namechanger', 'B_Part_No']

    # Find the index of each column based on the header row
    column_indices = {column: header_row.index(column) + 1 for column in columns}

    # Create a new sheet to save the rearranged data
    if 'AVL_GROUP_Rearranged' not in wb.sheetnames:
        rearranged_sheet = wb.create_sheet('AVL_GROUP_Rearranged')
    else:
        rearranged_sheet = wb['AVL_GROUP_Rearranged']

    # Write the new headers (with 'B_Part_No' renamed to 'PartNumberName')
    new_headers = ['Group', 'Priority', 'Namechanger', 'PartNumberName']
    rearranged_sheet.append(new_headers)

    # Step 2: Copy data from the old sheet in the new order and rename 'B_Part_No' to 'PartNumberName'
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rearranged_row = [
            row[column_indices['Group'] - 1],        # Group
            row[column_indices['Priority'] - 1],     # Priority
            row[column_indices['Namechanger'] - 1],  # Namechanger
            row[column_indices['B_Part_No'] - 1],    # PartNumberName (formerly B_Part_No)
        ]
        rearranged_sheet.append(rearranged_row)

    # Save the workbook
    wb.save(xlsx_path)

    print(f"Columns rearranged and saved in the sheet 'AVL_GROUP_Rearranged' with 'B_Part_No' renamed to 'PartNumberName'.")
    ################################################################################################################################
    # File path for the Excel workbook
    xlsx_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'

    # Load the workbook
    wb = openpyxl.load_workbook(xlsx_path)

    # Find the sheet that starts with "NeximPartNumber_"
    nexim_sheet_name = None
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("NeximPartNumber_"):
            nexim_sheet_name = sheet_name
            break

    if not nexim_sheet_name:
        print("No sheet found starting with 'NeximPartNumber_'.")
        exit()

    # Read data from both 'AVL_GROUP_Rearranged' and 'NeximPartNumber_' sheet
    sheet_avl = pd.read_excel(xlsx_path, sheet_name='AVL_GROUP_Rearranged')
    sheet_nexim = pd.read_excel(xlsx_path, sheet_name=nexim_sheet_name)

    # Ensure column names are consistent for comparison
    sheet_avl['PartNumberName'] = sheet_avl['PartNumberName'].astype(str)
    sheet_nexim['PartNumberName'] = sheet_nexim['PartNumberName'].astype(str)

    # Perform the lookup (left join) by merging based on 'PartNumberName'
    lookup_result = pd.merge(sheet_avl, sheet_nexim, on='PartNumberName', how='left', indicator=True)

    # Separate the data:
    # 1. Rows where 'PartNumberName' is in 'NeximPartNumber_' but not in 'AVL_GROUP_Rearranged'
    not_present_in_avl = sheet_nexim[~sheet_nexim['PartNumberName'].isin(sheet_avl['PartNumberName'])]

    # 2. Keep the rest of the lookup result and drop the '_merge' column
    present_in_both = lookup_result.drop(columns=['_merge'])

    # Step 1: Write the "not present" data to a new sheet in the workbook
    if 'NeximPartNumber_not_present' in wb.sheetnames:
        del wb['NeximPartNumber_not_present']
    ws_not_present = wb.create_sheet('NeximPartNumber_not_present')

    # Write the header for 'not present' data
    ws_not_present.append(not_present_in_avl.columns.tolist())

    # Write the data rows for 'not present' data
    for row in not_present_in_avl.itertuples(index=False, name=None):
        ws_not_present.append(row)

    # Step 2: Update 'AVL_GROUP_Rearranged' with the lookup results (present in both)
    # Delete existing rows except for the header
    ws_avl = wb['AVL_GROUP_Rearranged']
    ws_avl.delete_rows(2, ws_avl.max_row)

    # Write the header (all original columns and additional ones from NeximPartNumber_)
    ws_avl.append(present_in_both.columns.tolist())

    # Write the data rows for 'present in both'
    for row in present_in_both.itertuples(index=False, name=None):
        ws_avl.append(row)

    # Step 3: Remove the first row from 'AVL_GROUP_Rearranged'
    ws_avl.delete_rows(1)

    # Step 4: Save and close the workbook
    wb.save(xlsx_path)
    wb.close()

    print(f"Process completed. 'AVL_GROUP_Rearranged' updated and non-matching rows saved in 'NeximPartNumber_not_present'.")
    ################################################################################################################################
    # Load the Excel file
    file_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'
    sheet_name = 'AVL_GROUP_Rearranged'

    # Read the sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Ensure the 'Priority' column is treated as numeric
    df['Priority'] = pd.to_numeric(df['Priority'], errors='coerce')

    # Fill missing values in the DataFrame within each 'Group'
    def fill_missing_values(group):
        group = group.copy()
        # Fill missing values based on the first non-null entry in the group
        for column in group.columns:
            if group[column].isnull().any():
                # Forward fill for each column
                group[column] = group[column].fillna(method='ffill')
                # Backward fill in case there are still NaNs after forward fill
                group[column] = group[column].fillna(method='bfill')
        return group

    # Apply the fill_missing_values function to each group
    filled_df = df.groupby('Group').apply(fill_missing_values).reset_index(drop=True)

    # Retain rows with Priority zero and ensure they are filled properly
    priority_zero_df = df[df['Priority'] == 0].copy()
    # Update priority zero rows with filled values based on their group
    for index, row in priority_zero_df.iterrows():
        group_df = filled_df[filled_df['Group'] == row['Group']]
        if not group_df.empty:
            # Use the first non-zero priority row in the same group to fill missing values
            filled_values = group_df.iloc[0]
            for col in df.columns:
                if pd.isna(row[col]):
                    priority_zero_df.at[index, col] = filled_values[col]

    # Combine the filled data with priority zero rows
    final_df = pd.concat([filled_df[filled_df['Priority'] != 0], priority_zero_df])

    # Sort by 'Group' and 'Priority'
    final_df = final_df.sort_values(by=['Group', 'Priority']).reset_index(drop=True)

    # Write the updated DataFrame back to Excel
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        final_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("Data processing complete.")
    ################################################################################################################################
    # Define file path and sheet names
    file_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'
    source_sheet = 'NeximPartNumber_not_present'
    target_sheet = 'AVL_GROUP_Rearranged'

    # Load the Excel file
    with pd.ExcelFile(file_path, engine='openpyxl') as xls:
        # Read data from both sheets
        source_df = pd.read_excel(xls, sheet_name=source_sheet)
        target_df = pd.read_excel(xls, sheet_name=target_sheet)

    # Combine data from source sheet into target sheet
    combined_df = pd.concat([target_df, source_df], ignore_index=True)

    # Optional: Reset 'Priority' numbers if needed
    combined_df['Priority'] = pd.to_numeric(combined_df['Priority'], errors='coerce')
    combined_df['Priority'] = combined_df.groupby('Group').cumcount() + 1

    # Save combined data back to the same sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        combined_df.to_excel(writer, sheet_name=target_sheet, index=False)

    # Load the workbook and remove the source sheet
    wb = load_workbook(file_path)
    if source_sheet in wb.sheetnames:
        wb.remove(wb[source_sheet])

    # Save the workbook with the removed sheet
    wb.save(file_path)

    print("Data has been combined and the sheet 'NeximPartNumber_not_present' has been removed successfully.")
    ################################################################################################################################
     # Define file path and source sheet name
    file_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'
    source_sheet = 'AVL_GROUP_Rearranged'

    # Load the workbook
    wb = load_workbook(file_path)

    # Find the target sheet whose name starts with 'NeximPartNumber_'
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('NeximPartNumber_'):
            target_sheet = sheet_name
            break

    if target_sheet is None:
        raise ValueError("No sheet with name starting with 'NeximPartNumber_' found.")

    # Save the target sheet name for renaming later
    target_sheet_name = target_sheet

    # Remove the target sheet
    wb.remove(wb[target_sheet])

    # Load the source sheet data (AVL_GROUP_Rearranged)
    with pd.ExcelFile(file_path, engine='openpyxl') as xls:
        source_df = pd.read_excel(xls, sheet_name=source_sheet)

    # Drop the specified columns
    columns_to_remove = ['Group', 'Priority', 'Namechanger']
    cleaned_df = source_df.drop(columns=columns_to_remove, errors='ignore')

    # Write the cleaned data back to the AVL_GROUP_Rearranged sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        cleaned_df.to_excel(writer, sheet_name=source_sheet, index=False)

    # Rename the sheet from AVL_GROUP_Rearranged to NeximPartNumber_
    ws = wb[source_sheet]
    ws.title = target_sheet_name

    # Save the workbook after renaming and removing the sheet
    wb.save(file_path)

    print(f"Sheet '{source_sheet}' renamed to '{target_sheet_name}', and file saved successfully.")
    ################################################################################################################################
    # Define file path
    file_path = r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\BOM_List_OP.xlsx'

    # Load the workbook
    wb = load_workbook(file_path)

    # Find the target sheet whose name contains 'NeximPartNumber_'
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if 'NeximPartNumber_' in sheet_name:
            target_sheet = sheet_name
            break

    if target_sheet is None:
        raise ValueError("No sheet with name containing 'NeximPartNumber_' found.")

    # Load the target sheet data into a DataFrame
    with pd.ExcelFile(file_path, engine='openpyxl') as xls:
        df = pd.read_excel(xls, sheet_name=target_sheet)

    # Clean column names by stripping any whitespace
    df.columns = df.columns.str.strip()

    # Check and drop the specified columns
    columns_to_remove = ['Group', 'Priority', 'Namechanger']
    df_cleaned = df.drop(columns=[col for col in columns_to_remove if col in df.columns], errors='ignore')

    # Define the CSV output path using the sheet name
    csv_output_path = os.path.join(os.path.dirname(file_path), f'{target_sheet}.csv')

    # Save the cleaned data to a CSV file with values enclosed in double quotes and headers included
    df_cleaned.to_csv(csv_output_path, index=False, quoting=csv.QUOTE_ALL)

    # Confirmation message
    print(f"Data from '{target_sheet}' has been cleaned and saved to: {csv_output_path}")

    print("All processes completed successfully!")

    # Popup message to notify that the operation is complete
    sg.Popup('Operation Completed', 'All processes have been successfully completed.')

    # Exit the program after completion
    sys.exit()