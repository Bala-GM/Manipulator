import os
import sys
import shutil
import xml.etree.ElementTree as ET
import pandas as pd
from tkinter import filedialog
import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# Program 7: Database inspection interface_GUI/J0124-89P13
def program_7():

    # Define the root directory
    root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

    # Function to find all FeederSetup.xml files
    def find_feeder_setup_files(root_directory):
        feeder_setup_files = []
        for root, dirs, files in os.walk(root_directory):
            for file in files:
                if file == "FeederSetup.xml":
                    feeder_setup_files.append(os.path.join(root, file))
        return feeder_setup_files

    # Function to find the setup description from the file path
    def find_setup_description(file_path):
        parts = file_path.split(";")
        setup_description = parts[-2].strip()
        return setup_description

    # Function to rename FeederSetup.xml with a description-based name
    def rename_feeder_setup_with_description(file_path):
        setup_description = find_setup_description(file_path)
        if "[Top] Line1" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line1" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        elif "[Top] Line2" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line2" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        elif "[Top] Line3" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line3" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        elif "[Top] Line4" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line4" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        elif "[Top] Line-1" in setup_description:
            new_file_name = "FeederSetup_AVL_T.xml"
        elif "[Bottom] Line-1" in setup_description:
            new_file_name = "FeederSetup_AVL_B.xml"
        else:
            return None
        
        new_path = os.path.join(os.path.dirname(file_path), new_file_name)
        os.rename(file_path, new_path)
        print(f"FeederSetup.xml renamed to: {new_path}")
        return new_path

    # Function to convert XML to DataFrame
    def xml_to_dataframe(xml_file):
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        data = []
        for feeder in root.findall('Feeder'):
            feeder_dict = {}
            for child in feeder.iter():
                if child.tag != 'Feeder':
                    if child.text:
                        if child.tag == 'AVL':
                            feeder_dict[child.tag] = '"' + child.text.strip() + '"'
                        else:
                            feeder_dict[child.tag] = child.text.strip()
            if feeder_dict:  # Only append if the dictionary is not empty
                data.append(feeder_dict)
        
        df = pd.DataFrame(data)
        
        # Remove rows where PartNumber is blank
        df = df[df['PartNumber'].notna() & (df['PartNumber'].str.strip() != '')]

        # Sort DataFrame by PartNumber column
        df.sort_values(by=['PartNumber'], inplace=True)
        
        return df

    # Function to convert DataFrame to Excel file
    def dataframe_to_excel(df, excel_file):
        writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
        df.to_excel(writer, index=False)
        writer.close()

    # Function to process the Excel file and keep specified columns
    def process_excel_file(file_path):
        df = pd.read_excel(file_path)
        
        # Keep only the specified columns in the given order
        columns_to_keep = ['PartNumber', 'AVLNAME', 'AVL', 'PartShapeName', 'PartNumberImage', 'QTY']
        df = df[columns_to_keep]
        
        # Add the Side column based on the file name
        side = 'BOT' if 'AVL_B' in file_path else 'TOP'
        df['Side'] = side
        
        # Save the modified DataFrame back to Excel
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
        df.to_excel(writer, index=False)
        writer.close()

    # Processing FeederSetup files
    feeder_setup_files = find_feeder_setup_files(root_directory)
    for feeder_setup_file in feeder_setup_files:
        # Rename XML file based on description
        renamed_xml_file = rename_feeder_setup_with_description(feeder_setup_file)
        
        # Define file paths
        input_file = renamed_xml_file if renamed_xml_file else 'FeederSetup.xml'
        destination_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
        
        if os.path.exists(input_file):
            # Convert XML to DataFrame
            df = xml_to_dataframe(input_file)
            
            # Convert DataFrame to Excel
            output_excel = os.path.splitext(input_file)[0] + '.xlsx'
            dataframe_to_excel(df, output_excel)
            
            # Move the Excel file to the destination directory
            shutil.move(output_excel, destination_directory)
            
            # Process the Excel file to keep specific columns and add the Side column
            final_excel_path = os.path.join(destination_directory, os.path.basename(output_excel))
            process_excel_file(final_excel_path)
        else:
            print(f"File not found: {input_file}. Skipping...")

    # Additional functions to write data to Excel with images and formatting
    def extract_data(input_file, side):
        part_numbers = []
        extracted_data = []

        df = pd.read_excel(input_file)

        for index, row in df.iterrows():
            current_data = {
                "PartNumber": row['PartNumber'],
                "AVLNAME": row['AVLNAME'],
                "AVL": row['AVL'],
                "PartShapeName": row['PartShapeName'],
                "PartNumberImage": row['PartNumberImage'],
                "QTY": row['QTY'],
                "Side": side
            }
            extracted_data.append(current_data)
        
        return extracted_data

    #Default paths for the input files
    default_path_B= r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.xlsx"
    default_path_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.xlsx"
    default_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

    # Create a simple GUI for file and directory selection
    layout = [
        [sg.Checkbox('FeederSetup_AVL_B.xlsx', default=True, key='-USE_FILE_B-'), sg.Checkbox('FeederSetup_AVL_T.xlsx', default=True, key='-USE_FILE_T-')],
        [sg.Text('Select Image Directory for AVL_B', size=(30, 1)), sg.Input(key='-DIR_B-', size=(60, 1), readonly=True), sg.FolderBrowse(initial_folder=default_directory)],
        [sg.Text('Select Image Directory for AVL_T', size=(30, 1)), sg.Input(key='-DIR_T-', size=(60, 1), readonly=True), sg.FolderBrowse(initial_folder=default_directory)],
        [sg.Submit(), sg.Cancel()]
    ]

    window = sg.Window('Feeder Setup Data Extractor', layout)

    event, values = window.read()
    window.close()

    if event == 'Submit':
        use_file_B = values['-USE_FILE_B-']
        use_file_T = values['-USE_FILE_T-']
        image_directory_B = values['-DIR_B-']
        image_directory_T = values['-DIR_T-']
        
        if not (use_file_B or use_file_T):
            sg.popup("Error: At least one of FeederSetup_AVL_B.xlsx or FeederSetup_AVL_T.xlsx must be selected.")
        else:
            extracted_data = []
            
            if use_file_B:
                file_path_B = default_path_B
                side_B = 'BOT'
                extracted_data.extend(extract_data(file_path_B, side_B))
            
            if use_file_T:
                file_path_T = default_path_T
                side_T = 'TOP'
                extracted_data.extend(extract_data(file_path_T, side_T))

            # Check if at least one file is selected
            if not (use_file_B or use_file_T):
                sg.popup('Please select at least one file to process.')
            else:
                output_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
                output_excel = os.path.join(output_directory, 'FeederSetup_Data.xlsx')

            # Extract data from the selected files
            data_B = extract_data(default_path_B, "BOT") if use_file_B and os.path.exists(default_path_B) else []
            data_T = extract_data(default_path_T, "TOP") if use_file_T and os.path.exists(default_path_T) else []
            
            if extracted_data:
                df = pd.DataFrame(extracted_data)
                
                #Create the path
                os.getcwd()
                Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
                Chd = os.getcwd()

                yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"

                if not os.path.isdir(yourfolder):
                    print('Folder Not Exist')
                    os.makedirs(yourfolder)
                
                output_excel = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'
                writer = pd.ExcelWriter(output_excel, engine='openpyxl')
                df.to_excel(writer, sheet_name="AVLPOL", index=False)
                writer.close()
        
                wb = load_workbook(output_excel)
                ws = wb.active
                
                table = Table(displayName="FeederSetupTable", ref=ws.dimensions)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                table.tableStyleInfo = style
                ws.add_table(table)
                
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.font = header_font
                
                alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = alignment

                # Load or create Excel workbook
                if os.path.exists(output_excel):
                    wb = load_workbook(output_excel)
                else:
                    wb = Workbook()

                if 'FeederSetup_AVL_B' in wb.sheetnames:
                    ws_B = wb['FeederSetup_AVL_B']
                else:
                    ws_B = wb.create_sheet(title='FeederSetup_AVL_B')

                if 'FeederSetup_AVL_T' in wb.sheetnames:
                    ws_T = wb['FeederSetup_AVL_T']
                else:
                    ws_T = wb.create_sheet(title='FeederSetup_AVL_T')
            
                # Function to write data to a given worksheet
                def write_data_to_sheet(ws, data, image_directory, table_name):
                    # Set default row height
                    for row in range(1, len(data) + 2):  # +2 to account for header row and start index
                        ws.row_dimensions[row].height = 100

                    # Set column widths (old settings)
                    ws.column_dimensions['A'].width = 25
                    ws.column_dimensions['B'].width = 25
                    ws.column_dimensions['C'].width = 100
                    ws.column_dimensions['D'].width = 25
                    ws.column_dimensions['E'].width = 30
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 15

                    # Write headers
                    headers = ['PartNumber', 'AVLNAME', 'AVL', 'PartShapeName', 'PartNumberImage', 'QTY','Side']
                    ws.append(headers)

                    # Format headers
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Write data rows
                    for row in data:
                        # Ensure AVL and AVLNAME fields are present and properly quoted for Excel format
                        if 'AVL' not in row:
                            row["AVL"] = ""
                        else:
                            row["AVL"] = f"'{row['AVL']}'"
                        
                        if 'AVLNAME' not in row:
                            row["AVLNAME"] = ""
                        
                        ws.append([row["PartNumber"], row["AVLNAME"], row["AVL"], row["PartShapeName"], row["PartNumberImage"], row["QTY"], row["Side"]])

                    # Insert images into the Excel file
                    for index, row in enumerate(data, start=2):  # start=2 to account for header row
                        image_path = os.path.join(image_directory, os.path.basename(row["PartNumberImage"]))
                        if os.path.exists(image_path):
                            img = Image(image_path)
                            img.height = 130  # Adjust image height if necessary
                            img.width = 210   # Adjust image width if necessary
                            img.anchor = f"E{index}"
                            ws.add_image(img, f"E{index}")  # Adjust the column and row as per your requirement
                            ws[f"A{index}"].alignment = Alignment(horizontal='center', vertical='center')  # Center align the image column
                            ws[f"B{index}"].alignment = Alignment(horizontal='center', vertical='center')  # Center align the image column
                            ws[f"C{index}"].alignment = Alignment(horizontal='center', vertical='center')  # Center align the image column
                            ws[f"D{index}"].alignment = Alignment(horizontal='center', vertical='center')  # Center align the image column
                            #ws[f"E{index}"].alignment = Alignment(horizontal='center', vertical='center')  # Center align the image column
                            ws[f"F{index}"].alignment = Alignment(horizontal='center', vertical='center')  # Center align the image column
                            ws[f"G{index}"].alignment = Alignment(horizontal='center', vertical='center')  # Center align the image column
                            
                    # Create table for better formatting
                    table = Table(displayName=table_name, ref=f"A1:G{len(data) + 1}")
                    style = TableStyleInfo(
                        name="TableStyleMedium9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=True
                    )
                    table.tableStyleInfo = style
                    ws.add_table(table)

                # Write data to respective sheets
                if data_B:
                    write_data_to_sheet(ws_B, data_B, image_directory_B, "Table_B")
                if data_T:
                    write_data_to_sheet(ws_T, data_T, image_directory_T, "Table_T")

                # Save the Excel file
                wb.save(output_excel)
                
                sg.popup(f"Data extracted and saved to {output_excel}")
            else:
                sg.popup("No data extracted.")
    else:
        sg.popup("Operation cancelled.")

        sys.exit()

    program_7()

#pyinstaller -F -i "AVL.ico" --noconsole AVL-Checker.py