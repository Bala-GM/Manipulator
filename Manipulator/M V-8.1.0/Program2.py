import pandas as pd
import numpy as np
import os
from os import getcwd
import csv as csv
from io import StringIO
import re
import datetime
from io import BytesIO
import time
import sys
from datetime import datetime
import shutil
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import subprocess
import threading
import urllib
import urllib.parse
from sqlite3 import dbapi2 as sqlite
import sqlite3
import lxml
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
import PySimpleGUI as sg
import sqlalchemy
from sqlalchemy import create_engine
import sqlalchemy_access as sq_a
import sqlalchemy_pyodbc_mssql as sqlalchemy
from flask_sqlalchemy import SQLAlchemy
import pyodbc
import pypyodbc
import odbc
import psycopg2
import mysql.connector as sql
import MySQLdb as sql #pip install mysqlclient
from plyer import notification
import xlrd
import xml.etree.ElementTree as ET
from openpyxl.utils.exceptions import InvalidFileException
import psutil

# Program 2: BOM Manipulation
def program_2():

    print('\n')
    print("\033[92;4m*******BOM Manipulation--PY_V-1.3 interface_GUI/A2824-89P13*******\033[0m")
    print('\n')

    # Get the current date and time
    current_datetime = datetime.now()

    # Format the current date and time as a string
    #formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
    # Format the date and time in a 12-hour clock with AM/PM
    formatted_datetime = current_datetime.strftime("%Y-%m-%d %I:%M:%S %p")

    # Print the formatted date and time
    print(f"\033[31mCurrent Date and Time: {formatted_datetime}\033[0m") #\033[0;31m
    print('\n')

    '''print(f"Current Year: {current_datetime.year}")
    print(f"Current Month: {current_datetime.month}")
    print(f"Current Day: {current_datetime.day}")
    print(f"Current Hour: {current_datetime.hour}")
    print(f"Current Minute: {current_datetime.minute}")
    print(f"Current Second: {current_datetime.second}")'''


    dLbr1 = input("\033[93mEnter BOM Name :\033[0m")
    print('\n')

    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        if os.path.exists("BOM_List_OP.xlsx"):
            os.remove("BOM_List_OP.xlsx")
    else:
        print("The file does not exist")

    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        if os.path.exists("BOM_List_OP.xlsx"):
            os.remove("BOM_List_OP.xlsx")
    else:
        print("The file does not exist")

    try:
        # BOM MANIPULATION
        os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
        file_path = 'BOM.xlsx'

        if os.path.isfile(file_path):
            ds1 = pd.read_excel(file_path, index_col=False)
        else:
            # Try reading as '.xls' format if '.xlsx' fails
            file_path = 'BOM.xls'
            ds1 = pd.read_excel(file_path, index_col=False)

        dfbom1 = ds1

    except ValueError:
        dfbom1 = pd.read_excel(file_path,index_col=False) 

    except Exception as e:
        # Handle the exception gracefully
        error_message = f"An error occurred: {e}"

        # Show error message in a pop-up box
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    # Define your column lists
    column_list_1 = ['Material', 'AltItemGroup', 'Priority', 'Long. Description', 'Ref.Designator/Circuit Reference', 'Quantity', 'Material Group']
    column_list_2 = ['Internal P/N', 'Group', 'Priority', 'Description', 'Ref.Designator', 'Qty', 'SMT/THT/Mech']

    # Check which column list is present in the DataFrame
    if all(column in ds1.columns for column in column_list_1):
        columns_to_use = column_list_1
    elif all(column in ds1.columns for column in column_list_2):
        columns_to_use = column_list_2
    else:
        # Show error message if none of the column lists is present
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        missing_columns = [column for column_list in [column_list_1, column_list_2] for column in column_list if column not in ds1.columns]
        error_message = f"The following columns are missing: {', '.join(missing_columns)}"
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    # Continue with the rest of your code using 'columns_to_use'
    print(f"Using columns: {columns_to_use}")

    # Rest of your code here
    # ...

    ds1.rename(
        columns={'Material':"PartNumber", 'AltItemGroup':"Group", 'Priority':'Priority', 'Long. Description':'Long Des', 'Ref.Designator/Circuit Reference':'RefList', 'Quantity':'Qty','Material Group':'Shape'},
        inplace=True,
    )

    ds1.rename(
        columns={'Internal P/N':"PartNumber", 'Group':"Group", 'Priority':'Priority', 'Description':'Long Des', 'Ref.Designator':'RefList', 'Qty':'Qty','SMT/THT/Mech':'Shape'},
        inplace=True,
    )

    print(ds1)

    ds2 = ds1[ds1['Priority'].isin([0, 1])]

    # Assuming ds2 is your DataFrame and 'PartNumber' and 'RefList' are the columns you want to check
    part_number_column = ds2['PartNumber']
    ref_list_column = ds2['RefList']

    # Flag to check if an empty value is found
    empty_value_found = False

    # Iterate through both columns simultaneously using iterrows
    for index, (part_number_value, ref_list_value) in ds2[['PartNumber', 'RefList']].iterrows():
        # Check if the 'RefList' value is empty (NaN or None)
        if pd.isna(ref_list_value):
            print(f"Error: Empty value found in 'RefList' for 'PartNumber' {part_number_value}")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = f"Empty value found in 'RefList' for 'PartNumber' {part_number_value}. Program will stop."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code
            #raise ValueError("Empty value found in 'RefList'")
            #empty_value_found = True
            #break  # Stop the iteration when the first empty value is found

    # If no empty values are found, print the 'PartNumber' column
    if not empty_value_found:
        print(part_number_column)
        # Continue with the rest of your program

    #file_name ="output.xlsx"
    #ds1.to_excel(file_name, index=False)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM/BOM_List_OP.xlsx") as writer:
        ds1.to_excel(writer, sheet_name="Orginal_BOM", index=False)
        ds2.to_excel(writer, sheet_name="BOM", index=False)
        ds1.to_excel(writer, sheet_name="Orginal_BOM_SL", index=False)
        ds2.to_excel(writer, sheet_name="BOM_SL", index=False)

        pass
        print('The file does not exist.')

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
    Chd = os.getcwd()
    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        ds1 = pd.read_excel(file_path, sheet_name="BOM", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False) 
        dsn1 = pd.read_excel(file_path, sheet_name="Orginal_BOM", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False)
        dsplcr1 = pd.read_excel(file_path, sheet_name="BOM_SL", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False) 
        dnsplcr1 = pd.read_excel(file_path, sheet_name="Orginal_BOM_SL", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False)

        ds1 = ds1[['PartNumber', 'Group','Priority','Long Des','Qty','Shape','RefList']]
        ds1['RefList'] = ds1['RefList'].str.replace("_x000D_","")
        ds1['RefList'] = ds1['RefList'].str.replace(" ","")
        ds1['RefList'] = ds1['RefList'].str.replace("\n","")
        
        # Create a new column 'AVL_Name' based on the condition in 'Priority' column
        ds1['AVL_Name'] = ds1.apply(lambda row: row['PartNumber'] if row['Priority'] == 1 else '', axis=1)
        
        print(ds1)

    #ds1 = pd.read_excel('Filename_OP.xlsx','BOM', index_col=False)

        ds2 = ds1.explode('RefList')

        ds2['RefList'] = ds2['RefList'].str.replace(" "," ")

    #ds2.drop(ds2.iloc[:, 1:6], inplace=True, axis=1)

        print(ds2)

        ds2.rename(columns = {'PartNumber':'B_Part_No'}, inplace = True)

        ds2.rename(columns = {'RefList':'B_Ref_List'}, inplace = True)

        ds2['B_Ref_List'] = ds2['B_Ref_List'] .str.strip('[]').str.split(',')

        print(ds2)

        ds2.to_dict()

        ds2.explode ('B_Ref_List',ignore_index=True)

        ds3 = ds2.explode('B_Ref_List',ignore_index=True) # split the Ref below example code

        ds2 = ds2[['Group','Priority','B_Part_No']]
        dc1 = ds2[['B_Part_No']]
        dc1.rename(columns = {'B_Part_No':'PBARNO'}, inplace = True)
        dc1['PBARPTN'] = dc1['PBARNO']
        dc1['PBARBAR'] = dc1['PBARNO']
        dc1.insert(3,'PBARQTY', 10000)
        dc1.insert(4,'PBARFTYP', 3)

    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>#
        dfs2 = ds2[['Group','Priority','B_Part_No']]
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("15","A")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("14","B")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("13","C")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("12","D")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("11","E")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("10","F")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("9","PTN_9")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("8","PTN_8")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("7","PTN_7")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("6","PTN_6")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("5","PTN_5")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("4","PTN_4")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("3","PTN_3")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("2","PTN_2")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("1","PTN_1")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("0","PTN_0")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("A","PTN_15")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("B","PTN_14")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("C","PTN_13")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("D","PTN_12")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("E","PTN_11")
        dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("F","PTN_10")
        dfs2.dropna(subset=['Group'], inplace=True)
        #df2 = dfs2.pivot(index='Group',columns='Priority',values='B_Part_No')

        # Assuming 'dfs2' is the DataFrame with 'Group', 'Priority', and 'B_Part_No' columns
        # Check for duplicate entries in 'Group' and 'Priority'
        duplicate_entries = dfs2[dfs2.duplicated(subset=['Group', 'Priority'], keep=False)]

        if not duplicate_entries.empty:
            # Show an error message if duplicates are found
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = f"Duplicate entries found in 'Group' and 'Priority':\n{duplicate_entries}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        # If no duplicates, proceed with pivoting
        try:
            df2 = dfs2.pivot(index='Group', columns='Priority', values='B_Part_No')
        except ValueError as e:
            # Handle the exception gracefully
            error_message = f"An error occurred during pivoting: {e}"
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        ds3.head()

        T10_col = ds3.pop('B_Ref_List') # col-1

        ds3.insert(0, 'B_Ref_List', T10_col)

        ds3 = ds3[['B_Ref_List','B_Part_No','Long Des']]

        ds1.dropna(subset=['RefList'], inplace=True)
        ds3.dropna(subset=['B_Ref_List'], inplace=True)

        #dsco1 = ds1['PartNumber'].value_counts().reset_index()
            #OUTPUT
        '''``````````````````````````````````````````````
        Total count of XY side in the "S" column: S
        TOP    930
        BOT    791
        Name: count, dtype: int64
        ``````````````````````````````````````````````'''
        #dsco1 = ds1['PartNumber'].value_counts()
        '''``````````````````````````````````````````````
        Total count of XY DATA side in the "S" column:      S  count
        0  TOP    930
        1  BOT    791
        ``````````````````````````````````````````````'''  
        dsco2 = ds1['Qty'].sum()
        dsco3 = len(ds3['B_Ref_List'])
        print(f'Total count of Qty in the "B_Ref_List" column SUM: {dsco2}')
        print(f'Total count of rows in the "B_Ref_List" column: {dsco3}')
        
        ds3['Bom Ref'] = ds3['B_Ref_List']

    #ONLY AVL PARTMASTER AND GOUPING
        dsn1 = dsn1[['PartNumber', 'Group','Priority','Long Des','Qty','Shape','RefList']]
        dsn1['RefList'] = dsn1['RefList'].str.replace("_x000D_","")
        dsn1['RefList'] = dsn1['RefList'].str.replace(" ","")
        dsn1['RefList'] = dsn1['RefList'].str.replace("\n","")
        print(dsn1)

    #ds1 = pd.read_excel('Filename_OP.xlsx','BOM', index_col=False)

        dsn2 = dsn1.explode('RefList')

        dsn2['RefList'] = dsn2['RefList'].str.replace(" "," ")

        #ds2.drop(ds2.iloc[:, 1:6], inplace=True, axis=1)

        print(dsn2)

        dsn2.rename(columns = {'PartNumber':'B_Part_No'}, inplace = True)

        dsn2.rename(columns = {'RefList':'B_Ref_List'}, inplace = True)

        dsn2['B_Ref_List'] = dsn2['B_Ref_List'] .str.strip('[]').str.split(',')

        dsn2.to_dict()

        dsn2.explode ('B_Ref_List',ignore_index=True)

        dsn3 = dsn2.explode('B_Ref_List',ignore_index=True)

        dsn2 = dsn2[['Group','Priority','B_Part_No']]
        
        # Continue with your code if no duplicates are found
        dcn1 = dsn2[['B_Part_No']]
        duplicate_rows = dcn1[dcn1.duplicated(subset=['B_Part_No'], keep=False)]
        
        if not duplicate_rows.empty:
            # Show an error message if duplicates are found
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = f"Duplicate entries found in 'B_Part_No':\nCheck the BOM! PartNo Col.\n{duplicate_rows}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code
        
        print(dcn1)
        dcn1.rename(columns = {'B_Part_No':'PBARNO'}, inplace = True)
        dcn1['PBARPTN'] = dcn1['PBARNO']
        dcn1['PBARBAR'] = dcn1['PBARNO']
        dcn1.insert(3,'PBARQTY', 10000)
        dcn1.insert(4,'PBARFTYP', 3)

        #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>#
        dfsn2 = dsn2[['Group','Priority','B_Part_No']]
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("15","A")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("14","B")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("13","C")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("12","D")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("11","E")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("10","F")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("9","PTN_9")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("8","PTN_8")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("7","PTN_7")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("6","PTN_6")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("5","PTN_5")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("4","PTN_4")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("3","PTN_3")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("2","PTN_2")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("1","PTN_1")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("0","PTN_0")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("A","PTN_15")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("B","PTN_14")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("C","PTN_13")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("D","PTN_12")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("E","PTN_11")
        dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("F","PTN_10")
        dfsn2.dropna(subset=['Group'], inplace=True)
        #dfn2 = dfsn2.pivot(index='Group',columns='Priority',values='B_Part_No')

            # Assuming 'dfs2' is the DataFrame with 'Group', 'Priority', and 'B_Part_No' columns
        # Check for duplicate entries in 'Group' and 'Priority'
        duplicate_entries = dfsn2[dfsn2.duplicated(subset=['Group', 'Priority'], keep=False)]

        if not duplicate_entries.empty:
            # Show an error message if duplicates are found
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = f"Duplicate entries found in 'Group' and 'Priority':\n{duplicate_entries}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        # If no duplicates, proceed with pivoting
        try:
            dfn2 = dfsn2.pivot(index='Group', columns='Priority', values='B_Part_No')
        except ValueError as e:
            # Handle the exception gracefully
            error_message = f"An error occurred during pivoting: {e}"
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        # Desired column order
        desired_order = ['Group Name','PTN_1','PTN_2','PTN_3','PTN_4','PTN_5','PTN_6','PTN_7','PTN_8','PTN_9','PTN_10','PTN_11','PTN_12','PTN_13','PTN_14','PTN_15']
        #desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10','PTN_11','P_11','PTN_12','P_12','PTN_13','P_13','PTN_14','P_14','PTN_15','P_15']

        # Create a list of columns present in both DataFrame and desired_order
        common_columns = [col for col in desired_order if col in dfn2.columns]

        # Reorder the DataFrame based on the desired_order
        df_AL1 = dfn2[common_columns]

        '''    # Assuming df is your DataFrame
            column_to_check = 'PTN_15'

            # Check if the column is present
            if column_to_check not in df_AL1.columns:
                # Show a pop-up message if the column is not present
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showinfo("Notification", f"The column '{column_to_check}' is missing.")

            # Continue with the rest of your code
            print("Continuing with the rest of the code...")
            # Your next line of code here'''

        # Assuming df is your DataFrame
        column_to_check = 'PTN_11'

        # Check if the column is present
        if column_to_check in df_AL1.columns:
            # Show a pop-up message if the column is present
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showinfo("Notification", f"The column '{column_to_check}' is present.")

        # Continue with the rest of your code
        print("Continuing with the rest of the code...")
        # Your next line of code here

        dsn3.head()

        T10_col = dsn3.pop('B_Ref_List') # col-1

        dsn3.insert(0, 'B_Ref_List', T10_col)

        dsn3 = dsn3[['B_Ref_List','B_Part_No','Long Des']]

        dsn1.dropna(subset=['RefList'], inplace=True)
        dsn3.dropna(subset=['B_Ref_List'], inplace=True)

        yourfolder4 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified"

        if not os.path.isdir(yourfolder4):
            print('Folder Not Exist')
            os.makedirs(yourfolder4)

    #########################################################################################################################################################################
        print('\n')
        print("\033[92;4m*******XY Data Manipulation*******\033[0m")
        print('\n')
    #########################################################################################################################################################################

        try:
            # XY Data MANIPULATION
            os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
            file_path_xlsx = 'BOM.xlsx'
            file_path_xls = 'BOM.xls'

            if os.path.isfile(file_path_xlsx):
                dsxy1 = pd.read_excel(file_path_xlsx, sheet_name="XY DATA", usecols=['R', 'X', 'Y', 'A', 'S'], index_col=False)
            elif os.path.isfile(file_path_xls):
                dsxy1 = pd.read_excel(file_path_xls, sheet_name="XY DATA", usecols=['R', 'X', 'Y', 'A', 'S'], index_col=False)
            else:
                root = tk.Tk()
                root.withdraw()
                error_message = "BOM file not found. Please check the file path."
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

            dfXY1 = dsxy1

            # Define your column lists
            column_list_1 = ['R', 'X', 'Y', 'A', 'S']

            # Check which column list is present in the DataFrame
            if all(column in dsxy1.columns for column in column_list_1):
                columns_to_use = column_list_1
            else:
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in column_list_1 if column not in dsxy1.columns]
                error_message = f"The following columns are missing: {', '.join(missing_columns)}"
                error_msgbm1 = "The following columns are missing in BOM (EXCEL) File Sheet 'XY DATA':\n'Reference CRD as R'\n'X Coordinate as X'\n'Y Coordinate as Y'\n'Angle as A'\n'Side as S'"
                messagebox.showerror("Error", error_message)
                messagebox.showerror("Error", error_msgbm1)
                sys.exit(1)  # Exit the program with an error code

            dfXY1['R'] = dfXY1['R'].str.replace(" ","")
            dfXY1.rename(columns={'R': 'B_Ref_List'}, inplace=True)
            dfXY1['R'] = dfXY1['B_Ref_List']

            dfXYC1 = len(dfXY1['B_Ref_List'])
            print(f'Total count of rows in the "XY_Ref_List" column: {dfXYC1}') # TAKING XY REFERENCE COUNT FROM THE BEIGN
            dfXYC2 = dfXY1['S'].value_counts().reset_index()

            dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('0201', '')
            dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('0402', '')
            dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('0603', '')
            dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('0805', '')
            dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('1206', '')

        except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            error_msgbm1 = f"The following columns are missing in BOM Sheet 'XY DATA':\n'Reference CRD as R'\n'X Coordinate as X'\n'Y Coordinate as Y'\n'Angle as A'\n'Side as S'"
            messagebox.showerror("Error", error_msgbm1)
            sys.exit(1)  # Exit the program with an error code


        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx") as writer:

            ds1.to_excel(writer, sheet_name="BOM", index=False)
            dfXY1.to_excel(writer, sheet_name="XY DATA", index=False)
            dsn2.to_excel(writer, sheet_name="AVL GROUP", index=False)
            dcn1.to_excel(writer, sheet_name="PART MASTER", index=False)
            df_AL1.to_excel(writer, sheet_name="AVL SHEET", index=True)
            ds3.to_excel(writer, sheet_name="BOM DATA", index=False)
            count_df = pd.DataFrame({'BOM Ref Count': [dsco3]})
            count_df.to_excel(writer, sheet_name="BOM Ref Count", index=TRUE)
            count_df = pd.DataFrame({'XY Ref Count': [dfXYC1]})
            count_df.to_excel(writer, sheet_name="XY Ref Count", index=TRUE)
            dfXYC2.to_excel(writer, sheet_name="XY Side Counts", index=True)
            dsplcr1.to_excel(writer, sheet_name="BOM_SL", index=False)
            dnsplcr1.to_excel(writer, sheet_name="Orginal_BOM_SL", index=False)

        pass
        print('The file does not exist.')
    #########################################################################################################################################################################
    #########################################################################################################################################################################
        #@@ CRD Inspection @@#
        print('\n')
        print("\033[92;4m******CRD CHECK******\033[0m")
        print('\n')
    #########################################################################################################################################################################
    #########################################################################################################################################################################
        Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
        Chd = os.getcwd()
        file_path = 'BOM_List_OP.xlsx'
        directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

        # Assuming 'BOM DATA' sheet contains a column named 'Bom Ref'
        df_bom_data = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM DATA")
        duplicates_bom_data = df_bom_data[df_bom_data.duplicated(subset='Bom Ref', keep=False)]
        # Assuming 'XY DATA' sheet contains a column named 'R'
        df_xy_data = pd.read_excel("BOM_List_OP.xlsx", sheet_name="XY DATA")
        duplicates_xy_data = df_xy_data[df_xy_data.duplicated(subset='R', keep=False)]
        # Function to check for duplicates and display an error message
        def check_and_show_duplicates(df, column_name, sheet_name):
            duplicates = df[df.duplicated(subset=column_name, keep=False)]
            if not duplicates.empty:
                print("Duplicate values in 'Bom Ref' column of 'BOM DATA':")
                print(duplicates_bom_data['Bom Ref'].tolist())
                print("\nDuplicate values in 'R' column of 'XY DATA':")
                print(duplicates_xy_data['R'].tolist())
                root = tk.Tk()
                root.withdraw()
                message = f"Duplicate values found in '{column_name}' column of '{sheet_name}' sheet:\n{duplicates[column_name].tolist()}"
                messagebox.showerror("Error", message)
                sys.exit()

        # Check for duplicates in 'Bom Ref' column of 'BOM DATA'
        check_and_show_duplicates(df_bom_data, 'Bom Ref', 'BOM DATA')

        # Check for duplicates in 'R' column of 'XY DATA'
        check_and_show_duplicates(df_xy_data, 'R', 'XY DATA')

    #########################################################################################################################################################################
    #########################################################################################################################################################################
        #@@ AVL Inspection @@#
        print('\n')
        print("\033[92;4m******AVL LINE INSPECTION******\033[0m")
        print('\n')
    #########################################################################################################################################################################
    #########################################################################################################################################################################

        Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
        Chd = os.getcwd()
        file_path = 'BOM_List_OP.xlsx'
        directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

        print(os.path.isfile(file_path))
        print(os.path.isfile(directory_path))

        try:
            if os.path.isfile(file_path):
                df_Iav1 = pd.read_csv(file_path)

        except ValueError:
            df_Iav1 = pd.read_excel(file_path, sheet_name="AVL SHEET", index_col=False) 

            print(df_Iav1)

            # Function to check for missing values between two strings
        def check_missing_values(row):
            start_index = None
            end_index = None

            for i in range(1, len(row) + 1):  # Check up to the last column
                col_name = f'PTN_{i}'
                if col_name in row.index:  # Check if the column exists
                    current_value = row[col_name]

                    if pd.isna(current_value):
                        if start_index is None:
                            start_index = i
                        end_index = i
                    else:
                        if start_index is not None and end_index is not None:
                            show_error(row['Group'], start_index, end_index)
                            start_index = None
                            end_index = None

        # Function to show pop-up error message
        def show_error(group, start_index, end_index):
            root = tk.Tk()
            root.withdraw()
            error_message = f"Error: Missing values between PTN_{start_index} and PTN_{end_index} in group '{group}'."
            messagebox.showerror("Error", error_message)

        # Check for missing values row-wise
        for index, row in df_Iav1.iterrows():
            check_missing_values(row)

        # Display the DataFrame with styling
        print(df_Iav1)

        # Function to check for missing values between two strings
        def check_missing_values(row):
            start_index = None
            end_index = None

            for i in range(1, len(row) + 1):  # Check up to the last column
                col_name = f'PTN_{i}'
                if col_name in row.index:  # Check if the column exists
                    current_value = row[col_name]

                    if pd.isna(current_value):
                        if start_index is None:
                            start_index = i
                        end_index = i
                    else:
                        if start_index is not None and end_index is not None:
                            show_error(row.get('Group', 'Unknown Group'), start_index, end_index)
                            start_index = None
                            end_index = None

        # Function to show pop-up error message
        def show_error(group, start_index, end_index):
            root = tk.Tk()
            root.withdraw()
            
            error_message = f"Error: Missing values between PTN_{start_index} and PTN_{end_index} in group '{group}'.\nDo you want to stop the program?"
            response = messagebox.askquestion("Error", error_message)

            if response == 'yes':
                sys.exit(1)

        # ...

        # Check for missing values row-wise
        for index, row in df_Iav1.iterrows():
            check_missing_values(row)

        # Display the DataFrame with styling
        print(df_Iav1)
    #########################################################################################################################################################################
    #########################################################################################################################################################################
        print('\n')
        print("\033[92;4m*******BOM & XY Verification Progress*******\033[0m")
        print('\n')
    #########################################################################################################################################################################
    #########################################################################################################################################################################

    #BOM XY VERIFICATION CODE BOM AND XY DATA AS VISE VERSA

        os.getcwd()
        Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
        Chd = os.getcwd()

        # Excel file path
        dfvbxy1 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM DATA")
        dfvbxy2 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="XY DATA")
        
        dfvbxy3 = pd.merge(dfvbxy1, dfvbxy2, on='B_Ref_List', how='left')
        dfvbxy4 = pd.merge(dfvbxy2, dfvbxy1, on='B_Ref_List', how='left')

        dfvbxy3["BOM and XY Compare"] = (dfvbxy3["Bom Ref"] == dfvbxy3["R"])
        dfvbxy3['BOM and XY Compare'] = dfvbxy3['BOM and XY Compare'].replace('TRUE','MATCH')
        dfvbxy3['BOM and XY Compare'] = dfvbxy3['BOM and XY Compare'].replace('FALSE','MISS_MATCH')

        dfvbxy3 = dfvbxy3.copy()
        dfvbxy3['BOM and XY Compare'] = dfvbxy3['BOM and XY Compare'].map({True: 'Match', False: 'Miss_Match'})
        dfvbxy3.sort_values(by='BOM and XY Compare', inplace=True, ascending=False)

        dfvbxy4["XY and BOM Compare"] = (dfvbxy4["R"] == dfvbxy4["Bom Ref"])
        dfvbxy4['XY and BOM Compare'] = dfvbxy4['XY and BOM Compare'].replace('TRUE','MATCH')
        dfvbxy4['XY and BOM Compare'] = dfvbxy4['XY and BOM Compare'].replace('FALSE','MISS_MATCH')

        dfvbxy4 = dfvbxy4.copy()
        dfvbxy4['XY and BOM Compare'] = dfvbxy4['XY and BOM Compare'].map({True: 'Match Mount Part', False: 'Miss_Match No Mount Part'})
        dfvbxy4.sort_values(by='XY and BOM Compare', inplace=True, ascending=False)

        dfvbxyc3 = dfvbxy3['BOM and XY Compare'].value_counts().reset_index()
        dfvbxyc4 = dfvbxy4['XY and BOM Compare'].value_counts().reset_index()
        dfvbxycS3 = dfvbxy3['S'].value_counts().reset_index()

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/Bom_List-Verified.xlsx") as writer:
            dfvbxy3.to_excel(writer, sheet_name="Bom to XY", index=False)
            dfvbxy4.to_excel(writer, sheet_name="XY to Bom", index=False)
            #dfvbxyc3.to_excel(writer, sheet_name="BOM and XY Compare", index=True)
            #dfvbxyc4.to_excel(writer, sheet_name="XY and BOM Compare", index=True)

    # MOVE THE Bo_List-Verified to BOM_List_OP
        def copy_sheet(source_wb_path, source_sheet_name, target_wb_path, target_sheet_name):
            # Read the source sheet into a DataFrame
            df = pd.read_excel(source_wb_path, sheet_name=source_sheet_name)

            # Open the target workbook in append mode
            with pd.ExcelWriter(target_wb_path, engine='openpyxl', mode='a') as writer:
                # Write the DataFrame to the target sheet
                df.to_excel(writer, sheet_name=target_sheet_name, index=False)

        # File paths
        source_wb_path = 'Bom_List-Verified.xlsx'
        target_wb_path = 'BOM_List_OP.xlsx'

        # Sheet names
        source_sheet_name_1 = 'Bom to XY'
        source_sheet_name_2 = 'XY to Bom'

        target_sheet_name_1 = 'Bom to XY'
        target_sheet_name_2 = 'XY to Bom'
        
        # Copy sheets from source workbook to target workbook
        copy_sheet(source_wb_path, source_sheet_name_1, target_wb_path, target_sheet_name_1)
        copy_sheet(source_wb_path, source_sheet_name_2, target_wb_path, target_sheet_name_2)

        file_path = 'Bom_List-Verified.xlsx'
        directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

        print(os.path.isfile(file_path))
        print(os.path.isfile(directory_path))

        try:
            if os.path.isfile(file_path):
                dt_H1 = pd.read_csv(file_path)

        except ValueError:
            if os.path.exists("Bom_List-Verified.xlsx"):
                os.remove("Bom_List-Verified.xlsx")
        else:
            print("The file does not exist")

    #########################################################################################################################################################################
    #########################################################################################################################################################################
            print('\n')
            print("\033[92;4m*******BOM XY Verification Progress*******\033[0m")
            print('\n')
    #########################################################################################################################################################################
    #########################################################################################################################################################################

    # Arrange single file and Highlight

        os.getcwd()
        Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
        Chd = os.getcwd()

        dir_B_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

            # Excel file path
        dfah1 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="Bom to XY")
        dfah2 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="XY to Bom")
        dfah3 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM")
        dfah4 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="XY DATA")
        dfah5 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM DATA")
        dfah6 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="AVL GROUP")
        dfah7 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="AVL SHEET")
        dfah8 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="PART MASTER")
        dfah9 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM_SL")
        dfah10 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="Orginal_BOM_SL")

        # Define a function for row styling
        def highlight_row(row):
                return ['background-color: lightgreen' if 'Match' in row.values else
                        'background-color: yellow' if 'Miss_Match' in row.values else
                        '' for _ in row]
            
            # Apply the styling function to the DataFrame
        style_df1 = dfah1.style.apply(highlight_row, axis=1)

                # Define a function for row styling
        def highlight_row(row):
                return ['background-color: lightgreen' if 'Match Mount Part' in row.values else
                        'background-color: pink' if 'Miss_Match No Mount Part' in row.values else
                        '' for _ in row]
            
            # Apply the styling function to the DataFrame
        style_df2 = dfah2.style.apply(highlight_row, axis=1)

    #########################################################################################################################################################################
    #########################################################################################################################################################################
        print('\n') #jan18 old sph & lcr
        print("\033[92;4m*******SHAPE-PACKAGE Assigning Progress*******\033[0m")
        print('\n')
    #########################################################################################################################################################################
    #########################################################################################################################################################################

        ##SHAPE & PACKAGE##

        # Load the Excel file
        excel_file_path = 'BOM_List_OP.xlsx'  # Replace with your file path

        # Define your column lists
        column_list = ["PartNumber", "Group", 'Priority', 'RefList', 'Qty', 'Shape', 'Long Des']

        dfshp = pd.read_excel(excel_file_path, sheet_name='BOM' ,usecols=column_list, index_col= False)

        print(dfshp)

        dfshp = dfshp[['PartNumber', 'Long Des']]

        print(dfshp)

        dfshp.rename(columns = {'Long Des':'Description'}, inplace = True)

        #dfshp['Description'] = dfshp['Long Des']

        #dfshp = dfshp['PartNUmber', 'Description']

        dfshp = pd.DataFrame(dfshp)
        # Define desired shapes
        desired_shapes = ("0201", "0402", "0603", "0805", "1206")

        # Custom function to extract and separate the shape
        def extract_shape(description):
            shape_match = re.search(r'\b\d{4}\b', description)
            if shape_match:
                shape = shape_match.group()
                if shape in desired_shapes:
                    return shape
            return None

        # Apply the custom function to create the Shape column
        dfshp['Shape'] = dfshp['Description'].apply(extract_shape)

        # Display the result
        print(dfshp[['PartNumber', 'Description', 'Shape']])
        #--------------------------------------------------------------
        '''#Desire Tol
        desired_Tol = ("1%", "5%", "10%", "15%", "20%", "25%", "±1%", "±5%", "±10%", "±15%", "±20%", "±25%")

        # Custom function to extract and separate the shape
        def extract_tol(description):
            tol_match = re.search(r'\b\d{2}\b', description)
            if tol_match:
                tol = tol_match.group()
                if tol in desired_Tol:
                    return tol
            return None

        # Apply the custom function to create the Shape column
        dfshp['Tol'] = dfshp['Description'].apply(extract_tol)

        # Display the result
        print(dfshp[['Description', 'Tol']])'''

        # Custom function to extract and separate the tolerance
        def extract_tol(description):
            tol_match = re.search(r'[±]?\d+%', description)
            if tol_match:
                return tol_match.group()
            return None

        # Apply the custom function to create the Tol column
        dfshp['Tol'] = dfshp['Description'].apply(extract_tol)

        # Display the result
        print(dfshp[['PartNumber', 'Description', 'Tol']])

        #--------------------------------------------------------------
        # Define desired component types
        desired_COMP = ("CAP", "RES", "IND")

        # Custom function to extract and separate the component type
        def extract_COMPs(description):
            for comp_type in desired_COMP:
                    if comp_type.lower() in description.lower():
                        return comp_type
            return None

        # Apply the custom function to create the LCRTYPE column
        dfshp['LCRTYPE'] = dfshp['Description'].apply(extract_COMPs)

        # Define desired Special component types
        desired_SPERESCOMP = ["MELF"]

        # Custom function to check if the description contains special components
        def contains_special_res_component(description):
            for serescomp_type in desired_SPERESCOMP:
                if serescomp_type.lower() in description.lower():
                    return True
            return False

        # Apply the custom function to create the SPERESCOMP column
        dfshp['SPERESCOMP'] = dfshp['Description'].apply(contains_special_res_component)

        # Transform 'SPERESCOMP' column to 'MELF' when it's True
        dfshp['SPERESCOMP'] = np.where(dfshp['SPERESCOMP'], 'MELF', '')

        # Define desired Special component types
        desired_SPETHTCOMP = ["THT"]

        # Custom function to check if the description contains special components
        def contains_special_tht_component(description):
            for sethtcomp_type in desired_SPETHTCOMP:
                if sethtcomp_type.lower() in description.lower():
                    return True
            return False

        # Apply the custom function to create the SPERESCOMP column
        dfshp['SPETHTCOMP'] = dfshp['Description'].apply(contains_special_tht_component)

        # Transform 'SPERESCOMP' column to 'MELF' when it's True
        dfshp['SPETHTCOMP'] = np.where(dfshp['SPETHTCOMP'], 'THT', '')

        # Define desired Special component types
        desired_SPESODCOMP = ("ZENER", "DIODE", "SOD")

        # Custom function to extract and separate the component type
        def extract_SPESODCOMP(description):
            for sesodcomp_type in desired_SPESODCOMP:
                if sesodcomp_type.lower() in description.lower():
                    return sesodcomp_type
            return None

        # Apply the custom function to create the LCRTYPE column
        dfshp['SPESODCOMP'] = dfshp['Description'].apply(extract_SPESODCOMP)

        # Define desired Special component types
        desired_SPECAPCOMP = ("TAN", "Tantalum", "Aluminium","ALLUM", "ALUM", "Electrolytic" ,"ALU")

        # Custom function to extract and separate the component type
        def extract_SPECAPCOMP(description):
            for setancomp_type in desired_SPECAPCOMP:
                if setancomp_type.lower() in description.lower():
                    return setancomp_type
            return None

        # Apply the custom function to create the LCRTYPE column
        dfshp['SPECAPCOMP'] = dfshp['Description'].apply(extract_SPECAPCOMP)

            # Define desired Special component types
        desired_SPEFERINDCOMP = ("IND", "FERRITEBEAD", "FERRITE","BEAD", "INDUCTOR")

        # Custom function to extract and separate the component type
        def extract_SPEFERINDCOMP(description):
            for seferindcomp_type in desired_SPEFERINDCOMP:
                if seferindcomp_type.lower() in description.lower():
                    return seferindcomp_type
            return None

        # Apply the custom function to create the LCRTYPE column
        dfshp['SPEFERINDCOMP'] = dfshp['Description'].apply(extract_SPEFERINDCOMP)


        # Assuming df is your DataFrame
        dfshp['PACKAGE'] = dfshp['Shape'].replace({'0201': '0802P', '0402': '0802P', '0603': '0804P', '0805': '0804E', '1206': '0804E'})

        dfshp = dfshp.rename(columns={"Description": "Long Desp"})

        # Display the result
        print(dfshp[['Long Desp', 'Shape', 'LCRTYPE', 'SPERESCOMP', 'SPETHTCOMP', 'SPESODCOMP', 'SPECAPCOMP', 'SPEFERINDCOMP', 'PACKAGE']])

    #########################################################################################################################################################################
    #########################################################################################################################################################################
        print('\n')
        print("\033[92;4m*******LCR Generation Progress*******\033[0m")
        print('\n')
    #########################################################################################################################################################################
    #########################################################################################################################################################################

        # Function to extract LCR information
        def extract_component_info(component_line):
            # Regular expressions for CAPACITOR
            capacitor_type_match = re.search(r'\b(CAP|Capacitor|MLCC|CAPACITOR)\b', component_line, re.IGNORECASE)
            capacitor_value_match = re.search(r'(\d+(\.\d+)?)\s*([pnuμmkPUNKM]?\w?)F', component_line)
            capacitor_tolerance_match = re.search(r'[±](\d+)%', component_line)
            # Regular expressions for RESISTOR
            resistor_type_match = re.search(r'\b(Res|Resistor|FLIM|RESISTOR)\b', component_line, re.IGNORECASE)
            resistor_value_match = re.search(r'(\d+(\.\d+)?)\s*([pnuμmkPUNKM]?\w?)\s*(\d+%)?', component_line)
            resistor_tolerance_match = re.search(r'[±](\d+)%', component_line)
            # Assign default values
            LCR_Type = None
            LCR_Value = None
            LCR_Unit = None
            LCR_Tolerance = None

            if capacitor_type_match:
                LCR_Type = capacitor_type_match.group(1)

            if capacitor_tolerance_match:
                LCR_Tolerance = capacitor_tolerance_match.group(1)

            if capacitor_value_match:
                groups = capacitor_value_match.groups()
                LCR_Value = groups[1] if groups[1] is not None else None
                LCR_Unit = groups[2] if groups[2] is not None else None

            if resistor_type_match:
                LCR_Type = resistor_type_match.group(1)

            if resistor_value_match:
                groups = resistor_value_match.groups()
                LCR_Value = groups[0] if groups[0] is not None else None
                LCR_Unit = groups[2] if groups[2] is not None else None

            if resistor_tolerance_match:
                LCR_Tolerance = resistor_tolerance_match.group(1)

            return LCR_Type, LCR_Value, LCR_Unit, LCR_Tolerance

        # Set your working directory
        os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')

        # Load the Excel file
        excel_file_path = 'BOM_List_OP.xlsx'  # Replace with your file path

        # Define your column list
        column_names = ["Long Des", "PartNumber"]

        # Read the Excel file without index_col and add 'PartNumber' to the column_names list
        dflcr = pd.read_excel(excel_file_path, sheet_name='BOM_SL', usecols=column_names)

        # Add 'PartNumber' as a separate column
        dflcr['PartNumber'] = pd.read_excel(excel_file_path, sheet_name='BOM_SL', usecols=['PartNumber'])

        # Apply the extract_component_info function to create LCR columns
        dflcr[['LCR Type', 'LCR Value', 'LCR Unit', 'LCR Tolerance']] = dflcr['Long Des'].apply(extract_component_info).apply(pd.Series)

        # Display the result
        print(dflcr[['PartNumber', 'LCR Type', 'LCR Value', 'LCR Unit', 'LCR Tolerance']])

        # Create a new column 'AVL_Name' based on the condition in 'Priority' column
        dfah10['AVL Name'] = dfah10.apply(lambda row: row['PartNumber'] if row['Priority'] == 1 else '', axis=1)
        
        print(dfah10)

        #start Here to merge
            # Desired column order
        desired_order = ['PartNumber','Group','Priority','Long Des','Shape', 'Qty','RefList','AVL Name']

        # Create a list of columns present in both DataFrame and desired_order
        common_columns = [col for col in desired_order if col in dfah10.columns]

        # Reorder the DataFrame based on the desired_order
        dfah10 = dfah10[common_columns]

        dfah10.rename(columns = {'Long Des':'Description'}, inplace = True)
        dfah10.rename(columns = {'Shape':'Shape DTA'}, inplace = True)

        dfah10 = pd.merge(dfah10 , dfshp, on='PartNumber', how='left')
        print(dfah10)
        dfah10 = pd.merge(dfah10 , dflcr, on='PartNumber', how='left')
        print(dfah10)

        del dfah10['Long Desp']
        del dfah10['Long Des']

        dfah10.sort_values(by='Priority', inplace=True, ascending=True)
        dfah10.sort_values(by='Group', inplace=True, ascending=True)

        F1sl_col = dfah10.pop('PACKAGE') # col-14
        dfah10.insert(9, 'PACKAGE', F1sl_col)

        # Line are in Kstudy Import don't delete

        #dfah10["LCRTYPE1"] = dfah10['LCRTYPE'].astype(str) +"-"+ dfah10['SPERESCOMP'].astype(str) +"-"+ dfah10['SPETHTCOMP'].astype(str) +"-"+ dfah10["SPESODCOMP"].astype(str) +"-"+ dfah10["SPECAPCOMP"]
        # Assuming dfah10 is your DataFrame with columns LCRTYPE, SPERESCOMP, SPETHTCOMP, SPESODCOMP, SPECAPCOMP
        #dfah10["LCRTYPE1"] = dfah10.apply(lambda row: "-".join(str(value) for value in row), axis=1)
        # OK LINE #dfah10["LCRTYPE1"] = dfah10[['LCRTYPE', 'SPERESCOMP', 'SPETHTCOMP', 'SPESODCOMP', 'SPECAPCOMP']].astype(str).apply('-'.join, axis=1)
        # Assuming dfah10 is your DataFrame with columns LCRTYPE, SPERESCOMP, SPETHTCOMP, SPESODCOMP, SPECAPCOMP
        dfah10["LCRTYPE"] = dfah10[['LCRTYPE', 'SPERESCOMP', 'SPETHTCOMP', 'SPESODCOMP', 'SPECAPCOMP', 'SPEFERINDCOMP']].apply(lambda x: ''.join(x.dropna().astype(str)), axis=1) #x: '-'.jo #># x: ''.jo
        
        dfah10['LCRTYPE'] = dfah10['LCRTYPE'].str.replace('FERRITEBEAD','IND')
        dfah10['LCRTYPE'] = dfah10['LCRTYPE'].str.replace('BEAD','IND')
        dfah10['LCRTYPE'] = dfah10['LCRTYPE'].str.replace('FERRITE','IND')
        dfah10['LCRTYPE'] = dfah10['LCRTYPE'].str.replace('INDIND','IND')

        dfah10["Shape"] = dfah10[['LCRTYPE','Shape']].apply(lambda x: '-'.join(x.dropna().astype(str)), axis=1)
        # Display the resulting DataFrame
        print(dfah10)

        # Assuming dfah10 is your DataFrame
        columns_to_delete = ['SPERESCOMP', 'SPETHTCOMP', 'SPESODCOMP', 'SPECAPCOMP', 'SPEFERINDCOMP']

        # Use the drop method to delete the specified columns
        dfah10.drop(columns=columns_to_delete, inplace=True)

        dfah10['Shape'] = dfah10['Shape'].str.replace('CAP-0201','C0201')
        dfah10['Shape'] = dfah10['Shape'].str.replace('CAP-0402','C0402')
        dfah10['Shape'] = dfah10['Shape'].str.replace('CAP-0603','C0603')
        dfah10['Shape'] = dfah10['Shape'].str.replace('CAP-0805','C0805')
        dfah10['Shape'] = dfah10['Shape'].str.replace('CAP-1206','C1206')
        dfah10['Shape'] = dfah10['Shape'].str.replace('RES-0201','R0201')
        dfah10['Shape'] = dfah10['Shape'].str.replace('RES-0402','R0402')
        dfah10['Shape'] = dfah10['Shape'].str.replace('RES-0603','R0603')
        dfah10['Shape'] = dfah10['Shape'].str.replace('RES-0805','R0805')
        dfah10['Shape'] = dfah10['Shape'].str.replace('RES-1206','R1206')
        dfah10['Shape'] = dfah10['Shape'].str.replace('IND-0201','C0201')
        dfah10['Shape'] = dfah10['Shape'].str.replace('IND-0402','C0402')
        dfah10['Shape'] = dfah10['Shape'].str.replace('IND-0603','C0603')
        dfah10['Shape'] = dfah10['Shape'].str.replace('IND-0805','C0805')
        dfah10['Shape'] = dfah10['Shape'].str.replace('IND-1206','C1206')

        dfah10.rename(columns = {'Tol':'Tolerance'}, inplace = True)

        dfah10['LCRTYPE'] = dfah10['LCRTYPE'].str.replace('CAP','CAPACITOR')
        dfah10['LCRTYPE'] = dfah10['LCRTYPE'].str.replace('RES','RESISTOR')
        dfah10['LCRTYPE'] = dfah10['LCRTYPE'].str.replace('IND','INDUCTOR')
        
        # Desired column order
        desired_order = ['PartNumber','Group','Priority','Description','Shape DTA','Qty','RefList','AVL Name','Shape','PACKAGE','LCRTYPE','LCR Type','LCR Value','LCR Unit','LCR Tolerance','Tolerance']
        
        # Create a list of columns present in both DataFrame and desired_order
        common_columns = [col for col in desired_order if col in dfah10.columns]

        # Reorder the DataFrame based on the desired_order
        dfah10 = dfah10[common_columns]

        dfah9 = dfah10
        #BOM_SL & ORGBOM_SL Reference list change
        dfah9['RefList'] = dfah9['RefList'].str.replace("_x000D_","")
        dfah9['RefList'] = dfah9['RefList'].str.replace(" ","")
        dfah9['RefList'] = dfah9['RefList'].str.replace("\n","")

        dfah9 = dfah10

        dfah9 = dfah9[dfah9['Priority'].isin([0, 1])]

        dfah11 = dfah10.copy()

        dfah11 = dfah11

        # Assuming dfah10 is your DataFrame
        columns_to_delete = ['Group','Priority','Shape DTA','Qty','RefList','AVL Name','Shape','PACKAGE','LCRTYPE','LCR Type','LCR Tolerance']

        # Use the drop method to delete the specified columns
        dfah11.drop(columns=columns_to_delete, inplace=True)

            # Desired column order
        desired_order = ['Description','PartNumber','LCRTYPE','LCR Value','LCR Unit','Tolerance']
        
        # Create a list of columns present in both DataFrame and desired_order
        common_columns = [col for col in desired_order if col in dfah11.columns]

        # Reorder the DataFrame based on the desired_order
        dfah11 = dfah11[common_columns]

        dfah11.rename(columns = {'PartNumber':'PBARNO'}, inplace = True)
        dfah11['PBARPTN'] = dfah11['PBARNO']
        dfah11['PBARBAR'] = dfah11['PBARNO']
        dfah11.insert(3,'PBARQTY', 10000)
        dfah11.insert(4,'PBARFTYP', 3)
        dfah11.insert(5,'PBARFPIT', '')
        dfah11.insert(6,'PBARWDT', '')
        dfah11.insert(7,'PBARLOQ', '')
        dfah11.insert(8,'PBARSLQ', '')
        dfah11.insert(9,'PBARSLM', '')
        dfah11.insert(10,'PBARVND', '')
        dfah11.insert(11,'PBARLOT', '')
        dfah11.insert(12,'PBARDTE', '')
        dfah11.insert(13,'PBARLOC', '')
        dfah11.insert(14,'PBARJOB', '')
        dfah11.insert(15,'PBARSPP', '')
        dfah11.insert(16,'PBARLCR', '')
        dfah11.insert(17,'PBARLMT', '')
        dfah11.insert(18,'PBARNOR', '')
        dfah11.insert(19,'PBARUPP', '')
        dfah11.insert(20,'PBARLOW', '')
        dfah11.insert(21,'PBARFEQ', '')
        dfah11.insert(22,'PBARVOL', '')
        dfah11.insert(23,'PBARMEM', '')
        dfah11.insert(24,'PBARDRYTYP', '')
        dfah11.insert(25,'PBARMSLVL', '')
        dfah11.insert(26,'PBARTRYX', '')
        dfah11.insert(27,'PBARTRYY', '')
        dfah11.insert(28,'PBARTRYMATRIX', '')
        dfah11.insert(29,'PBARNOTE1', '')
        dfah11.insert(30,'PBARNOTE2', '')
        dfah11.insert(31,'PBARNOTE3', '')
        dfah11.insert(32,'PBARNOTE4', '')
        dfah11.insert(33,'PBARMDF', '')
        dfah11.insert(34,'PBARSAFETYCNT', '')
        dfah11.insert(35,'PBARTRAYPACKAGE', '')
        dfah11.insert(36,'PBARUSELIMIT', '')
        dfah11.insert(37,'PBARDTEFORMAT', '')
        dfah11.insert(38,'CPBARPTN', '')
        dfah11.insert(39,'CPBARBAR', '')
        dfah11.insert(40,'CPBARQTY', '')
        dfah11.insert(41,'CPBARLOQ', '')
        dfah11.insert(42,'CPBARSLQ', '')
        dfah11.insert(43,'CPBARSLM', '')
        dfah11.insert(44,'CPBARVND', '')
        dfah11.insert(45,'CPBARLOT', '')
        dfah11.insert(46,'CPBARDTE', '')
        dfah11.insert(47,'CPBARLOC', '')
        dfah11.insert(48,'CPBARMEM', '')
        dfah11.insert(49,'CPBARNOTE1', '')
        dfah11.insert(50,'CPBARNOTE2', '')
        dfah11.insert(51,'CPBARNOTE3', '')
        dfah11.insert(52,'CPBARNOTE4', '')
        dfah11.insert(53,'CPBARLIGHTING', '')
        dfah11.insert(54,'CPBARSAFETYCNT', '')
        dfah11.insert(55,'PBARUNPSTCHK', '')
        dfah11.insert(56,'PBARTRYQTY', '')
        dfah11.insert(57,'PBARPARTSCHG', '')
        dfah11.insert(58,'PBARSHAPE', '')
        dfah11.insert(59,'PBARPACKAGE', '')
        dfah11.insert(60,'PBARDIRECTION', '')

        desired_order = ['Description','PBARNO','PBARPTN','PBARBAR','PBARQTY','PBARFTYP','LCR Value','LCR Unit','Tolerance','PBARFPIT','PBARWDT','PBARLOQ','PBARSLQ','PBARSLM','PBARVND','PBARLOT','PBARDTE','PBARLOC','PBARJOB','PBARSPP','PBARLCR','PBARLMT','PBARNOR','PBARUPP','PBARLOW','PBARFEQ','PBARVOL','PBARMEM','PBARDRYTYP','PBARMSLVL','PBARTRYX','PBARTRYY','PBARTRYMATRIX','PBARNOTE1','PBARNOTE2','PBARNOTE3','PBARNOTE4','PBARMDF','PBARSAFETYCNT','PBARTRAYPACKAGE','PBARUSELIMIT','PBARDTEFORMAT','CPBARPTN','CPBARBAR','CPBARQTY','CPBARLOQ','CPBARSLQ','CPBARSLM','CPBARVND','CPBARLOT','CPBARDTE','CPBARLOC','CPBARMEM','CPBARNOTE1','CPBARNOTE2','CPBARNOTE3','CPBARNOTE4','CPBARLIGHTING','CPBARSAFETYCNT','PBARUNPSTCHK','PBARTRYQTY','PBARPARTSCHG','PBARSHAPE','PBARPACKAGE','PBARDIRECTION']
            
        # Create a list of columns present in both DataFrame and desired_order
        common_columns = [col for col in desired_order if col in dfah11.columns]

        # Reorder the DataFrame based on the desired_order
        dfah11 = dfah11[common_columns]

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/Bom_List-Verified.xlsx", engine='openpyxl') as writer:
            style_df1.to_excel(writer, sheet_name="Bom to XY", index=False)
            style_df2.to_excel(writer, sheet_name="XY to Bom", index=False)
            dfah3.to_excel(writer, sheet_name="BOM", index=False)
            dfah4.to_excel(writer, sheet_name="XY DATA", index=False)
            dfah5.to_excel(writer, sheet_name="BOM DATA", index=False)
            dfah6.to_excel(writer, sheet_name="AVL GROUP", index=False)
            dfah7.to_excel(writer, sheet_name="AVL SHEET", index=False)
            dfah8.to_excel(writer, sheet_name="PART MASTER", index=False)
            dfshp.to_excel(writer, sheet_name="SHP-PKG", index=False)
            dflcr.to_excel(writer, sheet_name="LCR", index=False)
            dfah9.to_excel(writer, sheet_name="BOM_SL", index=False)
            dfah10.to_excel(writer, sheet_name="Orginal_BOM_SL", index=False)
            dfah11.to_excel(writer, sheet_name="Part Master LCR", index=False)

            print('\n')
            print("\033[92;4m*******BOM XY Verification Progress Complete*******\033[0m")
            print('\n')
            
            print("``````````````````````````````````````````````")
            print(f"Current Date and Time: {formatted_datetime}")
            print("``````````````````````````````````````````````")
            print("\n")
            print("``````````````````````````````````````````````")
            print(dir_B_path,'\\__BOM__\\',dLbr1)
            print(f"Bom Name:{dLbr1}")
            print("``````````````````````````````````````````````")
            print('\n')
            print("``````````````````````````````````````````````")
            print(f'BOM Part-No Qty column SUM: {dsco2}')
            print("``````````````````````````````````````````````")
            print("\n")
            print("``````````````````````````````````````````````")
            print(f'Total count of BOM DATA in the "BOM_Ref_List" column: {dsco3}')
            print("``````````````````````````````````````````````")
            print("\n")
            print("``````````````````````````````````````````````")
            print(f'Total count of XY in the "Reference" column: {dfXYC1}')
            print("``````````````````````````````````````````````")
            print("\n")
            print("``````````````````````````````````````````````")
            print(f'Total count of XY DATA side in the "S" column: {dfXYC2}')
            print("``````````````````````````````````````````````")
            print("\n")
            print("``````````````````````````````````````````````")
            print(f'Total count of BOM TO XY DATA side in the "S" column: {dfvbxycS3}')
            print("``````````````````````````````````````````````")
            print("\n")
            print("``````````````````````````````````````````````")
            print(f'BOM and XY Compare: {dfvbxyc3}')
            print("``````````````````````````````````````````````")
            print("\n")
            print("``````````````````````````````````````````````")
            print(f'XY and BOM Compare: {dfvbxyc4}')
            print("``````````````````````````````````````````````")

    # Define the PySimpleGUI layout
    layout = [
        [sg.Multiline(size=(160, 40), font=('Courier', 9), key='-LOGWINDOW-')],
        [sg.Button('Save to Excel'), sg.Button('Quit')]
    ]

    # Create the window
    window = sg.Window("BOM XY VERIFICATION", layout, finalize=True)

    def print_to_log(*args, **kwargs):
        window['-LOGWINDOW-'].print(*args, **kwargs)
        window.Refresh()

    def save_to_excel(log_contents):
        # Get the current date and time
        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

        # Split log contents into lines
        lines = log_contents.strip().split('\n')

        # Create a DataFrame with each line in a new row
        df = pd.DataFrame({'LogContents': lines})

        # Save the DataFrame to the same Excel file with a new sheet
        excel_file_path = 'Bom_List-Verified.xlsx'
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=f'Log_{current_datetime}', index=False)

        return f"Log saved to {excel_file_path}, Sheet: Log_{current_datetime}"

    print_to_log("BOM XY VERIFICATION___Compelete $ PROCESS $")

    sys.stdout.write("\n")

    print_to_log("``````````````````````````````````````````````")
    print_to_log(f"Current Date and Time: {formatted_datetime}")
    print_to_log("``````````````````````````````````````````````")
    print_to_log("\n")
    print_to_log("``````````````````````````````````````````````")
    print_to_log(dir_B_path,'\\__BOM__\\',dLbr1)
    print_to_log(f"Bom Name:{dLbr1}")
    print_to_log("``````````````````````````````````````````````")
    print_to_log('\n')
    print_to_log("``````````````````````````````````````````````")
    print_to_log(f'BOM Part-No Qty column SUM: {dsco2}')
    print_to_log("``````````````````````````````````````````````")
    print_to_log("\n")
    print_to_log("``````````````````````````````````````````````")
    print_to_log(f'Total count of BOM in the "BOM_Ref_List" column: {dsco3}')
    print_to_log("``````````````````````````````````````````````")
    print_to_log("\n")
    print_to_log("``````````````````````````````````````````````")
    print_to_log(f'Total count of XY DATA in the "Reference" column: {dfXYC1}')
    print_to_log("``````````````````````````````````````````````")
    print_to_log("\n")
    print_to_log("``````````````````````````````````````````````")
    print_to_log(f'Total count of XY DATA side in the "S" column: {dfXYC2}')
    print_to_log("``````````````````````````````````````````````")
    print_to_log("\n")
    print_to_log("``````````````````````````````````````````````")
    print_to_log(f'Total count of BOM TO XY DATA side in the "S" column: {dfvbxycS3}')
    print_to_log("``````````````````````````````````````````````")
    print_to_log("\n")
    print_to_log("``````````````````````````````````````````````")
    print_to_log(f'BOM and XY Compare: {dfvbxyc3}')
    print_to_log("``````````````````````````````````````````````")
    print_to_log("\n")
    print_to_log("``````````````````````````````````````````````")
    print_to_log(f'XY and BOM Compare: {dfvbxyc4}')
    print_to_log("``````````````````````````````````````````````")

    # Create an event loop
    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Quit':
            break
        elif event == 'Save to Excel':
            # Get the contents of the log window
            log_contents = values['-LOGWINDOW-']

            # Save to Excel and get the log information
            log_info = save_to_excel(log_contents)

            print_to_log(log_info)

    time.sleep (2)
    # Close the window

    # Assuming feeder verification is completed
    BOM_XY_verification_completed = True

    if BOM_XY_verification_completed:
        root = tk.Tk()
        root.withdraw()  # Hide the main window

        messagebox.showinfo("BOM & XY Verification", "BOM & XY verification has been completed!")

    window.close()

    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        if os.path.exists("BOM_List_OP.xlsx"):
            os.remove("BOM_List_OP.xlsx")
    else:
        print("The file does not exist")

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
    Chd = os.getcwd()

    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        if os.path.exists("BOM_List_OP.xlsx"):
            os.remove("BOM_List_OP.xlsx")
    else:
        print("The file does not exist")

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()

    '''    src_1 = 'Bom_List-Verified.xlsx'
        os.rename(src_1, dLbr1 +"_XY"+".xlsx")
        time.sleep (2)
        window.close()
        time.sleep (5)'''

    # Check if the file exists before renaming
    src_1 = 'Bom_List-Verified.xlsx'
    if os.path.isfile(src_1):
        os.rename(src_1, f'{dLbr1}_XY.xlsx')
        print(f"File {src_1} renamed to {dLbr1}_XY.xlsx")
    else:
        print(f"File {src_1} does not exist.")

    time.sleep(5)

    print('\n')
    print("\033[92;4m BOM Generation Complete \033[0m")
    print('\n')
    print("\033[92;4m BOM and xy Verfication Found OK \033[0m")
    print('\n')

    # Notify when the process is completed
    BXY_List_Generation_Completed = True

    if BXY_List_Generation_Completed:
        root = tk.Tk()
        root.withdraw()  # Hide the main window

        messagebox.showinfo("Process Completed", "BOM & XY Verified\nShape, Package & LCR Value has been Generated!")

    sys.exit() #BOM Manipulation