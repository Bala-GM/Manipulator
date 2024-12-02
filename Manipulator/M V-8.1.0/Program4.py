import pandas as pd
import numpy as np
import os
from os import getcwd
import csv as csv
from io import StringIO
import re
import datetime
from io import BytesIO
import time
import sys
from datetime import datetime
import shutil
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import subprocess
import threading
import urllib
import urllib.parse
from sqlite3 import dbapi2 as sqlite
import sqlite3
import lxml
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
import PySimpleGUI as sg
import sqlalchemy
from sqlalchemy import create_engine
import sqlalchemy_access as sq_a
import sqlalchemy_pyodbc_mssql as sqlalchemy
from flask_sqlalchemy import SQLAlchemy
import pyodbc
import pypyodbc
import odbc
import psycopg2
import mysql.connector as sql
import MySQLdb as sql #pip install mysqlclient
from plyer import notification
import xlrd
import xml.etree.ElementTree as ET
from openpyxl.utils.exceptions import InvalidFileException
import psutil

# Program 4: FeederLoadingList: V-2.3 X
def program_4():

    print('\n')
    print("\033[32;4m*******Syrma Feeder LoadingList Version--PY_V-2.6 interface_GUI/J0124-89P13*******\033[0m")

    '''bil1 = pyfiglet.figlet_format("Version--PY-V1.5 interface_GUI/J0324", width = 300)print(bil1)'''

    # Get the current date and time
    current_datetime = datetime.now()

    # Format the current date and time as a string
    #formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
    # Format the date and time in a 12-hour clock with AM/PM
    formatted_datetime = current_datetime.strftime("%Y-%m-%d %I:%M:%S %p")

    print('\n')

    # Print the formatted date and time
    print(f"\033[31mCurrent Date and Time: {formatted_datetime}\033[0m")

    print('\n')

    '''print(f"Current Year: {current_datetime.year}")
    print(f"Current Month: {current_datetime.month}")
    print(f"Current Day: {current_datetime.day}")
    print(f"Current Hour: {current_datetime.hour}")
    print(f"Current Minute: {current_datetime.minute}")
    print(f"Current Second: {current_datetime.second}")'''

    ######################################################################################################
    ######################################################################################################
    ######################################################################################################
    ######################################################################################################
    ######################################################################################################
    print("\033[32;4m*******FeederSetup*******\033[0m")
    ######################################################################################################

    # FeederSetup

    def find_feeder_setup_files(root_directory):
        feeder_setup_files = []
        for root, dirs, files in os.walk(root_directory):
            for file in files:
                if file == "FeederSetup.csv":
                    feeder_setup_files.append(os.path.join(root, file))
        return feeder_setup_files

    # Function to find the setup description from file path
    def find_setup_description(file_path):
        parts = file_path.split(";")
        setup_description = parts[-2].strip()
        return setup_description

    def rename_feeder_setup_with_description(file_path):
        setup_description = find_setup_description(file_path)
        if "[Top] Line1" in setup_description:
            new_file_name = "FeederSetup_TL1.csv"
        elif "[Bottom] Line1" in setup_description:
            new_file_name = "FeederSetup_BL1.csv"
        elif "[Top] Line2" in setup_description:
            new_file_name = "FeederSetup_TL2.csv"
        elif "[Bottom] Line2" in setup_description:
            new_file_name = "FeederSetup_BL2.csv"
        elif "[Top] Line3" in setup_description:
            new_file_name = "FeederSetup_TL3.csv"
        elif "[Bottom] Line3" in setup_description:
            new_file_name = "FeederSetup_BL3.csv"
        elif "[Top] Line4" in setup_description:
            new_file_name = "FeederSetup_TL4.csv"
        elif "[Bottom] Line4" in setup_description:
            new_file_name = "FeederSetup_BL4.csv"
        elif "[Top] Line-1" in setup_description:
            new_file_name = "FeederSetup_TL4C.csv"
        elif "[Bottom] Line-1" in setup_description:
            new_file_name = "FeederSetup_BL4C.csv"
        else:
            return
        
        new_path = os.path.join(os.path.dirname(file_path), new_file_name)
        os.rename(file_path, new_path)
        print('\n')
        print(f"FeederSetup.csv renamed to: {new_path}")
        print('\n')
        return new_path

    # Define the root directory
    root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

    # Find FeederSetup.csv files
    feeder_setup_files = find_feeder_setup_files(root_directory)

    # Rename FeederSetup.csv files with setup descriptions and move them
    for file_path in feeder_setup_files:
        new_path = rename_feeder_setup_with_description(file_path)
        if new_path:
            # Create the FeederSetup directory if it doesn't exist
            feeder_setup_dir = os.path.join(os.path.dirname(root_directory), "FeederSetup")
            if not os.path.exists(feeder_setup_dir):
                os.makedirs(feeder_setup_dir)
            # Move the renamed file to the FeederSetup directory
            shutil.copy(new_path, os.path.join(feeder_setup_dir, os.path.basename(new_path)))

    # Function to log the usage and input
    def log_usage_and_input(dL1):
        # Get the current date and time
        #current_datetime1 = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create or open the log file
        os.getcwd()
        Chd= os.chdir('D:/NX_BACKWORK/Database_File/SMT_Log')
        Chd = os.getcwd()
        log_file_path = 'Loading_List_log.txt'
        with open(log_file_path, 'a') as log_file:
            # Write the usage information to the log file
            log_file.write(f"{formatted_datetime}: App used. Input: {dL1}\n")

    ######################################################################################################
    ######################################################################################################
    ######################################################################################################
    print("\033[32;4m*******FeederSetup Input*******\033[0m")
    ######################################################################################################
    # Input values for cell B3 and Rev A1
    print('\n')
    dL1 = input("\033[93mEnter BOM Name :\033[0m")
    print('\n')
    dL2 = input("\033[93mEnter Feeder Name :\033[0m")
    value_B3 = dL2[:12]  # Take only the first 12 characters from dL2
    print('\n')
    Revision = input("\033[93mEnter the program for Revision: \033[0m")
    print('\n')
    dLine123 = input("\033[93mEnter the Line: \033[0m")

    # Log the usage and input
    log_usage_and_input(dL1)

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
    Chd = os.getcwd()

    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Upload"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)

    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)

    os.getcwd()
    #Chd= os.chdir('D:\\NX_BACKWORK')
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

    if os.path.exists("Feeder_List_OPT.xlsx"):
        os.remove("Feeder_List_OPT.xlsx")
    else:
        print("The file does not exist")

    if os.path.exists("Feeder_List_OPB.xlsx"):
        os.remove("Feeder_List_OPB.xlsx")
    else:
        print("The file does not exist")

    if os.path.exists("Upload-Data.xlsx"):
        os.remove("Upload-Data.xlsx")
    else:
        print("The file does not exist")

    os.getcwd()
    #Chd= os.chdir('D:\\NX_BACKWORK')
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')
    Chd = os.getcwd()

    ##########################################################################################################################################

    #bil2 = pyfiglet.figlet_format("FeederSetup Progress", width = 150)
    print('\n')
    print('\033[92;4m******FeederSetup Progress******\033[0m')
    print('\n')
    ##########################################################################################################################################

    #LINE1T

    file_path = 'FeederSetup_TL1.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)
            
    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_TL1.csv', encoding="utf-8",index_col=False, skiprows=range(2))
        
        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 350:
            # Continue with the rest of your code
            print(f"dt_H1 line count: {len(dt_H1)}")
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 353.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-1 Slot Count in TOP Feeder '353'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_TL1.csv', encoding="utf-8",index_col=False, skiprows=range(2, 351), nrows=3)

            # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-1 Slot Count in TOP Feeder '351'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

        #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)
        
        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_TL1.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_TL1.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_TL1: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('3-1-2-','3-2-')

        df1['Location'] = df1['Location'].str.replace('3-1-1-','3-1-')

        df1['Location'] = df1['Location'].str.replace('2-1-2-','2-2-')

        df1['Location'] = df1['Location'].str.replace('2-1-1-','2-1-')

        df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

        df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

        df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

        df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)

        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
        
        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb1 = load_workbook(Feeder_List_OPT)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb2 = load_workbook(Feeder_List_OPT)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
            #copy from wb1
            c = ws1.cell(row=row, column=10)
            #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPT.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "TOP"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['TOP_Side'] = "TOP"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "TOP"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "TOP"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)    
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE1B

    file_path = 'FeederSetup_BL1.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_BL1.csv', encoding="utf-8",index_col=False, skiprows=range(2))
        
        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 350:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 353.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-1 Slot Count in BOT Feeder '353'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_BL1.csv', encoding="utf-8",index_col=False, skiprows=range(2, 351), nrows=3)

                # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-1 Slot Count in BOT Feeder '351'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
        
        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_BL1.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_BL1.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_BL1: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code


    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('3-1-2-','3-2-')

        df1['Location'] = df1['Location'].str.replace('3-1-1-','3-1-')

        df1['Location'] = df1['Location'].str.replace('2-1-2-','2-2-')

        df1['Location'] = df1['Location'].str.replace('2-1-1-','2-1-')

        df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

        df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

        df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

        df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)

        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
        
    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb1 = load_workbook(Feeder_List_OPB)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb2 = load_workbook(Feeder_List_OPB)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
        #copy from wb1
                c = ws1.cell(row=row, column=10)
        #paste in ws2
                ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPB.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "BOT"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['BOT_Side'] = "BOT"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "BOT"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "BOT"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)  
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE2T

    file_path = 'FeederSetup_TL2.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_TL2.csv', encoding="utf-8",index_col=False, skiprows=range(2))

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 400:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 403.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-2 Slot Count in TOP Feeder '403'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_TL2.csv', encoding="utf-8",index_col=False, skiprows=range(2, 401), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-2 Slot Count in TOP Feeder '401'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

        #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_TL2.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_TL2.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_TL2: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        #df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        #df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('3-1','3')

        df1['Location'] = df1['Location'].str.replace('2-1','2')

        df1['Location'] = df1['Location'].str.replace('1-1','1')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)

        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
        
        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb1 = load_workbook(Feeder_List_OPT)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb2 = load_workbook(Feeder_List_OPT)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
            #copy from wb1
            c = ws1.cell(row=row, column=10)
            #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPT.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "TOP"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['TOP_Side'] = "TOP"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "TOP"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "TOP"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)    
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE2B

    file_path = 'FeederSetup_BL2.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_BL2.csv', encoding="utf-8",index_col=False, skiprows=range(2))

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 400:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 403.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-2 Slot Count in BOT Feeder '403'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_BL2.csv', encoding="utf-8",index_col=False, skiprows=range(2, 401), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-2 Slot Count in BOT Feeder '401'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code
        
        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
        
        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_BL2.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_BL2.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_BL2: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        #df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        #df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('3-1','3')

        df1['Location'] = df1['Location'].str.replace('2-1','2')

        df1['Location'] = df1['Location'].str.replace('1-1','1')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)

        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
        
    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb1 = load_workbook(Feeder_List_OPB)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb2 = load_workbook(Feeder_List_OPB)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
        #copy from wb1
                c = ws1.cell(row=row, column=10)
        #paste in ws2
                ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPB.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "BOT"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['BOT_Side'] = "BOT"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "BOT"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "BOT"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)  
            #df1.to_excel(writer, sheet_name="FS_upload", index=False)
            #df2.to_excel(writer, sheet_name="FS_Count", index=False)
            #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
            #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)   
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE3T

    file_path = 'FeederSetup_TL3.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_TL3.csv', encoding="utf-8",index_col=False, skiprows=range(2))

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 170:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 173.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-3 Slot Count in TOP Feeder '173'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_TL3.csv', encoding="utf-8",index_col=False, skiprows=range(2, 171), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-3 Slot Count in TOP Feeder '171'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

        #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_TL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_TL3.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_TL3: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        #df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

        df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

        df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

        df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

        df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

        df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)

        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
        
        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)


        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb1 = load_workbook(Feeder_List_OPT)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb2 = load_workbook(Feeder_List_OPT)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
            #copy from wb1
            c = ws1.cell(row=row, column=10)
            #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPT.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "TOP"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['TOP_Side'] = "TOP"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "TOP"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "TOP"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)    
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE3B

    file_path = 'FeederSetup_BL3.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_BL3.csv', encoding="utf-8",index_col=False, skiprows=range(2))

            # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 170:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 173.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-3 Slot Count in BOT Feeder '173'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_BL3.csv', encoding="utf-8",index_col=False, skiprows=range(2, 171), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-3 Slot Count in BOT Feeder '171'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code
        
        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
        
        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_BL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_BL3.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_BL3: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        #df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

        df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

        df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

        df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

        df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

        df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)

        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
        
    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb1 = load_workbook(Feeder_List_OPB)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb2 = load_workbook(Feeder_List_OPB)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
        #copy from wb1
                c = ws1.cell(row=row, column=10)
        #paste in ws2
                ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPB.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "BOT"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['BOT_Side'] = "BOT"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "BOT"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "BOT"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)  
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE4T

    file_path = 'FeederSetup_TL4.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_TL4.csv', encoding="utf-8",index_col=False, skiprows=range(2))

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 82:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 85.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-4 Slot Count in TOP Feeder '85'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_TL4.csv', encoding="utf-8",index_col=False, skiprows=range(2, 83), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-4 Slot Count in TOP Feeder '83'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

        #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_TL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_TL4.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_TL4: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        #df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

        df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

        df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

        df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

        #df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

        #df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)
        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
        
        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)


        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb1 = load_workbook(Feeder_List_OPT)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb2 = load_workbook(Feeder_List_OPT)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
            #copy from wb1
            c = ws1.cell(row=row, column=10)
            #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPT.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "TOP"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['TOP_Side'] = "TOP"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "TOP"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "TOP"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)    
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE4B

    file_path = 'FeederSetup_BL4.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_BL4.csv', encoding="utf-8",index_col=False, skiprows=range(2))

            # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 82:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 85.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-4 Slot Count in BOT Feeder '85'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_BL4.csv', encoding="utf-8",index_col=False, skiprows=range(2, 83), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-4 Slot Count in BOT Feeder '85'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code
        
        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
        
        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_BL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_BL4.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_BL4: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        #df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

        df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

        df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

        df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

        #df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

        #df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)
        
        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
        
    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb1 = load_workbook(Feeder_List_OPB)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb2 = load_workbook(Feeder_List_OPB)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
        #copy from wb1
                c = ws1.cell(row=row, column=10)
        #paste in ws2
                ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPB.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "BOT"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['BOT_Side'] = "BOT"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "BOT"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "BOT"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)  
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE1T-C4

    file_path = 'FeederSetup_TL4C.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_TL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2))

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 227:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 230.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-4C Slot Count in TOP Feeder '230'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_TL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2, 228), nrows=3)

            # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-4C Slot Count in TOP Feeder '228'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

        #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_TL4C.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_TL4C.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_TL4C: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        #df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('2-1-1-','7-')

        df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

        df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

        df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

        df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

        df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

        df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)

        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
        
        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb1 = load_workbook(Feeder_List_OPT)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
        wb2 = load_workbook(Feeder_List_OPT)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
            #copy from wb1
            c = ws1.cell(row=row, column=10)
            #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPT.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "TOP"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['TOP_Side'] = "TOP"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "TOP"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "TOP"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)    
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #LINE1B-C4

    file_path = 'FeederSetup_BL4C.csv'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        dt_H1 = pd.read_csv('FeederSetup_BL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2))

        # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 227:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            print("dt_H1 is either None or its length is not equal to 230.")
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-4C Slot Count in BOT Feeder '230'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

        dt_H1 = pd.read_csv('FeederSetup_BL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2, 228), nrows=3)

            # Check if dt_H1 is defined and the line count is 351
        if dt_H1 is not None and len(dt_H1) == 3:
            # Continue with the rest of your code
            print(dt_H1)
        else:
            # Show error message
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            error_message = "Error: Either Check the FeederSetup.csv Line-4C Slot Count in BOT Feeder '228'."
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code
        
        dt_H1['TotalSlots'] = ''
        dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
        
        dt_H1['PlacedParts'] = ''
        dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

        dt_H1['Col1'] = dt_H1['JobName'].str[13:]

        dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

        print(dt_H1.drop(index=[1, 2]))

        dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

        dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

        dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

        dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

        dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

        dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

        N9_Col = dt_H1.pop('Side') # col-10

        dt_H1.insert(9, 'Side', N9_Col)

        dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

        dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

        dt_H1.insert(15, 'CATEGORY', '')
        dt_H1.insert(16, 'MODEL NAME', '')
        dt_H1.insert(17, 'CURRENT REVISION', '')
        dt_H1.insert(18, 'MODIFIED  DATE', '')
        dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
        dt_H1.insert(20, 'BOM ECO NUMBER', '')
        dt_H1['Verify-DateTime'] = datetime.now()
        dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
        dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
        dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
        dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

        dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
        #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

        print(dt_H1)
    #-----------------------------------------------------------------------------------------------------------------------#

    #print(dt_H1.drop(index=[1, 2]))

    # NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
    #df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

    #-----------------------------------------------------------------------------------------------------------------------#

        #df1 = pd.read_csv('FeederSetup_BL4C.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

    # Specify the columns you want to read
        columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

        try:
            df1 = pd.read_csv('FeederSetup_BL4C.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

            # Check if all the specified columns are present in the DataFrame
            if all(column in df1.columns for column in columns_to_read):
                print("All columns are present in the DataFrame.")
            else:
                # Show error message if any columns are missing
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                missing_columns = [column for column in columns_to_read if column not in df1.columns]
                error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

        except Exception as e:
                # Handle the exception gracefully
                error_message = f"An error occurred FeederSetup_BL4C: {e}"

                # Show error message in a pop-up box
                root = tk.Tk()
                root.withdraw()  # Hide the main window
                messagebox.showerror("Error", error_message)
                sys.exit(1)  # Exit the program with an error code

    #df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
    # NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

        df1.dropna(subset=['RefList'], inplace=True)

        df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

        df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

        df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

        print (df1)

        df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

        #df1['ModelName'] = df1['ModelName'].str.replace('1','')

        df1.rename(columns = {'Location':'Lock'}, inplace = True)

        df1.rename(columns = {'LineName':'Location'}, inplace = True)

        df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

        df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

        df1['Location'] = df1['Location'].str.replace('2-1-1-','7-')

        df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

        df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

        df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

        df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

        df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

        df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

        F1_col = df1.pop('PartNumber') # col-1

        df1.insert(1, 'PartNumber', F1_col)

        df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

        S2_col = df1.pop('FeederName') # col-2

        df1.insert(2, 'FeederName', S2_col)

        T3_col = df1.pop('FeederType') # col-3 rename to type

        df1.insert(3, 'FeederType', T3_col)

        df1.rename(columns = {'FeederType':'Type'}, inplace = True)

        F4_col = df1.pop('TapeWidth') # col-4 rename to size

        df1.insert(4, 'TapeWidth', F4_col)

        df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

        F5_col = df1.pop('FeedPitch') # col-5 

        df1.insert(5, 'FeedPitch', F5_col)

        S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

        df1.insert(6, 'PTPMNH', S6_col)

        df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

        S7_col = df1.pop('Status') # col-7

        df1.insert(7, 'Status', S7_col)

        E8_col = df1.pop('QTY') # col-8 

        df1.insert(8, 'QTY', E8_col)

        df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

        df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

        extracted_col = dt_H1["Side"] 

        df1.insert(9, "Side", extracted_col)

        #NEW PN# PARTNO
        df1['PartNO'] = "PN#"
        df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
        del df1['PartNO']
        df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
        
    #dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

            dt_H1.to_excel(writer, sheet_name="S1", index=False)
            df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb1 = load_workbook(Feeder_List_OPB)
        ws1 = wb1.active
        ws1 = wb1.worksheets[0]

        Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
        wb2 = load_workbook(Feeder_List_OPB)
        ws2 = wb2.active
        ws2 = wb2.worksheets[1]

        print(ws1)

        print(ws2)

        for row in range(1, 10):
        #copy from wb1
                c = ws1.cell(row=row, column=10)
        #paste in ws2
                ws2.cell(row=row-0, column=10, value=c.value)

        print(ws2)

        wb2.save(str('Feeder_List_OPB.xlsx'))

        df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

        df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

        df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

        df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

        df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

    # NOTE reflist column to next axis column this line df = nothing df1 split colum next to

    #print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

        df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

        df2.explode ('F_Ref_List')

        df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

        df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

        df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

        df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

        df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

    #df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    # NOTE df3.loc [variable] delete all col after reflist 

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

        df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

        df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

    #print (df3_1) # NOTE line to create dummy new page with old content

        df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

        df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

        df3_1.insert(9, 'Tray Dir','')
        df3_1.insert(10, 'PartComment','')
        df3_1.insert(11, 'Barcode Label','')

        df4_1 = df3_1['Size'].value_counts()
        df4_1['Feedersize'] = "BOT"
        df5_1 = df3_1['Side'].value_counts()
        df5_1['BOT_Side'] = "BOT"
        df6_1 = df3_1['FeederName'].value_counts()
        df6_1['FeederSize'] = "BOT"
        df7_1 = df3_1['Type'].value_counts()
        df7_1['FeederType'] = "BOT"

        df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

        df3.to_dict()

        df3.explode ('F_Ref_List',ignore_index=True)

        df4 = df3.explode('F_Ref_List',ignore_index=True)

        df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

        df4.head()

        #NEW PN# PARTNO
        #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
        #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

            dt_H1.to_excel(writer, sheet_name="Home", index=False)  
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
            df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
            df4.to_excel(writer, sheet_name="FL_Verify", index=False)
            df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
            df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
            df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
            df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

    pass
    print('The file does not exist.')

    ##########################################################################################################################################

    ##########################################################################################################################################

    #bil3 = pyfiglet.figlet_format("FeederSetup Progress Merge", width = 150)
    print('\n')
    print("\033[92;4m******FeederSetup Progress Merge******\033[0m")
    print('\n')
    ##########################################################################################################################################

    ##########################################################################################################################################

    #LEN MERGE TOP AND BOT

    data_file_folder = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

    df=[]
    for file in os.listdir(data_file_folder):
        if file.endswith('.xlsx'):
            print('Loading file {0}...'.format(file))
            df.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Home'))
    df1=[]
    for file in os.listdir(data_file_folder):
        if file.endswith('.xlsx'):
            print('Loading file {0}...'.format(file))
            df1.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FL_Upload'))
    df2=[]
    for file in os.listdir(data_file_folder):
        if file.endswith('.xlsx'):
            print('Loading file {0}...'.format(file))
            df2.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FL_Verify'))
    df3=[]
    for file in os.listdir(data_file_folder):
        if file.endswith('.xlsx'):
            print('Loading file {0}...'.format(file))
            df3.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Size'))
    df4=[]
    for file in os.listdir(data_file_folder):
        if file.endswith('.xlsx'):
            print('Loading file {0}...'.format(file))
            df4.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Side'))
    df5=[]
    for file in os.listdir(data_file_folder):
        if file.endswith('.xlsx'):
            print('Loading file {0}...'.format(file))
            df5.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FeederName'))

    df6=[]
    for file in os.listdir(data_file_folder):
        if file.endswith('.xlsx'):
            print('Loading file {0}...'.format(file))
            df6.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Type'))

    len(df)
    df_master1 = pd.concat(df, axis=0)
    len(df1)
    df_master2 = pd.concat(df1, axis=0)
    len(df2)
    df_master3 = pd.concat(df2, axis=0)
    len(df3)
    df_master4 = pd.concat(df3, axis=0)
    len(df4)
    df_master5 = pd.concat(df4, axis=0)
    len(df5)
    df_master6 = pd.concat(df5, axis=0)
    len(df6)
    df_master7 = pd.concat(df6, axis=0)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederSetup.xlsx") as writer:
            df_master2.to_excel(writer, sheet_name="FeederSetup", index=False)
#++df_master7#

    Chd = os.getcwd()
    Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')

    if os.path.exists("Feeder_List_OPT.xlsx"):
        os.rename("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx" , "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Feeder_List_OPT.xlsx")
    else:
        print("The file does not exist")

    Chd = os.getcwd()
    Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')

    if os.path.exists("Feeder_List_OPB.xlsx"):
        os.rename("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx" , "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Feeder_List_OPB.xlsx")
    else:
        print("The file does not exist")

    Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

    # Process Feeder_List_OPT.xlsx
    try:
        xls = pd.ExcelFile('Feeder_List_OPT.xlsx', engine='openpyxl')
        dffst11 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="Home")
        dffst12 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="FL_Upload")
        dffst13 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="FL_Verify")
        dffst14 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="Size")
        dffst15 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="Side")
        dffst16 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="FeederName")
        dffst17 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="Type")

        dffst12['F_Part_No'] = dffst12['F_Part_No'].astype(str).str.replace('PN#', '')
        dffst13['F_Part_No'] = dffst13['F_Part_No'].astype(str).str.replace('PN#', '')

        with pd.ExcelWriter('Feeder_List_OPT.xlsx', engine='openpyxl', mode='w') as writer:
            dffst12.to_excel(writer, sheet_name="FL_Upload", index=False)

    except FileNotFoundError:
        print("Feeder_List_OPT.xlsx not found. Skipping processing.")

    # Process Feeder_List_OPB.xlsx
    try:
        xls = pd.ExcelFile('Feeder_List_OPB.xlsx', engine='openpyxl')
        dffsb11 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="Home")
        dffsb12 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="FL_Upload")
        dffsb13 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="FL_Verify")
        dffsb14 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="Size")
        dffsb15 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="Side")
        dffsb16 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="FeederName")
        dffsb17 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="Type")

        dffsb12['F_Part_No'] = dffsb12['F_Part_No'].astype(str).str.replace('PN#', '')
        dffsb13['F_Part_No'] = dffsb13['F_Part_No'].astype(str).str.replace('PN#', '')

        with pd.ExcelWriter('Feeder_List_OPB.xlsx', engine='openpyxl', mode='w') as writer:
            dffsb12.to_excel(writer, sheet_name="FL_Upload", index=False)

    except FileNotFoundError:
        print("Feeder_List_OPB.xlsx not found. Skipping processing.")

    ##########################################################################################################################################

    ##########################################################################################################################################

    #bil5 = pyfiglet.figlet_format("BOM Manipulation", width = 100)
    print('\n')
    print("\033[92;4m******BOM Manipulation******\033[0m")
    print('\n')
    ##########################################################################################################################################

    ##########################################################################################################################################

    #BOM MANIPULATE

    try:
        # BOM MANIPULATION
        os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
        file_path = 'BOM.xlsx'

        if os.path.isfile(file_path):
            ds1 = pd.read_excel(file_path, index_col=False)
        else:
            # Try reading as '.xls' format if '.xlsx' fails
            file_path = 'BOM.xls'
            ds1 = pd.read_excel(file_path, index_col=False)

        dfbom1 = ds1

    except ValueError:
        dfbom1 = pd.read_excel(file_path,index_col=False) 

    except Exception as e:
        # Handle the exception gracefully
        error_message = f"An error occurred: {e}"

        # Show error message in a pop-up box
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    # Define your column lists
    column_list_1 = ['Material', 'AltItemGroup', 'Priority', 'Long. Description', 'Ref.Designator/Circuit Reference', 'Quantity', 'Material Group']
    column_list_2 = ['Internal P/N', 'Group', 'Priority', 'Description', 'Ref.Designator', 'Qty', 'SMT/THT/Mech']

    # Check which column list is present in the DataFrame
    if all(column in ds1.columns for column in column_list_1):
        columns_to_use = column_list_1
    elif all(column in ds1.columns for column in column_list_2):
        columns_to_use = column_list_2
    else:
        # Show error message if none of the column lists is present
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        missing_columns = [column for column_list in [column_list_1, column_list_2] for column in column_list if column not in ds1.columns]
        error_message = f"The following columns are missing: {', '.join(missing_columns)}"
        error_msgbm1 = f"The following columns are missing: in SAP BOM\n'Material'\n'AltItemGroup'\n'Priority'\n'Long. Description'\n'Ref.Designator/Circuit Reference'\n'Quantity'\n'Material Group'"
        error_msgbm2 = f"The following columns are missing: in Internal BOM\n'Internal P/N'\n'Group'\n'Priority'\n'Description'\n'Ref.Designator'\n'Qty'\n'SMT/THT/Mech'"
        messagebox.showerror("Error", error_message)
        messagebox.showerror("Error", error_msgbm1)
        messagebox.showerror("Error", error_msgbm2)
        sys.exit(1)  # Exit the program with an error code

    # Continue with the rest of your code using 'columns_to_use'
    print(f"Using columns: {columns_to_use}")

    # Rest of your code here
    # ...

    ds1.rename(
        columns={'Material':"PartNumber", 'AltItemGroup':"Group", 'Priority':'Priority', 'Long. Description':'Long Des', 'Ref.Designator/Circuit Reference':'RefList', 'Quantity':'Qty','Material Group':'Shape'},
        inplace=True,
    )

    ds1.rename(
        columns={'Internal P/N':"PartNumber", 'Group':"Group", 'Priority':'Priority', 'Description':'Long Des', 'Ref.Designator':'RefList', 'Qty':'Qty','SMT/THT/Mech':'Shape'},
        inplace=True,
    )

    print(ds1)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM/BOM_List_OP.xlsx") as writer:
        ds1.to_excel(writer, sheet_name="BOM", index=False)

    pass
    print('The file does not exist.')

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
    Chd = os.getcwd()
    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        ds1 = pd.read_excel(file_path, sheet_name="BOM", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False) 
        
        ds1 = ds1[['PartNumber', 'Group','Priority','Long Des','Qty','Shape','RefList']]
        ds1['RefList'] = ds1['RefList'].str.replace("_x000D_","")
        ds1['RefList'] = ds1['RefList'].str.replace(" ","")
        ds1['RefList'] = ds1['RefList'].str.replace("\n","")
        print(ds1)
        ds3 = ds1[['PartNumber','Long Des']]
        ds3.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)
        ds3 = ds3[['F_Part_No','Long Des']]

        #NEW PN# PARTNO
        ds1['PartNO'] = "PN#"
        ds1["PartNumber"] = ds1['PartNO'].astype(str) +""+ ds1['PartNumber'].astype(str)
        del ds1['PartNO']
        ds3['PartNO'] = "PN#"
        ds3["F_Part_No"] = ds3['PartNO'].astype(str) +""+ ds3['F_Part_No'].astype(str)
        del ds3['PartNO']

        with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx") as writer:
            ds1.to_excel(writer, sheet_name="BOM", index=False)
            ds3.to_excel(writer, sheet_name="BOM_Data", index=False)

        pass
        print('The file does not exist.')

    ##########################################################################################################################################
    ##########################################################################################################################################
    #FEEDER VERIFICATION CODE BOM AND FEEDER AS VISE VERSA

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    xls = pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
    df1 = pd.read_excel('BOM_List_OP.xlsx', sheet_name='BOM_Data')
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    xls = pd.ExcelFile('FeederSetup.xlsx',engine='openpyxl')
    df112 = pd.read_excel('FeederSetup.xlsx', sheet_name='FeederSetup')
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx") as writer:
        df1.to_excel(writer, sheet_name="BOM_Data", index=False)
        df112.to_excel(writer, sheet_name="FeederSetup", index=False)
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    from Program5 import program_5
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    xls = pd.ExcelFile('FeederVerify.xlsx',engine='openpyxl')
    df111 = pd.read_excel("FeederVerify.xlsx", sheet_name='BOM_Data')
    df112 = pd.read_excel("FeederVerify.xlsx", sheet_name='FeederSetup')

    df111 = df111[['F_Part_No','Long Des']]
    df113 = pd.merge(df112 , df111, on='F_Part_No', how='inner') # Merge on 'F_Part_No'
    df113.rename(columns = {'F_Part_No':'Part Number'}, inplace = True)
    df113.rename(columns = {'Location':'Feeder Location'}, inplace = True)
    df113.rename(columns = {'Long Des':'Part Description'}, inplace = True)
    df113.rename(columns = {'F_Ref_List':'Reference'}, inplace = True)
    df113 = df113[['Feeder Location','FeederName','Type','Size','FeedPitch','Part Height','Part Number','Part Description','Reference','QTY','Side','ModelName']]

    #NEW PN# PART NO
    df113['Part Number'] = df113['Part Number'].str.replace('PN#','')

    # Save the styled DataFrame to Excel
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx") as writer:
        df113.to_excel(writer, sheet_name="Upload_data", index=False)

    ###################################################################################################################
    ###################################################################################################################
    ###################################################################################################################
    ###################################################################################################################
        
    #CREATE & SEPRATE FEEDER LOADING LIST DATA

    os.getcwd()

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()

    df1 = pd.read_excel("FeederVerify.xlsx", sheet_name="Upload_data")

    df1.sort_values(by='Side', inplace=True, ascending=True)
    df2_1 = df1
    #del.TOP
    df1 = df1[df1["Side"].str.contains("TOP")==False]
    df2 = df1[df1["ModelName"].str.contains("AIMEX2|AIMEX3|AIMEX-IIIC_2|AIMEX-IIIC_3")==False]
    df2.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df3 = df1[df1["ModelName"].str.contains("NXT|AIMEX3|AIMEX-IIIC_1|AIMEX-IIIC_3")==False]
    df3.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df4 = df1[df1["ModelName"].str.contains("NXT|AIMEX2|AIMEX-IIIC_1|AIMEX-IIIC_2")==False]
    df4.sort_values(by='Feeder Location', inplace=True, ascending=True)
    #del.BOT
    df2_1 = df2_1[df2_1["Side"].str.contains("BOT")==False]
    df2_2 = df2_1[df2_1["ModelName"].str.contains("AIMEX2|AIMEX3|AIMEX-IIIC_2|AIMEX-IIIC_3")==False]
    df2_2.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df2_3 = df2_1[df2_1["ModelName"].str.contains("NXT|AIMEX3|AIMEX-IIIC_1|AIMEX-IIIC_3")==False]
    df2_3.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df2_4 = df2_1[df2_1["ModelName"].str.contains("NXT|AIMEX2|AIMEX-IIIC_1|AIMEX-IIIC_2")==False]
    df2_4.sort_values(by='Feeder Location', inplace=True, ascending=True)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data.xlsx") as writer:
        df2.to_excel(writer, sheet_name="NXT&AMX1_B", index=False)
        df3.to_excel(writer, sheet_name="AMX2_B", index=False)
        df4.to_excel(writer, sheet_name="AMX3_B", index=False)

        df2_2.to_excel(writer, sheet_name="NXT&AMX1_T", index=False)
        df2_3.to_excel(writer, sheet_name="AMX2_T", index=False)
        df2_4.to_excel(writer, sheet_name="AMX3_T", index=False)

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    ##########################################################################################################################################

    #Upload data to merge and del side and Module

    os.getcwd()

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

    df1 = pd.read_excel("Upload-Data.xlsx", sheet_name="NXT&AMX1_B")
    df1["Remarks"] = df1['Side'].astype(str) +"--"+ df1['ModelName']
    del df1['Side']
    del df1['ModelName']

    df2 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX2_B")
    df2["Remarks"] = df2['Side'].astype(str) +"--"+ df2['ModelName']
    del df2['Side']
    del df2['ModelName']

    df3 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX3_B")
    df3["Remarks"] = df3['Side'].astype(str) +"--"+ df3['ModelName']
    del df3['Side']
    del df3['ModelName']

    df4 = pd.read_excel("Upload-Data.xlsx", sheet_name="NXT&AMX1_T")
    df4["Remarks"] = df4['Side'].astype(str) +"--"+ df4['ModelName']
    del df4['Side']
    del df4['ModelName']

    df5 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX2_T")
    df5["Remarks"] = df5['Side'].astype(str) +"--"+ df5['ModelName']
    del df5['Side']
    del df5['ModelName']

    df6 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX3_T")
    df6["Remarks"] = df6['Side'].astype(str) +"--"+ df6['ModelName']
    del df6['Side']
    del df6['ModelName']

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx") as writer:
        df1.to_excel(writer, sheet_name="NXT&AMX1_B", index=False)
        df2.to_excel(writer, sheet_name="AMX2_B", index=False)
        df3.to_excel(writer, sheet_name="AMX3_B", index=False)
        df4.to_excel(writer, sheet_name="NXT&AMX1_T", index=False)
        df5.to_excel(writer, sheet_name="AMX2_T", index=False)
        df6.to_excel(writer, sheet_name="AMX3_T", index=False)

    ##########################################################################################################################################

    ##########################################################################################################################################

    #bil10 = pyfiglet.figlet_format("Feeder Loading List Progress", width = 200)
    print('\n')
    print("\033[1;92;4m******Feeder Loading List Progress******\033[0m")
    print('\n')
    ##########################################################################################################################################

    ##########################################################################################################################################

    #Feeder List change

    shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_FeederSetup/Line X Sample.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx')
    shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_FeederSetup/Line X Sample.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx')

        ##BOT FEEDER LIST

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'NXT&AMX1_B'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
    destination_sheet_name = 'NXT'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "NXT&AMX1_B", "NXT", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'AMX2_B'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
    destination_sheet_name = 'AIMEX 2'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "AMX2_B", "AIMEX 2", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'AMX3_B'

        # Destination Excel file 
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
    destination_sheet_name = 'AIMEX 3'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "AMX3_B", "AIMEX 3", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

        ##TOP FEEDER LIST

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'NXT&AMX1_T'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
    destination_sheet_name = 'NXT'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "NXT&AMX1_T", "NXT", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'AMX2_T'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
    destination_sheet_name = 'AIMEX 2'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "AMX2_T", "AIMEX 2", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'AMX3_T'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
    destination_sheet_name = 'AIMEX 3'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "AMX3_T", "AIMEX 3", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    ##########################################################################################################################################

    #CREATEBACKUPFOLDER

    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)

    yourfolder1 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\BOM"

    if not os.path.isdir(yourfolder1):
        print('Folder Not Exist')
        os.makedirs(yourfolder1)

    yourfolder2 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\FeederSetup"

    if not os.path.isdir(yourfolder2):
        print('Folder Not Exist')
        os.makedirs(yourfolder2)

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
    Chd = os.getcwd()

    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        if os.path.exists("BOM_List_OP.xlsx"):
            os.remove("BOM_List_OP.xlsx")
    else:
        print("The file does not exist")

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

    if os.path.exists("Feeder_List_OPT.xlsx"):
        os.remove("Feeder_List_OPT.xlsx")
    else:
        print("The file does not exist")

    if os.path.exists("Feeder_List_OPB.xlsx"):
        os.remove("Feeder_List_OPB.xlsx")
    else:
        print("The file does not exist")

    ##########################################################################################################################################

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')

    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Line_X"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)

    # Get the current working directory
    os.getcwd()

    # Change directory to the location of your Excel files
    os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')

    # Get the current working directory
    Chd = os.getcwd()

    # Load the workbooks
    workbook_T = openpyxl.load_workbook('Line X Sample T.xlsx')
    workbook_B = openpyxl.load_workbook('Line X Sample B.xlsx')

    # Specify the worksheet names
    worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

    # Input values for cell B3 and Revision A1
    #print('\n')
    #value_B3 = input("\033[93mEnter Feeder Name (12 characters): \033[0m").strip()[:12]  # Take only the first 12 characters and remove extra spaces
    #print('\n')
    #Revision = input("\033[93mEnter Revision A1: \033[0m")

    # Iterate over each workbook
    for workbook, workbook_name in [(workbook_T, 'Line X Sample T.xlsx'), (workbook_B, 'Line X Sample B.xlsx')]:
        if 'T' in workbook_name:
            location = 'T'
        elif 'B' in workbook_name:
            location = 'B'
        else:
            location = 'T/B'

        # Iterate over each sheet in the workbook
        for sheet_name in worksheet_names:
            # Select the worksheet
            worksheet = workbook[sheet_name]

            # Iterate over all rows in the worksheet
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                # Check if the row is empty
                if all(cell.value is None for cell in row):
                    # If the row is empty, hide it
                    worksheet.row_dimensions[row[0].row].hidden = True

            # Combine the input values for B3 & K4 cell
            worksheet['B3'] = value_B3 + " " + location + " " + Revision
            worksheet['K4'] = dLine123

    # Save the workbooks
    workbook_T.save('Line X Sample T.xlsx')
    workbook_B.save('Line X Sample B.xlsx')

    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/Line X Sample T.xlsx')
    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/Line X Sample B.xlsx')

    ##########################################################################################################################################
    ##########################################################################################################################################

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X')
    Chd = os.getcwd()

    src_8 = 'Line X Sample T.xlsx'
    os.rename(src_8, dL1 +"_T_"+ Revision +".xlsx")

    src_9 = 'Line X Sample B.xlsx'
    os.rename(src_9, dL1 +"_B_"+ Revision +".xlsx")

    # Specify the current name of the folder
    cfn1 = "Line_X"

    # Rename the folder
    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
    Chd = os.getcwd()
    os.rename(cfn1, dL1 +"-"+ Revision)

    print(f"\033[92mFolder '{cfn1}' renamed successfully to '{dL1}'.\033[0m")

    # Function to move a folder to a specified destination
    def move_folder(src, dst):
        try:
            shutil.move(src, dst)
            sg.popup(f"Folder '{os.path.basename(src)}' moved successfully to '{dst}'")
        except Exception as e:
            sg.popup_error(f"Error occurred while moving folder: {e}")

    # Function to create a new customer name
    def create_customer_name(line):
        new_customer = sg.popup_get_text(f"Enter the name of the new customer for '{line}':", title="Create New Customer")
        if new_customer:
            line_path = os.path.join(destination_path, line)
            new_customer_path = os.path.join(line_path, new_customer)
            os.makedirs(new_customer_path, exist_ok=True)
            return new_customer
        else:
            return None

    # Define the root directory
    root_dir = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output'

    # Define the destination path
    destination_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#_Feeder Loading Line List'

    # Get a list of lines from the destination path
    lines = [name for name in os.listdir(destination_path) if os.path.isdir(os.path.join(destination_path, name))]

    # Define the layout for the PySimpleGUI window
    layout = [
        [sg.Text("Select the folder to move:")],
        [sg.Listbox(values=os.listdir(root_dir), size=(50, 6), key='-FOLDER LIST-', enable_events=True)],
        [sg.Text("Select the line:")],
        [sg.Listbox(values=lines, size=(50, 3), key='-LINE-', enable_events=True)],
        [sg.Text("Select or create a customer name:"), sg.Button("New Customer")],
        [sg.Listbox(values=[], size=(50, 6), key='-CUSTOMER-', enable_events=True)],
        #[sg.Listbox(values=[], size=(50, 6), key='-CUSTOMER-', enable_events=True), sg.Button("New Customer")],
        [sg.Button("Move"), sg.Button("Cancel")]
    ]

    # Create the PySimpleGUI window
    window = sg.Window("Move Folder", layout)

    # Event loop for the PySimpleGUI window
    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == "Cancel":
            break
        elif event == "-FOLDER LIST-":
            selected_folder = values['-FOLDER LIST-'][0]
            window['-CUSTOMER-'].update(values=[])
        elif event == "-LINE-":
            line = values['-LINE-'][0]
            customers = os.listdir(os.path.join(destination_path, line)) if os.path.exists(os.path.join(destination_path, line)) else []
            window['-CUSTOMER-'].update(values=customers)
        elif event == "New Customer":
            line = values['-LINE-'][0]
            new_customer = create_customer_name(line)
            if new_customer:
                window['-CUSTOMER-'].update(values=[new_customer])
        elif event == "Move":
            selected_folder = values['-FOLDER LIST-'][0]
            line = values['-LINE-'][0]
            customer = values['-CUSTOMER-'][0]
            if not selected_folder:
                sg.popup_error("Please select a folder to move")
            elif not line:
                sg.popup_error("Please select a line")
            elif not customer:
                sg.popup_error("Please select or create a customer")
            else:
                # Move the folder
                old_folder_path = os.path.join(root_dir, selected_folder)
                new_folder_path = os.path.join(destination_path, line, customer)
                move_folder(old_folder_path, new_folder_path)
                break

    # Close the PySimpleGUI window
    window.close()

    time.sleep (2)

    time.sleep (5)

    print('\n')
    print('\033[92;3mFeeder Setup Loading List Generation Complete\033[0m')
    print('\n')

    # Assuming feeder verification is completed
    Feeder_List_Generation_Completed = True

    if Feeder_List_Generation_Completed:
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        time.sleep (5)
        messagebox.showinfo("Feeder Loading List", "Feeder Loading List has been Generated!")
        
        sys.exit() #LoadingList X