import pandas as pd
import numpy as np
import os
from os import getcwd
import csv as csv
from io import StringIO
import re
import datetime
from io import BytesIO
import time
import sys
from datetime import datetime
import shutil
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import subprocess
import threading
import urllib
import urllib.parse
from sqlite3 import dbapi2 as sqlite
import sqlite3
import lxml
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
import PySimpleGUI as sg
import sqlalchemy
from sqlalchemy import create_engine
import sqlalchemy_access as sq_a
import sqlalchemy_pyodbc_mssql as sqlalchemy
from flask_sqlalchemy import SQLAlchemy
import pyodbc
import pypyodbc
import odbc
import psycopg2
import mysql.connector as sql
import MySQLdb as sql #pip install mysqlclient
from plyer import notification
import xlrd
import xml.etree.ElementTree as ET
from openpyxl.utils.exceptions import InvalidFileException
import psutil

print('\n')
print("\033[32;4m*******Syrma Feeder LoadingList Version--PY_V-2.5 interface_GUI/M1124-89P13*******\033[0m")

'''bil1 = pyfiglet.figlet_format("Version--PY-V1.5 interface_GUI/J0324", width = 300)print(bil1)'''

# Get the current date and time
current_datetime = datetime.now()

# Format the current date and time as a string
#formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
# Format the date and time in a 12-hour clock with AM/PM
formatted_datetime = current_datetime.strftime("%Y-%m-%d %I:%M:%S %p")

print('\n')

# Print the formatted date and time
print(f"\033[31mCurrent Date and Time: {formatted_datetime}\033[0m")

print('\n')

'''print(f"Current Year: {current_datetime.year}")
print(f"Current Month: {current_datetime.month}")
print(f"Current Day: {current_datetime.day}")
print(f"Current Hour: {current_datetime.hour}")
print(f"Current Minute: {current_datetime.minute}")
print(f"Current Second: {current_datetime.second}")'''

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print("\033[32;4m*******FeederSetup*******\033[0m")
######################################################################################################



# FeederSetup

def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_TL1.xml"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_BL1.xml"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_TL2.xml"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_BL2.xml"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_TL3.xml"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_BL3.xml"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_TL4.xml"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_BL4.xml"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_TL4C.xml"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_BL4C.xml"
    else:
        return
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print('\n')
    print(f"FeederSetup.xml renamed to: {new_path}")
    print('\n')
    return new_path

    
# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Find FeederSetup.csv files
feeder_setup_files = find_feeder_setup_files(root_directory)



'''# Function to log the usage and input
def log_usage_and_input(dL1):
    # Get the current date and time
    #current_datetime1 = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Create or open the log file
    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Database_File/SMT_Log')
    Chd = os.getcwd()
    log_file_path = 'Loading_List_log.txt'
    with open(log_file_path, 'a') as log_file:
        # Write the usage information to the log file
        log_file.write(f"{formatted_datetime}: App used. Input: {dL1}\n")'''

'''import csv

# Sample input data
input_data = """
LineName: Line1
        OrderNum: 2
        ModelName: AIMEX2
        ModuleNumber: 1
        SideNo: 1
        SetupName: Original setup
            AVLNAME: 1110036257
            AVL: 1110036257,1110036950,1110039007
            PartNumber: 1110036257
            FeederName: KT-0800F-180
            Status: Fixed
            Location: 32
            FeederID: 22032
            PartNumberImage: Images\\1110036257.bmp
            PartShapeName: 1110036257
            PackageName: 0802P
            FeederType: Paper
            TapeWidth: 8mm
            FeedPitch: 2
            PTPMNH: 0.28
            QTY: 1
            RefList: 1:RN2
            PMADC: 0
        LineName: Line1
        OrderNum: 2
        ModelName: AIMEX2
        ModuleNumber: 1
        SideNo: 1
        SetupName: Original setup
            AVLNAME: 1100028722
            AVL: 1100028722,1110038821
            PartNumber: 1100028722
            FeederName: KT-0800F-180
            Status: Fixed
            Location: 33
            FeederID: 22033
            PartNumberImage: Images\\1100028722.bmp
            PartShapeName: 1100028722
            PackageName: 0802P
            FeederType: Paper
            TapeWidth: 8mm
            FeedPitch: 2
            PTPMNH: 0.45
            QTY: 3
            RefList: 1:RN1 1:RN3 1:RN4
            PMADC: 0
"""

# Split the input data into lines
lines = input_data.splitlines()

# Initialize an empty list to store extracted data
extracted_data = []

# Temporary storage for the current set of values
current_data = {}

# Iterate through each line to extract relevant fields
for line in lines:
    line = line.strip()
    if line.startswith("AVLNAME:"):
        current_data["AVLNAME"] = line.split(":")[1].strip()
    elif line.startswith("AVL:"):
        current_data["AVL"] = f"'{line.split(':')[1].strip()}'"
    elif line.startswith("PartNumber:"):
        current_data["PartNumber"] = line.split(":")[1].strip()
    elif line.startswith("PartNumberImage:"):
        current_data["PartNumberImage"] = line.split(":")[1].strip()
        # Append current_data to extracted_data and reset current_data
        extracted_data.append(current_data)
        current_data = {}

# Define the output CSV file name
output_csv = 'output.csv'

# Write the results to the CSV file
with open(output_csv, mode='w', newline='') as file:
    writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
    writer.writeheader()
    for data in extracted_data:
        writer.writerow(data)

print(f"Data has been saved to {output_csv}")'''



import os
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import xml.etree.ElementTree as ET

def xml_to_txt(xml_file, txt_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    def extract_text(element, level=0):
        text = ''
        indent = ' ' * (level * 4)  # Indent each level by 4 spaces
        if element.text and element.text.strip():
            text += f'{indent}{element.tag}: {element.text.strip()}\n'
        for child in element:
            text += extract_text(child, level + 1)
        return text

    with open(txt_file, 'w', encoding='utf-8') as file:
        file.write(extract_text(root))

# Example usage
xml_file = 'FeederSetup.xml'  # Replace with your XML file path
txt_file = 'FeederSetup.txt'  # Replace with your desired TXT file path
xml_to_txt(xml_file, txt_file)


# Define the file paths
input_file = 'FeederSetup.txt'
output_csv = 'output.csv'
image_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\ECH1ESN00017 T&B;B1;Bottom;MODEL1;[Bottom] Line1;Original setup\Images"
output_excel = 'output.xlsx'

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Extract the data
data = extract_data(input_file)

# Write the results to the CSV file
with open(output_csv, mode='w', newline='') as file:
    writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
    writer.writeheader()
    for data_row in data:
        # Ensure AVL field is properly quoted for CSV format
        data_row["AVL"] = f"'{data_row['AVL']}'"
        writer.writerow(data_row)

# Write the data to an Excel file
wb = Workbook()
ws = wb.active
ws.title = "Feeder Setup Data"

# Write headers
headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
ws.append(headers)

# Write data rows
for row in data:
    # Ensure AVL field is properly quoted for Excel format
    row["AVL"] = f"'{row['AVL']}'"
    ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

# Insert images into the Excel file
for index, row in enumerate(data, start=2):  # start=2 to account for header row
    image_path = os.path.join(image_directory, os.path.basename(row["PartNumberImage"]))
    if os.path.exists(image_path):
        img = Image(image_path)
        ws.add_image(img, f"D{index}")  # Adjust the column and row as per your requirement

# Save the Excel file
wb.save(output_excel)

print(f"Data has been saved to {output_csv} and {output_excel}")

import os
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import xml.etree.ElementTree as ET

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
image_directory = root_directory

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_TL1.txt"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_BL1.txt"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_TL2.txt"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_BL2.txt"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_TL3.txt"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_BL3.txt"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_TL4.txt"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_BL4.txt"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_TL4C.txt"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_BL4C.txt"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to TXT
def xml_to_txt(xml_file, txt_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    def extract_text(element, level=0):
        text = ''
        indent = ' ' * (level * 4)  # Indent each level by 4 spaces
        if element.text and element.text.strip():
            text += f'{indent}{element.tag}: {element.text.strip()}\n'
        for child in element:
            text += extract_text(child, level + 1)
        return text

    with open(txt_file, 'w', encoding='utf-8') as file:
        file.write(extract_text(root))

# Processing FeederSetup files
feeder_setup_files = find_feeder_setup_files(root_directory)
for feeder_setup_file in feeder_setup_files:
    # Convert XML to TXT
    txt_file = feeder_setup_file.replace(".xml", ".txt")
    xml_to_txt(feeder_setup_file, txt_file)
    
    # Rename TXT file based on description
    renamed_txt_file = rename_feeder_setup_with_description(txt_file)

# Define file paths
input_file = renamed_txt_file if renamed_txt_file else 'FeederSetup.txt'
output_csv = 'output.csv'
output_excel = 'output.xlsx'

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Extract the data
data = extract_data(input_file)

# Write the results to the CSV file
with open(output_csv, mode='w', newline='') as file:
    writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
    writer.writeheader()
    for data_row in data:
        # Ensure AVL field is properly quoted for CSV format
        data_row["AVL"] = f"'{data_row['AVL']}'"
        writer.writerow(data_row)

# Write the data to an Excel file
wb = Workbook()
ws = wb.active
ws.title = "Feeder Setup Data"

# Set default row height
for row in range(1, len(data) + 2):  # +2 to account for header row and start index
    ws.row_dimensions[row].height = 100

# Set column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 120
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25

# Write headers
headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
ws.append(headers)

# Write data rows
for row in data:
    # Ensure AVL field is properly quoted for Excel format
    row["AVL"] = f"'{row['AVL']}'"
    ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

# Insert images into the Excel file
for index, row in enumerate(data, start=2):  # start=2 to account for header row
    image_path = os.path.join(image_directory, os.path.basename(row["PartNumberImage"]))
    if os.path.exists(image_path):
        img = Image(image_path)
        img.height = 100  # Adjust image height if necessary
        img.width = 100   # Adjust image width if necessary
        ws.add_image(img, f"D{index}")  # Adjust the column and row as per your requirement

# Save the Excel file
wb.save(output_excel)

print(f"Data has been saved to {output_csv} and {output_excel}")






import os
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import xml.etree.ElementTree as ET

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
image_directory = root_directory

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_TL1.txt"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_BL1.txt"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_TL2.txt"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_BL2.txt"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_TL3.txt"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_BL3.txt"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_TL4.txt"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_BL4.txt"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_TL4C.txt"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_BL4C.txt"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to TXT
def xml_to_txt(xml_file, txt_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    def extract_text(element, level=0):
        text = ''
        indent = ' ' * (level * 4)  # Indent each level by 4 spaces
        if element.text and element.text.strip():
            text += f'{indent}{element.tag}: {element.text.strip()}\n'
        for child in element:
            text += extract_text(child, level + 1)
        return text

    with open(txt_file, 'w', encoding='utf-8') as file:
        file.write(extract_text(root))

# Processing FeederSetup files
feeder_setup_files = find_feeder_setup_files(root_directory)
renamed_txt_file = None
for feeder_setup_file in feeder_setup_files:
    # Convert XML to TXT
    txt_file = feeder_setup_file.replace(".xml", ".txt")
    xml_to_txt(feeder_setup_file, txt_file)
    
    # Rename TXT file based on description
    renamed_txt_file = rename_feeder_setup_with_description(txt_file)

# Define file paths
input_file = renamed_txt_file if renamed_txt_file else 'FeederSetup.txt'

################################################################################

# Extract directory and base name from input file path
output_dir = os.path.dirname(input_file)
output_base_name = os.path.splitext(os.path.basename(input_file))[0]

output_csv = os.path.join(output_dir, f'{output_base_name}.csv')
output_excel = os.path.join(output_dir, f'{output_base_name}.xlsx')

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Extract the data
data = extract_data(input_file)

# Write the results to the CSV file
with open(output_csv, mode='w', newline='') as file:
    writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
    writer.writeheader()
    for data_row in data:
        # Ensure AVL field is properly quoted for CSV format
        data_row["AVL"] = f"'{data_row['AVL']}'"
        writer.writerow(data_row)

# Write the data to an Excel file
wb = Workbook()
ws = wb.active
ws.title = "Feeder Setup Data"

# Set default row height
for row in range(1, len(data) + 2):  # +2 to account for header row and start index
    ws.row_dimensions[row].height = 100

# Set column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 100
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25

# Write headers
headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
ws.append(headers)

# Write data rows
for row in data:
    # Ensure AVL field is properly quoted for Excel format
    row["AVL"] = f"'{row['AVL']}'"
    ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

# Insert images into the Excel file
for index, row in enumerate(data, start=2):  # start=2 to account for header row
    image_path = os.path.join(image_directory, os.path.basename(row["PartNumberImage"]))
    if os.path.exists(image_path):
        img = Image(image_path)
        img.height = 100  # Adjust image height if necessary
        img.width = 100   # Adjust image width if necessary
        ws.add_image(img, f"D{index}")  # Adjust the column and row as per your requirement

# Save the Excel file
wb.save(output_excel)

print(f"Data has been saved to {output_csv} and {output_excel}")






import os
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import xml.etree.ElementTree as ET
import shutil

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
image_directory = root_directory

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to TXT
def xml_to_txt(xml_file, txt_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    def extract_text(element, level=0):
        text = ''
        indent = ' ' * (level * 4)  # Indent each level by 4 spaces
        if element.text and element.text.strip():
            text += f'{indent}{element.tag}: {element.text.strip()}\n'
        for child in element:
            text += extract_text(child, level + 1)
        return text

    with open(txt_file, 'w', encoding='utf-8') as file:
        file.write(extract_text(root))

# Processing FeederSetup files
feeder_setup_files = find_feeder_setup_files(root_directory)
renamed_txt_file = None
for feeder_setup_file in feeder_setup_files:
    # Convert XML to TXT
    txt_file = feeder_setup_file.replace(".xml", ".txt")
    xml_to_txt(feeder_setup_file, txt_file)
    
    # Rename TXT file based on description
    renamed_txt_file = rename_feeder_setup_with_description(txt_file)
    
    # Define file paths
    input_file = renamed_txt_file if renamed_txt_file else 'FeederSetup.txt'
    destination_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
    
    if os.path.exists(input_file):
        # Move the file to the destination directory
        shutil.move(input_file, destination_directory)
    else:
        print(f"File not found: {input_file}. Skipping...")


#save the data in same sheet

import os
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Define the file paths
input_file_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.txt"
input_file_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.txt"
output_csv = 'FeederSetup_Data.csv'
output_excel = 'FeederSetup_Data.xlsx'

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Extract data from both files
data = []
if os.path.exists(input_file_B):
    data.extend(extract_data(input_file_B))

if os.path.exists(input_file_T):
    data.extend(extract_data(input_file_T))

# Write the results to the CSV file
with open(output_csv, mode='w', newline='') as file:
    writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
    writer.writeheader()
    for data_row in data:
        # Ensure AVL field is properly quoted for CSV format
        data_row["AVL"] = f"'{data_row['AVL']}'"
        writer.writerow(data_row)

# Write the data to an Excel file
wb = Workbook()
ws = wb.active
ws.title = "Feeder Setup Data"

# Set default row height
for row in range(1, len(data) + 2):  # +2 to account for header row and start index
    ws.row_dimensions[row].height = 100

# Set column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 100
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25

# Write headers
headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
ws.append(headers)

# Write data rows
for row in data:
    # Ensure AVL field is properly quoted for Excel format
    row["AVL"] = f"'{row['AVL']}'"
    ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

# Save the Excel file
wb.save(output_excel)

print(f"Data has been saved to {output_csv} and {output_excel}")




#############################

#merge file code

import os
import csv
from openpyxl import Workbook

# Define the file paths
input_file_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.txt"
input_file_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.txt"
output_csv = 'FeederSetup_Data.csv'
output_excel = 'FeederSetup_Data.xlsx'

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Extract data from both files
data_B = extract_data(input_file_B) if os.path.exists(input_file_B) else []
data_T = extract_data(input_file_T) if os.path.exists(input_file_T) else []

# Write the results to the CSV file
with open(output_csv, mode='w', newline='') as file:
    writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
    writer.writeheader()
    for data_row in data_B + data_T:
        # Ensure AVL field is properly quoted for CSV format
        data_row["AVL"] = f"'{data_row['AVL']}'"
        writer.writerow(data_row)

# Create a new Excel workbook and add sheets for B and T data
wb = Workbook()
ws_B = wb.active
ws_B.title = "FeederSetup_AVL_B"

ws_T = wb.create_sheet(title="FeederSetup_AVL_T")

# Function to write data to a given worksheet
def write_data_to_sheet(ws, data):
    # Set default row height
    for row in range(1, len(data) + 2):  # +2 to account for header row and start index
        ws.row_dimensions[row].height = 100

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25

    # Write headers
    headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
    ws.append(headers)

    # Write data rows
    for row in data:
        # Ensure AVL field is properly quoted for Excel format
        row["AVL"] = f"'{row['AVL']}'"
        ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

# Write data to respective sheets
write_data_to_sheet(ws_B, data_B)
write_data_to_sheet(ws_T, data_T)

# Save the Excel file
wb.save(output_excel)

print(f"Data has been saved to {output_csv} and {output_excel}") 





#################################################
#split file excel

import os
import csv
from openpyxl import Workbook

# Define the file paths
input_file_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.txt"
input_file_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.txt"
output_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
output_csv = os.path.join(output_directory, 'FeederSetup_Data.csv')
output_excel = os.path.join(output_directory, 'FeederSetup_Data.xlsx')

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Extract data from both files
data_B = extract_data(input_file_B) if os.path.exists(input_file_B) else []
data_T = extract_data(input_file_T) if os.path.exists(input_file_T) else []

# Write the results to the CSV file
with open(output_csv, mode='w', newline='') as file:
    writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
    writer.writeheader()
    for data_row in data_B + data_T:
        # Ensure AVL field is properly quoted for CSV format
        data_row["AVL"] = f"'{data_row['AVL']}'"
        writer.writerow(data_row)

# Create a new Excel workbook and add sheets for B and T data
wb = Workbook()
ws_B = wb.active
ws_B.title = "FeederSetup_AVL_B"

ws_T = wb.create_sheet(title="FeederSetup_AVL_T")

# Function to write data to a given worksheet
def write_data_to_sheet(ws, data):
    # Set default row height
    for row in range(1, len(data) + 2):  # +2 to account for header row and start index
        ws.row_dimensions[row].height = 100

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25

    # Write headers
    headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
    ws.append(headers)

    # Write data rows
    for row in data:
        # Ensure AVL field is properly quoted for Excel format
        row["AVL"] = f"'{row['AVL']}'"
        ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

# Write data to respective sheets
write_data_to_sheet(ws_B, data_B)
write_data_to_sheet(ws_T, data_T)

# Save the Excel file
wb.save(output_excel)

print(f"Data has been saved to {output_csv} and {output_excel}")


##########################################################
#working Image Split in excel

import os
import csv
import PySimpleGUI as sg
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Create a simple GUI for file and directory selection
layout = [
    [sg.Text('Select FeederSetup_AVL_B.txt File'), sg.Input(), sg.FileBrowse(key='-FILE_B-')],
    [sg.Text('Select FeederSetup_AVL_T.txt File'), sg.Input(), sg.FileBrowse(key='-FILE_T-')],
    [sg.Text('Select Image Directory for AVL_B'), sg.Input(), sg.FolderBrowse(key='-DIR_B-')],
    [sg.Text('Select Image Directory for AVL_T'), sg.Input(), sg.FolderBrowse(key='-DIR_T-')],
    [sg.Submit(), sg.Cancel()]
]

window = sg.Window('Feeder Setup Data Extractor', layout)

event, values = window.read()
window.close()

if event == 'Submit':
    input_file_B = values['-FILE_B-']
    input_file_T = values['-FILE_T-']
    image_directory_B = values['-DIR_B-']
    image_directory_T = values['-DIR_T-']
    
    # Check if paths are provided
    if not input_file_B or not input_file_T or not image_directory_B or not image_directory_T:
        sg.popup('Please provide all required file paths and directories.')
    else:
        output_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
        output_csv = os.path.join(output_directory, 'FeederSetup_Data.csv')
        output_excel = os.path.join(output_directory, 'FeederSetup_Data.xlsx')

        # Extract data from both files
        data_B = extract_data(input_file_B) if os.path.exists(input_file_B) else []
        data_T = extract_data(input_file_T) if os.path.exists(input_file_T) else []

        # Write the results to the CSV file
        with open(output_csv, mode='w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
            writer.writeheader()
            for data_row in data_B + data_T:
                # Ensure AVL field is properly quoted for CSV format
                data_row["AVL"] = f"'{data_row['AVL']}'"
                writer.writerow(data_row)

        # Load or create Excel workbook
        if os.path.exists(output_excel):
            wb = load_workbook(output_excel)
        else:
            wb = Workbook()
        
        # Create or get sheets for B and T data
        if 'FeederSetup_AVL_B' in wb.sheetnames:
            ws_B = wb['FeederSetup_AVL_B']
        else:
            ws_B = wb.create_sheet(title='FeederSetup_AVL_B')

        if 'FeederSetup_AVL_T' in wb.sheetnames:
            ws_T = wb['FeederSetup_AVL_T']
        else:
            ws_T = wb.create_sheet(title='FeederSetup_AVL_T')

        # Function to write data to a given worksheet
        def write_data_to_sheet(ws, data, image_directory):
            # Set default row height
            for row in range(1, len(data) + 2):  # +2 to account for header row and start index
                ws.row_dimensions[row].height = 100

            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 100
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25

            # Write headers
            headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
            ws.append(headers)

            # Write data rows
            for row in data:
                # Ensure AVL field is properly quoted for Excel format
                row["AVL"] = f"'{row['AVL']}'"
                ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

            # Insert images into the Excel file
            for index, row in enumerate(data, start=2):  # start=2 to account for header row
                image_path = os.path.join(image_directory, os.path.basename(row["PartNumberImage"]))
                if os.path.exists(image_path):
                    img = Image(image_path)
                    img.height = 100  # Adjust image height if necessary
                    img.width = 100   # Adjust image width if necessary
                    ws.add_image(img, f"D{index}")  # Adjust the column and row as per your requirement

        # Write data to respective sheets
        write_data_to_sheet(ws_B, data_B, image_directory_B)
        write_data_to_sheet(ws_T, data_T, image_directory_T)

        # Save the Excel file
        wb.save(output_excel)

        sg.popup(f"Data has been saved to {output_csv} and {output_excel}")

else:
    sg.popup('Operation cancelled.')





############################################
#EXcel Split fike in default path

import os
import csv
import PySimpleGUI as sg
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Default paths for the input files
default_path_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.txt"
default_path_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.txt"
default_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Create a simple GUI for file and directory selection
layout = [
    [sg.Text('FeederSetup_AVL_B.txt File', size=(30, 1)), sg.Input(default_path_B, key='-FILE_B-', size=(60, 1)), sg.FileBrowse(file_types=(("Text Files", "*.txt"),))],
    [sg.Text('FeederSetup_AVL_T.txt File', size=(30, 1)), sg.Input(default_path_T, key='-FILE_T-', size=(60, 1)), sg.FileBrowse(file_types=(("Text Files", "*.txt"),))],
    [sg.Text('Select Image Directory for AVL_B', size=(30, 1)), sg.Input(default_text=default_directory, key='-DIR_B-', size=(60, 1)), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Text('Select Image Directory for AVL_T', size=(30, 1)), sg.Input(default_text=default_directory, key='-DIR_T-', size=(60, 1)), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Submit(), sg.Cancel()]
]

window = sg.Window('Feeder Setup Data Extractor', layout)

event, values = window.read()
window.close()

if event == 'Submit':
    input_file_B = values['-FILE_B-']
    input_file_T = values['-FILE_T-']
    image_directory_B = values['-DIR_B-']
    image_directory_T = values['-DIR_T-']
    
    # Check if paths are provided
    if not input_file_B or not input_file_T or not image_directory_B or not image_directory_T:
        sg.popup('Please provide all required file paths and directories.')
    else:
        output_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
        output_csv = os.path.join(output_directory, 'FeederSetup_Data.csv')
        output_excel = os.path.join(output_directory, 'FeederSetup_Data.xlsx')

        # Extract data from both files
        data_B = extract_data(input_file_B) if os.path.exists(input_file_B) else []
        data_T = extract_data(input_file_T) if os.path.exists(input_file_T) else []

        # Write the results to the CSV file
        with open(output_csv, mode='w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
            writer.writeheader()
            for data_row in data_B + data_T:
                # Ensure AVL field is properly quoted for CSV format
                data_row["AVL"] = f"'{data_row['AVL']}'"
                writer.writerow(data_row)

        # Load or create Excel workbook
        if os.path.exists(output_excel):
            wb = load_workbook(output_excel)
        else:
            wb = Workbook()
        
        # Create or get sheets for B and T data
        if 'FeederSetup_AVL_B' in wb.sheetnames:
            ws_B = wb['FeederSetup_AVL_B']
        else:
            ws_B = wb.create_sheet(title='FeederSetup_AVL_B')

        if 'FeederSetup_AVL_T' in wb.sheetnames:
            ws_T = wb['FeederSetup_AVL_T']
        else:
            ws_T = wb.create_sheet(title='FeederSetup_AVL_T')

        # Function to write data to a given worksheet
        def write_data_to_sheet(ws, data, image_directory):
            # Set default row height
            for row in range(1, len(data) + 2):  # +2 to account for header row and start index
                ws.row_dimensions[row].height = 100

            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 100
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25

            # Write headers
            headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
            ws.append(headers)

            # Write data rows
            for row in data:
                # Ensure AVL field is properly quoted for Excel format
                row["AVL"] = f"'{row['AVL']}'"
                ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

            # Insert images into the Excel file
            for index, row in enumerate(data, start=2):  # start=2 to account for header row
                image_path = os.path.join(image_directory, os.path.basename(row["PartNumberImage"]))
                if os.path.exists(image_path):
                    img = Image(image_path)
                    img.height = 100  # Adjust image height if necessary
                    img.width = 100   # Adjust image width if necessary
                    ws.add_image(img, f"D{index}")  # Adjust the column and row as per your requirement

        # Write data to respective sheets
        write_data_to_sheet(ws_B, data_B, image_directory_B)
        write_data_to_sheet(ws_T, data_T, image_directory_T)

        # Save the Excel file
        wb.save(output_excel)

        sg.popup(f"Data has been saved to {output_csv} and {output_excel}")

else:
    sg.popup('Operation cancelled.')



######################################################################
#Split File SEither option


import os
import csv
import PySimpleGUI as sg
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

# Function to extract required data from the input file
def extract_data(input_file):
    extracted_data = []
    current_data = {}
    
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
            elif line.startswith("PartNumber:"):
                current_data["PartNumber"] = line.split(":")[1].strip()
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
                # Ensure all keys are present before appending
                if "AVLNAME" in current_data and "AVL" in current_data and "PartNumber" in current_data and "PartNumberImage" in current_data:
                    extracted_data.append(current_data)
                current_data = {}
    
    return extracted_data

# Default paths for the input files
default_path_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.txt"
default_path_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.txt"
default_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Create a simple GUI for file and directory selection
layout = [
    [sg.Checkbox('FeederSetup_AVL_B.txt', default=True, key='-USE_FILE_B-')],
    [sg.Text('FeederSetup_AVL_B.txt File', size=(30, 1)), sg.Input(default_path_B, key='-FILE_B-', size=(60, 1)), sg.FileBrowse(file_types=(("Text Files", "*.txt"),))],
    [sg.Checkbox('FeederSetup_AVL_T.txt', default=True, key='-USE_FILE_T-')],
    [sg.Text('FeederSetup_AVL_T.txt File', size=(30, 1)), sg.Input(default_path_T, key='-FILE_T-', size=(60, 1)), sg.FileBrowse(file_types=(("Text Files", "*.txt"),))],
    [sg.Text('Select Image Directory for AVL_B', size=(30, 1)), sg.Input(default_text=default_directory, key='-DIR_B-', size=(60, 1)), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Text('Select Image Directory for AVL_T', size=(30, 1)), sg.Input(default_text=default_directory, key='-DIR_T-', size=(60, 1)), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Submit(), sg.Cancel()]
]

window = sg.Window('Feeder Setup Data Extractor', layout)

event, values = window.read()
window.close()

if event == 'Submit':
    use_file_B = values['-USE_FILE_B-']
    use_file_T = values['-USE_FILE_T-']
    input_file_B = values['-FILE_B-']
    input_file_T = values['-FILE_T-']
    image_directory_B = values['-DIR_B-']
    image_directory_T = values['-DIR_T-']
    
    # Check if at least one file is selected
    if not (use_file_B or use_file_T):
        sg.popup('Please select at least one file to process.')
    else:
        output_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
        output_csv = os.path.join(output_directory, 'FeederSetup_Data.csv')
        output_excel = os.path.join(output_directory, 'FeederSetup_Data.xlsx')

        # Extract data from the selected files
        data_B = extract_data(input_file_B) if use_file_B and os.path.exists(input_file_B) else []
        data_T = extract_data(input_file_T) if use_file_T and os.path.exists(input_file_T) else []

        # Write the results to the CSV file
        with open(output_csv, mode='w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=["AVLNAME", "AVL", "PartNumber", "PartNumberImage"])
            writer.writeheader()
            for data_row in data_B + data_T:
                # Ensure AVL field is properly quoted for CSV format
                data_row["AVL"] = f"'{data_row['AVL']}'"
                writer.writerow(data_row)

        # Load or create Excel workbook
        if os.path.exists(output_excel):
            wb = load_workbook(output_excel)
        else:
            wb = Workbook()
        
        # Create or get sheets for B and T data
        if 'FeederSetup_AVL_B' in wb.sheetnames:
            ws_B = wb['FeederSetup_AVL_B']
        else:
            ws_B = wb.create_sheet(title='FeederSetup_AVL_B')

        if 'FeederSetup_AVL_T' in wb.sheetnames:
            ws_T = wb['FeederSetup_AVL_T']
        else:
            ws_T = wb.create_sheet(title='FeederSetup_AVL_T')

        # Function to write data to a given worksheet
        def write_data_to_sheet(ws, data, image_directory):
            # Set default row height
            for row in range(1, len(data) + 2):  # +2 to account for header row and start index
                ws.row_dimensions[row].height = 100

            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 100
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25

            # Write headers
            headers = ["AVLNAME", "AVL", "PartNumber", "PartNumberImage"]
            ws.append(headers)

            # Write data rows
            for row in data:
                # Ensure AVL field is properly quoted for Excel format
                row["AVL"] = f"'{row['AVL']}'"
                ws.append([row["AVLNAME"], row["AVL"], row["PartNumber"], row["PartNumberImage"]])

            # Insert images into the Excel file
            for index, row in enumerate(data, start=2):  # start=2 to account for header row
                image_path = os.path.join(image_directory, os.path.basename(row["PartNumberImage"]))
                if os.path.exists(image_path):
                    img = Image(image_path)
                    img.height = 100  # Adjust image height if necessary
                    img.width = 100   # Adjust image width if necessary
                    ws.add_image(img, f"D{index}")  # Adjust the column and row as per your requirement

        # Write data to respective sheets
        if data_B:
            write_data_to_sheet(ws_B, data_B, image_directory_B)
        if data_T:
            write_data_to_sheet(ws_T, data_T, image_directory_T)

        # Save the Excel file
        wb.save(output_excel)

        sg.popup(f"Data has been saved to {output_csv} and {output_excel}")

else:
    sg.popup('Operation cancelled.')




########################################################################

#NOte taker

import PySimpleGUI as sg
import ctypes

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout
note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 12))],
    [sg.Button('Save Note'), sg.Button('Hide')]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True
)

note_window.hide()

# Arrow button layout
arrow_layout = [
    [sg.Button('⯈', key='-SHOW-', font=('Arial', 12), size=(2, 1))]
]

arrow_window = sg.Window(
    '',
    arrow_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width - 50, screen_height - 200),
    finalize=True
)

note_visible = False

while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WIN_CLOSED:
        break

    if window == arrow_window and event == '-SHOW-':
        if note_visible:
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
        else:
            note_window.move(screen_width - 400, screen_height - 450)
            note_window.un_hide()
            arrow_window['-SHOW-'].update('⯇')
        note_visible = not note_visible

    if window == note_window:
        if event == 'Hide':
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
            note_visible = False

        if event == 'Save Note':
            with open('quick_note.txt', 'w') as f:
                f.write(values['-NOTE-'])
            sg.popup('Note Saved!', keep_on_top=True)

note_window.close()
arrow_window.close()



####################################################################################
#tracker and QuickNote second Trial

import PySimpleGUI as sg
import ctypes
import os
import datetime
from pynput import keyboard, mouse

# Define file paths
note_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Note.txt"
tracker_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Tracker.txt"

# Ensure the directory exists
os.makedirs(os.path.dirname(note_file_path), exist_ok=True)
os.makedirs(os.path.dirname(tracker_file_path), exist_ok=True)

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout
note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 12))],
    [sg.Button('Save Note'), sg.Button('Hide'), sg.Button('Play', key='-PLAY-')]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True
)

note_window.hide()

# Arrow button layout
arrow_layout = [
    [sg.Button('⯈', key='-SHOW-', font=('Arial', 12), size=(2, 1))]
]

arrow_window = sg.Window(
    '',
    arrow_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width - 50, screen_height - 200),
    finalize=True
)

note_visible = False

# Function to log activities
def log_activity(message):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    with open(tracker_file_path, 'a') as tracker_file:
        tracker_file.write(f'{timestamp}: {message}\n')

# Function to handle key press and mouse events using pynput
def on_key_press(key):
    try:
        log_activity(f'Key Pressed: {key.char}')
    except AttributeError:
        log_activity(f'Special Key Pressed: {key}')

def on_click(x, y, button, pressed):
    if pressed:
        log_activity(f'Mouse Clicked: {button} at ({x}, {y})')

keyboard_listener = keyboard.Listener(on_press=on_key_press)
mouse_listener = mouse.Listener(on_click=on_click)

keyboard_listener.start()
mouse_listener.start()

while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WIN_CLOSED:
        break

    if window == arrow_window and event == '-SHOW-':
        if note_visible:
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
        else:
            note_window.move(screen_width - 400, screen_height - 450)
            note_window.un_hide()
            arrow_window['-SHOW-'].update('⯇')
        note_visible = not note_visible
        log_activity('Note window toggled')

    if window == note_window:
        if event == 'Hide':
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
            note_visible = False
            log_activity('Note window hidden')

        if event == 'Save Note':
            with open(note_file_path, 'a') as note_file:
                timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
                note_file.write(f'{timestamp}\n{values["-NOTE-"]}\n\n')
            sg.popup('Note Saved!', keep_on_top=True)
            log_activity('Note saved')

        if event == '-PLAY-':
            note_window.move(screen_width - 400, screen_height - 450)
            log_activity('Note window moved to play position')

note_window.close()
arrow_window.close()

keyboard_listener.stop()
mouse_listener.stop()


###########################################################################

#NOte Saved Notification eliminitated no the next code 

import os
import datetime
import ctypes
import PySimpleGUI as sg
from pynput import keyboard, mouse

# Define file paths
note_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Note.txt"
tracker_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Tracker.txt"

# Ensure the directory exists
os.makedirs(os.path.dirname(note_file_path), exist_ok=True)
os.makedirs(os.path.dirname(tracker_file_path), exist_ok=True)

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout
note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 12))],
    [sg.Button('Save Note'), sg.Button('Hide'), sg.Button('Quit')]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True
)

note_window.hide()

# Arrow button layout
arrow_layout = [
    [sg.Button('⯈', key='-SHOW-', font=('Arial', 12), size=(2, 1))]
]

arrow_window = sg.Window(
    '',
    arrow_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width - 50, screen_height - 200),
    finalize=True
)

note_visible = False

# Function to log activities
def log_activity(message):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    with open(tracker_file_path, 'a') as tracker_file:
        tracker_file.write(f'{timestamp}: {message}\n')

# Function to handle key press and mouse events using pynput
def on_key_press(key):
    try:
        log_activity(f'Key Pressed: {key.char}')
    except AttributeError:
        log_activity(f'Special Key Pressed: {key}')

def on_click(x, y, button, pressed):
    if pressed:
        log_activity(f'Mouse Clicked: {button} at ({x}, {y})')

keyboard_listener = keyboard.Listener(on_press=on_key_press)
mouse_listener = mouse.Listener(on_click=on_click)

keyboard_listener.start()
mouse_listener.start()

while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WIN_CLOSED:
        break

    if window == arrow_window and event == '-SHOW-':
        if note_visible:
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
        else:
            note_window.move(screen_width - 400, screen_height - 450)
            note_window.un_hide()
            arrow_window['-SHOW-'].update('⯇')
        note_visible = not note_visible
        log_activity('Note window toggled')

    if window == note_window:
        if event == 'Hide':
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
            note_visible = False
            log_activity('Note window hidden')

        if event == 'Save Note':
            with open(note_file_path, 'a') as note_file:
                timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
                note_file.write(f'{timestamp}\n{values["-NOTE-"]}\n\n')
            sg.popup('Note Saved!', keep_on_top=True)
            log_activity('Note saved')
            values["-NOTE-"] = ''  # Clear the note area after saving

        if event == 'Quit':
            log_activity('Application quit')
            break

note_window.close()
arrow_window.close()

keyboard_listener.stop()
mouse_listener.stop()

# Hide the Tracker.txt file
if os.path.exists(tracker_file_path):
    os.system(f'attrib +h "{tracker_file_path}"')



    ########################################################################

    #Current working file 


    import os
import time
import datetime
import ctypes
import PySimpleGUI as sg
from pynput import keyboard, mouse

# Define file paths
note_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Note.txt"
tracker_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Tracker.txt"

# Ensure the directory exists
os.makedirs(os.path.dirname(note_file_path), exist_ok=True)
os.makedirs(os.path.dirname(tracker_file_path), exist_ok=True)

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout with a dark theme
sg.theme('DarkGrey13')

note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 10))],
    [sg.Button('Save Note'), sg.Button('Hide'), sg.Button('Quit')]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True
)

note_window.hide()

# Arrow button layout
arrow_layout = [
    [sg.Button('⯈', key='-SHOW-', font=('Arial', 10), size=(0, 0))]
]

arrow_window = sg.Window(
    '',
    arrow_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width - 50, screen_height - 200),
    finalize=True
)

note_visible = False

# Function to log activities
def log_activity(message):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    with open(tracker_file_path, 'a') as tracker_file:
        tracker_file.write(f'{timestamp}: {message}\n')

# Function to handle key press and mouse events using pynput
def on_key_press(key):
    try:
        log_activity(f'Key Pressed: {key.char}')
    except AttributeError:
        log_activity(f'Special Key Pressed: {key}')

def on_click(x, y, button, pressed):
    if pressed:
        log_activity(f'Mouse Clicked: {button} at ({x}, {y})')

keyboard_listener = keyboard.Listener(on_press=on_key_press)
mouse_listener = mouse.Listener(on_click=on_click)

keyboard_listener.start()
mouse_listener.start()

def save_note_to_file(note_text):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
    with open(note_file_path, 'a') as note_file:
        note_file.write(f'\n{timestamp}\n{note_text}\n')
    log_activity('Note saved')

def autosave(note_text):
    while True:
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
        with open(note_file_path, 'a') as note_file:
            note_file.write(f'{"AutoSave"}\n{timestamp}\n{note_text}\n')
        log_activity('Note saved')
        time.sleep(60 * 30)


while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WIN_CLOSED:
        break

    if window == arrow_window and event == '-SHOW-':
        if note_visible:
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
        else:
            note_window.move(screen_width - 300, screen_height - 400)
            note_window.un_hide()
            arrow_window['-SHOW-'].update('⯇')
        note_visible = not note_visible
        log_activity('Note window toggled')

    if window == note_window:
        if event == 'Hide':
            save_note_to_file(values['-NOTE-'])
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
            note_visible = False
            log_activity('Note window hidden')

        if event == 'Save Note':
            save_note_to_file(values['-NOTE-'])
            #sg.popup_no_wait('Note Saved!', keep_on_top=True)
            log_activity('Note saved')

        if event == 'Quit':
            #save_note_to_file(values['-NOTE-'])
            log_activity('Application quit')
            break

keyboard_listener.stop()
mouse_listener.stop()

note_window.close()
arrow_window.close()

# Hide the Tracker.txt file
if os.path.exists(tracker_file_path):
    os.system(f'attrib +h "{tracker_file_path}"')





####################################################################

#Testing Code Not working Tracker QuitNote

'''Sure, here's the complete script with password protection for Tracker.txt using AES encryption:

python
Copy code'''

import os
import time
import datetime
import ctypes
import PySimpleGUI as sg
from pynput import keyboard, mouse
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from base64 import urlsafe_b64encode, urlsafe_b64decode

# Define file paths
note_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Note.txt"
tracker_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Tracker.txt"

# Ensure the directory exists
os.makedirs(os.path.dirname(note_file_path), exist_ok=True)
os.makedirs(os.path.dirname(tracker_file_path), exist_ok=True)

# Function to generate encryption key from password
def generate_key_from_password(password, salt=b'salt_'):
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=100000,
        backend=default_backend()
    )
    return kdf.derive(password.encode())

# Function to encrypt data
def encrypt_data(data, key):
    iv = os.urandom(16)
    cipher = Cipher(algorithms.AES(key), modes.CFB(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    encrypted_data = encryptor.update(data) + encryptor.finalize()
    return urlsafe_b64encode(iv + encrypted_data)

# Function to decrypt data
def decrypt_data(encrypted_data, key):
    encrypted_data = urlsafe_b64decode(encrypted_data)
    iv = encrypted_data[:16]
    data = encrypted_data[16:]
    cipher = Cipher(algorithms.AES(key), modes.CFB(iv), backend=default_backend())
    decryptor = cipher.decryptor()
    decrypted_data = decryptor.update(data) + decryptor.finalize()
    return decrypted_data

# Function to save encrypted data to file
def save_encrypted_data(file_path, data, key):
    encrypted_data = encrypt_data(data.encode(), key)
    with open(file_path, 'wb') as encrypted_file:
        encrypted_file.write(encrypted_data)

# Function to read encrypted data from file
def read_encrypted_data(file_path, key):
    with open(file_path, 'rb') as encrypted_file:
        encrypted_data = encrypted_file.read()
        decrypted_data = decrypt_data(encrypted_data, key).decode()
    return decrypted_data

# Function to log activities
def log_activity(message, key):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    encrypted_message = encrypt_data(f'{timestamp}: {message}', key)
    with open(tracker_file_path, 'ab') as tracker_file:
        tracker_file.write(encrypted_message + b'\n')

# Function to get password from user and generate key
def get_password_and_key():
    password = sg.popup_get_text('Enter password to protect Tracker.txt:', password_char='*')
    if not password:
        sg.popup_error('Password cannot be empty. Exiting.')
        raise SystemExit
    key = generate_key_from_password(password)
    return key

# Get password and generate key
key = get_password_and_key()

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout with a dark theme
sg.theme('DarkGrey13')

note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 10))],
    [sg.Button('Save Note'), sg.Button('Hide'), sg.Button('Quit')]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True
)

note_window.hide()

# Arrow button layout
arrow_layout = [
    [sg.Button('⯈', key='-SHOW-', font=('Arial', 10), size=(0, 0))]
]

arrow_window = sg.Window(
    '',
    arrow_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width - 50, screen_height - 200),
    finalize=True
)

note_visible = False

# Function to handle key press and mouse events using pynput
def on_key_press(key):
    try:
        log_activity(f'Key Pressed: {key.char}', key)
    except AttributeError:
        log_activity(f'Special Key Pressed: {key}', key)

def on_click(x, y, button, pressed):
    if pressed:
        log_activity(f'Mouse Clicked: {button} at ({x}, {y})', key)

keyboard_listener = keyboard.Listener(on_press=on_key_press)
mouse_listener = mouse.Listener(on_click=on_click)

keyboard_listener.start()
mouse_listener.start()

def save_note_to_file(note_text):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
    with open(note_file_path, 'a') as note_file:
        note_file.write(f'\n{timestamp}\n{note_text}\n')
    log_activity('Note saved', key)

def autosave(note_text):
    while True:
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
        with open(note_file_path, 'a') as note_file:
            note_file.write(f'{"AutoSave"}\n{timestamp}\n{note_text}\n')
        log_activity('Note autosaved', key)
        time.sleep(60 * 30)

while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WIN_CLOSED:
        break

    if window == arrow_window and event == '-SHOW-':
        if note_visible:
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
        else:
            note_window.move(screen_width - 300, screen_height - 400)
            note_window.un_hide()
            arrow_window['-SHOW-'].update('⯇')
        note_visible = not note_visible
        log_activity('Note window toggled', key)

    if window == note_window:
        if event == 'Hide':
            save_note_to_file(values['-NOTE-'])
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
            note_visible = False
            log_activity('Note window hidden', key)

        if event == 'Save Note':
            save_note_to_file(values['-NOTE-'])
            log_activity('Note saved', key)

        if event == 'Quit':
            log_activity('Application quit', key)
            break

keyboard_listener.stop()
mouse_listener.stop()

note_window.close()
arrow_window.close()

# Hide the Tracker.txt file
if os.path.exists(tracker_file_path):
    os.system(f'attrib +h "{tracker_file_path}"')
'''This script does the following:

Prompts the user to enter a password to protect Tracker.txt.
Uses the password to generate an encryption key.
Encrypts all log entries before writing them to Tracker.txt.
Decrypts Tracker.txt entries when reading from the file.
Handles keyboard and mouse events, logging activities in the encrypted Tracker.txt.
Provides a GUI for taking notes, which can be shown or hidden using a toggle button.
To use this script, ensure you have the required packages installed:

bash
Copy code
pip install PySimpleGUI pynput cryptography
Make sure to keep the password safe, as it's necessary for decrypting the log file.'''




########################################################################################

#working Code for QuickNote

import os
import time
import datetime
import ctypes
import PySimpleGUI as sg
from pynput import keyboard, mouse

# Define file paths
note_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Note.txt"
tracker_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Tracker.txt"

# Ensure the directory exists
os.makedirs(os.path.dirname(note_file_path), exist_ok=True)
os.makedirs(os.path.dirname(tracker_file_path), exist_ok=True)

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout with a dark theme
sg.theme('DarkGrey13')

note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 10))],
    [sg.Button('Save Note'), sg.Button('Hide'), sg.Button('Quit')]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True
)

note_window.hide()

# Arrow button layout
arrow_layout = [
    [sg.Button('⯈', key='-SHOW-', font=('Arial', 10), size=(0, 0))]
]

arrow_window = sg.Window(
    '',
    arrow_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width - 50, screen_height - 200),
    finalize=True
)

note_visible = False

# Function to log activities
def log_activity(message):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    with open(tracker_file_path, 'a') as tracker_file:
        tracker_file.write(f'{timestamp}: {message}\n')

# Function to handle key press and mouse events using pynput
def on_key_press(key):
    try:
        log_activity(f'Key Pressed: {key.char}')
    except AttributeError:
        log_activity(f'Special Key Pressed: {key}')

def on_click(x, y, button, pressed):
    if pressed:
        log_activity(f'Mouse Clicked: {button} at ({x}, {y})')

keyboard_listener = keyboard.Listener(on_press=on_key_press)
mouse_listener = mouse.Listener(on_click=on_click)

keyboard_listener.start()
mouse_listener.start()

def save_note_to_file(note_text):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
    with open(note_file_path, 'a') as note_file:
        note_file.write(f'\n{timestamp}\n{note_text}\n')
    log_activity('Note saved')

def autosave(note_text):
    while True:
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
        with open(note_file_path, 'a') as note_file:
            note_file.write(f'{"AutoSave"}\n{timestamp}\n{note_text}\n')
        log_activity('Note saved')
        time.sleep(60 * 30)


while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WIN_CLOSED:
        break

    if window == arrow_window and event == '-SHOW-':
        if note_visible:
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
        else:
            note_window.move(screen_width - 300, screen_height - 400)
            note_window.un_hide()
            arrow_window['-SHOW-'].update('⯇')
        note_visible = not note_visible
        log_activity('Note window toggled')

    if window == note_window:
        if event == 'Hide':
            save_note_to_file(values['-NOTE-'])
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
            note_visible = False
            log_activity('Note window hidden')

        if event == 'Save Note':
            save_note_to_file(values['-NOTE-'])
            #sg.popup_no_wait('Note Saved!', keep_on_top=True)
            log_activity('Note saved')

        if event == 'Quit':
            #save_note_to_file(values['-NOTE-'])
            log_activity('Application quit')
            break

keyboard_listener.stop()
mouse_listener.stop()

note_window.close()
arrow_window.close()

# Hide the Tracker.txt file
if os.path.exists(tracker_file_path):
    os.system(f'attrib +h "{tracker_file_path}"')


####################################################

#code work need to correct some line 

import os
import time
import datetime
import ctypes
from tkinter import *
from PIL import Image, ImageTk
import PySimpleGUI as sg
from pynput import keyboard, mouse

# Define file paths
note_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Note.txt"
tracker_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Tracker.txt"

# Ensure the directory exists
os.makedirs(os.path.dirname(note_file_path), exist_ok=True)
os.makedirs(os.path.dirname(tracker_file_path), exist_ok=True)

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout with a dark theme
sg.theme('DarkGrey13')

note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 12), pad=(10, 10), border_width=0, background_color='black', text_color='white')],
    [sg.Button('Save Note', button_color=('DarkGrey', 'White'), font=('Arial', 10)),
     sg.Button('Hide', button_color=('DarkGrey', 'White'), font=('Arial', 10)),
     sg.Button('Quit', button_color=('DarkGrey', 'red'), font=('Arial', 10))]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True,
    background_color='black',
    element_justification='center'
)

note_window.hide()

note_visible = False

# Function to log activities
def log_activity(message):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    with open(tracker_file_path, 'a') as tracker_file:
        tracker_file.write(f'{timestamp}: {message}\n')

# Function to handle key press and mouse events using pynput
def on_key_press(key):
    try:
        log_activity(f'Key Pressed: {key.char}')
    except AttributeError:
        log_activity(f'Special Key Pressed: {key}')

def on_click(x, y, button, pressed):
    if pressed:
        log_activity(f'Mouse Clicked: {button} at ({x}, {y})')

keyboard_listener = keyboard.Listener(on_press=on_key_press)
mouse_listener = mouse.Listener(on_click=on_click)

keyboard_listener.start()
mouse_listener.start()

def save_note_to_file(note_text):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
    with open(note_file_path, 'a') as note_file:
        note_file.write(f'\n{timestamp}\n{note_text}\n')
    log_activity('Note saved')

def create_tkinter_window():
    global play_button  # Declare play_button as global
    root = Tk()
    root.geometry(f'50x50+{screen_width - 50}+{screen_height // 2 - 25}')  # Set initial position
    root.overrideredirect(1)
    root.wm_attributes("-topmost", 1)
    root.configure(bg='orange')

    def on_enter(event):
        root.geometry(f'50x50+{screen_width - 100}+{screen_height // 2 - 25}')  # Adjust position on hover

    def on_leave(event):
        root.geometry(f'50x50+{screen_width - 50}+{screen_height // 2 - 25}')  # Adjust position on hover

    def toggle_note_window():
        global note_visible
        if note_visible:
            note_window.hide()
            play_button.config(text='⯈')
        else:
            note_window.move(screen_width - 400, screen_height // 2 - 200)
            note_window.un_hide()
            play_button.config(text='⯇')
        note_visible = not note_visible
        log_activity('Note window toggled')

    play_button = Button(
        root, text='⯈', font=('Arial', 18), bg='orange', fg='white', relief='flat', command=toggle_note_window,
        borderwidth=0, highlightthickness=0, padx=10, pady=10
    )
    play_button.pack(fill=BOTH, expand=TRUE)
    play_button.bind("<Enter>", on_enter)
    play_button.bind("<Leave>", on_leave)

    return root

root = create_tkinter_window()

while True:
    window, event, values = sg.read_all_windows(timeout=100)

    if event == sg.WIN_CLOSED:
        break

    if window == note_window:
        if event == 'Hide':
            save_note_to_file(values['-NOTE-'])
            note_window.hide()
            play_button.config(text='⯈')
            note_visible = False
            log_activity('Note window hidden')

        if event == 'Save Note':
            save_note_to_file(values['-NOTE-'])
            sg.popup_no_wait('Note Saved!', keep_on_top=True)
            log_activity('Note saved')

        if event == 'Quit':
            save_note_to_file(values['-NOTE-'])
            log_activity('Application quit')
            break

keyboard_listener.stop()
mouse_listener.stop()

note_window.close()
root.destroy()

# Hide the Tracker.txt file
if os.path.exists(tracker_file_path):
    os.system(f'attrib +h "{tracker_file_path}"')



#########################################################################


#code note work smooth transition used for study can use line

import os
import time
import datetime
import ctypes
from tkinter import *
from PIL import Image, ImageTk
import PySimpleGUI as sg
from pynput import keyboard, mouse

# Define file paths
note_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Note.txt"
tracker_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Tracker.txt"

# Ensure the directory exists
os.makedirs(os.path.dirname(note_file_path), exist_ok=True)
os.makedirs(os.path.dirname(tracker_file_path), exist_ok=True)

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout with a dark theme
sg.theme('DarkGrey13')

note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 10), pad=(2, 2), border_width=0, background_color='black', text_color='white')],
    [sg.Button('Save Note', font=('Arial', 10)),
     sg.Button('Hide', font=('Arial', 10)),
     sg.Button('Quit', font=('Arial', 10))]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True,
    background_color='black',
    element_justification='center'
)

note_window.hide()

note_visible = False

# Function to log activities
def log_activity(message):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    with open(tracker_file_path, 'a') as tracker_file:
        tracker_file.write(f'{timestamp}: {message}\n')

# Function to handle key press and mouse events using pynput
def on_key_press(key):
    try:
        log_activity(f'Key Pressed: {key.char}')
    except AttributeError:
        log_activity(f'Special Key Pressed: {key}')

def on_click(x, y, button, pressed):
    if pressed:
        log_activity(f'Mouse Clicked: {button} at ({x}, {y})')

keyboard_listener = keyboard.Listener(on_press=on_key_press)
mouse_listener = mouse.Listener(on_click=on_click)

keyboard_listener.start()
mouse_listener.start()

def save_note_to_file(note_text):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
    with open(note_file_path, 'a') as note_file:
        note_file.write(f'\n{timestamp}\n{note_text}\n')
    log_activity('Note saved')

def create_tkinter_window():
    global play_button  # Declare play_button as global
    root = Tk()
    root.geometry(f'25x25+{screen_width - 50}+{screen_height // 2 - 25}')  # Set initial position
    root.overrideredirect(1)
    root.wm_attributes("-topmost", 1)
    root.configure(bg='orange')

    def on_enter(event):
        global play_button
        target_x = screen_width - 100
        current_x = root.winfo_x()
        step = 5  # Adjust the step size for smoother transition
        while current_x > target_x:
            current_x -= step
            root.geometry(f'25x25+{current_x}+{screen_height // 2 - 25}')
            root.update()
            time.sleep(0.05)  # Adjust sleep time for smoother transition

    def on_leave(event):
        global play_button
        target_x = screen_width - 50
        current_x = root.winfo_x()
        step = 3  # Adjust the step size for smoother transition
        while current_x < target_x:
            current_x += step
            root.geometry(f'25x25+{current_x}+{screen_height // 2 - 25}')
            root.update()
            time.sleep(0.05)  # Adjust sleep time for smoother transition

    def toggle_note_window():
        global note_visible
        if note_visible:
            note_window.hide()
            play_button.config(text='⯈')
        else:
            note_window.move(screen_width - 360, screen_height // 2 - 50)
            note_window.un_hide()
            play_button.config(text='⯇')
        note_visible = not note_visible
        log_activity('Note window toggled')

    play_button = Button(
        root, text='⯈', font=('Arial', 10), bg='orange', fg='white', relief='flat', command=toggle_note_window,
        borderwidth=0, highlightthickness=0, padx=1, pady=1
    )
    play_button.pack(fill=BOTH, expand=TRUE)
    play_button.bind("<Enter>", on_enter)
    play_button.bind("<Leave>", on_leave)

    return root

root = create_tkinter_window()

while True:
    window, event, values = sg.read_all_windows(timeout=100)

    if event == sg.WIN_CLOSED:
        break

    if window == note_window:
        if event == 'Hide':
            save_note_to_file(values['-NOTE-'])
            note_window.hide()
            play_button.config(text='⯈')
            note_visible = False
            log_activity('Note window hidden')

        if event == 'Save Note':
            save_note_to_file(values['-NOTE-'])
            sg.popup_no_wait('Note Saved!', keep_on_top=True)
            log_activity('Note saved')

        if event == 'Quit':
            save_note_to_file(values['-NOTE-'])
            log_activity('Application quit')
            break

keyboard_listener.stop()
mouse_listener.stop()

note_window.close()
root.destroy()

# Hide the Tracker.txt file
if os.path.exists(tracker_file_path):
    os.system(f'attrib +h "{tracker_file_path}"')



#######################################################################

#terminal to no to use
import os
import time
import datetime
import ctypes
import PySimpleGUI as sg
from pynput import keyboard, mouse

# Define file paths
note_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Note.txt"
tracker_file_path = r"D:\NX_BACKWORK\Database_File\Process_QuickNOTE\Tracker.txt"

# Ensure the directory exists
os.makedirs(os.path.dirname(note_file_path), exist_ok=True)
os.makedirs(os.path.dirname(tracker_file_path), exist_ok=True)

# Get screen width and height
user32 = ctypes.windll.user32
screen_width, screen_height = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)

# Define the note window layout with a dark theme
sg.theme('DarkGrey13')

note_layout = [
    [sg.Multiline(size=(40, 20), key='-NOTE-', font=('Arial', 10))],
    [sg.Button('Save Note'), sg.Button('Hide'), sg.Button('Quit')]
]

note_window = sg.Window(
    'Quick Note',
    note_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width, screen_height - 300),  # Initial position off screen
    finalize=True
)

note_window.hide()

# Arrow button layout
arrow_layout = [
    [sg.Button('⯈', key='-SHOW-', font=('Arial', 10), size=(0, 0))]
]

arrow_window = sg.Window(
    '',
    arrow_layout,
    no_titlebar=True,
    keep_on_top=True,
    alpha_channel=0.95,
    grab_anywhere=True,
    location=(screen_width - 50, screen_height - 200),
    finalize=True
)

note_visible = False

# Function to log activities
def log_activity(message):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')
    with open(tracker_file_path, 'a') as tracker_file:
        tracker_file.write(f'{timestamp}: {message}\n')

# Function to handle key press and mouse events using pynput
def on_key_press(key):
    try:
        log_activity(f'Key Pressed: {key.char}')
    except AttributeError:
        log_activity(f'Special Key Pressed: {key}')

def on_click(x, y, button, pressed):
    if pressed:
        log_activity(f'Mouse Clicked: {button} at ({x}, {y})')

keyboard_listener = keyboard.Listener(on_press=on_key_press)
mouse_listener = mouse.Listener(on_click=on_click)

keyboard_listener.start()
mouse_listener.start()

def save_note_to_file(note_text):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
    with open(note_file_path, 'a') as note_file:
        note_file.write(f'\n{timestamp}\n{note_text}\n')
    log_activity('Note saved')

while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WIN_CLOSED:
        break

    if window == arrow_window and event == '-SHOW-':
        if note_visible:
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
        else:
            note_window.move(screen_width - 300, screen_height - 400)
            note_window.un_hide()
            arrow_window['-SHOW-'].update('⯇')
        note_visible = not note_visible
        log_activity('Note window toggled')

    if window == note_window:
        if event == 'Hide':
            save_note_to_file(values['-NOTE-'])
            note_window.hide()
            arrow_window['-SHOW-'].update('⯈')
            note_visible = False
            log_activity('Note window hidden')

        if event == 'Save Note':
            save_note_to_file(values['-NOTE-'])
            log_activity('Note saved')

        if event == 'Quit':
            log_activity('Application quit')
            break

keyboard_listener.stop()
mouse_listener.stop()

note_window.close()
arrow_window.close()

# Hide the Tracker.txt file
if os.path.exists(tracker_file_path):
    os.system(f'attrib +h "{tracker_file_path}"')

#pyinstaller -F -i "AVL.ico" --noconsole AVL-Checker.py





import os
import shutil
import pandas as pd
import xml.etree.ElementTree as ET

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to DataFrame
def xml_to_dataframe(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    data = []
    for feeder in root.findall('Feeder'):
        feeder_dict = {}
        for child in feeder.iter():
            if child.tag != 'Feeder':
                if child.text:
                    if child.tag == 'AVL':
                        feeder_dict[child.tag] = '"' + child.text.strip() + '"'
                    else:
                        feeder_dict[child.tag] = child.text.strip()
                else:
                    feeder_dict[child.tag] = ''
        data.append(feeder_dict)
    
    df = pd.DataFrame(data)
    return df


# Function to convert DataFrame to Excel file
def dataframe_to_excel(df, excel_file):
    writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()

# Processing FeederSetup files
feeder_setup_files = find_feeder_setup_files(root_directory)
for feeder_setup_file in feeder_setup_files:
    # Rename XML file based on description
    renamed_xml_file = rename_feeder_setup_with_description(feeder_setup_file)
    
    # Define file paths
    input_file = renamed_xml_file if renamed_xml_file else 'FeederSetup.xml'
    destination_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
    
    if os.path.exists(input_file):
        # Convert XML to DataFrame
        df = xml_to_dataframe(input_file)
        
        # Convert DataFrame to Excel
        output_excel = os.path.splitext(input_file)[0] + '.xlsx'
        dataframe_to_excel(df, output_excel)
        
        # Move the Excel file to the destination directory
        shutil.move(output_excel, destination_directory)
    else:
        print(f"File not found: {input_file}. Skipping...")

print("Conversion complete.")






##Working COde AVL Checker##

import os
import shutil
import pandas as pd
import xml.etree.ElementTree as ET

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to DataFrame
def xml_to_dataframe(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    data = []
    for feeder in root.findall('Feeder'):
        feeder_dict = {}
        for child in feeder.iter():
            if child.tag != 'Feeder':
                if child.text:
                    if child.tag == 'AVL':
                        feeder_dict[child.tag] = '"' + child.text.strip() + '"'
                    else:
                        feeder_dict[child.tag] = child.text.strip()
        if feeder_dict:  # Only append if the dictionary is not empty
            data.append(feeder_dict)
    
    df = pd.DataFrame(data)
    
    # Remove rows where PartNumber is blank
    df = df[df['PartNumber'].notna() & (df['PartNumber'].str.strip() != '')]

    # Sort DataFrame by PartNumber column
    df.sort_values(by=['PartNumber'], inplace=True)
    
    return df

# Function to convert DataFrame to Excel file
def dataframe_to_excel(df, excel_file):
    writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()

# Processing FeederSetup files
feeder_setup_files = find_feeder_setup_files(root_directory)
for feeder_setup_file in feeder_setup_files:
    # Rename XML file based on description
    renamed_xml_file = rename_feeder_setup_with_description(feeder_setup_file)
    
    # Define file paths
    input_file = renamed_xml_file if renamed_xml_file else 'FeederSetup.xml'
    destination_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
    
    if os.path.exists(input_file):
        # Convert XML to DataFrame
        df = xml_to_dataframe(input_file)
        
        # Convert DataFrame to Excel
        output_excel = os.path.splitext(input_file)[0] + '.xlsx'
        dataframe_to_excel(df, output_excel)
        
        # Move the Excel file to the destination directory
        shutil.move(output_excel, destination_directory)
    else:
        print(f"File not found: {input_file}. Skipping...")

print("Conversion complete.")




##Working Code AVL Checker


import os
import shutil
import pandas as pd
import xml.etree.ElementTree as ET

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to DataFrame
def xml_to_dataframe(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    data = []
    for feeder in root.findall('Feeder'):
        feeder_dict = {}
        for child in feeder.iter():
            if child.tag != 'Feeder':
                if child.text:
                    if child.tag == 'AVL':
                        feeder_dict[child.tag] = '"' + child.text.strip() + '"'
                    else:
                        feeder_dict[child.tag] = child.text.strip()
        if feeder_dict:  # Only append if the dictionary is not empty
            data.append(feeder_dict)
    
    df = pd.DataFrame(data)
    
    # Remove rows where PartNumber is blank
    df = df[df['PartNumber'].notna() & (df['PartNumber'].str.strip() != '')]

    # Sort DataFrame by PartNumber column
    df.sort_values(by=['PartNumber'], inplace=True)
    
    return df

# Function to convert DataFrame to Excel file
def dataframe_to_excel(df, excel_file):
    writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()

# Function to process the Excel file and keep specified columns
def process_excel_file(file_path):
    df = pd.read_excel(file_path)
    
    # Keep only the specified columns in the given order
    columns_to_keep = ['PartNumber', 'AVLNAME', 'AVL', 'PartShapeName', 'PartNumberImage', 'QTY']
    df = df[columns_to_keep]
    
    # Add the Side column based on the file name
    side = 'BOT' if 'AVL_B' in file_path else 'TOP'
    df['Side'] = side
    
    # Save the modified DataFrame back to Excel
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()

# Processing FeederSetup files
feeder_setup_files = find_feeder_setup_files(root_directory)
for feeder_setup_file in feeder_setup_files:
    # Rename XML file based on description
    renamed_xml_file = rename_feeder_setup_with_description(feeder_setup_file)
    
    # Define file paths
    input_file = renamed_xml_file if renamed_xml_file else 'FeederSetup.xml'
    destination_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
    
    if os.path.exists(input_file):
        # Convert XML to DataFrame
        df = xml_to_dataframe(input_file)
        
        # Convert DataFrame to Excel
        output_excel = os.path.splitext(input_file)[0] + '.xlsx'
        dataframe_to_excel(df, output_excel)
        
        # Move the Excel file to the destination directory
        shutil.move(output_excel, destination_directory)
        
        # Process the Excel file to keep specific columns and add the Side column
        final_excel_path = os.path.join(destination_directory, os.path.basename(output_excel))
        process_excel_file(final_excel_path)
    else:
        print(f"File not found: {input_file}. Skipping...")

print("Conversion complete.")




##working code 1st 
import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import xml.etree.ElementTree as ET
import shutil
import PySimpleGUI as sg

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
image_directory = root_directory

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.txt"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.txt"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to TXT
def xml_to_txt(xml_file, txt_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    def extract_text(element, level=0):
        text = ''
        indent = ' ' * (level * 4)  # Indent each level by 4 spaces
        if element.text and element.text.strip():
            text += f'{indent}{element.tag}: {element.text.strip()}\n'
        for child in element:
            text += extract_text(child, level + 1)
        return text

    with open(txt_file, 'w', encoding='utf-8') as file:
        file.write(extract_text(root))

# Processing FeederSetup files
feeder_setup_files = find_feeder_setup_files(root_directory)
renamed_txt_file = None
for feeder_setup_file in feeder_setup_files:
    # Convert XML to TXT
    txt_file = feeder_setup_file.replace(".xml", ".txt")
    xml_to_txt(feeder_setup_file, txt_file)
    
    # Rename TXT file based on description
    renamed_txt_file = rename_feeder_setup_with_description(txt_file)
    
    # Define file paths
    input_file = renamed_txt_file if renamed_txt_file else 'FeederSetup.txt'
    destination_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
    
    if os.path.exists(input_file):
        # Move the file to the destination directory
        shutil.move(input_file, destination_directory)
    else:
        print(f"File not found: {input_file}. Skipping...")

# Function to extract required data from the input file
def extract_data(input_file, side):
    part_numbers = []
    extracted_data = []

    # First pass: extract all part numbers
    with open(input_file, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("PartNumber:"):
                part_number = line.split(":")[1].strip()
                if part_number:  # Skip blank part numbers
                    part_numbers.append(part_number)

    # Second pass: extract data for each valid part number
    with open(input_file, 'r') as file:
        current_data = {}
        for line in file:
            line = line.strip()
            if line.startswith("PartNumber:"):
                if current_data and current_data.get("PartNumber") in part_numbers:
                    if "AVL" not in current_data:
                        current_data["AVL"] = ""
                    if "AVLNAME" not in current_data:
                        current_data["AVLNAME"] = ""
                    extracted_data.append(current_data)
                current_data = {"PartNumber": line.split(":")[1].strip(), "Side": side}
            elif line.startswith("PartNumberImage:"):
                current_data["PartNumberImage"] = line.split(":")[1].strip()
            elif line.startswith("AVLNAME:"):
                current_data["AVLNAME"] = line.split(":")[1].strip()
            elif line.startswith("AVL:"):
                current_data["AVL"] = line.split(":")[1].strip()
        if current_data and current_data.get("PartNumber") in part_numbers:
            if "AVL" not in current_data:
                current_data["AVL"] = ""
            if "AVLNAME" not in current_data:
                current_data["AVLNAME"] = ""
            extracted_data.append(current_data)
    
    return extracted_data

# Default paths for the input files
default_path_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.txt"
default_path_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.txt"
default_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Create a simple GUI for file and directory selection
layout = [
    [sg.Checkbox('FeederSetup_AVL_B.txt', default=True, key='-USE_FILE_B-'), sg.Checkbox('FeederSetup_AVL_T.txt', default=True, key='-USE_FILE_T-')],
    [sg.Text('Select Image Directory for AVL_B', size=(30, 1)), sg.Input(key='-DIR_B-', size=(60, 1), readonly=True), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Text('Select Image Directory for AVL_T', size=(30, 1)), sg.Input(key='-DIR_T-', size=(60, 1), readonly=True), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Submit(), sg.Cancel()]
]

window = sg.Window('Feeder Setup Data Extractor', layout)

event, values = window.read()
window.close()

os.getcwd()
Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
Chd = os.getcwd()

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)

if event == 'Submit':
    use_file_B = values['-USE_FILE_B-']
    use_file_T = values['-USE_FILE_T-']
    image_directory_B = values['-DIR_B-']
    image_directory_T = values['-DIR_T-']
    
    # Check if at least one file is selected
    if not (use_file_B or use_file_T):
        sg.popup('Please select at least one file to process.')
    else:
        output_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
        output_csv = os.path.join(output_directory, 'FeederSetup_Data.csv')
        output_excel = os.path.join(output_directory, 'FeederSetup_Data.xlsx')

        # Extract data from the selected files
        data_B = extract_data(default_path_B, "BOT") if use_file_B and os.path.exists(default_path_B) else []
        data_T = extract_data(default_path_T, "TOP") if use_file_T and os.path.exists(default_path_T) else []

        # Write the results to the CSV file
        with open(output_csv, mode='w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=["PartNumber", "PartNumberImage", "AVLNAME", "AVL", "Side"])
            writer.writeheader()
            for data_row in data_B + data_T:
                # Ensure AVL field is properly quoted for CSV format
                data_row["AVL"] = f'"{data_row["AVL"]}"'
                writer.writerow(data_row)

        # Load or create Excel workbook
if os.path.exists(output_excel):
    wb = load_workbook(output_excel)
else:
    wb = Workbook()

if 'FeederSetup_AVL_B' in wb.sheetnames:
    ws_B = wb['FeederSetup_AVL_B']
else:
    ws_B = wb.create_sheet(title='FeederSetup_AVL_B')

if 'FeederSetup_AVL_T' in wb.sheetnames:
    ws_T = wb['FeederSetup_AVL_T']
else:
    ws_T = wb.create_sheet(title='FeederSetup_AVL_T')

# Function to write data to a given worksheet
def write_data_to_sheet(ws, data, image_directory, table_name):
    # Set default row height
    for row in range(1, len(data) + 2):  # +2 to account for header row and start index
        ws.row_dimensions[row].height = 100

    # Set column widths (old settings)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 100
    ws.column_dimensions['E'].width = 15

    # Write headers
    headers = ["PartNumber", "PartNumberImage", "AVLNAME", "AVL", "Side"]
    ws.append(headers)

    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data rows
    for row in data:
        # Ensure AVL and AVLNAME fields are present and properly quoted for Excel format
        if 'AVL' not in row:
            row["AVL"] = ""
        else:
            row["AVL"] = f"'{row['AVL']}'"
        
        if 'AVLNAME' not in row:
            row["AVLNAME"] = ""
        
        ws.append([row["PartNumber"], row["PartNumberImage"], row["AVLNAME"], row["AVL"], row["Side"]])

    # Insert images into the Excel file
    for index, row in enumerate(data, start=2):  # start=2 to account for header row
        image_path = os.path.join(image_directory, os.path.basename(row["PartNumberImage"]))
        if os.path.exists(image_path):
            img = Image(image_path)
            img.height = 100  # Adjust image height if necessary
            img.width = 200   # Adjust image width if necessary
            ws.add_image(img, f"B{index}")  # Adjust the column and row as per your requirement

    # Create table for better formatting
    table = Table(displayName=table_name, ref=f"A1:E{len(data) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    ws.add_table(table)

# Write data to respective sheets
if data_B:
    write_data_to_sheet(ws_B, data_B, image_directory_B, "Table_B")
if data_T:
    write_data_to_sheet(ws_T, data_T, image_directory_T, "Table_T")

# Save the Excel file
wb.save(output_excel)

sg.popup(f"Data has been saved to {output_csv} and {output_excel}")





###########################

#working test column format

import os
import shutil
import pandas as pd
import xml.etree.ElementTree as ET
import PySimpleGUI as sg

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to DataFrame
def xml_to_dataframe(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    data = []
    for feeder in root.findall('Feeder'):
        feeder_dict = {}
        for child in feeder.iter():
            if child.tag != 'Feeder':
                if child.text:
                    if child.tag == 'AVL':
                        feeder_dict[child.tag] = '"' + child.text.strip() + '"'
                    else:
                        feeder_dict[child.tag] = child.text.strip()
        if feeder_dict:  # Only append if the dictionary is not empty
            data.append(feeder_dict)
    
    df = pd.DataFrame(data)
    
    # Remove rows where PartNumber is blank
    df = df[df['PartNumber'].notna() & (df['PartNumber'].str.strip() != '')]

    # Sort DataFrame by PartNumber column
    df.sort_values(by=['PartNumber'], inplace=True)
    
    return df

# Function to convert DataFrame to Excel file
def dataframe_to_excel(df, excel_file):
    writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()

# Function to process the Excel file and keep specified columns
def process_excel_file(file_path):
    df = pd.read_excel(file_path)
    
    # Keep only the specified columns in the given order
    columns_to_keep = ['PartNumber', 'AVLNAME', 'AVL', 'PartShapeName', 'PartNumberImage', 'QTY']
    df = df[columns_to_keep]
    
    # Add the Side column based on the file name
    side = 'BOT' if 'AVL_B' in file_path else 'TOP'
    df['Side'] = side
    
    # Save the modified DataFrame back to Excel
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()

# Default paths for the input files
default_path_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.txt"
default_path_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.txt"
default_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Create a simple GUI for file and directory selection
layout = [
    [sg.Checkbox('FeederSetup_AVL_B.txt', default=True, key='-USE_FILE_B-'), sg.Checkbox('FeederSetup_AVL_T.txt', default=True, key='-USE_FILE_T-')],
    [sg.Text('Select Image Directory for AVL_B', size=(30, 1)), sg.Input(key='-DIR_B-', size=(60, 1), readonly=True), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Text('Select Image Directory for AVL_T', size=(30, 1)), sg.Input(key='-DIR_T-', size=(60, 1), readonly=True), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Submit(), sg.Cancel()]
]

window = sg.Window('Feeder Setup Data Extractor', layout)

event, values = window.read()
window.close()

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"

if not os.path.isdir(yourfolder):
    os.makedirs(yourfolder)

if event == 'Submit':
    use_file_B = values['-USE_FILE_B-']
    use_file_T = values['-USE_FILE_T-']
    image_directory_B = values['-DIR_B-']
    image_directory_T = values['-DIR_T-']
    
    # Check if at least one file is selected
    if not (use_file_B or use_file_T):
        sg.popup('Please select at least one file to process.')
    else:
        output_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"
        
        # Processing FeederSetup files
        feeder_setup_files = find_feeder_setup_files(root_directory)
        for feeder_setup_file in feeder_setup_files:
            # Rename XML file based on description
            renamed_xml_file = rename_feeder_setup_with_description(feeder_setup_file)
            
            # Define file paths
            input_file = renamed_xml_file if renamed_xml_file else 'FeederSetup.xml'
            
            if os.path.exists(input_file):
                # Convert XML to DataFrame
                df = xml_to_dataframe(input_file)
                
                # Convert DataFrame to Excel
                output_excel = os.path.splitext(input_file)[0] + '.xlsx'
                dataframe_to_excel(df, output_excel)
                
                # Move the Excel file to the destination directory
                shutil.move(output_excel, output_directory)
                
                # Process the Excel file to keep specific columns and add the Side column
                final_excel_path = os.path.join(output_directory, os.path.basename(output_excel))
                process_excel_file(final_excel_path)
            else:
                print(f"File not found: {input_file}. Skipping...")

        sg.popup(f"Data has been saved to the Excel files in {output_directory}")

print("Conversion complete.")

#############################################

#working 

import os
import shutil
import xml.etree.ElementTree as ET
import pandas as pd
from tkinter import filedialog  # Add this import statement
import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Function to find all FeederSetup.xml files
def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.xml":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

# Function to find the setup description from the file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename FeederSetup.xml with a description-based name
def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line4" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_T.xml"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_AVL_B.xml"
    else:
        return None
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print(f"FeederSetup.xml renamed to: {new_path}")
    return new_path

# Function to convert XML to DataFrame
def xml_to_dataframe(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    data = []
    for feeder in root.findall('Feeder'):
        feeder_dict = {}
        for child in feeder.iter():
            if child.tag != 'Feeder':
                if child.text:
                    if child.tag == 'AVL':
                        feeder_dict[child.tag] = '"' + child.text.strip() + '"'
                    else:
                        feeder_dict[child.tag] = child.text.strip()
        if feeder_dict:  # Only append if the dictionary is not empty
            data.append(feeder_dict)
    
    df = pd.DataFrame(data)
    
    # Remove rows where PartNumber is blank
    df = df[df['PartNumber'].notna() & (df['PartNumber'].str.strip() != '')]

    # Sort DataFrame by PartNumber column
    df.sort_values(by=['PartNumber'], inplace=True)
    
    return df

# Function to convert DataFrame to Excel file
def dataframe_to_excel(df, excel_file):
    writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()

# Function to process the Excel file and keep specified columns
def process_excel_file(file_path):
    df = pd.read_excel(file_path)
    
    # Keep only the specified columns in the given order
    columns_to_keep = ['PartNumber', 'AVLNAME', 'AVL', 'PartShapeName', 'PartNumberImage', 'QTY']
    df = df[columns_to_keep]
    
    # Add the Side column based on the file name
    side = 'BOT' if 'AVL_B' in file_path else 'TOP'
    df['Side'] = side
    
    # Save the modified DataFrame back to Excel
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()

# Processing FeederSetup files
feeder_setup_files = find_feeder_setup_files(root_directory)
for feeder_setup_file in feeder_setup_files:
    # Rename XML file based on description
    renamed_xml_file = rename_feeder_setup_with_description(feeder_setup_file)
    
    # Define file paths
    input_file = renamed_xml_file if renamed_xml_file else 'FeederSetup.xml'
    destination_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"
    
    if os.path.exists(input_file):
        # Convert XML to DataFrame
        df = xml_to_dataframe(input_file)
        
        # Convert DataFrame to Excel
        output_excel = os.path.splitext(input_file)[0] + '.xlsx'
        dataframe_to_excel(df, output_excel)
        
        # Move the Excel file to the destination directory
        shutil.move(output_excel, destination_directory)
        
        # Process the Excel file to keep specific columns and add the Side column
        final_excel_path = os.path.join(destination_directory, os.path.basename(output_excel))
        process_excel_file(final_excel_path)
    else:
        print(f"File not found: {input_file}. Skipping...")

# Additional functions to write data to Excel with images and formatting
def extract_data(input_file, side):
    part_numbers = []
    extracted_data = []

    df = pd.read_excel(input_file)

    for index, row in df.iterrows():
        current_data = {
            "PartNumber": row['PartNumber'],
            "AVLNAME": row['AVLNAME'],
            "AVL": row['AVL'],
            "PartShapeName": row['PartShapeName'],
            "PartNumberImage": row['PartNumberImage'],
            "QTY": row['QTY'],
            "Side": side
        }
        extracted_data.append(current_data)
    
    return extracted_data

# Default paths for the input files
default_path_B = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_B.xlsx"
default_path_T = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup\FeederSetup_AVL_T.xlsx"
default_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Create a simple GUI for file and directory selection
layout = [
    [sg.Checkbox('FeederSetup_AVL_B.xlsx', default=True, key='-USE_FILE_B-'), sg.Checkbox('FeederSetup_AVL_T.xlsx', default=True, key='-USE_FILE_T-')],
    [sg.Text('Select Image Directory for AVL_B', size=(30, 1)), sg.Input(key='-DIR_B-', size=(60, 1), readonly=True), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Text('Select Image Directory for AVL_T', size=(30, 1)), sg.Input(key='-DIR_T-', size=(60, 1), readonly=True), sg.FolderBrowse(initial_folder=default_directory)],
    [sg.Submit(), sg.Cancel()]
]

window = sg.Window('Feeder Setup Data Extractor', layout)

event, values = window.read()
window.close()

os.getcwd()
Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
Chd = os.getcwd()

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)

if event == 'Submit':
    use_file_B = values['-USE_FILE_B-']
    use_file_T = values['-USE_FILE_T-']
    image_directory_B = values['-DIR_B-']
    image_directory_T = values['-DIR_T-']
    
    if not (use_file_B or use_file_T):
        sg.popup("Error: At least one of FeederSetup_AVL_B.xlsx or FeederSetup_AVL_T.xlsx must be selected.")
    else:
        extracted_data = []
        
        if use_file_B:
            file_path_B = default_path_B
            side_B = 'BOT'
            extracted_data.extend(extract_data(file_path_B, side_B))
        
        if use_file_T:
            file_path_T = default_path_T
            side_T = 'TOP'
            extracted_data.extend(extract_data(file_path_T, side_T))
        
        if extracted_data:
            df = pd.DataFrame(extracted_data)
            
            output_excel = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check/AVL_Polarity_Check.xlsx'
            writer = pd.ExcelWriter(output_excel, engine='openpyxl')
            df.to_excel(writer, index=False)
            writer.close()  # Close the writer
    
    # Load the workbook and continue with the remaining code...

            
            wb = load_workbook(output_excel)
            ws = wb.active
            
            table = Table(displayName="FeederSetupTable", ref=ws.dimensions)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)
            
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
            
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = alignment
            
            max_image_width = 150
            max_image_height = 150
            
            def resize_image(image_path, max_width, max_height):
                from PIL import Image as PILImage
                with PILImage.open(image_path) as img:
                    width_ratio = max_width / img.width
                    height_ratio = max_height / img.height
                    new_width = int(img.width * min(width_ratio, height_ratio))
                    new_height = int(img.height * min(width_ratio, height_ratio))
                    resized_img = img.resize((new_width, new_height))
                    return resized_img
            
            for row in range(2, ws.max_row + 1):
                image_name = str(ws.cell(row=row, column=5).value)  # Ensure image_name is treated as a string
                side = ws.cell(row=row, column=7).value
                if image_name:
                    image_name = image_name.strip().replace('"', '')
                    image_path = os.path.join(image_directory_B if side == 'BOT' else image_directory_T, image_name + '.bmp')
                    if os.path.exists(image_path):
                        resized_img = resize_image(image_path, max_image_width, max_image_height)
                        resized_img_path = os.path.join("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/AVL & Polarity Check", image_name + '_resized.bmp')
                        resized_img.save(resized_img_path)
                        img = Image(resized_img_path)
                        img.anchor = 'E' + str(row)
                        ws.add_image(img)
            
            for column in ['A', 'B', 'C', 'D', 'F', 'G']:
                max_length = 0
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output_excel)
            sg.popup(f"Data extracted and saved to {output_excel}")
        else:
            sg.popup("No data extracted.")
else:
    sg.popup("Operation cancelled.")
