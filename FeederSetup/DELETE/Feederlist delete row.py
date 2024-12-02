import pandas as pd
import os
from spire.xls import *
from spire.xls.common import *

Chd=os.getcwd
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Load the Excel file into a DataFrame
#df = pd.read_excel('your_excel_file.xlsx')

#Specify the input and output file paths
inputFile = "125852_T.xlsx"
#outputFile = "DeleteBlankRows.xlsx"

#Create a workbook instance
workbook = Workbook()
#Load an Excel file
workbook.LoadFromFile(inputFile)

#Get the first worksheet
sheet1 = workbook.Worksheets[1]
#Delete blank rows from the worksheet
for i in range(sheet1.Rows.Length - 1, -1, -1):
    if sheet1.Rows[i].IsBlank:
        sheet1.DeleteRow(i + 1)

sheet2 = workbook.Worksheets[2]
#Delete blank rows from the worksheet
for i in range(sheet2.Rows.Length - 1, -1, -1):
    if sheet2.Rows[i].IsBlank:
        sheet2.DeleteRow(i + 1)

sheet3 = workbook.Worksheets[3]
#Delete blank rows from the worksheet
for i in range(sheet3.Rows.Length - 1, -1, -1):
    if sheet3.Rows[i].IsBlank:
        sheet3.DeleteRow(i + 1)

#Save the result file
workbook.SaveToFile(inputFile, ExcelVersion.Version2013)
workbook.Dispose()





import openpyxl
import os

# Change directory to the location of your Excel file
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Load the workbook
workbook = openpyxl.load_workbook('125852_T.xlsx')

# Select the worksheet
worksheet = workbook['NXT']  # Replace 'Your_Worksheet_Name' with your sheet name

# Iterate over all rows in the worksheet
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    # Check if the row is empty
    if all(cell.value is None for cell in row):
        # If the row is empty, hide it
        worksheet.row_dimensions[row[0].row].hidden = True

# Save the workbook
workbook.save('125852_T.xlsx')



import openpyxl
import os

# Change directory to the location of your Excel file
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Load the workbook
workbook = openpyxl.load_workbook('125852_T.xlsx')

# Specify the worksheet names
worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

# Iterate over each worksheet
for sheet_name in worksheet_names:
    # Select the worksheet
    worksheet = workbook[sheet_name]

    # Iterate over all rows in the worksheet
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        # Check if the row is empty
        if all(cell.value is None for cell in row):
            # If the row is empty, hide it
            worksheet.row_dimensions[row[0].row].hidden = True

# Save the workbook
workbook.save('125852_T.xlsx')


import openpyxl
import os

# Change directory to the location of your Excel file
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Load the workbook
workbook = openpyxl.load_workbook('125852_T.xlsx')

# Specify the worksheet names
worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

# Iterate over each worksheet
for sheet_name in worksheet_names:
    # Select the worksheet
    worksheet = workbook[sheet_name]

    # Iterate over all rows in the worksheet
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        # Check if the row is empty
        if all(cell.value is None for cell in row):
            # If the row is empty, hide it
            worksheet.row_dimensions[row[0].row].hidden = True

    # Input value into cell B3
    worksheet['B3'] = 'Your Value'  # Replace 'Your Value' with the value you want to input

# Save the workbook
workbook.save('125852_T.xlsx')


import openpyxl
import os

# Change directory to the location of your Excel files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Load the workbooks
workbook_T = openpyxl.load_workbook('Line X Sample T.xlsx')
workbook_B = openpyxl.load_workbook('Line X Sample B.xlsx')

# Specify the worksheet names
worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

# Input value for cell B3
value_B3 = input("Enter the value for cell B3: ")

# Process the input to ensure it matches the specified format
input_data = value_B3[:12].strip()  # Take only the first 12 characters and remove extra spaces
if len(input_data) < 12:
    print("Warning: Input data contains less than 12 characters.")
    input_data += ' ' * (12 - len(input_data))  # Pad with spaces if the input is less than 12 characters

# Prompt for Rev A1 input
rev_input = input("Enter the value for Rev A1: ")

# Iterate over each worksheet in both workbooks
for workbook in [workbook_T, workbook_B]:
    for sheet_name in worksheet_names:
        # Select the worksheet
        worksheet = workbook[sheet_name]

        # Iterate over all rows in the worksheet
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            # Check if the row is empty
            if all(cell.value is None for cell in row):
                # If the row is empty, hide it
                worksheet.row_dimensions[row[0].row].hidden = True

        # Input value into cell B3
        worksheet['B3'] = input_data

        # Input value into Rev A1
        worksheet['A1'] = rev_input

# Save the workbooks
workbook_T.save('Line X Sample T.xlsx')
workbook_B.save('Line X Sample B.xlsx')


import openpyxl
import os

# Change directory to the location of your Excel files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Load the workbooks
workbook_T = openpyxl.load_workbook('Line X Sample T.xlsx')
workbook_B = openpyxl.load_workbook('Line X Sample B.xlsx')

# Specify the worksheet names
worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

# Input values for cell B3 and Rev A1
value_B3 = input("Enter the value for cell B3 (12 characters, followed by T or B, followed by Revision): ").strip()[:15]  # Take only the first 15 characters and remove extra spaces
rev_input = input("Enter the value for Revision A1: ")

# Iterate over each workbook
for workbook in [workbook_T, workbook_B]:
    for sheet_name in worksheet_names:
        # Select the worksheet
        worksheet = workbook[sheet_name]

        # Iterate over all rows in the worksheet
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            # Check if the row is empty
            if all(cell.value is None for cell in row):
                # If the row is empty, hide it
                worksheet.row_dimensions[row[0].row].hidden = True

        # Combine the input values for B3 cell
        worksheet['B3'] = value_B3 + " " + rev_input

# Save the workbooks
workbook_T.save('Line X Sample T.xlsx')
workbook_B.save('Line X Sample B.xlsx')



import openpyxl
import os

# Change directory to the location of your Excel files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Load the workbooks
workbook_T = openpyxl.load_workbook('Line X Sample T.xlsx')
workbook_B = openpyxl.load_workbook('Line X Sample B.xlsx')

# Specify the worksheet names
worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

# Input values for cell B3 and Rev A1
value_B3 = input("Enter the value for cell B3 (12 characters, followed by Revision): ").strip()[:12]  # Take only the first 12 characters and remove extra spaces
rev_input = input("Enter the value for Revision A1: ")

# Iterate over each workbook
for workbook in [workbook_T, workbook_B]:
    for sheet_name in worksheet_names:
        # Select the worksheet
        worksheet = workbook[sheet_name]

        # Check if the workbook name contains 'T' or 'B'
        if 'T' in workbook_T:
            # If 'T' is in the workbook name, set the value as 'T'
            location = 'T'
        elif 'B' in workbook_B:
            # If 'B' is in the workbook name, set the value as 'B'
            location = 'B'
        else:
            # Default to 'T' if neither 'T' nor 'B' is found
            location = 'T'

        # Combine the input values for B3 cell
        worksheet['B3'] = value_B3 + " " + location + " " + rev_input

# Save the workbooks
workbook_T.save('Line X Sample T.xlsx')
workbook_B.save('Line X Sample B.xlsx')



import openpyxl
import os

# Change directory to the location of your Excel files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Load the workbooks
workbook_T = openpyxl.load_workbook('Line X Sample T.xlsx')
workbook_B = openpyxl.load_workbook('Line X Sample B.xlsx')

# Specify the worksheet names
worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

# Input values for cell B3 and Rev A1
value_B3 = input("Enter the value for cell B3 (12 characters, followed by Revision): ").strip()[:12]  # Take only the first 12 characters and remove extra spaces
rev_input = input("Enter the value for Revision A1: ")

# Iterate over each workbook
for workbook in [workbook_T, workbook_B]:
    for sheet_name in worksheet_names:
        # Select the worksheet
        worksheet = workbook[sheet_name]

        # Collect indices of empty rows
        empty_rows = [row[0].row for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row) if all(cell.value is None for cell in row)]

        # Iterate over the collected indices in reverse order and delete rows
        for row_index in reversed(empty_rows):
            worksheet.delete_rows(row_index)

        # Check if the workbook name contains 'T' or 'B'
        if 'T' in workbook_T:
            # If 'T' is in the workbook name, set the value as 'T'
            location = 'T'
        elif 'B' in workbook_B:
            # If 'B' is in the workbook name, set the value as 'B'
            location = 'B'
        else:
            # Default to 'T' if neither 'T' nor 'B' is found
            location = 'T'

        # Combine the input values for B3 cell
        worksheet['B3'] = value_B3 + " " + location + " " + rev_input

# Save the workbooks
workbook_T.save('Line X Sample T.xlsx')
workbook_B.save('Line X Sample B.xlsx')



'''# Load the workbooks
    workbook_T = openpyxl.load_workbook('Line X Sample T.xlsx')
    workbook_B = openpyxl.load_workbook('Line X Sample B.xlsx')

    # Specify the worksheet names
    worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

    # Input values for cell B3 and Rev A1
    #value_B3 = input("Enter the value for cell B3 (12 characters, followed by T or B, followed by Revision): ").strip()[:15]  # Take only the first 15 characters and remove extra spaces
    #Revision = input("Enter the value for Revision A1: ")

    # Iterate over each workbook
    for workbook in [workbook_T, workbook_B]:
        for sheet_name in worksheet_names:
            # Select the worksheet
            worksheet = workbook[sheet_name]

            # Iterate over all rows in the worksheet
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                # Check if the row is empty
                if all(cell.value is None for cell in row):
                    # If the row is empty, hide it
                    worksheet.row_dimensions[row[0].row].hidden = True

                    # Check if the workbook name contains 'T' or 'B'
            if 'T' in workbook_T:
                # If 'T' is in the workbook name, set the value as 'T'
                location = 'T'
            elif 'B' in workbook_B:
                # If 'B' is in the workbook name, set the value as 'B'
                location = 'B'
            else:
                # Default to 'T' if neither 'T' nor 'B' is found
                location = 'T'

            # Combine the input values for B3 cell
            worksheet['B3'] = value_B3 + " " + location + " " + Revision

    # Save the workbooks
    workbook_T.save('Line X Sample T.xlsx')
    workbook_B.save('Line X Sample B.xlsx')'''


import openpyxl
import os

# Get the current working directory
os.getcwd()

# Change directory to the location of your Excel files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')

# Get the current working directory
Chd = os.getcwd()

# Load the workbooks
workbook_T = openpyxl.load_workbook('Line X Sample T.xlsx')
workbook_B = openpyxl.load_workbook('Line X Sample B.xlsx')

# Specify the worksheet names
worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

# Input values for cell B3 and Revision A1
print('\n')
value_B3 = input("\033[93mEnter Feeder Name (12 characters): \033[0m").strip()[:12]  # Take only the first 12 characters and remove extra spaces
print('\n')
Revision = input("\033[93mEnter Revision A1: \033[0m")

# Iterate over each workbook
for workbook, workbook_name in [(workbook_T, 'Line X Sample T.xlsx'), (workbook_B, 'Line X Sample B.xlsx')]:
    if 'T' in workbook_name:
        location = 'T'
    elif 'B' in workbook_name:
        location = 'B'
    else:
        location = 'T/B'

    # Iterate over each sheet in the workbook
    for sheet_name in worksheet_names:
        # Select the worksheet
        worksheet = workbook[sheet_name]

        # Iterate over all rows in the worksheet
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            # Check if the row is empty
            if all(cell.value is None for cell in row):
                # If the row is empty, hide it
                worksheet.row_dimensions[row[0].row].hidden = True

        # Combine the input values for B3 cell
        worksheet['B3'] = value_B3 + " " + location + " " + Revision

# Save the workbooks
workbook_T.save('Line X Sample T.xlsx')
workbook_B.save('Line X Sample B.xlsx')
