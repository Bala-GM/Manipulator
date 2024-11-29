import pandas as pd
import numpy as np
import lxml
import os
from os import getcwd
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
from io import StringIO
import csv as csv
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
from io import BytesIO
import time
import sys
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
import PySimpleGUI as sg
import subprocess
import threading
import urllib
import urllib.parse
import sqlalchemy 
import sqlalchemy_access as sq_a
import sqlalchemy_pyodbc_mssql as sqlalchemy
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
import shutil
import pyodbc
import pypyodbc
import odbc
import psycopg2
import MySQLdb as sql
from sqlite3 import dbapi2 as sqlite
import sqlite3

print("*******Feeder and BOM data Verification version--py_V-1.1.0 interface_GUI/D1823*******")

dL1 = input("Enter BOM Name :")
dL2 = input("Enter Feeder Name :")

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
Chd = os.getcwd()

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Upload"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)

os.getcwd()
#Chd= os.chdir('D:\\NX_BACKWORK')
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

if os.path.exists("Feeder_List_OPT.xlsx"):
  os.remove("Feeder_List_OPT.xlsx")
else:
  print("The file does not exist")

if os.path.exists("Feeder_List_OPB.xlsx"):
  os.remove("Feeder_List_OPB.xlsx")
else:
  print("The file does not exist")

if os.path.exists("Upload-Data.xlsx"):
  os.remove("Upload-Data.xlsx")
else:
  print("The file does not exist")

os.getcwd()
#Chd= os.chdir('D:\\NX_BACKWORK')
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')
Chd = os.getcwd()

##########################################################################################################################################

##########################################################################################################################################

#LINE1T

file_path = 'FeederSetup_TL1.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'
print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))
try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)
except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL1.csv', encoding="utf-8",index_col=False, skiprows=range(2, 351), nrows=3)

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']
 
    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)
    
    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    df1 = pd.read_csv('FeederSetup_TL1.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('3-1-2-','3-2-')

    df1['Location'] = df1['Location'].str.replace('3-1-1-','3-1-')

    df1['Location'] = df1['Location'].str.replace('2-1-2-','2-2-')

    df1['Location'] = df1['Location'].str.replace('2-1-1-','2-1-')

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
   
    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')
 
    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE1B

file_path = 'FeederSetup_BL1.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'
print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))
try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)
except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL1.csv', encoding="utf-8",index_col=False, skiprows=range(2, 351), nrows=3)
      
    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']
 
    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER


    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    df1 = pd.read_csv('FeederSetup_BL1.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('3-1-2-','3-2-')

    df1['Location'] = df1['Location'].str.replace('3-1-1-','3-1-')

    df1['Location'] = df1['Location'].str.replace('2-1-2-','2-2-')

    df1['Location'] = df1['Location'].str.replace('2-1-1-','2-1-')

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')
 
    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

    pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE2T

file_path = 'FeederSetup_TL2.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'
print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))
try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)
except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL2.csv', encoding="utf-8",index_col=False, skiprows=range(2, 401), nrows=3)

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']
 
    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER


    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    df1 = pd.read_csv('FeederSetup_TL2.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    #df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('3-1','3')

    df1['Location'] = df1['Location'].str.replace('2-1','2')

    df1['Location'] = df1['Location'].str.replace('1-1','1')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

   
    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)


    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')
 
    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE2B

file_path = 'FeederSetup_BL2.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL2.csv', encoding="utf-8",index_col=False, skiprows=range(2, 401), nrows=3)
      
    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']
 
    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    df1 = pd.read_csv('FeederSetup_BL2.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    #df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('3-1','3')

    df1['Location'] = df1['Location'].str.replace('2-1','2')

    df1['Location'] = df1['Location'].str.replace('1-1','1')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')
 
    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)   
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
    pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE3T

file_path = 'FeederSetup_TL3.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'
print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))
try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)
except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL3.csv', encoding="utf-8",index_col=False, skiprows=range(2, 171), nrows=3)

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']
 
    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    df1 = pd.read_csv('FeederSetup_TL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

   
    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')
 
    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE3B

file_path = 'FeederSetup_BL3.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL3.csv', encoding="utf-8",index_col=False, skiprows=range(2, 171), nrows=3)
      
    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']
 
    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    df1 = pd.read_csv('FeederSetup_BL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')
 
    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

    pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE1T-C4

file_path = 'FeederSetup_TL4C.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'
print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))
try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)
except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2, 228), nrows=3)

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']
 
    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    df1 = pd.read_csv('FeederSetup_TL4C.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('2-1-1-','7-')

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')
 
    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE1B-C4

file_path = 'FeederSetup_BL4C.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2, 228), nrows=3)
      
    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']
 
    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    df1 = pd.read_csv('FeederSetup_BL4C.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('2-1-1-','7-')

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')
 
    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

    pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LEN MERGE TOP AND BOT

data_file_folder = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

df=[]
for file in os.listdir(data_file_folder):
     if file.endswith('.xlsx'):
          print('Loading file {0}...'.format(file))
          df.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Home'))
df1=[]
for file in os.listdir(data_file_folder):
     if file.endswith('.xlsx'):
          print('Loading file {0}...'.format(file))
          df1.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FL_Upload'))
df2=[]
for file in os.listdir(data_file_folder):
     if file.endswith('.xlsx'):
          print('Loading file {0}...'.format(file))
          df2.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FL_Verify'))
df3=[]
for file in os.listdir(data_file_folder):
     if file.endswith('.xlsx'):
          print('Loading file {0}...'.format(file))
          df3.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Size'))
df4=[]
for file in os.listdir(data_file_folder):
     if file.endswith('.xlsx'):
          print('Loading file {0}...'.format(file))
          df4.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Side'))
df5=[]
for file in os.listdir(data_file_folder):
     if file.endswith('.xlsx'):
          print('Loading file {0}...'.format(file))
          df5.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FeederName'))

df6=[]
for file in os.listdir(data_file_folder):
     if file.endswith('.xlsx'):
          print('Loading file {0}...'.format(file))
          df6.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Type'))

len(df)
df_master1 = pd.concat(df, axis=0)
len(df1)
df_master2 = pd.concat(df1, axis=0)
len(df2)
df_master3 = pd.concat(df2, axis=0)
len(df3)
df_master4 = pd.concat(df3, axis=0)
len(df4)
df_master5 = pd.concat(df4, axis=0)
len(df5)
df_master6 = pd.concat(df5, axis=0)
len(df6)
df_master7 = pd.concat(df6, axis=0)

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederSetup.xlsx") as writer:
    #df_master.to_excel('masterfile.xlsx',index=False)
        df_master1.to_excel(writer, sheet_name="Home", index=False)
        df_master2.to_excel(writer, sheet_name="FeederSetup", index=False)
        df_master3.to_excel(writer, sheet_name="FeederCol", index=False)
        df_master4.to_excel(writer, sheet_name="FeederSize", index=False)
        df_master5.to_excel(writer, sheet_name="Total side Count", index=False)
        df_master6.to_excel(writer, sheet_name="FeederName", index=False)
        df_master7.to_excel(writer, sheet_name="Type", index=False)

if os.path.exists("Feeder_List_OPT.xlsx"):
  os.rename("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx" , "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Feeder_List_OPT.xlsx")
else:
  print("The file does not exist")

if os.path.exists("Feeder_List_OPB.xlsx"):
  os.rename("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx" , "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Feeder_List_OPB.xlsx")
else:
  print("The file does not exist")

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#Master list program update and LOG

os.getcwd()
Chd= os.chdir("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified")
Chd = os.getcwd()
xls = pd.ExcelFile('FeederSetup.xlsx',engine='openpyxl')
dfpm1 = pd.read_excel('FeederSetup.xlsx', sheet_name='Home')

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/Pro_ML.xlsx") as writer:
    #df_master.to_excel('masterfile.xlsx',index=False)
        dfpm1.to_excel(writer, sheet_name="Home", index=False)

if os.path.exists("Pro_ML.xlsx"):
    os.rename("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/Pro_ML.xlsx" , "D:/NX_BACKWORK/Database_File/SMT_Master_List/Pro_ML.xlsx")
else:
     print("The file does not exist")

data_file_folder = 'D:/NX_BACKWORK/Database_File/SMT_Master_List'

dspm1=[]
for file in os.listdir(data_file_folder):
     if file.endswith('.xlsx'):
          print('Loading file {0}...'.format(file))
          dspm1.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Home'))

len(dspm1)
dsf_master1 = pd.concat(dspm1, axis=0)

with pd.ExcelWriter("D:/NX_BACKWORK/Database_File/SMT_Master_List/Program_Master_List.xlsx") as writer:
    dsf_master1.to_excel(writer, sheet_name="Home", index=False)

print("Tranfer Complete...")

os.getcwd()
Chd= os.chdir("D:/NX_BACKWORK/Database_File/SMT_Master_List")
Chd = os.getcwd()

print("Del Start....")

if os.path.exists("Pro_ML.xlsx"):
  os.remove("Pro_ML.xlsx")
else:
  print("The file does not exist")

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#BOM MANIPULATE

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
Chd = os.getcwd()

file_path = 'BOM.xlsx'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    ds1 = pd.read_excel(file_path,index_col=False) 

    column_list =['Material', 'AltItemGroup', 'Priority', 'Long. Description', 'Ref.Designator/Circuit Reference', 'Quantity','Material Group']
    column_list =['Internal P/N', 'Group', 'Priority', 'Description', 'Ref.Designator', 'Qty','SMT/THT/Mech']

    ds1.rename(
        columns={'Material':"PartNumber", 'AltItemGroup':"Group", 'Priority':'Priority', 'Long. Description':'Long Des', 'Ref.Designator/Circuit Reference':'RefList', 'Quantity':'Qty','Material Group':'Shape'},
        inplace=True,
    )

    ds1.rename(
        columns={'Internal P/N':"PartNumber", 'Group':"Group", 'Priority':'Priority', 'Description':'Long Des', 'Ref.Designator':'RefList', 'Qty':'Qty','SMT/THT/Mech':'Shape'},
        inplace=True,
    )

    print(ds1)

#file_name ="output.xlsx"
#ds1.to_excel(file_name)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM/BOM_List_OP.xlsx") as writer:
        ds1.to_excel(writer, sheet_name="BOM", index=False)

    pass
    print('The file does not exist.')

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
Chd = os.getcwd()
file_path = 'BOM_List_OP.xlsx'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    ds1 = pd.read_excel(file_path, usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False) 

    ds1 = ds1[['PartNumber', 'Group','Priority','Long Des','Qty','Shape','RefList']]
    ds1['RefList'] = ds1['RefList'].str.replace("_x000D_","")
    ds1['RefList'] = ds1['RefList'].str.replace(" ","")
    ds1['RefList'] = ds1['RefList'].str.replace("\n","")
    print(ds1)

#ds1 = pd.read_excel('Filename_OP.xlsx','BOM', index_col=False)

    ds2 = ds1.explode('RefList')

    ds2['RefList'] = ds2['RefList'].str.replace(" "," ")

#ds2.drop(ds2.iloc[:, 1:6], inplace=True, axis=1)

    print(ds2)

    ds2.rename(columns = {'PartNumber':'B_Part_No'}, inplace = True)

    ds2.rename(columns = {'RefList':'B_Ref_List'}, inplace = True)

    ds2['B_Ref_List'] = ds2['B_Ref_List'] .str.strip('[]').str.split(',')

    ds2.to_dict()

    ds2.explode ('B_Ref_List',ignore_index=True)

    ds3 = ds2.explode('B_Ref_List',ignore_index=True)

    ds2 = ds2[['Group','Priority','B_Part_No']]
    dc1 = ds2[['B_Part_No']]
    dc1.rename(columns = {'B_Part_No':'PBARNO'}, inplace = True)
    dc1['PBARPTN'] = dc1['PBARNO']
    dc1['PBARBAR'] = dc1['PBARNO']
    dc1.insert(3,'PBARQTY', 10000)
    dc1.insert(4,'PBARFTYP', 3)

    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>#
    dfs2 = ds2[['Group','Priority','B_Part_No']]
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("15","A")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("14","B")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("13","C")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("12","D")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("11","E")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("10","F")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("9","PTN_9")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("8","PTN_8")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("7","PTN_7")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("6","PTN_6")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("5","PTN_5")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("4","PTN_4")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("3","PTN_3")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("2","PTN_2")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("1","PTN_1")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("0","PTN_0")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("A","PTN_15")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("B","PTN_14")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("C","PTN_13")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("D","PTN_12")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("E","PTN_11")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("F","PTN_10")
    dfs2.dropna(subset=['Group'], inplace=True)
    df2 = dfs2.pivot(index='Group',columns='Priority',values='B_Part_No')

    ds3.head()

    T10_col = ds3.pop('B_Ref_List') # col-1

    ds3.insert(0, 'B_Ref_List', T10_col)

    ds3 = ds3[['B_Ref_List','B_Part_No','Long Des']]

    ds1.dropna(subset=['RefList'], inplace=True)
    ds3.dropna(subset=['B_Ref_List'], inplace=True)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx") as writer:

    #dt_H.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
    #df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
    #df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        ds1.to_excel(writer, sheet_name="BOM", index=False)
        ds2.to_excel(writer, sheet_name="AVL GROUP", index=False)
        dc1.to_excel(writer, sheet_name="Part Master", index=False)
        df2.to_excel(writer, sheet_name="AVL_SHEET", index=True)
        ds3.to_excel(writer, sheet_name="BOM_Data", index=False)

    pass
    print('The file does not exist.')
 
##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#FEEDERVERIFICATIONCODEBOMANDFEEDER

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
xls = pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
df1 = pd.read_excel('BOM_List_OP.xlsx', sheet_name='BOM_Data')
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
xls = pd.ExcelFile('FeederSetup.xlsx',engine='openpyxl')
df2 = pd.read_excel('FeederSetup.xlsx', sheet_name='FeederCol')
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
xls = pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
df111 = pd.read_excel('BOM_List_OP.xlsx', sheet_name='BOM')
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
xls = pd.ExcelFile('FeederSetup.xlsx',engine='openpyxl')
df112 = pd.read_excel('FeederSetup.xlsx', sheet_name='FeederSetup')
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx") as writer:
    df1.to_excel(writer, sheet_name="BOM_Data", index=False)
    df2.to_excel(writer, sheet_name="FeederCol", index=False)
    df111.to_excel(writer, sheet_name="BOM", index=False)
    df112.to_excel(writer, sheet_name="FeederSetup", index=False)

xls = pd.ExcelFile('FeederVerify.xlsx',engine='openpyxl')
df1 = pd.read_excel("FeederVerify.xlsx", sheet_name='BOM_Data')
df2 = pd.read_excel("FeederVerify.xlsx", sheet_name='FeederCol')
df111 = pd.read_excel("FeederVerify.xlsx", sheet_name='BOM')
df112 = pd.read_excel("FeederVerify.xlsx", sheet_name='FeederSetup')
    
df2['Feeder Reference'] = df2['F_Ref_List']
df1.rename(columns = {'B_Ref_List':'F_Ref_List'}, inplace = True)
    #df1['B_Ref.List'] = df1['F_Ref_List']
df3 = pd.merge(df1 , df2, on='F_Ref_List', how='left')
df3.rename(columns = {'F_Ref_List':'BOM Reference'}, inplace = True)
df1.rename(columns = {'F_Ref_List':'B_Ref_List'}, inplace = True)
print(df1,df2)

df111.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)
df111 = df111[['F_Part_No','Long Des']]
df113 = pd.merge(df111 , df112, on='F_Part_No', how='inner')
df113.rename(columns = {'F_Part_No':'Part Number'}, inplace = True)
df113.rename(columns = {'Location':'Feeder Location'}, inplace = True)
df113.rename(columns = {'Long Des':'Part Description'}, inplace = True)
df113.rename(columns = {'F_Ref_List':'Reference'}, inplace = True)
df113 = df113[['Feeder Location','FeederName','Type','Size','FeedPitch','Part Height','Part Number','Part Description','Reference','QTY','Side','ModelName']]

df3["BOM and Feeder Compare"] = (df3["B_Part_No"] == df3["F_Part_No"])
df3['BOM and Feeder Compare'] = df3['BOM and Feeder Compare'].replace('TRUE','MATCH')
df3['BOM and Feeder Compare'] = df3['BOM and Feeder Compare'].replace('FALSE','MISS_MATCH')

df3 = df3.copy()
df3['BOM and Feeder Compare'] = df3['BOM and Feeder Compare'].map({True: 'Match', False: 'Miss_Match'})
df3.sort_values(by='BOM and Feeder Compare', inplace=True, ascending=False)

df4 = df3['BOM and Feeder Compare'].value_counts()
#df4 = df3['Size'].value_counts()
df5 = df3['Side'].value_counts()
#df6 = df3['F_Ref_List'].value_counts()
df7 = df3['B_Part_No'].value_counts()
#df8 = df3['B_Ref.List'].value_counts()
df9 = df3['F_Part_No'].value_counts()
#df10 = df3['FeederName'].value_counts()
print('***')
dbf1 = print(df1)
print('***')
dbf2 = print(df2)
print('***')
dbf2 = df2.copy()
dbf1 = df1.copy()
dbf2_col = dbf2.pop('Feeder Reference')
dbf2.insert(1, 'Feeder Reference', dbf2_col)
dbf2.rename(columns = {'F_Ref_List':'B_Ref_List'}, inplace = True)
dbf3 = pd.merge(dbf2 , dbf1, on='B_Ref_List', how='left')

dbf3["Feeder and BOM Compare"] = (dbf3["F_Part_No"] == dbf3["B_Part_No"])
dbf3['Feeder and BOM Compare'] = dbf3['Feeder and BOM Compare'].replace('TRUE','MATCH')
dbf3['Feeder and BOM Compare'] = dbf3['Feeder and BOM Compare'].replace('FALSE','MISS_MATCH')

dbf3 = dbf3.copy()
dbf3['Feeder and BOM Compare'] = dbf3['Feeder and BOM Compare'].map({True: 'Match', False: 'Miss_Match'})
dbf3.sort_values(by='Feeder and BOM Compare', inplace=True, ascending=False)

dbf4 = dbf3['Feeder and BOM Compare'].value_counts() 

# Define a function for row styling
def highlight_row(row):
    return ['background-color: lightgreen' if 'Match' in row.values else
            'background-color: yellow' if 'Miss_Match' in row.values else
            '' for _ in row]

# Apply the styling function to the DataFrame
styled_df3 = df3.style.apply(highlight_row, axis=1)

# Define a function for row styling
def highlight_row(row):
    return ['background-color: lightgreen' if 'Match' in row.values else
            'background-color: yellow' if 'Miss_Match' in row.values else
            '' for _ in row]

# Apply the styling function to the DataFrame
styled_dbf3 = dbf3.style.apply(highlight_row, axis=1)

# Save the styled DataFrame to Excel
with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx") as writer:

        styled_df3.to_excel(writer, sheet_name="Verify_data_BAF", index=False, engine='openpyxl')
        df4.to_excel(writer, sheet_name="BOM and Feeder Compare",index=TRUE)
        #dbf2.to_excel(writer, sheet_name="Verify_data_FAB", index=False)
        styled_dbf3.to_excel(writer, sheet_name="Verify_data_FAB", index=False, engine='openpyxl')
        dbf4.to_excel(writer, sheet_name="Feeder and BOM Compare",index=TRUE)
        df113.to_excel(writer, sheet_name="Upload_data", index=False)
        df1.to_excel(writer, sheet_name="BOM_data", index=False)
        df2.to_excel(writer, sheet_name="Feeder_data", index=False)
        df5.to_excel(writer, sheet_name="Side",index=TRUE)
        #df6.to_excel(writer, sheet_name="F_Ref_List",index=TRUE)
        df7.to_excel(writer, sheet_name="B_Part_No",index=TRUE)
        #df8.to_excel(writer, sheet_name="B_Ref.List",index=TRUE)
        df9.to_excel(writer, sheet_name="F_Part_No",index=TRUE)
        #df10.to_excel(writer, sheet_name="FeederName",index=TRUE)

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#SEPRATEFEEDERLISTDATA

os.getcwd()

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

df1 = pd.read_excel("FeederVerify.xlsx", sheet_name="Upload_data")

df1.sort_values(by='Side', inplace=True, ascending=True)
df2_1 = df1
#del.TOP
df1 = df1[df1["Side"].str.contains("TOP")==False]
df2 = df1[df1["ModelName"].str.contains("AIMEX2|AIMEX3|AIMEX-IIIC_2|AIMEX-IIIC_3")==False]
df2.sort_values(by='Feeder Location', inplace=True, ascending=True)
df3 = df1[df1["ModelName"].str.contains("NXT|AIMEX3|AIMEX-IIIC_1|AIMEX-IIIC_3")==False]
df3.sort_values(by='Feeder Location', inplace=True, ascending=True)
df4 = df1[df1["ModelName"].str.contains("NXT|AIMEX2|AIMEX-IIIC_1|AIMEX-IIIC_2")==False]
df4.sort_values(by='Feeder Location', inplace=True, ascending=True)
#del.BOT
df2_1 = df2_1[df2_1["Side"].str.contains("BOT")==False]
df2_2 = df2_1[df2_1["ModelName"].str.contains("AIMEX2|AIMEX3|AIMEX-IIIC_2|AIMEX-IIIC_3")==False]
df2_2.sort_values(by='Feeder Location', inplace=True, ascending=True)
df2_3 = df2_1[df2_1["ModelName"].str.contains("NXT|AIMEX3|AIMEX-IIIC_1|AIMEX-IIIC_3")==False]
df2_3.sort_values(by='Feeder Location', inplace=True, ascending=True)
df2_4 = df2_1[df2_1["ModelName"].str.contains("NXT|AIMEX2|AIMEX-IIIC_1|AIMEX-IIIC_2")==False]
df2_4.sort_values(by='Feeder Location', inplace=True, ascending=True)

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data.xlsx") as writer:
    df2.to_excel(writer, sheet_name="NXT&AMX1_B", index=False)
    df3.to_excel(writer, sheet_name="AMX2_B", index=False)
    df4.to_excel(writer, sheet_name="AMX3_B", index=False)
    
    df2_2.to_excel(writer, sheet_name="NXT&AMX1_T", index=False)
    df2_3.to_excel(writer, sheet_name="AMX2_T", index=False)
    df2_4.to_excel(writer, sheet_name="AMX3_T", index=False)

##########################################################################################################################################

##########################################################################################################################################

#Upload data to merge and del side and Module
    
os.getcwd()

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

df1 = pd.read_excel("Upload-Data.xlsx", sheet_name="NXT&AMX1_B")
df1["Remarks"] = df1['Side'].astype(str) +"--"+ df1['ModelName']
del df1['Side']
del df1['ModelName']

df2 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX2_B")
df2["Remarks"] = df2['Side'].astype(str) +"--"+ df2['ModelName']
del df2['Side']
del df2['ModelName']

df3 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX3_B")
df3["Remarks"] = df3['Side'].astype(str) +"--"+ df3['ModelName']
del df3['Side']
del df3['ModelName']

df4 = pd.read_excel("Upload-Data.xlsx", sheet_name="NXT&AMX1_T")
df4["Remarks"] = df4['Side'].astype(str) +"--"+ df4['ModelName']
del df4['Side']
del df4['ModelName']

df5 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX2_T")
df5["Remarks"] = df5['Side'].astype(str) +"--"+ df5['ModelName']
del df5['Side']
del df5['ModelName']

df6 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX3_T")
df6["Remarks"] = df6['Side'].astype(str) +"--"+ df6['ModelName']
del df6['Side']
del df6['ModelName']

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx") as writer:
    df1.to_excel(writer, sheet_name="NXT&AMX1_B", index=False)
    df2.to_excel(writer, sheet_name="AMX2_B", index=False)
    df3.to_excel(writer, sheet_name="AMX3_B", index=False)
    df4.to_excel(writer, sheet_name="NXT&AMX1_T", index=False)
    df5.to_excel(writer, sheet_name="AMX2_T", index=False)
    df6.to_excel(writer, sheet_name="AMX3_T", index=False)

##########################################################################################################################################

##########################################################################################################################################

#PART MASTER

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

# Excel file path
df_PM1 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="Part Master")
with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/PartMaster.xlsx") as writer:
    df_PM1.to_excel(writer, sheet_name="T_PBAR", index=False)
excel_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/PartMaster.xlsx'

# Access database connection parameters
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_Part Master/MODEL.mdb', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb')
access_db_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb'
driver = 'Microsoft Access Driver (*.mdb, *.accdb)'
user = ''
password = ''

# Set up the connection string
conn_str = f"DRIVER={{{driver}}};DBQ={access_db_path};UID={user};PWD={password};"

# Connect to the Access database
print(pyodbc.drivers())
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()

# Read Excel data into a pandas DataFrame
print('Open Excel....')
df = pd.read_excel(excel_file_path)
print(df.head(10))

# Define the table name in the Access database
print('open MS Access....')
table_name = 'T_PBAR'

# Check if the table exists
existing_tables = [table[2] for table in cursor.tables(tableType='TABLE')]
if table_name in existing_tables:
    # Append data to the existing table
    for _, row in df.iterrows():
        insert_query = f'''
        INSERT INTO {table_name} ({', '.join(df.columns)})
        VALUES ({', '.join(map(lambda x: f"'{row[x]}'", df.columns))})
        '''
        cursor.execute(insert_query)
        conn.commit()
        print('writing to access')
else:
    print(f"The table '{table_name}' does not exist in the Access database.")

# Close the database connection
conn.close()
print('write complete')

##########################################################################################################################################

##########################################################################################################################################

#AVL#@@#

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')

print(pyodbc.drivers())

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

xls = pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
df1 = pd.read_excel('BOM_List_OP.xlsx', sheet_name='AVL_SHEET')
print(df1.head(10))
with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
    df1.to_excel(writer, sheet_name="AVL_SHEET", index=False)
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()
df_AL1 = pd.read_excel('AVL.xlsx', sheet_name='AVL_SHEET')

#new_column_name  = df_AL1.insert(1,'Group Name', '') #new_column_name = ('Group Name') # Get user input for the new column name
# Desired column name
desired_column_name = 'Group'

# Get user input for the new column name
new_column_name = ("Group Name")

# Get user input for the new column value
new_column_value = input(f"Enter the value for the new column '{new_column_name}': ")
#dL1 = new_column_value
# Check if the desired column name exists
if desired_column_name in df_AL1.columns:
    # Find the index of the desired column
    index_of_desired_column = df_AL1.columns.get_loc(desired_column_name)
    
    # Insert the new column next to the desired column
    df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value) #dL1 = new_column_value

print(df_AL1)

df_AL1['AVL Name']=df_AL1['PTN_1']

first_column = df_AL1.pop('AVL Name')

df_AL1.insert(2, 'AVL Name', first_column)

df_AL1.insert(3, 'Comment', '')

df_AL1['Comment'] = df_AL1['Group']

#PTN_1>>

# Desired column name
desired_column_name1 = 'PTN_1'

# New column to insert
new_column_name = 'P_1'
new_column_value = '1'

# Check if the desired column name exists
if desired_column_name1 in df_AL1.columns:
    # Find the index of the desired column
    index_of_desired_column = df_AL1.columns.get_loc(desired_column_name1)
    
    # Insert the new column next to the desired column
    df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value)

#PTN_2>>

desired_column_name2 = 'PTN_2'

# New column to insert
new_column_name = 'P_2'
new_column_value = '0'

# Check if the desired column name exists
if desired_column_name2 in df_AL1.columns:
    # Find the index of the desired column
    index_of_desired_column = df_AL1.columns.get_loc(desired_column_name2)
    
    # Insert the new column next to the desired column
    df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value)

#PTN_3>>

desired_column_name3 = 'PTN_3'
column_to_replace3 = 'PTN_3'

# Check if the column exists in the DataFrame
if column_to_replace3 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace3] = df_AL1[column_to_replace3].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name3 = 'P_3'
    new_column_value3 = '0'

    # Check if the desired column name exists
    if desired_column_name3 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name3)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name3,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name3].any():
            print(f"Values are present in '{desired_column_name3}' column:")
            print(df_AL1[desired_column_name3])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name3] != "$", new_column_name3] = 0
        else:
            print(f"No values are present in '{desired_column_name3}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name3}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace3}' does not exist in the DataFrame.")

column_to_replace3 = 'PTN_3'

# Check if the column exists in the DataFrame
if column_to_replace3 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace3] = df_AL1[column_to_replace3].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace3}' does not exist in the DataFrame.")

#PTN_4>>

desired_column_name4 = 'PTN_4'
column_to_replace4 = 'PTN_4'

# Check if the column exists in the DataFrame
if column_to_replace4 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace4] = df_AL1[column_to_replace4].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name4 = 'P_4'
    new_column_value4 = '0'

    # Check if the desired column name exists
    if desired_column_name4 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name4)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name4,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name4].any():
            print(f"Values are present in '{desired_column_name4}' column:")
            print(df_AL1[desired_column_name4])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name4] != "$", new_column_name4] = 0
        else:
            print(f"No values are present in '{desired_column_name4}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name4}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace4}' does not exist in the DataFrame.")

column_to_replace4 = 'PTN_4'

# Check if the column exists in the DataFrame
if column_to_replace4 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace4] = df_AL1[column_to_replace4].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace4}' does not exist in the DataFrame.")

#PTN_5>>

desired_column_name5 = 'PTN_5'
column_to_replace5 = 'PTN_5'

# Check if the column exists in the DataFrame
if column_to_replace5 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace5] = df_AL1[column_to_replace5].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name5 = 'P_5'
    new_column_value5 = '0'

    # Check if the desired column name exists
    if desired_column_name5 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name5)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name5,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name5].any():
            print(f"Values are present in '{desired_column_name5}' column:")
            print(df_AL1[desired_column_name5])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name5] != "$", new_column_name5] = 0
        else:
            print(f"No values are present in '{desired_column_name5}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name5}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace5}' does not exist in the DataFrame.")

column_to_replace5 = 'PTN_5'

# Check if the column exists in the DataFrame
if column_to_replace5 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace5] = df_AL1[column_to_replace5].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace5}' does not exist in the DataFrame.")

#PTN_6>>

desired_column_name6 = 'PTN_6'
column_to_replace6 = 'PTN_6'

# Check if the column exists in the DataFrame
if column_to_replace6 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace6] = df_AL1[column_to_replace6].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name6 = 'P_6'
    new_column_value6 = '0'

    # Check if the desired column name exists
    if desired_column_name6 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name6)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name6,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name6].any():
            print(f"Values are present in '{desired_column_name6}' column:")
            print(df_AL1[desired_column_name6])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name6] != "$", new_column_name6] = 0
        else:
            print(f"No values are present in '{desired_column_name6}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name6}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace6}' does not exist in the DataFrame.")

column_to_replace6 = 'PTN_6'

# Check if the column exists in the DataFrame
if column_to_replace6 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace6] = df_AL1[column_to_replace6].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace6}' does not exist in the DataFrame.")

#PTN_7>>

desired_column_name7 = 'PTN_7'
column_to_replace7 = 'PTN_7'

# Check if the column exists in the DataFrame
if column_to_replace7 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace7] = df_AL1[column_to_replace7].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name7 = 'P_7'
    new_column_value7 = '0'

    # Check if the desired column name exists
    if desired_column_name7 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name7)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name7,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name7].any():
            print(f"Values are present in '{desired_column_name7}' column:")
            print(df_AL1[desired_column_name7])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name7] != "$", new_column_name7] = 0
        else:
            print(f"No values are present in '{desired_column_name7}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name7}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace7}' does not exist in the DataFrame.")

column_to_replace7 = 'PTN_7'

# Check if the column exists in the DataFrame
if column_to_replace7 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace7] = df_AL1[column_to_replace7].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace7}' does not exist in the DataFrame.")

#PTN_8>>

desired_column_name8 = 'PTN_8'
column_to_replace8 = 'PTN_8'

# Check if the column exists in the DataFrame
if column_to_replace8 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace8] = df_AL1[column_to_replace8].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name8 = 'P_8'
    new_column_value8 = '0'

    # Check if the desired column name exists
    if desired_column_name8 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name8)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name8,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name8].any():
            print(f"Values are present in '{desired_column_name8}' column:")
            print(df_AL1[desired_column_name8])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name8] != "$", new_column_name8] = 0
        else:
            print(f"No values are present in '{desired_column_name8}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name8}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace8}' does not exist in the DataFrame.")

column_to_replace8 = 'PTN_8'

# Check if the column exists in the DataFrame
if column_to_replace8 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace8] = df_AL1[column_to_replace8].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace8}' does not exist in the DataFrame.")

#PTN_9>>

desired_column_name9 = 'PTN_9'
column_to_replace9 = 'PTN_9'

# Check if the column exists in the DataFrame
if column_to_replace9 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace9] = df_AL1[column_to_replace9].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name9 = 'P_9'
    new_column_value9 = '0'

    # Check if the desired column name exists
    if desired_column_name9 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name9)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name9,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name9].any():
            print(f"Values are present in '{desired_column_name9}' column:")
            print(df_AL1[desired_column_name9])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name9] != "$", new_column_name9] = 0
        else:
            print(f"No values are present in '{desired_column_name9}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name9}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace9}' does not exist in the DataFrame.")

column_to_replace9 = 'PTN_9'

# Check if the column exists in the DataFrame
if column_to_replace9 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace9] = df_AL1[column_to_replace9].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace9}' does not exist in the DataFrame.")

#PTN_10>>

desired_column_name10 = 'PTN_10'
column_to_replace10 = 'PTN_10'

# Check if the column exists in the DataFrame
if column_to_replace10 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace10] = df_AL1[column_to_replace10].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name10 = 'P_10'
    new_column_value10 = '0'

    # Check if the desired column name exists
    if desired_column_name10 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name10)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name10,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name10].any():
            print(f"Values are present in '{desired_column_name10}' column:")
            print(df_AL1[desired_column_name10])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name10] != "$", new_column_name10] = 0
        else:
            print(f"No values are present in '{desired_column_name10}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name10}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace10}' does not exist in the DataFrame.")

column_to_replace10 = 'PTN_10'

# Check if the column exists in the DataFrame
if column_to_replace10 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace10] = df_AL1[column_to_replace10].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace10}' does not exist in the DataFrame.")

#PTN_11>>

desired_column_name11 = 'PTN_11'
column_to_replace11 = 'PTN_11'

# Check if the column exists in the DataFrame
if column_to_replace11 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace11] = df_AL1[column_to_replace11].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name11 = 'P_11'
    new_column_value11 = '0'

    # Check if the desired column name exists
    if desired_column_name11 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name11)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name11,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name11].any():
            print(f"Values are present in '{desired_column_name11}' column:")
            print(df_AL1[desired_column_name11])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name11] != "$", new_column_name11] = 0
        else:
            print(f"No values are present in '{desired_column_name11}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name11}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace11}' does not exist in the DataFrame.")

column_to_replace11 = 'PTN_11'

# Check if the column exists in the DataFrame
if column_to_replace11 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace11] = df_AL1[column_to_replace11].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace11}' does not exist in the DataFrame.")

#PTN_12>>

desired_column_name12 = 'PTN_12'
column_to_replace12 = 'PTN_12'

# Check if the column exists in the DataFrame
if column_to_replace12 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace12] = df_AL1[column_to_replace12].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name12 = 'P_12'
    new_column_value12 = '0'

    # Check if the desired column name exists
    if desired_column_name12 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name12)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name12,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name12].any():
            print(f"Values are present in '{desired_column_name12}' column:")
            print(df_AL1[desired_column_name12])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name12] != "$", new_column_name12] = 0
        else:
            print(f"No values are present in '{desired_column_name12}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name12}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace12}' does not exist in the DataFrame.")

column_to_replace12 = 'PTN_12'

# Check if the column exists in the DataFrame
if column_to_replace12 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace12] = df_AL1[column_to_replace12].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace12}' does not exist in the DataFrame.")

#PTN_13>>

desired_column_name13 = 'PTN_13'
column_to_replace13 = 'PTN_13'

# Check if the column exists in the DataFrame
if column_to_replace13 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace13] = df_AL1[column_to_replace13].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name13 = 'P_13'
    new_column_value13 = '0'

    # Check if the desired column name exists
    if desired_column_name13 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name13)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name13,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name13].any():
            print(f"Values are present in '{desired_column_name13}' column:")
            print(df_AL1[desired_column_name13])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name13] != "$", new_column_name13] = 0
        else:
            print(f"No values are present in '{desired_column_name13}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name13}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace13}' does not exist in the DataFrame.")

column_to_replace13 = 'PTN_13'

# Check if the column exists in the DataFrame
if column_to_replace13 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace13] = df_AL1[column_to_replace13].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace13}' does not exist in the DataFrame.")

#PTN_14>>

desired_column_name14 = 'PTN_14'
column_to_replace14 = 'PTN_14'

# Check if the column exists in the DataFrame
if column_to_replace14 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace14] = df_AL1[column_to_replace14].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name14 = 'P_14'
    new_column_value14 = '0'

    # Check if the desired column name exists
    if desired_column_name14 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name14)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name14,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name14].any():
            print(f"Values are present in '{desired_column_name14}' column:")
            print(df_AL1[desired_column_name14])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name14] != "$", new_column_name14] = 0
        else:
            print(f"No values are present in '{desired_column_name14}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name14}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace14}' does not exist in the DataFrame.")

column_to_replace14 = 'PTN_14'

# Check if the column exists in the DataFrame
if column_to_replace14 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace14] = df_AL1[column_to_replace14].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace14}' does not exist in the DataFrame.")

#PTN_15>>

desired_column_name15 = 'PTN_15'
column_to_replace15 = 'PTN_15'

# Check if the column exists in the DataFrame
if column_to_replace15 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace15] = df_AL1[column_to_replace15].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name15 = 'P_15'
    new_column_value15 = '0'

    # Check if the desired column name exists
    if desired_column_name15 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name15)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name15,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name15].any():
            print(f"Values are present in '{desired_column_name15}' column:")
            print(df_AL1[desired_column_name15])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name15] != "$", new_column_name15] = 0
        else:
            print(f"No values are present in '{desired_column_name15}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name15}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace15}' does not exist in the DataFrame.")

column_to_replace15 = 'PTN_15'

# Check if the column exists in the DataFrame
if column_to_replace15 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace15] = df_AL1[column_to_replace15].replace("$", np.NaN)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace15}' does not exist in the DataFrame.")

print(df_AL1.head(5))

del df_AL1['Group']

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
    df_AL1.to_excel(writer, sheet_name="AVL_SHEET", index=False)

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

# Example DataFrame
data = pd.read_excel('AVL.xlsx', sheet_name='AVL_SHEET')
df_AL1 = pd.DataFrame(data)

# Desired column order
desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10']
#desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10','PTN_11','P_11','PTN_12','P_12','PTN_13','P_13','PTN_14','P_14','PTN_15','P_15']

# Create a list of columns present in both DataFrame and desired_order
common_columns = [col for col in desired_order if col in df_AL1.columns]

# Reorder the DataFrame based on the desired_order
df_AL1 = df_AL1[common_columns]

# Display the reordered DataFrame
print(df_AL1)

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
    df_AL1.to_excel(writer, sheet_name="AVL_SHEET", index=False)
    df_AL1.T.reset_index().T.to_excel(writer, sheet_name="AVL_SHEET", header=False ,index=False)
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')

#read_file = pd.read_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.xlsx',skiprows=0)

read_file = pd.read_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.xlsx')

read_file.to_csv (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.txt', index = None, header= None)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()

# Replace '0.0' with '0' in the content
modified_content = content.replace('0.0', '0')

# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(modified_content)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC20 = content.replace(',,,,,,,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC20)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC19 = content.replace(',,,,,,,,,,,,,,,,,,,', '')
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC19)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC18 = content.replace(',,,,,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC18)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC17 = content.replace(',,,,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC17)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC16 = content.replace(',,,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC16)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC15 = content.replace(',,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC15)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC14 = content.replace(',,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC14)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC13 = content.replace(',,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC13)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC12 = content.replace(',,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC12)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC11 = content.replace(',,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC11)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC10 = content.replace(',,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC10)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC9 = content.replace(',,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC9)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC8 = content.replace(',,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC8)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC7 = content.replace(',,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC7)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC6 = content.replace(',,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC6)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC5 = content.replace(',,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC5)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC4 = content.replace(',,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC4)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC3 = content.replace(',,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC3)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC2 = content.replace(',,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC2)
    #output_file.write('MC2','MC3','MC4','MC5','MC6','MC7','MC8','MC9','MC10','MC11','MC12','MC13','MC14','MC15','MC16','MC17','MC18','MC19','MC20')

'''# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC1 = content.replace('+', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC1)'''

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC0 = content.replace('.0', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC0)

print(f"AVL CREATED: D:/NX_BACKWORK/r'AVL.txt")

# Specify the path to your text file
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
df = pd.read_table(txt_file_path, delimiter='\t', quoting=3)  # 3 corresponds to QUOTE_NONE
csv_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.csv'
df.to_csv(csv_file_path, index=False, sep='\t')  # 0 corresponds to QUOTE_NONE

#read_file = pd.read_table (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.txt', sep='"')
#read_file.to_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVList.xlsx', index=None)
#read_file.to_csv (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.csv', index = None, header= None)

#Feeder List change

shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_FeederSetup/Line X Sample.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx')
shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_FeederSetup/Line X Sample.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx')

##BOT FEEDER LIST

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

# Source Excel file
source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
source_sheet_name = 'NXT&AMX1_B'

# Destination Excel file
destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
destination_sheet_name = 'NXT'

def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet_name]

    # Load the destination workbook
    destination_workbook = openpyxl.load_workbook(destination_file)
    destination_sheet = destination_workbook[destination_sheet_name]

    # Iterate through the source sheet and copy data to the destination sheet with an offset
    for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        # Offset the row index by the specified offset
        destination_row = row_index + offset

        # Copy data to the destination sheet
        for col_index, value in enumerate(row, start=1):
            destination_sheet.cell(row=destination_row, column=col_index, value=value)

    # Save the changes to the destination workbook
    destination_workbook.save(destination_file)

# Example usage:
copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "NXT&AMX1_B", "NXT", offset=5)
# Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

# Source Excel file
source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
source_sheet_name = 'AMX2_B'

# Destination Excel file
destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
destination_sheet_name = 'AIMEX 2'

def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet_name]

    # Load the destination workbook
    destination_workbook = openpyxl.load_workbook(destination_file)
    destination_sheet = destination_workbook[destination_sheet_name]

    # Iterate through the source sheet and copy data to the destination sheet with an offset
    for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        # Offset the row index by the specified offset
        destination_row = row_index + offset

        # Copy data to the destination sheet
        for col_index, value in enumerate(row, start=1):
            destination_sheet.cell(row=destination_row, column=col_index, value=value)

    # Save the changes to the destination workbook
    destination_workbook.save(destination_file)

# Example usage:
copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "AMX2_B", "AIMEX 2", offset=5)
# Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

# Source Excel file
source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
source_sheet_name = 'AMX3_B'

# Destination Excel file 
destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
destination_sheet_name = 'AIMEX 3'

def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet_name]

    # Load the destination workbook
    destination_workbook = openpyxl.load_workbook(destination_file)
    destination_sheet = destination_workbook[destination_sheet_name]

    # Iterate through the source sheet and copy data to the destination sheet with an offset
    for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        # Offset the row index by the specified offset
        destination_row = row_index + offset

        # Copy data to the destination sheet
        for col_index, value in enumerate(row, start=1):
            destination_sheet.cell(row=destination_row, column=col_index, value=value)

    # Save the changes to the destination workbook
    destination_workbook.save(destination_file)

# Example usage:
copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "AMX3_B", "AIMEX 3", offset=5)
# Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

##TOP FEEDER LIST

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

# Source Excel file
source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
source_sheet_name = 'NXT&AMX1_T'

# Destination Excel file
destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
destination_sheet_name = 'NXT'

def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet_name]

    # Load the destination workbook
    destination_workbook = openpyxl.load_workbook(destination_file)
    destination_sheet = destination_workbook[destination_sheet_name]

    # Iterate through the source sheet and copy data to the destination sheet with an offset
    for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        # Offset the row index by the specified offset
        destination_row = row_index + offset

        # Copy data to the destination sheet
        for col_index, value in enumerate(row, start=1):
            destination_sheet.cell(row=destination_row, column=col_index, value=value)

    # Save the changes to the destination workbook
    destination_workbook.save(destination_file)

# Example usage:
copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "NXT&AMX1_T", "NXT", offset=5)
# Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

# Source Excel file
source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
source_sheet_name = 'AMX2_T'

# Destination Excel file
destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
destination_sheet_name = 'AIMEX 2'

def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet_name]

    # Load the destination workbook
    destination_workbook = openpyxl.load_workbook(destination_file)
    destination_sheet = destination_workbook[destination_sheet_name]

    # Iterate through the source sheet and copy data to the destination sheet with an offset
    for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        # Offset the row index by the specified offset
        destination_row = row_index + offset

        # Copy data to the destination sheet
        for col_index, value in enumerate(row, start=1):
            destination_sheet.cell(row=destination_row, column=col_index, value=value)

    # Save the changes to the destination workbook
    destination_workbook.save(destination_file)

# Example usage:
copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "AMX2_T", "AIMEX 2", offset=5)
# Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

# Source Excel file
source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
source_sheet_name = 'AMX3_T'

# Destination Excel file
destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
destination_sheet_name = 'AIMEX 3'

def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet_name]

    # Load the destination workbook
    destination_workbook = openpyxl.load_workbook(destination_file)
    destination_sheet = destination_workbook[destination_sheet_name]

    # Iterate through the source sheet and copy data to the destination sheet with an offset
    for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        # Offset the row index by the specified offset
        destination_row = row_index + offset

        # Copy data to the destination sheet
        for col_index, value in enumerate(row, start=1):
            destination_sheet.cell(row=destination_row, column=col_index, value=value)

    # Save the changes to the destination workbook
    destination_workbook.save(destination_file)

# Example usage:
copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "AMX3_T", "AIMEX 3", offset=5)
# Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#CREATEBACKUPFOLDER

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)

yourfolder1 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\BOM"

if not os.path.isdir(yourfolder1):
    print('Folder Not Exist')
    os.makedirs(yourfolder1)

yourfolder2 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\FeederSetup"

if not os.path.isdir(yourfolder2):
    print('Folder Not Exist')
    os.makedirs(yourfolder2)

yourfolder3 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\Upload"

if not os.path.isdir(yourfolder3):
    print('Folder Not Exist')
    os.makedirs(yourfolder3)

yourfolder4 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\Verified"

if not os.path.isdir(yourfolder4):
    print('Folder Not Exist')
    os.makedirs(yourfolder4)

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
Chd = os.getcwd()

file_path = 'BOM_List_OP.xlsx'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    if os.path.exists("BOM_List_OP.xlsx"):
        os.remove("BOM_List_OP.xlsx")
else:
    print("The file does not exist")

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

if os.path.exists("Feeder_List_OPT.xlsx"):
  os.remove("Feeder_List_OPT.xlsx")
else:
  print("The file does not exist")

if os.path.exists("Feeder_List_OPB.xlsx"):
  os.remove("Feeder_List_OPB.xlsx")
else:
  print("The file does not exist")

for i in range(100):
    row = "="*i + ">"
    sys.stdout.write("%s\r %d%%\r" %(row, i + 1))
    sys.stdout.flush()
    time.sleep(0.1)

logwindow = sg.Multiline(size=(160, 40), font=('Courier', 9))
print = logwindow.print
layout = [[logwindow],[sg.Button('Quit')]]

# Create the window
window = sg.Window("FeederSetup", layout, finalize=True)

print("FeederSetup___Compelete $ PROCESS $")

sys.stdout.write("\n")

#################################################################

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

xls=pd.ExcelFile('FeederVerify.xlsx',engine='openpyxl')
df1 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM_data")
dfs1 = pd.read_excel("FeederVerify.xlsx", sheet_name="Feeder_data")
df2 = pd.read_excel("FeederVerify.xlsx", sheet_name="Side")
df3 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM and Feeder Compare")
dfs21 = pd.read_excel('Feederverify.xlsx', sheet_name="Feeder_data", usecols=['Location','F_Part_No','FeederName','Type','Size','FeedPitch','Part Height','Status','QTY','Side','ModelName','F_Ref_List','Feeder Reference'],index_col=False)
dfs22 = pd.read_excel('Feederverify.xlsx', sheet_name="BOM_data", usecols=['B_Ref_List','B_Part_No','Long Des'],index_col=False)
dfs3 = pd.read_excel("FeederVerify.xlsx", sheet_name="Verify_data_BAF")
dbf3 = pd.read_excel("FeederVerify.xlsx", sheet_name="Verify_data_FAB")
dfsg21 = dfs21[dfs21['Feeder Reference'].duplicated() == True]
dfsg22 = dfs22[dfs22['B_Ref_List'].duplicated() == True]
dfsg31 = dfs3[dfs3['BOM and Feeder Compare'].str.contains('Miss_Match')]
dfsg32 = dbf3[dbf3['Feeder and BOM Compare'].str.contains('Miss_Match')]

rc = len(df1)
rc1 = len(dfs1)

print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print(Chd,'\\__BOM__\\',dL1)
print(Chd,'\\__FeederSetup__\\',dL2)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

ds1 = print("BOM Count:",rc)
print('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx, Sheetname=BOM_data')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

ds9 = print("Feeder Count:",rc1)
print('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederSetup.xlsx, Sheetname=Feedercol')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

ds2 = print("BOT & TOP Count:")
ds2 = print(df2)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

ds3 = print("Compare Count:")
ds3 = print(df3)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

print("Feeder duplicate Reference")
print(dfsg21)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

print("BOM duplicate Reference")
print(dfsg22)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

print("Miss Match Row BOM to Feeder")
print(dfsg31) 
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

print("Miss Match Row Feeder to BOM")
print(dfsg32) 
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

##########################################################################################################################################

# Create an event loop
while True:
    event, values = window.read(timeout=1)
    if event == sg.WIN_CLOSED or event == 'Quit':
        break

##########################################################################################################################################

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Line_X"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

#shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/Upload-Data.xlsx')
#shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/PartMaster.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/PartMaster.xlsx')
shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/MODEL.mdb')
shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/Line X Sample T.xlsx')
shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/Line X Sample B.xlsx')

##########################################################################################################################################

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

#shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/BOM_List_OP.xlsx')
#shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederSetup.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/FeederSetup.xlsx')
shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/FeederVerify.xlsx')
shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.csv', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/AVL.CSV')

##########################################################################################################################################

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X')
Chd = os.getcwd()

#src_1 = 'Upload-Data.xlsx'
#os.rename(src_1, dL2 +"_UD"+".xlsx")

#src_2 = 'PartMaster.xlsx'
#os.rename(src_2, dL2 +"_PM"+".xlsx")

src_3 = 'MODEL.mdb'
os.rename(src_3, dL2 +"_PM-Model"+".mdb")

#src_4 = 'BOM_List_OP.xlsx'
#os.rename(src_4, dL1 +"_BOM"+".xlsx")

#src_5 = 'FeederSetup.xlsx'
#os.rename(src_5, dL2 +"_FS"+".xlsx")

src_6 = 'FeederVerify.xlsx'
os.rename(src_6, dL2 +"_FV"+".xlsx")

src_7 = 'AVL.csv'
os.rename(src_7, dL1 +"_AVL"+".csv")

src_8 = 'Line X Sample T.xlsx'
os.rename(src_8, dL1 +"_T"+".xlsx")

src_9 = 'Line X Sample B.xlsx'
os.rename(src_9, dL1 +"_B"+".xlsx")
time.sleep (2)

window.close()

time.sleep (5)